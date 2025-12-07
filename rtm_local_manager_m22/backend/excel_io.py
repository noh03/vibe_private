"""
excel_io.py - Excel import/export helpers for RTM Local Manager.

- Export:
    export_project_to_excel(conn, project_id, file_path)

- Import:
    import_project_from_excel(conn, project_id, file_path)

주의사항:
- .xlsx 처리를 위해 openpyxl 라이브러리를 사용한다. (사용 PC에 설치되어 있어야 함)
- 이 모듈은 main import 시점에는 openpyxl 을 import 하지 않는다.
  (함수 내부에서 import 하므로, openpyxl 미설치 환경에서도 프로그램 실행 자체는 가능)
"""

from __future__ import annotations

from typing import Any, Dict, List

import sqlite3


def _ensure_openpyxl():
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover - runtime 오류 표시용
        raise RuntimeError(
            "openpyxl 이 설치되어 있지 않아 Excel 기능을 사용할 수 없습니다.\n"
            "pip install openpyxl 로 설치 후 다시 시도해 주세요."
        ) from e
    return openpyxl


# --------------------------------------------------------------------------- Export helpers


def export_project_to_excel(conn: sqlite3.Connection, project_id: int, file_path: str) -> None:
    """
    현재 project_id 에 속한 주요 엔티티들을 하나의 Excel 워크북(.xlsx)으로 내보낸다.

    시트 구성:
      - Issues
      - TestcaseSteps
      - Relations
      - TestPlanTestcases
      - TestExecutions
      - TestcaseExecutions

    각 시트는 최소한의 식별자(jira_key, summary 등)를 포함하여,
    나중에 Import 시 jira_key 를 기준으로 매핑할 수 있도록 설계한다.
    """
    openpyxl = _ensure_openpyxl()
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # default sheet 제거
    default_sheet = wb.active
    wb.remove(default_sheet)

    cur = conn.cursor()

    # --- Issues 시트
    ws = wb.create_sheet("Issues")
    # 컬럼 정의 (DB 스키마의 주요 필드를 대부분 노출하여 엑셀만으로 이슈 메타를 관리할 수 있도록 확장)
    issue_cols = [
        "id",
        "jira_key",
        "issue_type",
        "folder_path",  # folders 트리 구조를 문자열 경로로 표현
        "summary",
        "description",
        "status",
        "priority",
        "assignee",
        "reporter",
        "labels",
        "components",
        "security_level",
        "fix_versions",
        "affects_versions",
        "rtm_environment",
        "due_date",
        "created",
        "updated",
        "attachments",
        "epic_link",
        "sprint",
        # 엑셀 전용 참조 키 (DB에는 저장되지 않지만, 시트 간 매핑용으로 사용)
        "excel_key",
    ]
    ws.append(issue_cols)
    cur.execute(
        """
        SELECT id,
               jira_key,
               issue_type,
               folder_id,
               summary,
               description,
               status,
               priority,
               assignee,
               reporter,
               labels,
               components,
               security_level,
               fix_versions,
               affects_versions,
               rtm_environment,
               due_date,
               created,
               updated,
               attachments,
               epic_link,
               sprint
          FROM issues
         WHERE project_id = ? AND is_deleted = 0
         ORDER BY id
        """,
        (project_id,),
    )
    from backend.db import get_folder_path
    for row in cur.fetchall():
        row = list(row)
        folder_id = row[3]
        row[3] = get_folder_path(conn, folder_id)
        # excel_key 는 DB 에 없으므로, export 시에는 빈 값으로 채워 템플릿만 제공
        row.append("")
        ws.append(row)

    # --- TestcaseSteps 시트
    ws = wb.create_sheet("TestcaseSteps")
    # SW 사양 기준으로 Test Case Steps 구조 확장:
    # - Preconditions: Test Case 단위 필드
    # - group_no: Step 그룹 번호 (각 그룹 내에서 order_no 가 1부터 시작)
    #
    # sheet 간 명시적 매핑을 위해:
    #   - issue_id     : 로컬 DB issues.id (Issues 시트의 id 와 1:1 대응)
    #   - issue_jira_key: 해당 Test Case 의 Jira Key (있는 경우)
    step_cols = [
        "issue_id",
        "issue_jira_key",
        "preconditions",
        "group_no",
        "order_no",
        "action",
        "input",
        "expected",
        # 엑셀 전용 참조 키 (Issues.excel_key 와 매핑)
        "excel_key",
    ]
    ws.append(step_cols)
    cur.execute(
        """
        SELECT t.issue_id,
               i.jira_key,
               i.preconditions,
               COALESCE(t.group_no, 1) AS group_no,
               t.order_no,
               t.action,
               t.input,
               t.expected
          FROM testcase_steps t
          JOIN issues i ON t.issue_id = i.id
         WHERE i.project_id = ? AND i.is_deleted = 0
         ORDER BY t.issue_id, group_no, order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        # DB 에는 excel_key 가 없으므로, export 시에는 빈 값으로 확장
        row = list(row)
        row.append("")
        ws.append(row)

    # --- Relations 시트
    ws = wb.create_sheet("Relations")
    rel_cols = ["src_jira_key", "dst_jira_key", "relation_type"]
    ws.append(rel_cols)
    cur.execute(
        """
        SELECT s.jira_key AS src_key,
               d.jira_key AS dst_key,
               r.relation_type
          FROM relations r
          JOIN issues s ON r.src_issue_id = s.id
          JOIN issues d ON r.dst_issue_id = d.id
         WHERE s.project_id = ? AND s.is_deleted = 0 AND d.is_deleted = 0
         ORDER BY src_key, dst_key
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        ws.append(list(row))

    # --- TestPlanTestcases 시트
    ws = wb.create_sheet("TestPlanTestcases")
    tp_cols = ["testplan_jira_key", "testcase_jira_key", "order_no"]
    ws.append(tp_cols)
    cur.execute(
        """
        SELECT tp.jira_key AS testplan_key,
               tc.jira_key AS testcase_key,
               t.order_no
          FROM testplan_testcases t
          JOIN issues tp ON t.testplan_id = tp.id
          JOIN issues tc ON t.testcase_id = tc.id
         WHERE tp.project_id = ? AND tp.is_deleted = 0 AND tc.is_deleted = 0
         ORDER BY testplan_key, t.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        ws.append(list(row))

    # --- TestExecutions 시트
    ws = wb.create_sheet("TestExecutions")
    te_cols = ["testexecution_jira_key", "environment", "start_date", "end_date", "result", "executed_by"]
    ws.append(te_cols)
    cur.execute(
        """
        SELECT i.jira_key,
               t.environment,
               t.start_date,
               t.end_date,
               t.result,
               t.executed_by
          FROM testexecutions t
          JOIN issues i ON t.issue_id = i.id
         WHERE i.project_id = ? AND i.is_deleted = 0
         ORDER BY i.jira_key
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        ws.append(list(row))

    # --- TestcaseExecutions 시트
    ws = wb.create_sheet("TestcaseExecutions")
    tce_cols = [
        "testexecution_jira_key",
        "testcase_jira_key",
        "order_no",
        "assignee",
        "result",
        "actual_time",
        "rtm_environment",
        "defects",
        "tce_test_key",
    ]
    ws.append(tce_cols)
    cur.execute(
        """
        SELECT tei.jira_key AS te_key,
               tci.jira_key AS tc_key,
               t.order_no,
               t.assignee,
               t.result,
               t.actual_time,
               t.rtm_environment,
               t.defects,
               t.tce_test_key
          FROM testcase_executions t
          JOIN testexecutions te ON t.testexecution_id = te.id
          JOIN issues tei ON te.issue_id = tei.id
          JOIN issues tci ON t.testcase_id = tci.id
         WHERE tei.project_id = ? AND tei.is_deleted = 0 AND tci.is_deleted = 0
         ORDER BY te_key, t.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        ws.append(list(row))

    # --- TestcaseStepExecutions 시트 (옵션)
    ws = wb.create_sheet("TestcaseStepExecutions")
    tse_cols = [
        "testexecution_jira_key",
        "testcase_jira_key",
        "group_no",
        "order_no",
        "status",
        "actual_result",
        "evidence",
    ]
    ws.append(tse_cols)
    cur.execute(
        """
        SELECT tei.jira_key    AS te_key,
               tci.jira_key    AS tc_key,
               ts.group_no     AS group_no,
               ts.order_no     AS order_no,
               tse.status      AS status,
               tse.actual_result AS actual_result,
               tse.evidence    AS evidence
          FROM testcase_step_executions tse
          JOIN testcase_executions tce ON tse.testcase_execution_id = tce.id
          JOIN testexecutions te ON tce.testexecution_id = te.id
          JOIN issues tei ON te.issue_id = tei.id
          JOIN issues tci ON tce.testcase_id = tci.id
          JOIN testcase_steps ts ON tse.testcase_step_id = ts.id
         WHERE tei.project_id = ? AND tei.is_deleted = 0 AND tci.is_deleted = 0
         ORDER BY te_key, tc_key, ts.group_no, ts.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        ws.append(list(row))

    # 약간의 자동 너비 조정 (간단히)
    for ws in wb.worksheets:
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                except Exception:
                    val = ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    wb.save(file_path)


# --------------------------------------------------------------------------- Import helpers


def import_project_from_excel(conn: sqlite3.Connection, project_id: int, file_path: str) -> None:
    """
    Excel 파일(.xlsx)에서 주요 엔티티를 읽어와 로컬 DB에 반영한다.

    처리 범위:
      - Issues 시트:
          - jira_key 기준으로 존재 여부 확인
          - 있으면 주요 필드 update, 없으면 신규 issue insert (folder_id 는 None)
      - TestcaseSteps 시트:
          - issue_jira_key 기준으로 대상 issue 찾은 뒤, 해당 issue 의 steps 전체 교체
      - Relations 시트:
          - src_jira_key / dst_jira_key 기준으로 양쪽 issue 찾고 relations 교체
      - TestPlanTestcases 시트:
          - testplan_jira_key / testcase_jira_key 기준으로 testplan_testcases 교체
      - TestExecutions 시트:
          - testexecution_jira_key 기준으로 testexecution 메타 정보 update/신규 생성
      - TestcaseExecutions 시트:
          - testexecution_jira_key / testcase_jira_key 기준으로 testcase_executions 교체

    Excel 컬럼 헤더는 export_project_to_excel()에서 생성한 것과 동일해야 한다.
    """
    openpyxl = _ensure_openpyxl()
    from backend.db import (
        get_issue_by_jira_key,
        get_issue_by_id,
        update_issue_fields,
        replace_steps_for_issue,
        replace_relations_for_issue,
        replace_testplan_testcases,
        get_or_create_testexecution_for_issue,
        update_testexecution_for_issue,
        replace_testcase_executions,
        get_steps_for_issue,
        replace_step_executions_for_tce,
        create_local_issue,
        ensure_folder_path,
    )

    wb = openpyxl.load_workbook(file_path, data_only=True)

    cur = conn.cursor()

    # Issues 시트와 다른 시트(TestcaseSteps 등) 사이를 연결하기 위한
    # 엑셀 전용 레퍼런스 키 매핑 (예: excel_key -> issue_id)
    # - DB에는 저장되지 않고, "한 번에 신규 TC + Steps 생성"을 위한 임시 매핑용이다.
    excel_issue_ref: Dict[str, int] = {}

    # --- Issues 시트 처리
    if "Issues" in wb.sheetnames:
        ws = wb["Issues"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            for row in rows[1:]:
                if not any(row):
                    continue

                # 0) id 컬럼이 있고 값이 있으면, 로컬 ID 기준 "수정"으로 처리
                local_id = None
                id_idx = col_idx.get("id")
                if id_idx is not None and 0 <= id_idx < len(row):
                    raw_id = row[id_idx]
                    if raw_id not in (None, ""):
                        try:
                            # 엑셀에서 숫자가 float 로 넘어오는 경우까지 고려
                            local_id = int(str(raw_id).split(".")[0])
                        except Exception:
                            local_id = None

                # 엑셀 전용 레퍼런스 키 (optional)
                excel_ref: str | None = None
                excel_ref_idx = col_idx.get("excel_key")
                if excel_ref_idx is not None and 0 <= excel_ref_idx < len(row):
                    raw_ref = row[excel_ref_idx]
                    if raw_ref not in (None, ""):
                        excel_ref = str(raw_ref).strip()

                # 공통 필드 추출 (id/jira_key 를 제외한 나머지)
                # jira_key 컬럼이 있으면 JIRA 이슈 기준으로 upsert,
                # jira_key 가 비어 있으면 "로컬 전용 이슈" 로 신규 생성한다.
                jira_key_idx = col_idx.get("jira_key")
                jira_key: str | None
                if jira_key_idx is None or jira_key_idx < 0 or jira_key_idx >= len(row):
                    jira_key = None
                else:
                    raw_key = row[jira_key_idx]
                    jira_key = str(raw_key) if raw_key not in (None, "") else None

                fields: Dict[str, Any] = {}
                for name in [
                    "issue_type",
                    "summary",
                    "description",
                    "status",
                    "priority",
                    "assignee",
                    "reporter",
                    "labels",
                    "components",
                    "security_level",
                    "fix_versions",
                    "affects_versions",
                    "rtm_environment",
                    "due_date",
                    "created",
                    "updated",
                    "attachments",
                    "epic_link",
                    "sprint",
                ]:
                    idx = col_idx.get(name)
                    if idx is None or idx < 0 or idx >= len(row):
                        continue
                    val = row[idx]
                    if val is not None:
                        fields[name] = str(val)

                # 폴더 경로 (optional): folder_path 컬럼이 있으면 이를 기반으로 folder_id 를 계산
                folder_id = None
                folder_idx = col_idx.get("folder_path")
                if folder_idx is not None and 0 <= folder_idx < len(row):
                    raw_path = row[folder_idx]
                    if raw_path not in (None, ""):
                        issue_type_for_folder = fields.get("issue_type")
                        folder_id = ensure_folder_path(
                            conn,
                            project_id=project_id,
                            path=str(raw_path),
                            issue_type=issue_type_for_folder,
                        )
                        if folder_id:
                            fields["folder_id"] = folder_id

                # 0-1) 이 행에서 실제로 업데이트/생성한 이슈의 id (엑셀 ref 매핑에 사용)
                target_issue_id: int | None = None

                # 0-2) local_id 로 업데이트 가능한 경우: 해당 행은 "수정"으로만 처리
                if local_id is not None:
                    issue = get_issue_by_id(conn, local_id)
                    if issue and issue.get("project_id") == project_id:
                        update_issue_fields(conn, local_id, fields)
                        target_issue_id = local_id
                        # 아래 jira_key / 신규 생성 로직은 건너뛴다.
                        pass
                    else:
                        # ID 가 없거나 다른 프로젝트인 경우에는 아래 jira_key / 신규 생성 로직으로 넘어간다.
                        target_issue_id = None

                # local_id 기반 업데이트가 안 된 경우에만 jira_key / 신규 생성 로직 수행
                if target_issue_id is None:
                    if jira_key:
                        # JIRA Key 가 있는 경우: 기존 이슈 update 또는 신규 JIRA-연동 이슈 insert
                        issue = get_issue_by_jira_key(conn, project_id, jira_key)
                        if issue:
                            update_issue_fields(conn, issue["id"], fields)
                            target_issue_id = int(issue["id"])
                        else:
                            cur.execute(
                                """
                                INSERT INTO issues (
                                    project_id, folder_id, jira_key, issue_type, summary,
                                    status, priority, assignee, reporter, labels, components,
                                    security_level, fix_versions, affects_versions,
                                    rtm_environment, due_date, created, updated, is_deleted
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'), 0)
                                """,
                                (
                                    project_id,
                                    folder_id,
                                    jira_key,
                                    fields.get("issue_type") or "",
                                    fields.get("summary") or "",
                                    fields.get("status") or "",
                                    fields.get("priority") or "",
                                    fields.get("assignee") or "",
                                    fields.get("reporter") or "",
                                    fields.get("labels") or "",
                                    fields.get("components") or "",
                                    fields.get("security_level") or "",
                                    fields.get("fix_versions") or "",
                                    fields.get("affects_versions") or "",
                                    fields.get("rtm_environment") or "",
                                    fields.get("due_date") or "",
                                ),
                            )
                            conn.commit()
                            try:
                                target_issue_id = int(cur.lastrowid)
                            except Exception:
                                # lastrowid 를 얻지 못하는 경우, jira_key 로 재조회
                                issue = get_issue_by_jira_key(conn, project_id, jira_key)
                                if issue:
                                    target_issue_id = int(issue["id"])
                    else:
                        # jira_key が 비어 있는 경우:
                        #   - 엑셀에 local ID 가 없어도, issue_type / summary 정보만 있으면
                        #     로컬 전용 이슈를 자동으로 생성한다.
                        issue_type = fields.get("issue_type")
                        summary = fields.get("summary", "")
                        if not issue_type:
                            # issue_type 이 없는 행은 스킵 (스키마상 NOT NULL)
                            continue

                        new_issue_id = create_local_issue(
                            conn,
                            project_id=project_id,
                            issue_type=issue_type,
                            folder_id=folder_id,
                            summary=summary,
                        )
                        # create_local_issue 로 기본 값은 들어갔으므로,
                        # 나머지 필드가 있다면 추가로 업데이트한다.
                        # (issue_type / summary 는 동일 값으로 한 번 더 세팅해도 무방)
                        update_issue_fields(conn, new_issue_id, fields)
                        target_issue_id = new_issue_id

                # 이 행에 excel_key 가 있다면, 해당 키로 issue_id 를 매핑해 둔다.
                if excel_ref and target_issue_id is not None:
                    excel_issue_ref[excel_ref] = target_issue_id

    # --- TestcaseSteps 시트 처리
    if "TestcaseSteps" in wb.sheetnames:
        ws = wb["TestcaseSteps"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            # sheet 간 매핑을 위해 issue_id 를 우선 사용, 없으면 issue_jira_key 로 fallback
            steps_by_issue_id: Dict[int, List[Dict[str, Any]]] = {}
            preconditions_by_issue_id: Dict[int, str] = {}
            for row in rows[1:]:
                if not any(row):
                    continue

                issue = None
                issue_id_val: int | None = None

                # 1) issue_id 로 우선 매핑 (Issues 시트의 id 와 동일)
                id_idx = col_idx.get("issue_id")
                if id_idx is not None and 0 <= id_idx < len(row):
                    raw_id = row[id_idx]
                    if raw_id not in (None, ""):
                        try:
                            issue_id_val = int(str(raw_id).split(".")[0])
                        except Exception:
                            issue_id_val = None
                        if issue_id_val is not None:
                            issue = get_issue_by_id(conn, issue_id_val)
                            if issue and issue.get("project_id") != project_id:
                                issue = None

                # 2) excel_key 로 매핑 (Issues 시트에서 excel_issue_ref 로 기록된 경우)
                if issue is None:
                    excel_ref_idx = col_idx.get("excel_key")
                    if excel_ref_idx is not None and 0 <= excel_ref_idx < len(row):
                        raw_ref = row[excel_ref_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                issue = get_issue_by_id(conn, mapped_id)
                                if issue and issue.get("project_id") != project_id:
                                    issue = None

                # 3) fallback: issue_jira_key 로 매핑 (기존 파일 호환)
                if issue is None:
                    key_idx = col_idx.get("issue_jira_key")
                    if key_idx is not None and 0 <= key_idx < len(row):
                        jira_key = row[key_idx]
                        if jira_key:
                            jira_key = str(jira_key)
                            issue = get_issue_by_jira_key(conn, project_id, jira_key)

                if not issue:
                    continue

                issue_id = int(issue["id"])

                order_no = None
                action = None
                inp = None
                expected = None
                group_no = None

                # Preconditions (Test Case 단위, 마지막 값이 우선)
                idx_pre = col_idx.get("preconditions")
                if idx_pre is not None and 0 <= idx_pre < len(row):
                    pre_val = row[idx_pre]
                    if pre_val not in (None, ""):
                        preconditions_by_issue_id[issue_id] = str(pre_val)

                idx_group = col_idx.get("group_no")
                if idx_group is not None and 0 <= idx_group < len(row):
                    group_no = row[idx_group]

                idx_order = col_idx.get("order_no")
                if idx_order is not None and 0 <= idx_order < len(row):
                    order_no = row[idx_order]
                idx_action = col_idx.get("action")
                if idx_action is not None and 0 <= idx_action < len(row):
                    action = row[idx_action]
                idx_input = col_idx.get("input")
                if idx_input is not None and 0 <= idx_input < len(row):
                    inp = row[idx_input]
                idx_expected = col_idx.get("expected")
                if idx_expected is not None and 0 <= idx_expected < len(row):
                    expected = row[idx_expected]

                steps_by_issue_id.setdefault(issue_id, []).append(
                    {
                        "group_no": int(group_no) if group_no is not None else 1,
                        "order_no": int(order_no) if order_no is not None else 0,
                        "action": str(action) if action is not None else "",
                        "input": str(inp) if inp is not None else "",
                        "expected": str(expected) if expected is not None else "",
                    }
                )

            for issue_id, steps in steps_by_issue_id.items():
                replace_steps_for_issue(conn, issue_id, steps)
                pre = preconditions_by_issue_id.get(issue_id)
                if pre is not None:
                    update_issue_fields(conn, issue_id, {"preconditions": pre})

    # --- Relations 시트 처리
    if "Relations" in wb.sheetnames:
        ws = wb["Relations"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            rels_by_src: Dict[str, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue
                src_idx = col_idx.get("src_jira_key")
                dst_idx = col_idx.get("dst_jira_key")
                type_idx = col_idx.get("relation_type")
                if src_idx is None or dst_idx is None:
                    continue
                src_key = row[src_idx]
                dst_key = row[dst_idx]
                if not src_key or not dst_key:
                    continue
                src_key = str(src_key)
                dst_key = str(dst_key)
                rel_type = ""
                if type_idx is not None and 0 <= type_idx < len(row):
                    val = row[type_idx]
                    if val is not None:
                        rel_type = str(val)
                rels_by_src.setdefault(src_key, []).append(
                    {"dst_jira_key": dst_key, "relation_type": rel_type}
                )

            for src_key, rel_list in rels_by_src.items():
                src_issue = get_issue_by_jira_key(conn, project_id, src_key)
                if not src_issue:
                    continue
                records: List[Dict[str, Any]] = []
                for rel in rel_list:
                    dst_issue = get_issue_by_jira_key(conn, project_id, rel["dst_jira_key"])
                    if not dst_issue:
                        continue
                    records.append(
                        {
                            "dst_issue_id": dst_issue["id"],
                            "relation_type": rel.get("relation_type") or "",
                        }
                    )
                if records:
                    replace_relations_for_issue(conn, src_issue["id"], records)

    # --- TestPlanTestcases 시트 처리
    if "TestPlanTestcases" in wb.sheetnames:
        ws = wb["TestPlanTestcases"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            by_tp: Dict[str, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue
                tp_idx = col_idx.get("testplan_jira_key")
                tc_idx = col_idx.get("testcase_jira_key")
                order_idx = col_idx.get("order_no")
                if tp_idx is None or tc_idx is None:
                    continue
                tp_key = row[tp_idx]
                tc_key = row[tc_idx]
                if not tp_key or not tc_key:
                    continue
                tp_key = str(tp_key)
                tc_key = str(tc_key)
                order_no = 0
                if order_idx is not None and 0 <= order_idx < len(row):
                    val = row[order_idx]
                    if val is not None:
                        try:
                            order_no = int(val)
                        except Exception:
                            order_no = 0
                by_tp.setdefault(tp_key, []).append(
                    {"testcase_jira_key": tc_key, "order_no": order_no}
                )

            for tp_key, mappings in by_tp.items():
                tp_issue = get_issue_by_jira_key(conn, project_id, tp_key)
                if not tp_issue:
                    continue
                records: List[Dict[str, Any]] = []
                for m in mappings:
                    tc_issue = get_issue_by_jira_key(conn, project_id, m["testcase_jira_key"])
                    if not tc_issue:
                        continue
                    records.append(
                        {
                            "testplan_id": tp_issue["id"],
                            "testcase_id": tc_issue["id"],
                            "order_no": m["order_no"],
                        }
                    )
                if records:
                    replace_testplan_testcases(conn, tp_issue["id"], records)

    # --- TestExecutions 시트 처리
    if "TestExecutions" in wb.sheetnames:
        ws = wb["TestExecutions"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            for row in rows[1:]:
                if not any(row):
                    continue
                key_idx = col_idx.get("testexecution_jira_key")
                if key_idx is None or key_idx < 0 or key_idx >= len(row):
                    continue
                te_key = row[key_idx]
                if not te_key:
                    continue
                te_key = str(te_key)
                te_issue = get_issue_by_jira_key(conn, project_id, te_key)
                if not te_issue:
                    continue
                env = None
                start_date = None
                end_date = None
                result_val = None
                executed_by = None

                idx_env = col_idx.get("environment")
                if idx_env is not None and 0 <= idx_env < len(row):
                    env = row[idx_env]
                idx_start = col_idx.get("start_date")
                if idx_start is not None and 0 <= idx_start < len(row):
                    start_date = row[idx_start]
                idx_end = col_idx.get("end_date")
                if idx_end is not None and 0 <= idx_end < len(row):
                    end_date = row[idx_end]
                idx_result = col_idx.get("result")
                if idx_result is not None and 0 <= idx_result < len(row):
                    result_val = row[idx_result]
                idx_exec = col_idx.get("executed_by")
                if idx_exec is not None and 0 <= idx_exec < len(row):
                    executed_by = row[idx_exec]

                te_row = get_or_create_testexecution_for_issue(conn, te_issue["id"])
                meta = {}
                if env is not None:
                    meta["environment"] = str(env)
                if start_date is not None:
                    meta["start_date"] = str(start_date)
                if end_date is not None:
                    meta["end_date"] = str(end_date)
                if result_val is not None:
                    meta["result"] = str(result_val)
                if executed_by is not None:
                    meta["executed_by"] = str(executed_by)
                if meta:
                    update_testexecution_for_issue(conn, te_issue["id"], meta)

    # --- TestcaseExecutions 시트 처리
    if "TestcaseExecutions" in wb.sheetnames:
        ws = wb["TestcaseExecutions"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}
            by_te: Dict[str, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue
                te_idx = col_idx.get("testexecution_jira_key")
                tc_idx = col_idx.get("testcase_jira_key")
                if te_idx is None or tc_idx is None:
                    continue
                te_key = row[te_idx]
                tc_key = row[tc_idx]
                if not te_key or not tc_key:
                    continue
                te_key = str(te_key)
                tc_key = str(tc_key)

                order_no = 0
                assignee = ""
                result_val = ""
                env = ""
                time_val = None
                defects = ""
                tce_test_key = ""

                idx_order = col_idx.get("order_no")
                if idx_order is not None and 0 <= idx_order < len(row):
                    val = row[idx_order]
                    if val is not None:
                        try:
                            order_no = int(val)
                        except Exception:
                            order_no = 0
                idx_assignee = col_idx.get("assignee")
                if idx_assignee is not None and 0 <= idx_assignee < len(row):
                    val = row[idx_assignee]
                    if val is not None:
                        assignee = str(val)
                idx_res = col_idx.get("result")
                if idx_res is not None and 0 <= idx_res < len(row):
                    val = row[idx_res]
                    if val is not None:
                        result_val = str(val)
                idx_env = col_idx.get("rtm_environment")
                if idx_env is not None and 0 <= idx_env < len(row):
                    val = row[idx_env]
                    if val is not None:
                        env = str(val)
                idx_time = col_idx.get("actual_time")
                if idx_time is not None and 0 <= idx_time < len(row):
                    val = row[idx_time]
                    if val not in (None, ""):
                        try:
                            time_val = int(str(val).split(".")[0])
                        except Exception:
                            time_val = None
                idx_def = col_idx.get("defects")
                if idx_def is not None and 0 <= idx_def < len(row):
                    val = row[idx_def]
                    if val is not None:
                        defects = str(val)
                idx_tce_key = col_idx.get("tce_test_key")
                if idx_tce_key is not None and 0 <= idx_tce_key < len(row):
                    val = row[idx_tce_key]
                    if val is not None:
                        tce_test_key = str(val)

                by_te.setdefault(te_key, []).append(
                    {
                        "testcase_jira_key": tc_key,
                        "order_no": order_no,
                        "assignee": assignee,
                        "result": result_val,
                        "actual_time": time_val,
                        "rtm_environment": env,
                        "defects": defects,
                        "tce_test_key": tce_test_key,
                    }
                )

            for te_key, items in by_te.items():
                te_issue = get_issue_by_jira_key(conn, project_id, te_key)
                if not te_issue:
                    continue
                te_row = get_or_create_testexecution_for_issue(conn, te_issue["id"])
                records: List[Dict[str, Any]] = []
                for item in items:
                    tc_issue = get_issue_by_jira_key(conn, project_id, item["testcase_jira_key"])
                    if not tc_issue:
                        continue
                    rec: Dict[str, Any] = {
                        "testcase_id": tc_issue["id"],
                        "order_no": item["order_no"],
                        "assignee": item["assignee"],
                        "result": item["result"],
                        "rtm_environment": item["rtm_environment"],
                        "defects": item["defects"],
                    }
                    if item.get("actual_time") is not None:
                        rec["actual_time"] = item["actual_time"]
                    if item.get("tce_test_key"):
                        rec["tce_test_key"] = item["tce_test_key"]
                    records.append(rec)
                if records:
                    replace_testcase_executions(conn, te_row["id"], records)

    # --- TestcaseStepExecutions 시트 처리 (옵션)
    if "TestcaseStepExecutions" in wb.sheetnames:
        ws = wb["TestcaseStepExecutions"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header)}

            # 캐시: (testcase_id) -> {(group_no, order_no): step_id}
            step_cache: Dict[int, Dict[tuple[int, int], int]] = {}
            # 캐시: (te_issue_id, tc_issue_id) -> testcase_execution_id
            tce_cache: Dict[tuple[int, int], int] = {}
            records_by_tce: Dict[int, List[Dict[str, Any]]] = {}

            cur = conn.cursor()

            for row in rows[1:]:
                if not any(row):
                    continue
                te_idx = col_idx.get("testexecution_jira_key")
                tc_idx = col_idx.get("testcase_jira_key")
                g_idx = col_idx.get("group_no")
                o_idx = col_idx.get("order_no")
                if te_idx is None or tc_idx is None or g_idx is None or o_idx is None:
                    continue

                te_key = row[te_idx]
                tc_key = row[tc_idx]
                if not te_key or not tc_key:
                    continue
                te_key = str(te_key)
                tc_key = str(tc_key)

                try:
                    group_no = int(row[g_idx]) if row[g_idx] not in (None, "") else 1
                except Exception:
                    group_no = 1
                try:
                    order_no = int(row[o_idx]) if row[o_idx] not in (None, "") else 1
                except Exception:
                    order_no = 1

                # 상태/결과/증빙
                status = ""
                actual_result = ""
                evidence = ""
                s_idx = col_idx.get("status")
                if s_idx is not None and 0 <= s_idx < len(row):
                    val = row[s_idx]
                    if val is not None:
                        status = str(val)
                ar_idx = col_idx.get("actual_result")
                if ar_idx is not None and 0 <= ar_idx < len(row):
                    val = row[ar_idx]
                    if val is not None:
                        actual_result = str(val)
                ev_idx = col_idx.get("evidence")
                if ev_idx is not None and 0 <= ev_idx < len(row):
                    val = row[ev_idx]
                    if val is not None:
                        evidence = str(val)

                # TE/TC 이슈 찾기
                te_issue = get_issue_by_jira_key(conn, project_id, te_key)
                tc_issue = get_issue_by_jira_key(conn, project_id, tc_key)
                if not te_issue or not tc_issue:
                    continue

                te_row = get_or_create_testexecution_for_issue(conn, te_issue["id"])
                te_id = te_row["id"]
                tc_id = tc_issue["id"]

                # TestcaseExecution id 찾기/캐시
                pair = (te_id, tc_id)
                tce_id = tce_cache.get(pair)
                if tce_id is None:
                    cur.execute(
                        """
                        SELECT id FROM testcase_executions
                         WHERE testexecution_id = ? AND testcase_id = ?
                         ORDER BY order_no ASC, id ASC
                        """,
                        (te_id, tc_id),
                    )
                    r = cur.fetchone()
                    if not r:
                        continue
                    tce_id = int(r[0])
                    tce_cache[pair] = tce_id

                # testcase_step_id 찾기/캐시
                step_map = step_cache.get(tc_id)
                if step_map is None:
                    step_map = {}
                    steps = get_steps_for_issue(conn, tc_id)
                    for s in steps:
                        try:
                            g = int(s.get("group_no", 1) or 1)
                            o = int(s.get("order_no", 1) or 1)
                            sid = int(s.get("id"))
                        except Exception:
                            continue
                        step_map[(g, o)] = sid
                    step_cache[tc_id] = step_map
                step_id = step_map.get((group_no, order_no))
                if not step_id:
                    continue

                records_by_tce.setdefault(tce_id, []).append(
                    {
                        "testcase_step_id": step_id,
                        "status": status,
                        "actual_result": actual_result,
                        "evidence": evidence,
                    }
                )

            # TCE 별로 Step 실행 상태 저장
            for tce_id, recs in records_by_tce.items():
                replace_step_executions_for_tce(conn, tce_id, recs)
