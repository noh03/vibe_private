import os

# ==========================================
# 설정: 프로젝트 폴더 이름
# ==========================================
ROOT_DIR = "JiraRTM_Full_Client"
UI_DIR = os.path.join(ROOT_DIR, "ui")

# 디렉토리 생성
os.makedirs(UI_DIR, exist_ok=True)
os.makedirs(os.path.join(ROOT_DIR, "temp_evidence"), exist_ok=True)
os.makedirs(os.path.join(ROOT_DIR, "temp_extracted_images"), exist_ok=True)

# 파일 내용을 저장할 딕셔너리
files = {}

# ==========================================
# 1. 환경 설정 및 라이브러리 (requirements.txt)
# ==========================================
files[os.path.join(ROOT_DIR, "requirements.txt")] = """
PyQt6
requests
pandas
openpyxl
openpyxl-image-loader
Pillow
"""

files[os.path.join(ROOT_DIR, "config.py")] = """
# Jira 환경 설정
JIRA_BASE_URL = "https://your-jira-datacenter.com"
PERSONAL_ACCESS_TOKEN = "YOUR_PAT_TOKEN_HERE"
PROJECT_KEY = "KVHSICCU"
PROJECT_ID = "41500"
DB_FILENAME = "local_rtm.db"
"""

# ==========================================
# 2. 데이터베이스 (database.py)
# ==========================================
files[os.path.join(ROOT_DIR, "database.py")] = """
import sqlite3
from config import DB_FILENAME

class DatabaseManager:
    def __init__(self):
        self.conn = sqlite3.connect(DB_FILENAME)
        self.cursor = self.conn.cursor()
        self.init_db()

    def init_db(self):
        # 이슈 기본 정보
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                jira_key TEXT UNIQUE,
                issue_type TEXT,
                summary TEXT,
                description TEXT,
                priority TEXT,
                assignee TEXT,
                status TEXT,
                parent_folder_id INTEGER
            )
        ''')
        # 테스트 스텝
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS test_steps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                issue_key TEXT,
                step_order INTEGER,
                action TEXT,
                input_data TEXT,
                expected_result TEXT,
                FOREIGN KEY(issue_key) REFERENCES issues(jira_key) ON DELETE CASCADE
            )
        ''')
        # 링크/관계
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS relations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_key TEXT,
                target_key TEXT,
                link_type TEXT,
                FOREIGN KEY(source_key) REFERENCES issues(jira_key)
            )
        ''')
        self.conn.commit()

    def close(self):
        self.conn.close()
"""

# ==========================================
# 3. API 매니저 - 핵심 로직 (api_manager.py)
# ==========================================
files[os.path.join(ROOT_DIR, "api_manager.py")] = """
import requests
import json
import os
from config import JIRA_BASE_URL, PERSONAL_ACCESS_TOKEN, PROJECT_ID

class JiraAPIManager:
    def __init__(self):
        self.base_url = JIRA_BASE_URL
        self.headers = {
            "Authorization": f"Bearer {PERSONAL_ACCESS_TOKEN}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

    def get_tree_structure(self):
        url = f"{self.base_url}/rest/rtm/1.0/api/tree/{PROJECT_ID}"
        try:
            resp = requests.get(url, headers=self.headers, verify=False)
            return resp.json() if resp.status_code == 200 else []
        except Exception as e:
            print(f"Tree Fetch Error: {e}")
            return []

    def create_issue(self, data):
        url = f"{self.base_url}/rest/api/2/issue"
        payload = {
            "fields": {
                "project": {"key": data.get('project_key')},
                "summary": data.get('summary'),
                "description": data.get('description'),
                "issuetype": {"name": data.get('issue_type')},
                "priority": {"name": data.get('priority')}
            }
        }
        try:
            resp = requests.post(url, headers=self.headers, json=payload, verify=False)
            return resp.json() if resp.status_code == 201 else None
        except: return None

    def update_test_case_execution(self, execution_key, tc_key, steps_data):
        # RTM 테스트 실행 결과 업데이트 (스텝 포함)
        url = f"{self.base_url}/rest/rtm/1.0/test-execution/{execution_key}/test-case/{tc_key}"
        
        overall_status = "PASS"
        payload_steps = []
        
        for step in steps_data:
            # 상태 계산
            if step['status'] == 'FAIL': overall_status = 'FAIL'
            elif step['status'] == 'TODO' and overall_status != 'FAIL': overall_status = 'IN PROGRESS'
            
            # 증적 파일 업로드 (로컬 경로가 있고 파일이 존재할 경우)
            evidence_ids = []
            ev_path = step.get('evidence_path')
            if ev_path and os.path.exists(ev_path):
                print(f"Uploading evidence: {ev_path}")
                att_id = self.upload_attachment(execution_key, ev_path)
                if att_id: evidence_ids.append(att_id)

            payload_steps.append({
                "row": step['order'],
                "status": step['status'],
                "actualResult": step['actual_result'],
                "evidenceIds": evidence_ids # RTM API 필드명 확인 필요 (attachments 등일 수 있음)
            })

        body = {
            "status": overall_status,
            "steps": payload_steps
        }
        
        try:
            resp = requests.put(url, headers=self.headers, json=body, verify=False)
            if resp.status_code in [200, 204]:
                print(f"Synced {tc_key}: Success")
                return True
            else:
                print(f"Sync Failed {tc_key}: {resp.text}")
                return False
        except Exception as e:
            print(f"API Error: {e}")
            return False

    def upload_attachment(self, issue_key, file_path):
        url = f"{self.base_url}/rest/api/2/issue/{issue_key}/attachments"
        headers_no_type = self.headers.copy()
        del headers_no_type["Content-Type"]
        headers_no_type["X-Atlassian-Token"] = "no-check"
        
        try:
            with open(file_path, 'rb') as f:
                files = {'file': f}
                resp = requests.post(url, headers=headers_no_type, files=files, verify=False)
                if resp.status_code == 200:
                    return resp.json()[0]['id']
        except Exception as e:
            print(f"Upload Error: {e}")
        return None

    def download_attachment(self, url, save_path):
        try:
            resp = requests.get(url, headers=self.headers, verify=False, stream=True)
            if resp.status_code == 200:
                with open(save_path, 'wb') as f:
                    for chunk in resp.iter_content(1024):
                        f.write(chunk)
                return True
        except: pass
        return False
"""

# ==========================================
# 4. 엑셀 매니저 - 이미지 객체 추출 포함 (excel_manager.py)
# ==========================================
files[os.path.join(ROOT_DIR, "excel_manager.py")] = """
import pandas as pd
import os
import uuid
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self):
        self.temp_img_dir = "temp_extracted_images"
        os.makedirs(self.temp_img_dir, exist_ok=True)

    def import_from_excel(self, file_path):
        if not os.path.exists(file_path): return None
        
        xls = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        wb = load_workbook(file_path)
        
        parsed_data = {
            "requirements": [], "test_cases": [], "test_plans": [], 
            "test_executions": [], "defects": []
        }

        # 1. Requirements
        if "Requirements" in xls:
            parsed_data["requirements"] = xls["Requirements"].where(pd.notnull(xls["Requirements"]), None).to_dict(orient='records')

        # 2. Test Cases (Flattened structure parsing)
        if "Test Cases" in xls:
            df = xls["Test Cases"].where(pd.notnull(xls["Test Cases"]), None)
            curr_tc = None
            for _, row in df.iterrows():
                if row.get('Summary'):
                    if curr_tc: parsed_data["test_cases"].append(curr_tc)
                    curr_tc = {
                        "summary": row.get('Summary'), 
                        "priority": row.get('Priority'), 
                        "steps": []
                    }
                if row.get('Step Action') and curr_tc:
                    curr_tc["steps"].append({
                        "order": row.get('Step Order'),
                        "action": row.get('Step Action'), 
                        "input": row.get('Step Input'), 
                        "result": row.get('Step Expected Result')
                    })
            if curr_tc: parsed_data["test_cases"].append(curr_tc)

        # 3. Test Executions (With Image Extraction Logic)
        if "Test Executions" in xls and "Test Executions" in wb.sheetnames:
            df = xls["Test Executions"].where(pd.notnull(xls["Test Executions"]), None)
            ws = wb["Test Executions"]
            
            # Image Loader setup
            try:
                image_loader = SheetImageLoader(ws)
            except:
                image_loader = None # 이미지가 없는 경우 등 예외 처리

            # Find Evidence Column Letter
            ev_col = None
            for cell in ws[1]:
                if cell.value == "Evidence": 
                    ev_col = cell.column_letter
                    break
            
            exec_map = {}
            curr_ex, curr_tc = None, None
            
            for idx, row in df.iterrows():
                excel_row = idx + 2 # Header is row 1, data starts row 2
                
                if row.get('Execution Key (or Summary)'): curr_ex = row.get('Execution Key (or Summary)')
                if row.get('TC Key'): curr_tc = row.get('TC Key')
                
                if curr_ex and curr_tc:
                    if curr_ex not in exec_map: exec_map[curr_ex] = {}
                    if curr_tc not in exec_map[curr_ex]: exec_map[curr_ex][curr_tc] = []
                    
                    if row.get('Step Order'):
                        ev_path = row.get('Evidence')
                        
                        # 셀에 텍스트가 없고, 이미지가 존재하면 이미지 추출
                        if not ev_path and ev_col and image_loader and image_loader.image_in(f"{ev_col}{excel_row}"):
                            try:
                                img = image_loader.get(f"{ev_col}{excel_row}")
                                tmp_name = f"{uuid.uuid4()}.png"
                                tmp_path = os.path.join(self.temp_img_dir, tmp_name)
                                img.save(tmp_path)
                                ev_path = tmp_path
                                print(f"Image extracted: {tmp_path}")
                            except Exception as e:
                                print(f"Image extract failed row {excel_row}: {e}")
                            
                        exec_map[curr_ex][curr_tc].append({
                            "order": row.get('Step Order'),
                            "status": row.get('Step Status', 'TODO'),
                            "actual_result": row.get('Actual Result', ''),
                            "evidence_path": ev_path,
                            "defects": row.get('Defects')
                        })
            
            # Convert map to list
            for ek, tcs in exec_map.items():
                parsed_data["test_executions"].append({
                    "execution_key": ek, 
                    "test_cases": [{"tc_key": k, "steps": v} for k,v in tcs.items()]
                })
        
        return parsed_data

    def export_executions_with_images(self, data, file_path, api_manager):
        # 1. 텍스트 데이터 내보내기 (기본)
        # (상세 구현은 생략하나, pd.DataFrame.to_excel 사용)
        pass 
"""

# ==========================================
# 5. UI Components (ui/)
# ==========================================
files[os.path.join(UI_DIR, "__init__.py")] = ""

# --- Login Dialog ---
files[os.path.join(UI_DIR, "login_dialog.py")] = """
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox
import config

class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Connect to Jira RTM")
        self.setFixedSize(300, 200)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.url = QLineEdit(config.JIRA_BASE_URL)
        self.key = QLineEdit(config.PROJECT_KEY)
        self.token = QLineEdit(config.PERSONAL_ACCESS_TOKEN)
        self.token.setEchoMode(QLineEdit.EchoMode.Password)
        self.token.setPlaceholderText("Personal Access Token")
        
        layout.addWidget(QLabel("Jira Base URL")); layout.addWidget(self.url)
        layout.addWidget(QLabel("Project Key")); layout.addWidget(self.key)
        layout.addWidget(QLabel("Token")); layout.addWidget(self.token)
        
        btn = QPushButton("Connect")
        btn.clicked.connect(self.save_and_accept)
        layout.addWidget(btn)

    def save_and_accept(self):
        config.JIRA_BASE_URL = self.url.text()
        config.PROJECT_KEY = self.key.text()
        config.PERSONAL_ACCESS_TOKEN = self.token.text()
        self.accept()
"""

# --- Issue Selector ---
files[os.path.join(UI_DIR, "issue_selector.py")] = """
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLineEdit, QListWidget, QPushButton, QHBoxLayout

class IssueSelectorDialog(QDialog):
    def __init__(self, filter_type=None):
        super().__init__()
        self.filter = filter_type
        self.selected_issue = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Search Issue Key or Summary...")
        self.list = QListWidget()
        self.search.textChanged.connect(self.do_search)
        
        btn_ok = QPushButton("Select")
        btn_ok.clicked.connect(self.on_select)
        
        layout.addWidget(self.search)
        layout.addWidget(self.list)
        layout.addWidget(btn_ok)
        self.load_mock()

    def load_mock(self):
        # 실제로는 API/DB 연동 필요. 여기선 Mock
        self.data = [
            {"key": "KV-1", "type": "Test Case", "summary": "Login Test"},
            {"key": "KV-99", "type": "Defect", "summary": "Crash on Start"},
            {"key": "KV-10", "type": "Requirement", "summary": "User Auth"}
        ]
        self.do_search("")

    def do_search(self, txt):
        self.list.clear()
        for d in self.data:
            if (not self.filter or d['type'] == self.filter) and txt.lower() in d['summary'].lower():
                self.list.addItem(f"[{d['type']}] {d['key']}: {d['summary']}")

    def on_select(self):
        if self.list.currentItem():
            # "[Type] KEY: Summary" 포맷 파싱
            self.selected_issue = self.list.currentItem().text().split(":")[0].split("] ")[1]
            self.accept()
"""

# --- Details Tab (Upload Logic Included) ---
files[os.path.join(UI_DIR, "details_tab.py")] = """
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QFormLayout, QLineEdit, QTextEdit, QPushButton, QMessageBox, QHBoxLayout

class DetailsTab(QWidget):
    def __init__(self, api_manager=None):
        super().__init__()
        self.api = api_manager
        self.data = {}
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        
        self.summary = QLineEdit()
        self.desc = QTextEdit()
        self.priority = QLineEdit()
        
        form.addRow("Summary", self.summary)
        form.addRow("Description", self.desc)
        form.addRow("Priority", self.priority)
        
        layout.addLayout(form)
        
        # 버튼 영역
        btn_layout = QHBoxLayout()
        self.btn_upload = QPushButton("☁ Upload to Jira")
        self.btn_upload.clicked.connect(self.upload)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_upload)
        layout.addLayout(btn_layout)

    def set_data(self, data):
        self.data = data
        self.summary.setText(data.get('summary', ''))
        self.desc.setText(data.get('description', ''))
        self.priority.setText(data.get('priority', 'Medium'))

    def upload(self):
        if not self.api: return
        payload = {
            "summary": self.summary.text(),
            "description": self.desc.toPlainText(),
            "priority": self.priority.text(),
            "issue_type": self.data.get('type', 'Task'),
            "project_key": "KVHSICCU" # config에서 가져오기
        }
        res = self.api.create_issue(payload)
        if res: QMessageBox.information(self, "Success", f"Created: {res['key']}")
        else: QMessageBox.warning(self, "Fail", "Upload failed")
"""

# --- Steps Tab (Dynamic Table) ---
files[os.path.join(UI_DIR, "steps_tab.py")] = """
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout, QHeaderView

class StepsTab(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        btn_layout = QHBoxLayout()
        btn_add = QPushButton("➕ Add Step")
        btn_remove = QPushButton("➖ Remove Step")
        btn_add.clicked.connect(self.add_step)
        btn_remove.clicked.connect(self.remove_step)
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_remove)
        
        layout.addLayout(btn_layout)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Action", "Input", "Expected Result"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

    def add_step(self):
        self.table.insertRow(self.table.rowCount())

    def remove_step(self):
        cur = self.table.currentRow()
        if cur >= 0: self.table.removeRow(cur)
"""

# --- Runner Dialog (Defect Workflow Included) ---
files[os.path.join(UI_DIR, "runner_dialog.py")] = """
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, 
                             QComboBox, QPushButton, QFileDialog, QMessageBox, QInputDialog, QHeaderView)
from ui.issue_selector import IssueSelectorDialog

class TestRunnerDialog(QDialog):
    def __init__(self, key, summary, steps):
        super().__init__()
        self.setWindowTitle(f"Executing: {key}")
        self.resize(900, 500)
        self.steps = steps
        self.init_ui(summary)

    def init_ui(self, summary):
        layout = QVBoxLayout(self)
        
        self.table = QTableWidget(len(self.steps), 6)
        self.table.setHorizontalHeaderLabels(["Action", "Input", "Expected", "Actual", "Status", "Evidence"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        
        for i, s in enumerate(self.steps):
            self.table.setItem(i, 0, QTableWidgetItem(s['action']))
            self.table.setItem(i, 1, QTableWidgetItem(s['input']))
            self.table.setItem(i, 2, QTableWidgetItem(s['result']))
            self.table.setItem(i, 3, QTableWidgetItem("")) # Actual Result
            
            combo = QComboBox()
            combo.addItems(["TODO", "PASS", "FAIL"])
            combo.currentTextChanged.connect(lambda t, r=i: self.check_fail(t, r))
            self.table.setCellWidget(i, 4, combo)
            
            btn = QPushButton("📎 Attach")
            btn.clicked.connect(lambda _, r=i: self.attach(r))
            self.table.setCellWidget(i, 5, btn)

        btn_save = QPushButton("Save & Finish")
        btn_save.clicked.connect(self.save)
        layout.addWidget(self.table)
        layout.addWidget(btn_save)

    def attach(self, row):
        f, _ = QFileDialog.getOpenFileName(self, "Select Evidence")
        if f: 
            # 사용자 데이터로 파일 경로 저장
            self.table.item(row, 0).setData(100, f) 
            self.table.cellWidget(row, 5).setText("✅ Added")

    def check_fail(self, text, row):
        if text == "FAIL":
            reply = QMessageBox.question(self, "Test Failed", "Create or Link Defect?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.handle_defect(row)

    def handle_defect(self, row):
        opts = ["Create New", "Link Existing"]
        item, ok = QInputDialog.getItem(self, "Defect", "Action:", opts, 0, False)
        if ok and item == "Link Existing":
            dlg = IssueSelectorDialog("Defect")
            if dlg.exec():
                # Actual Result에 결함 키 추가
                current = self.table.item(row, 3).text()
                self.table.setItem(row, 3, QTableWidgetItem(f"{current} [Defect: {dlg.selected_issue}]"))
        elif ok and item == "Create New":
            # 간소화된 생성 로직
            summ, ok_s = QInputDialog.getText(self, "New Defect", "Summary:")
            if ok_s:
                self.table.setItem(row, 3, QTableWidgetItem(f"{self.table.item(row, 3).text()} [New Defect: {summ}]"))

    def save(self):
        self.accept()
    
    def get_results(self):
        res = {"overall_status": "PASS", "step_results": []}
        for i in range(self.table.rowCount()):
            status = self.table.cellWidget(i, 4).currentText()
            actual = self.table.item(i, 3).text()
            # 저장해둔 파일 경로 가져오기
            ev_path = self.table.item(i, 0).data(100)
            
            if status == "FAIL": res["overall_status"] = "FAIL"
            
            res["step_results"].append({
                "order": i+1,
                "status": status,
                "actual_result": actual,
                "evidence_path": ev_path
            })
        return res
"""

# --- Execution Tab (Dashboard, Import) ---
files[os.path.join(UI_DIR, "execution_tab.py")] = """
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QFileDialog, QGroupBox, QLabel, QHBoxLayout, QMessageBox
from ui.runner_dialog import TestRunnerDialog

class ExecutionTab(QWidget):
    def __init__(self, api_manager=None, excel_manager=None):
        super().__init__()
        self.api = api_manager
        self.excel = excel_manager
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Dashboard
        dash = QGroupBox("Dashboard")
        d_layout = QHBoxLayout()
        self.lbl_stats = QLabel("Pass: 0 | Fail: 0")
        d_layout.addWidget(self.lbl_stats)
        dash.setLayout(d_layout)
        layout.addWidget(dash)

        # Actions
        btn_layout = QHBoxLayout()
        btn_import = QPushButton("📤 Import Results from Excel & Sync")
        btn_import.clicked.connect(self.import_excel)
        btn_run = QPushButton("▶ Run Selected")
        btn_run.clicked.connect(self.run_manual)
        
        btn_layout.addWidget(btn_import)
        btn_layout.addWidget(btn_run)
        layout.addLayout(btn_layout)

        # Table
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Key", "Summary", "Status", "Defects", "Action"])
        layout.addWidget(self.table)

    def set_data(self, data):
        # Data format: {'test_cases': [{'key':..., 'summary':..., 'status':...}]}
        self.table.setRowCount(0)
        for tc in data.get('test_cases', []):
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(tc['key']))
            self.table.setItem(row, 1, QTableWidgetItem(tc['summary']))
            self.table.setItem(row, 2, QTableWidgetItem(tc.get('status', 'TODO')))
            self.table.setItem(row, 3, QTableWidgetItem(""))

    def run_manual(self):
        cur = self.table.currentRow()
        if cur < 0: return
        key = self.table.item(cur, 0).text()
        summary = self.table.item(cur, 1).text()
        
        # Mock Steps (실제론 DB 조회 필요)
        steps = [{"action": "Step 1", "input": "-", "result": "Success"}]
        
        dlg = TestRunnerDialog(key, summary, steps)
        if dlg.exec():
            res = dlg.get_results()
            self.table.item(cur, 2).setText(res['overall_status'])
            # API Sync Logic could go here

    def import_excel(self):
        f, _ = QFileDialog.getOpenFileName(self, "Excel File")
        if f and self.excel:
            data = self.excel.import_from_excel(f)
            if not data or 'test_executions' not in data: return
            
            count = 0
            for exc in data['test_executions']:
                ex_key = exc['execution_key']
                for tc in exc['test_cases']:
                    # Call API to sync
                    if self.api.update_test_case_execution(ex_key, tc['tc_key'], tc['steps']):
                        count += 1
            QMessageBox.information(self, "Sync", f"Synced {count} Test Cases to Jira.")
"""

# --- Test Plan Tab ---
files[os.path.join(UI_DIR, "test_plan_tab.py")] = """
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QPushButton, QMessageBox
from ui.issue_selector import IssueSelectorDialog

class TestPlanTab(QWidget):
    def __init__(self):
        super().__init__()
        self.tcs = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Key", "Summary"])
        
        btn_add = QPushButton("Add Test Case")
        btn_add.clicked.connect(self.add_tc)
        btn_create = QPushButton("Create Execution")
        btn_create.clicked.connect(self.create_exec)
        
        layout.addWidget(btn_add)
        layout.addWidget(self.table)
        layout.addWidget(btn_create)

    def add_tc(self):
        dlg = IssueSelectorDialog("Test Case")
        if dlg.exec():
            row = self.table.rowCount()
            self.table.insertRow(row)
            # 여기선 키만 넣음 (실제론 상세 조회)
            self.table.setItem(row, 0, QTableWidgetItem(dlg.selected_issue))

    def create_exec(self):
        QMessageBox.information(self, "Create", "Execution Created (Mock)")
"""

# --- Relations Tab ---
files[os.path.join(UI_DIR, "relations_tab.py")] = """
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QPushButton
class RelationsTab(QWidget):
    def __init__(self):
        super().__init__()
        QVBoxLayout(self).addWidget(QTableWidget(0, 3))
"""

# ==========================================
# 6. 메인 윈도우 (Main Window)
# ==========================================
files[os.path.join(UI_DIR, "main_window.py")] = """
from PyQt6.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QSplitter, QTreeWidget, QTabWidget, QTreeWidgetItem, QLabel
from PyQt6.QtCore import Qt
from ui.details_tab import DetailsTab
from ui.steps_tab import StepsTab
from ui.test_plan_tab import TestPlanTab
from ui.execution_tab import ExecutionTab
from ui.relations_tab import RelationsTab
from excel_manager import ExcelManager
from api_manager import JiraAPIManager

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Jira RTM Client (Full)")
        self.resize(1200, 800)
        self.api = JiraAPIManager()
        self.excel = ExcelManager()
        self.setup_ui()

    def setup_ui(self):
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left Panel
        left_widget = QWidget()
        l_layout = QVBoxLayout(left_widget)
        l_layout.addWidget(QLabel("RTM Tree"))
        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self.on_click)
        l_layout.addWidget(self.tree)
        splitter.addWidget(left_widget)
        
        # Right Panel (Tabs)
        self.tabs = QTabWidget()
        
        self.details = DetailsTab(self.api)
        self.steps = StepsTab()
        self.plan = TestPlanTab()
        self.exec = ExecutionTab(self.api, self.excel)
        self.rel = RelationsTab()
        
        self.tabs.addTab(self.details, "Details")
        self.tabs.addTab(self.steps, "Steps")
        self.tabs.addTab(self.plan, "Plan")
        self.tabs.addTab(self.exec, "Execution")
        self.tabs.addTab(self.rel, "Relations")
        
        splitter.addWidget(self.tabs)
        splitter.setSizes([300, 900])
        self.setCentralWidget(splitter)

    def on_click(self, item, col):
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if not data: return
        
        node_type = data.get('type')
        text = item.text(0)
        
        # 탭 표시 제어 Logic
        self.tabs.setTabVisible(1, node_type == 'testCase') # Steps
        self.tabs.setTabVisible(2, node_type == 'testPlan') # Plan
        self.tabs.setTabVisible(3, node_type == 'testExecution') # Execution
        
        # Details Set
        self.details.set_data({"summary": text, "type": node_type})
        
        if node_type == 'testExecution':
            self.tabs.setCurrentWidget(self.exec)
            # Mock Data for Execution Table
            self.exec.set_data({'test_cases': [{'key': 'KV-10', 'summary': 'Login', 'status': 'TODO'}]})

    def populate_server_tree(self, data):
        self.tree.clear()
        if not data: return
        for node in data:
            self._add_node(self.tree, node)

    def _add_node(self, parent, node):
        item = QTreeWidgetItem(parent)
        item.setText(0, node.get('text', 'Node'))
        item.setData(0, Qt.ItemDataRole.UserRole, {'type': node.get('type'), 'id': node.get('id')})
        
        if 'children' in node:
            for child in node['children']:
                self._add_node(item, child)
"""

# ==========================================
# 7. 실행 파일 (main.py)
# ==========================================
files[os.path.join(ROOT_DIR, "main.py")] = """
import sys
from PyQt6.QtWidgets import QApplication
from ui.main_window import MainWindow
from ui.login_dialog import LoginDialog
from database import DatabaseManager

def main():
    app = QApplication(sys.argv)
    
    # 1. Login
    login = LoginDialog()
    if login.exec() == 1:
        # 2. DB Init
        db = DatabaseManager()
        
        # 3. Main Window
        window = MainWindow()
        window.show()
        
        # 4. Fetch Tree (Mock if API fails)
        data = window.api.get_tree_structure()
        if not data:
            # Mock Data for Demo if API not connected
            data = [
                {"text": "Requirement Folder", "type": "folder", "children": [
                    {"text": "Login Req", "type": "requirement", "id": 100}
                ]},
                {"text": "Test Plan Folder", "type": "folder", "children": [
                    {"text": "Regression Plan", "type": "testPlan", "id": 200}
                ]},
                {"text": "Execution Folder", "type": "folder", "children": [
                    {"text": "Cycle 1 Execution", "type": "testExecution", "id": 300}
                ]}
            ]
        window.populate_server_tree(data)
        
        sys.exit(app.exec())

if __name__ == "__main__":
    main()
"""

# ==========================================
# 파일 생성 실행
# ==========================================
for path, content in files.items():
    with open(path, "w", encoding="utf-8") as f:
        f.write(content.strip())
    print(f"✅ Created: {path}")

print("\n" + "="*50)
print(f"🎉 PROJECT GENERATED: {os.path.abspath(ROOT_DIR)}")
print("="*50)
print("1. cd JiraRTM_Full_Client")
print("2. pip install -r requirements.txt")
print("3. python main.py")
print("="*50)
