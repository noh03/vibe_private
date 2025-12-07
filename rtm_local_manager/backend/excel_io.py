"""
excel_io.py - Excel import/export helpers for RTM Local Manager.

- Export:
    export_project_to_excel(conn, project_id, file_path)

- Import:
    import_project_from_excel(conn, project_id, file_path)

주의사항:
- .xlsx 처리를 위해 openpyxl 라이브러리를 사용한다. (사용 PC에 설치되어 있어야 함)
- 이 모듈은 main import 시점에는 openpyxl 을 import 하지 않는다.
  (함수 내부에서 import 하므로, openpyxl 미설치 환경에서도 프로그램 실행 자체는 가능)
"""

from __future__ import annotations

from typing import Any, Dict, List

import sqlite3

from backend.excel_mapping import load_mapping as _load_excel_mapping


def _get_col_index(sheet_name: str, header: List[Any]) -> Dict[str, int]:
    """
    주어진 시트 이름과 헤더 행을 기반으로, '논리 컬럼 이름 -> 실제 열 index' 매핑을 생성한다.

    - 기본적으로 헤더 문자열 그대로를 key 로 사용하는 dict 를 만든다.
    - 추가로 excel_mapping.json 에서 불러온 매핑을 적용하여,
      "논리 컬럼 이름" -> "엑셀 상의 실제 헤더 텍스트" 를 해석해 index 를 보강한다.
    - 이렇게 하면:
        - 열 순서와 관계없이, 열 이름 기반으로 import 가 동작하고
        - 사용자가 열 이름을 바꾸더라도 매핑 설정만 맞추면 import 가 가능하다.
    """
    # 원본 헤더 기준 인덱스
    raw_idx: Dict[str, int] = {}
    for i, h in enumerate(header):
        name = str(h) if h is not None else ""
        if name:
            raw_idx[name] = i

    col_idx: Dict[str, int] = dict(raw_idx)

    # 매핑 파일 로드 (매 호출 시 최신 설정을 반영)
    try:
        mapping_all = _load_excel_mapping()
    except Exception:
        mapping_all = {}

    sheet_map = (mapping_all or {}).get(sheet_name, {})
    if isinstance(sheet_map, dict):
        for logical_name, excel_header in sheet_map.items():
            if not excel_header:
                continue
            idx = raw_idx.get(str(excel_header))
            if idx is not None:
                col_idx[logical_name] = idx

    return col_idx


def _ensure_openpyxl():
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover - runtime 오류 표시용
        raise RuntimeError(
            "openpyxl 이 설치되어 있지 않아 Excel 기능을 사용할 수 없습니다.\n"
            "pip install openpyxl 로 설치 후 다시 시도해 주세요."
        ) from e
    return openpyxl


# --------------------------------------------------------------------------- Export helpers


def export_project_to_excel(conn: sqlite3.Connection, project_id: int, file_path: str) -> None:
    """
    현재 project_id 에 속한 주요 엔티티들을 하나의 Excel 워크북(.xlsx)으로 내보낸다.

    시트 구성:
      - Issues
      - TestcaseSteps
      - Relations
      - TestPlanTestcases
      - TestExecutions
      - TestcaseExecutions

    각 시트는 최소한의 식별자(jira_key, summary 등)를 포함하여,
    나중에 Import 시 jira_key 를 기준으로 매핑할 수 있도록 설계한다.
    """
    openpyxl = _ensure_openpyxl()
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment

    wb = openpyxl.Workbook()
    # default sheet 제거
    default_sheet = wb.active
    wb.remove(default_sheet)

    cur = conn.cursor()

    # --- Issues 시트
    ws = wb.create_sheet("Issues")
    # 컬럼 정의 (DB 스키마의 주요 필드를 대부분 노출하여 엑셀만으로 이슈 메타를 관리할 수 있도록 확장)
    # excel_key 를 id 바로 옆에 배치하여, 신규 설계 시 작성 편의성을 높인다.
    issue_cols = [
        "id",
        "excel_key",  # 엑셀 전용 참조 키 (DB에는 저장되지 않지만, 시트 간 매핑용으로 사용)
        "jira_key",
        "issue_type",
        "folder_path",  # folders 트리 구조를 문자열 경로로 표현
        "summary",
        "description",
        "status",
        "priority",
        "assignee",
        "reporter",
        "labels",
        "components",
        "security_level",
        "fix_versions",
        "affects_versions",
        "rtm_environment",
        "due_date",
        "created",
        "updated",
        "attachments",
        "epic_link",
        "sprint",
    ]
    ws.append(issue_cols)
    cur.execute(
        """
        SELECT id,
               jira_key,
               issue_type,
               folder_id,
               summary,
               description,
               status,
               priority,
               assignee,
               reporter,
               labels,
               components,
               security_level,
               fix_versions,
               affects_versions,
               rtm_environment,
               due_date,
               created,
               updated,
               attachments,
               epic_link,
               sprint
          FROM issues
         WHERE project_id = ? AND is_deleted = 0
         ORDER BY id
        """,
        (project_id,),
    )
    from backend.db import get_folder_path
    # DB select 컬럼 순서에 대한 이름 매핑
    db_issue_cols = [
        "id",
        "jira_key",
        "issue_type",
        "folder_path",
        "summary",
        "description",
        "status",
        "priority",
        "assignee",
        "reporter",
        "labels",
        "components",
        "security_level",
        "fix_versions",
        "affects_versions",
        "rtm_environment",
        "due_date",
        "created",
        "updated",
        "attachments",
        "epic_link",
        "sprint",
    ]
    for row in cur.fetchall():
        row = list(row)
        folder_id = row[3]
        row[3] = get_folder_path(conn, folder_id)
        row_dict = {name: row[i] for i, name in enumerate(db_issue_cols)}
        # excel_key 는 DB 에 없으므로, export 시에는 빈 값으로 채워 템플릿만 제공
        row_dict["excel_key"] = ""
        excel_row = [row_dict.get(col, "") for col in issue_cols]
        ws.append(excel_row)

    # --- TestcaseSteps 시트
    ws = wb.create_sheet("Testcase_Steps")
    # SW 사양 기준으로 Test Case Steps 구조 확장:
    # - Preconditions: Test Case 단위 필드
    # - group_no: Step 그룹 번호 (각 그룹 내에서 order_no 가 1부터 시작)
    #
    # sheet 간 명시적 매핑을 위해:
    #   - issue_id      : 로컬 DB issues.id (Issues 시트의 id 와 1:1 대응)
    #   - excel_key     : Issues.excel_key 와 매핑되는 엑셀 전용 키
    #   - issue_jira_key: 해당 Test Case 의 Jira Key (있는 경우)
    step_cols = [
        "issue_id",
        "excel_key",
        "issue_jira_key",
        "preconditions",
        "group_no",
        "order_no",
        "action",
        "input",
        "expected",
    ]
    ws.append(step_cols)
    cur.execute(
        """
        SELECT t.issue_id,
               i.jira_key,
               i.preconditions,
               COALESCE(t.group_no, 1) AS group_no,
               t.order_no,
               t.action,
               t.input,
               t.expected
          FROM testcase_steps t
          JOIN issues i ON t.issue_id = i.id
         WHERE i.project_id = ? AND i.is_deleted = 0
         ORDER BY t.issue_id, group_no, order_no
        """,
        (project_id,),
    )
    db_step_cols = [
        "issue_id",
        "issue_jira_key",
        "preconditions",
        "group_no",
        "order_no",
        "action",
        "input",
        "expected",
    ]
    for row in cur.fetchall():
        row = list(row)
        row_dict = {name: row[i] for i, name in enumerate(db_step_cols)}
        # DB 에는 excel_key 가 없으므로, export 시에는 빈 값으로 제공
        row_dict["excel_key"] = ""
        excel_row = [row_dict.get(col, "") for col in step_cols]
        ws.append(excel_row)

    # --- Relations 시트
    ws = wb.create_sheet("Relations")
    # jira_key 가 없는 이슈도 excel_key 로 링크를 걸 수 있도록 확장
    rel_cols = ["src_jira_key", "src_excel_key", "dst_jira_key", "dst_excel_key", "relation_type"]
    ws.append(rel_cols)
    cur.execute(
        """
        SELECT s.jira_key AS src_key,
               d.jira_key AS dst_key,
               r.relation_type
          FROM relations r
          JOIN issues s ON r.src_issue_id = s.id
          JOIN issues d ON r.dst_issue_id = d.id
         WHERE s.project_id = ? AND s.is_deleted = 0 AND d.is_deleted = 0
         ORDER BY src_key, dst_key
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        src_key, dst_key, rel_type = row
        row_dict = {
            "src_jira_key": src_key,
            "src_excel_key": "",
            "dst_jira_key": dst_key,
            "dst_excel_key": "",
            "relation_type": rel_type,
        }
        excel_row = [row_dict.get(col, "") for col in rel_cols]
        ws.append(excel_row)

    # --- TestPlanTestcases 시트
    ws = wb.create_sheet("TestPlan_Testcases")
    tp_cols = [
        "testplan_jira_key",
        "testplan_excel_key",
        "testcase_jira_key",
        "testcase_excel_key",
        "order_no",
    ]
    ws.append(tp_cols)
    cur.execute(
        """
        SELECT tp.jira_key AS testplan_key,
               tc.jira_key AS testcase_key,
               t.order_no
          FROM testplan_testcases t
          JOIN issues tp ON t.testplan_id = tp.id
          JOIN issues tc ON t.testcase_id = tc.id
         WHERE tp.project_id = ? AND tp.is_deleted = 0 AND tc.is_deleted = 0
         ORDER BY testplan_key, t.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        testplan_key, testcase_key, order_no = row
        row_dict = {
            "testplan_jira_key": testplan_key,
            "testplan_excel_key": "",
            "testcase_jira_key": testcase_key,
            "testcase_excel_key": "",
            "order_no": order_no,
        }
        excel_row = [row_dict.get(col, "") for col in tp_cols]
        ws.append(excel_row)

    # --- TestExecutions 시트
    ws = wb.create_sheet("Test_Executions")
    te_cols = [
        "testexecution_jira_key",
        "testexecution_excel_key",
        "environment",
        "start_date",
        "end_date",
        "result",
        "executed_by",
    ]
    ws.append(te_cols)
    cur.execute(
        """
        SELECT i.jira_key,
               t.environment,
               t.start_date,
               t.end_date,
               t.result,
               t.executed_by
          FROM testexecutions t
          JOIN issues i ON t.issue_id = i.id
         WHERE i.project_id = ? AND i.is_deleted = 0
         ORDER BY i.jira_key
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        te_key, env, start_date, end_date, result_val, executed_by = row
        row_dict = {
            "testexecution_jira_key": te_key,
            "testexecution_excel_key": "",
            "environment": env,
            "start_date": start_date,
            "end_date": end_date,
            "result": result_val,
            "executed_by": executed_by,
        }
        excel_row = [row_dict.get(col, "") for col in te_cols]
        ws.append(excel_row)

    # --- TestcaseExecutions 시트
    ws = wb.create_sheet("Testcase_Executions")
    tce_cols = [
        "testexecution_jira_key",
        "testexecution_excel_key",
        "testcase_jira_key",
        "testcase_excel_key",
        "order_no",
        "assignee",
        "result",
        "actual_time",
        "rtm_environment",
        "defects",
        "tce_test_key",
    ]
    ws.append(tce_cols)
    cur.execute(
        """
        SELECT tei.jira_key AS te_key,
               tci.jira_key AS tc_key,
               t.order_no,
               t.assignee,
               t.result,
               t.actual_time,
               t.rtm_environment,
               t.defects,
               t.tce_test_key
          FROM testcase_executions t
          JOIN testexecutions te ON t.testexecution_id = te.id
          JOIN issues tei ON te.issue_id = tei.id
          JOIN issues tci ON t.testcase_id = tci.id
         WHERE tei.project_id = ? AND tei.is_deleted = 0 AND tci.is_deleted = 0
         ORDER BY te_key, t.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        te_key, tc_key, order_no, assignee, result_val, actual_time, env, defects, tce_test_key = row
        row_dict = {
            "testexecution_jira_key": te_key,
            "testexecution_excel_key": "",
            "testcase_jira_key": tc_key,
            "testcase_excel_key": "",
            "order_no": order_no,
            "assignee": assignee,
            "result": result_val,
            "actual_time": actual_time,
            "rtm_environment": env,
            "defects": defects,
            "tce_test_key": tce_test_key,
        }
        excel_row = [row_dict.get(col, "") for col in tce_cols]
        ws.append(excel_row)

    # --- TestcaseStepExecutions 시트 (옵션)
    ws = wb.create_sheet("TestcaseStep_Executions")
    tse_cols = [
        "testexecution_jira_key",
        "testexecution_excel_key",
        "testcase_jira_key",
        "testcase_excel_key",
        "group_no",
        "order_no",
        "status",
        "actual_result",
        "evidence",
    ]
    ws.append(tse_cols)
    cur.execute(
        """
        SELECT tei.jira_key    AS te_key,
               tci.jira_key    AS tc_key,
               ts.group_no     AS group_no,
               ts.order_no     AS order_no,
               tse.status      AS status,
               tse.actual_result AS actual_result,
               tse.evidence    AS evidence
          FROM testcase_step_executions tse
          JOIN testcase_executions tce ON tse.testcase_execution_id = tce.id
          JOIN testexecutions te ON tce.testexecution_id = te.id
          JOIN issues tei ON te.issue_id = tei.id
          JOIN issues tci ON tce.testcase_id = tci.id
          JOIN testcase_steps ts ON tse.testcase_step_id = ts.id
         WHERE tei.project_id = ? AND tei.is_deleted = 0 AND tci.is_deleted = 0
         ORDER BY te_key, tc_key, ts.group_no, ts.order_no
        """,
        (project_id,),
    )
    for row in cur.fetchall():
        te_key, tc_key, group_no, order_no, status, actual_result, evidence = row
        row_dict = {
            "testexecution_jira_key": te_key,
            "testexecution_excel_key": "",
            "testcase_jira_key": tc_key,
            "testcase_excel_key": "",
            "group_no": group_no,
            "order_no": order_no,
            "status": status,
            "actual_result": actual_result,
            "evidence": evidence,
        }
        excel_row = [row_dict.get(col, "") for col in tse_cols]
        ws.append(excel_row)

    # --- Manual 시트: 엑셀 사용 요약 설명 (코드 내 정의된 텍스트 사용)
    manual_ws = wb.create_sheet("Manual")

    manual_lines = [
        ("1. 개요", "1. Overview"),
        ("", ""),
        (
            "이 시트는 RTM Local Manager 에서 사용하는 Excel 파일 구조와 각 시트의 역할을 요약한 설명입니다.",
            "This sheet summarizes the Excel structure used by RTM Local Manager and the role of each worksheet.",
        ),
        (
            "이 파일은 로컬 SQLite DB 의 전체 스냅샷이며, 수정 후 Import 를 통해 DB 를 다시 갱신할 수 있습니다.",
            "This file is a full snapshot of the local SQLite DB, and you can update the DB by editing and importing it.",
        ),
        ("", ""),
        ("2. 공통 규칙", "2. Common rules"),
        ("", ""),
        ("- 빈 행: 모든 셀이 비어 있는 행은 무시됩니다.", "- Empty rows: rows where all cells are empty are ignored."),
        (
            "- ID / jira_key:",
            "- ID / jira_key:",
        ),
        (
            "  - id: 로컬 DB 의 issues.id (Export 전용, 일반적으로 수정하지 않습니다).",
            "  - id: issues.id in the local DB (export-only, usually do not edit).",
        ),
        (
            "  - jira_key: JIRA/RTM 이슈 키 (예: PROJ-123).",
            "  - jira_key: JIRA/RTM issue key (e.g. PROJ-123).",
        ),
        (
            "- issue_type: REQUIREMENT, TEST_CASE, TEST_PLAN, TEST_EXECUTION, DEFECT 중 하나의 상수 값입니다.",
            "- issue_type: one of REQUIREMENT, TEST_CASE, TEST_PLAN, TEST_EXECUTION, DEFECT.",
        ),
        (
            "- 여러 값 리스트: labels, components, fix_versions, affects_versions, attachments, defects 등은 쉼표(,)로 구분된 문자열입니다. 예: Backend, API",
            "- Multi-value fields such as labels, components, fix_versions, affects_versions, attachments, defects are comma-separated strings (e.g. Backend, API).",
        ),
        ("", ""),
        ("3. Import 동작 요약 (중요)", "3. Import behavior summary (important)"),
        ("", ""),
        (
            "- Issues 시트:",
            "- Issues sheet:",
        ),
        (
            "  1) id 가 있는 행 → 해당 id 레코드를 그대로 업데이트합니다.",
            "  1) Row with id → updates the record with that id.",
        ),
        (
            "  2) id 가 없고 jira_key 가 있는 행 → jira_key 로 기존 이슈를 찾아 업데이트하거나, 없으면 새로 생성합니다.",
            "  2) Row without id but with jira_key → updates the issue found by jira_key, or creates a new one.",
        ),
        (
            "  3) id / jira_key 가 모두 없고 issue_type 이 있는 행 → 새 로컬 전용 이슈를 생성합니다.",
            "  3) Row without id and jira_key but with issue_type → creates a new local-only issue.",
        ),
        (
            "- Testcase_Steps / Relations / TestPlan_Testcases / Test_Executions / Testcase_Executions / TestcaseStep_Executions:",
            "- Testcase_Steps / Relations / TestPlan_Testcases / Test_Executions / Testcase_Executions / TestcaseStep_Executions:",
        ),
        (
            "  - 각 시트는 Import 시 해당 이슈/플랜/실행에 대한 기존 내용을 덮어쓰는(overwrite) 방식으로 동작합니다.",
            "  - Each of these sheets overwrites existing data for the corresponding issues/plans/executions.",
        ),
        (
            "  - 필요 없는 행은 Import 전에 엑셀에서 삭제해야 합니다.",
            "  - Delete rows you no longer need in Excel before importing.",
        ),
        ("", ""),
        ("4. 시나리오별 사용 예시", "4. Usage scenarios"),
        ("", ""),
        (
            "- 새 Test Case + Steps 설계:",
            "- Designing a new Test Case with Steps:",
        ),
        (
            "  1) Issues 시트에 id / jira_key 없이 issue_type=TEST_CASE, summary, excel_key 를 입력합니다.",
            "  1) In Issues, enter issue_type=TEST_CASE, summary and excel_key without id / jira_key.",
        ),
        (
            "  2) Testcase_Steps 시트에 같은 excel_key 로 여러 Step 행을 작성합니다.",
            "  2) In Testcase_Steps, add multiple step rows with the same excel_key.",
        ),
        (
            "  3) Import 하면 새 TC 와 Steps 가 함께 생성됩니다.",
            "  3) After import, the new Test Case and its Steps are created together.",
        ),
        (
            "- 기존 JIRA 이슈 메타 보정:",
            "- Adjusting metadata of existing JIRA issues:",
        ),
        (
            "  1) Issues 시트에 jira_key 와 수정하고 싶은 필드(status, priority 등)만 채웁니다.",
            "  1) In Issues, fill jira_key and only the fields you want to change (status, priority, etc.).",
        ),
        (
            "  2) Import 하면 로컬 DB 의 해당 이슈 메타만 업데이트됩니다.",
            "  2) Import updates only those fields in the local DB.",
        ),
        ("", ""),
        ("5. Issues 시트", "5. Issues sheet"),
        ("", ""),
        (
            "- 각 행은 하나의 이슈(Requirement / Test Case / Test Plan / Test Execution / Defect)를 나타냅니다.",
            "- Each row represents one issue (Requirement / Test Case / Test Plan / Test Execution / Defect).",
        ),
        (
            "- 주요 컬럼: id, jira_key, issue_type, folder_path, summary, description, status, priority, assignee, reporter, labels, components, security_level, fix_versions, affects_versions, rtm_environment, due_date, created, updated, attachments, epic_link, sprint, excel_key.",
            "- Main columns: id, jira_key, issue_type, folder_path, summary, description, status, priority, assignee, reporter, labels, components, security_level, fix_versions, affects_versions, rtm_environment, due_date, created, updated, attachments, epic_link, sprint, excel_key.",
        ),
        ("", ""),
        ("6. Testcase_Steps 시트", "6. Testcase_Steps sheet"),
        ("", ""),
        (
            "- 각 행은 하나의 Test Case Step 을 나타냅니다.",
            "- Each row represents one Test Case Step.",
        ),
        (
            "- 주요 컬럼: issue_id, issue_jira_key, preconditions, group_no, order_no, action, input, expected, excel_key.",
            "- Main columns: issue_id, issue_jira_key, preconditions, group_no, order_no, action, input, expected, excel_key.",
        ),
        ("", ""),
        ("7. Relations 시트", "7. Relations sheet"),
        ("", ""),
        (
            "- 이슈 간 관계(링크)를 정의합니다.",
            "- Defines relationships(links) between issues.",
        ),
        (
            "- 주요 컬럼: src_jira_key, dst_jira_key, relation_type (Tests, Relates, Blocks 등).",
            "- Main columns: src_jira_key, dst_jira_key, relation_type (Tests, Relates, Blocks, etc.).",
        ),
        ("", ""),
        ("8. TestPlan_Testcases 시트", "8. TestPlan_Testcases sheet"),
        ("", ""),
        (
            "- Test Plan 과 Test Case 의 매핑을 정의합니다.",
            "- Defines mappings between Test Plans and Test Cases.",
        ),
        (
            "- 주요 컬럼: testplan_jira_key, testcase_jira_key, order_no.",
            "- Main columns: testplan_jira_key, testcase_jira_key, order_no.",
        ),
        ("", ""),
        ("9. Test_Executions 시트", "9. Test_Executions sheet"),
        ("", ""),
        (
            "- Test Execution 이슈의 메타 정보를 정의합니다.",
            "- Defines metadata of Test Execution issues.",
        ),
        (
            "- 주요 컬럼: testexecution_jira_key, environment, start_date, end_date, result, executed_by.",
            "- Main columns: testexecution_jira_key, environment, start_date, end_date, result, executed_by.",
        ),
        ("", ""),
        ("10. Testcase_Executions 시트", "10. Testcase_Executions sheet"),
        ("", ""),
        (
            "- 각 Test Execution 내의 Test Case Execution(TCE) 정보를 정의합니다.",
            "- Defines Test Case Execution (TCE) information under each Test Execution.",
        ),
        (
            "- 주요 컬럼: testexecution_jira_key, testcase_jira_key, order_no, assignee, result, actual_time, rtm_environment, defects, tce_test_key.",
            "- Main columns: testexecution_jira_key, testcase_jira_key, order_no, assignee, result, actual_time, rtm_environment, defects, tce_test_key.",
        ),
        ("", ""),
        ("11. TestcaseStep_Executions 시트", "11. TestcaseStep_Executions sheet"),
        ("", ""),
        (
            "- 각 TCE 에 대한 Step 실행 결과를 정의합니다.",
            "- Defines step execution results for each TCE.",
        ),
        (
            "- 주요 컬럼: testexecution_jira_key, testcase_jira_key, group_no, order_no, status, actual_result, evidence.",
            "- Main columns: testexecution_jira_key, testcase_jira_key, group_no, order_no, status, actual_result, evidence.",
        ),
        ("", ""),
        (
            "이 설명을 참고하여 각 시트/컬럼에 의미 있는 값을 작성하면, Excel 하나로 요구사항–테스트–실행–결함 데이터를 설계·검토하고, Import/Export 기능을 통해 로컬 DB 및 JIRA/RTM 과 동기화할 수 있습니다.",
            "Using this guide, you can design and review requirements–tests–executions–defects in a single Excel file and sync them with the local DB and JIRA/RTM via Import/Export.",
        ),
    ]

    for i, (ko, en) in enumerate(manual_lines, start=1):
        c_ko = manual_ws.cell(row=i, column=1, value=ko)
        c_en = manual_ws.cell(row=i, column=2, value=en)
        c_ko.alignment = Alignment(wrap_text=True, vertical="top")
        c_en.alignment = Alignment(wrap_text=True, vertical="top")

    # 적당한 컬럼 폭 설정
    manual_ws.column_dimensions["A"].width = 60
    manual_ws.column_dimensions["B"].width = 60

    # Manual 시트를 맨 왼쪽(첫 번째 탭)으로 이동
    try:
        wb._sheets.remove(manual_ws)
        wb._sheets.insert(0, manual_ws)
    except Exception:
        # openpyxl 내부 구조 변경 등에 대비한 방어 코드
        pass

    # FAQ 시트를 생성하고 Manual 오른쪽(두 번째 탭)에 배치
    faq_ws = wb.create_sheet("FAQ")
    faq_lines = [
        ("FAQ – 자주 묻는 질문", "FAQ – Frequently Asked Questions"),
        ("", ""),
        (
            "Q1. Export 한 엑셀을 바로 수정해도 되나요?",
            "Q1. Can I edit the exported Excel file directly?",
        ),
        (
            "A1. 네. 이 파일은 로컬 DB 의 전체 스냅샷입니다. Issues / Testcase_Steps / Relations 등 필요한 시트를 수정한 후 Import 하면 DB 가 갱신됩니다.",
            "A1. Yes. This file is a full snapshot of the local DB. Edit sheets such as Issues / Testcase_Steps / Relations and then import to update the DB.",
        ),
        ("", ""),
        (
            "Q2. 어떤 컬럼은 절대 바꾸면 안 되나요?",
            "Q2. Are there columns I should avoid changing?",
        ),
        (
            "A2. id 컬럼은 일반적으로 수정하지 않는 것이 안전합니다. jira_key 도 기존 JIRA 이슈를 다른 키로 바꾸는 용도로는 사용하지 않는 것을 권장합니다.",
            "A2. It is safest not to edit the id column. Also, do not use jira_key to arbitrarily change an existing issue key in JIRA.",
        ),
        ("", ""),
        (
            "Q3. Import 할 때 데이터가 사라질 수 있나요?",
            "Q3. Can data be lost during import?",
        ),
        (
            "A3. 네. Testcase_Steps, Relations, TestPlan_Testcases, Test_Executions, Testcase_Executions, TestcaseStep_Executions 시트는 기존 데이터를 덮어쓰는 방식입니다.",
            "A3. Yes. The Testcase_Steps, Relations, TestPlan_Testcases, Test_Executions, Testcase_Executions and TestcaseStep_Executions sheets overwrite existing data.",
        ),
        (
            "    필요 없는 행은 Import 전에 엑셀에서 삭제해야 합니다.",
            "    Remove any rows you no longer need in Excel before importing.",
        ),
        ("", ""),
        (
            "Q4. 여러 시트를 한 번에 수정해도 되나요?",
            "Q4. Can I modify multiple sheets at once?",
        ),
        (
            "A4. 네. 한 번의 Import 로 모든 시트를 함께 반영할 수 있습니다. 다만 서로 연관된 excel_key / jira_key 값이 일관되게 작성되었는지 확인해야 합니다.",
            "A4. Yes. A single import can apply changes from all sheets, but ensure related excel_key / jira_key values are consistent.",
        ),
        ("", ""),
        (
            "Q5. 새 Test Case 를 설계할 때 최소로 채워야 하는 컬럼은 무엇인가요?",
            "Q5. What are the minimum columns required when designing a new Test Case?",
        ),
        (
            "A5. Issues 시트에서는 issue_type=TEST_CASE, summary, (필요시 excel_key). Steps 를 함께 설계하려면 Testcase_Steps 시트에 같은 excel_key 로 group_no / order_no / action / expected 를 입력합니다.",
            "A5. In Issues, set issue_type=TEST_CASE, summary and optionally excel_key. To design steps together, use the same excel_key in Testcase_Steps and fill group_no / order_no / action / expected.",
        ),
        ("", ""),
        (
            "Q6. Import 후 JIRA 와는 언제 동기화되나요?",
            "Q6. When is data synchronized with JIRA after import?",
        ),
        (
            "A6. Excel Import 는 로컬 DB 만 갱신합니다. JIRA/RTM 과의 실제 동기화는 GUI 의 Pull/Push, Create in JIRA, Execute Test Plan 기능 등을 통해 별도로 수행합니다.",
            "A6. Excel import updates only the local DB. Synchronization with JIRA/RTM is done separately via GUI actions such as Pull/Push, Create in JIRA and Execute Test Plan.",
        ),
        ("", ""),
        (
            "Q7. 예전 버전에서 Export 한 파일도 그대로 Import 할 수 있나요?",
            "Q7. Can I import files exported from older versions?",
        ),
        (
            "A7. 기본적인 시트/컬럼 구조가 유지된 경우 대부분 호환됩니다. 새로 추가된 컬럼은 비어 있어도 괜찮지만, 매우 오래된 파일은 사양서를 참고해 컬럼을 보정하는 것을 권장합니다.",
            "A7. As long as the basic sheets and columns are compatible, most older files can be imported. Newer columns may be empty; for very old files, adjust columns according to the specification document.",
        ),
    ]

    for i, (ko, en) in enumerate(faq_lines, start=1):
        c_ko = faq_ws.cell(row=i, column=1, value=ko)
        c_en = faq_ws.cell(row=i, column=2, value=en)
        c_ko.alignment = Alignment(wrap_text=True, vertical="top")
        c_en.alignment = Alignment(wrap_text=True, vertical="top")

    faq_ws.column_dimensions["A"].width = 60
    faq_ws.column_dimensions["B"].width = 60

    try:
        wb._sheets.remove(faq_ws)
        wb._sheets.insert(1, faq_ws)
    except Exception:
        pass

    # 컬럼 헤더에 간단한 설명 Comment 추가
    def _add_header_comments(ws_name: str, comments: Dict[str, str]) -> None:
        ws_local = wb[ws_name]
        header_row = list(ws_local.iter_rows(min_row=1, max_row=1))[0]
        name_to_cell = {str(cell.value): cell for cell in header_row if cell.value}
        for col_name, text in comments.items():
            cell = name_to_cell.get(col_name)
            if cell is not None and not cell.comment:
                cell.comment = Comment(text, "RTM Local Manager")

    issues_comments: Dict[str, str] = {
        "id": "로컬 DB issues.id\nExport 전용. 일반적으로 수정하지 않습니다.\n예: 101",
        "excel_key": "엑셀 전용 임시 키\n새 이슈와 다른 시트를 연결할 때 사용합니다.\n예: TC_LOGIN_001",
        "jira_key": "JIRA/RTM 이슈 키\n이미 존재하는 이슈를 수정할 때 사용합니다.\n예: PROJ-123",
        "issue_type": "이슈 타입\nREQUIREMENT / TEST_CASE / TEST_PLAN / TEST_EXECUTION / DEFECT 중 하나.",
        "folder_path": "트리 상의 폴더 경로\n없으면 기본 폴더에 생성됩니다.\n예: Requirements/Release1/ModuleA",
        "summary": "이슈 제목\n예: 로그인 성공 시 메인 화면 이동",
        "description": "이슈 상세 설명(본문).",
        "status": "JIRA 상태 이름\n예: To Do, In Progress, Done",
        "priority": "우선순위\n예: Highest, High, Medium, Low",
        "assignee": "담당자 표시 이름.",
        "reporter": "보고자 표시 이름.",
        "labels": "라벨 리스트(쉼표로 구분)\n예: regression, smoke",
        "components": "컴포넌트 리스트(쉼표로 구분)\n예: Backend, API",
        "security_level": "보안 레벨 이름.",
        "fix_versions": "Fix Versions(쉼표로 구분)\n예: 1.0.0, 1.0.1",
        "affects_versions": "Affects Versions(쉼표로 구분).",
        "rtm_environment": "RTM 실행 환경 태그\n예: DEV, QA, PROD",
        "due_date": "마감일(옵션)\n예: 2025-12-31",
        "created": "생성 일시(참고용).",
        "updated": "수정 일시(참고용).",
        "attachments": "첨부파일 정보 문자열.\n실제 첨부는 GUI/REST 로 처리하는 것을 권장.",
        "epic_link": "연결된 Epic 이슈 키\n예: PROJ-1",
        "sprint": "스프린트 이름 또는 Sprint 필드 문자열.",
    }
    _add_header_comments("Issues", issues_comments)

    steps_comments: Dict[str, str] = {
        "issue_id": "이 Step 이 속한 Test Case 의 로컬 ID\nExport 된 값은 그대로 두는 것을 권장합니다.",
        "excel_key": "Issues.excel_key 와 매핑되는 키\njira_key 가 없어도 TC 와 Step 을 연결할 수 있습니다.\n예: TC_LOGIN_001",
        "issue_jira_key": "이 Step 이 속한 Test Case 의 JIRA 키\n예: PROJ-200",
        "preconditions": "Test Case 전제조건 텍스트.",
        "group_no": "Step 그룹 번호\n없으면 1로 처리됩니다.",
        "order_no": "그룹 내 Step 순번(정수).",
        "action": "수행할 동작(What to do).",
        "input": "입력 값(With what).",
        "expected": "기대 결과(What is expected).",
    }
    _add_header_comments("Testcase_Steps", steps_comments)

    relations_comments: Dict[str, str] = {
        "src_jira_key": "관계 출발 이슈의 JIRA 키.\n예: REQ-1",
        "src_excel_key": "관계 출발 이슈의 excel_key\njira_key 가 없을 때 사용.\n예: REQ_LOGIN",
        "dst_jira_key": "관계 도착 이슈의 JIRA 키.\n예: TC-1",
        "dst_excel_key": "관계 도착 이슈의 excel_key\njira_key 가 없을 때 사용.\n예: TC_LOGIN_001",
        "relation_type": "관계 타입 문자열\n예: Tests, Is tested by, Relates",
    }
    _add_header_comments("Relations", relations_comments)

    tp_comments: Dict[str, str] = {
        "testplan_jira_key": "Test Plan 이슈의 JIRA 키.",
        "testplan_excel_key": "Test Plan 이슈의 excel_key\njira_key 가 없을 때 사용.",
        "testcase_jira_key": "포함될 Test Case 의 JIRA 키.",
        "testcase_excel_key": "포함될 Test Case 의 excel_key\njira_key 가 없을 때 사용.",
        "order_no": "Test Plan 내 실행 순서(정수).",
    }
    _add_header_comments("TestPlan_Testcases", tp_comments)

    te_comments: Dict[str, str] = {
        "testexecution_jira_key": "Test Execution 이슈의 JIRA 키.",
        "testexecution_excel_key": "Test Execution 이슈의 excel_key\njira_key 가 없을 때 사용.",
        "environment": "실행 환경\n예: QA, PROD",
        "start_date": "실행 시작 일시(문자열).",
        "end_date": "실행 종료 일시(문자열).",
        "result": "실행 결과\n예: In progress, Pass, Fail",
        "executed_by": "실행자 이름.",
    }
    _add_header_comments("Test_Executions", te_comments)

    tce_comments: Dict[str, str] = {
        "testexecution_jira_key": "상위 Test Execution 의 JIRA 키.",
        "testexecution_excel_key": "상위 Test Execution 의 excel_key\njira_key 가 없을 때 사용.",
        "testcase_jira_key": "실행 대상 Test Case 의 JIRA 키.",
        "testcase_excel_key": "실행 대상 Test Case 의 excel_key\njira_key 가 없을 때 사용.",
        "order_no": "해당 TE 내에서 Test Case 의 순번.",
        "assignee": "이 TCE 의 담당자.",
        "result": "TCE 결과\n예: Not Executed, Pass, Fail",
        "actual_time": "실제 소요 시간(분 단위 정수).",
        "rtm_environment": "RTM 실행 환경 태그\n예: DEV, QA, PROD",
        "defects": "연결된 Defect Jira Key 목록(쉼표로 구분).\n예: PROJ-101, PROJ-102",
        "tce_test_key": "RTM 상의 Test Case Execution key.",
    }
    _add_header_comments("Testcase_Executions", tce_comments)

    tse_comments: Dict[str, str] = {
        "testexecution_jira_key": "상위 Test Execution 의 JIRA 키.",
        "testexecution_excel_key": "상위 Test Execution 의 excel_key\njira_key 가 없을 때 사용.",
        "testcase_jira_key": "실행 대상 Test Case 의 JIRA 키.",
        "testcase_excel_key": "실행 대상 Test Case 의 excel_key\njira_key 가 없을 때 사용.",
        "group_no": "설계된 Step 의 그룹 번호.",
        "order_no": "설계된 Step 의 순번.",
        "status": "Step 실행 상태\n예: Not Executed, Pass, Fail",
        "actual_result": "Step 실제 실행 결과(텍스트).",
        "evidence": "증거(파일명/로그 경로/URL 등 텍스트).",
    }
    _add_header_comments("TestcaseStep_Executions", tse_comments)

    # 약간의 자동 너비 조정 (간단히)
    for ws in wb.worksheets:
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                except Exception:
                    val = ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    wb.save(file_path)


# --------------------------------------------------------------------------- Import helpers


def import_project_from_excel(
    conn: sqlite3.Connection,
    project_id: int,
    file_path: str,
    progress_cb: Any | None = None,
) -> None:
    """
    Excel 파일(.xlsx)에서 주요 엔티티를 읽어와 로컬 DB에 반영한다.

    처리 범위:
      - Issues 시트:
          - jira_key 기준으로 존재 여부 확인
          - 있으면 주요 필드 update, 없으면 신규 issue insert (folder_id 는 None)
      - TestcaseSteps 시트:
          - issue_jira_key 기준으로 대상 issue 찾은 뒤, 해당 issue 의 steps 전체 교체
      - Relations 시트:
          - src_jira_key / dst_jira_key 기준으로 양쪽 issue 찾고 relations 교체
      - TestPlanTestcases 시트:
          - testplan_jira_key / testcase_jira_key 기준으로 testplan_testcases 교체
      - TestExecutions 시트:
          - testexecution_jira_key 기준으로 testexecution 메타 정보 update/신규 생성
      - TestcaseExecutions 시트:
          - testexecution_jira_key / testcase_jira_key 기준으로 testcase_executions 교체

    Excel 컬럼 헤더는 기본적으로 export_project_to_excel()에서 생성한 것과 동일해야 하며,
    Settings > Excel Column Mapping... 에서 정의한 매핑을 통해 다른 헤더 이름도 인식할 수 있다.

    progress_cb 인자가 주어지면, 시트 단위로 진행 상황을 콜백한다:
        progress_cb(message: str, current_step: int, total_steps: int)
    """
    openpyxl = _ensure_openpyxl()
    from backend.db import (
        get_issue_by_jira_key,
        get_issue_by_id,
        update_issue_fields,
        replace_steps_for_issue,
        replace_relations_for_issue,
        replace_testplan_testcases,
        get_or_create_testexecution_for_issue,
        update_testexecution_for_issue,
        replace_testcase_executions,
        get_steps_for_issue,
        replace_step_executions_for_tce,
        create_local_issue,
        ensure_folder_path,
    )

    wb = openpyxl.load_workbook(file_path, data_only=True)

    cur = conn.cursor()

    # 진행 상황 보고용 헬퍼
    def _progress(message: str, current: int, total: int) -> None:
        if not progress_cb:
            return
        try:
            progress_cb(message, int(current), int(total))
        except Exception:
            # 콜백 오류는 전체 import 를 방해하지 않도록 무시
            pass

    # 처리 대상 시트 순서 및 총 단계 수 계산
    ordered_sheets = [
        "Issues",
        "TestcaseSteps",
        "Relations",
        "TestPlanTestcases",
        "TestExecutions",
        "TestcaseExecutions",
        "TestcaseStepExecutions",
    ]
    present_sheets = [name for name in ordered_sheets if name in wb.sheetnames]
    total_steps = max(1, len(present_sheets))
    step_index = 0

    _progress("Loading workbook...", 0, total_steps)

    # Issues 시트와 다른 시트(TestcaseSteps 등) 사이를 연결하기 위한
    # 엑셀 전용 레퍼런스 키 매핑 (예: excel_key -> issue_id)
    # - DB에는 저장되지 않고, "한 번에 신규 TC + Steps 생성"을 위한 임시 매핑용이다.
    excel_issue_ref: Dict[str, int] = {}

    # --- Issues 시트 처리
    if "Issues" in wb.sheetnames:
        ws = wb["Issues"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"Issues 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            # 열 순서와 무관하게, 이름/매핑 기반으로 index 를 해석
            col_idx = _get_col_index("Issues", header)
            for row in rows[1:]:
                if not any(row):
                    continue

                # 0) id 컬럼이 있고 값이 있으면, 로컬 ID 기준 "수정"으로 처리
                local_id = None
                id_idx = col_idx.get("id")
                if id_idx is not None and 0 <= id_idx < len(row):
                    raw_id = row[id_idx]
                    if raw_id not in (None, ""):
                        try:
                            # 엑셀에서 숫자가 float 로 넘어오는 경우까지 고려
                            local_id = int(str(raw_id).split(".")[0])
                        except Exception:
                            local_id = None

                # 엑셀 전용 레퍼런스 키 (optional)
                excel_ref: str | None = None
                excel_ref_idx = col_idx.get("excel_key")
                if excel_ref_idx is not None and 0 <= excel_ref_idx < len(row):
                    raw_ref = row[excel_ref_idx]
                    if raw_ref not in (None, ""):
                        excel_ref = str(raw_ref).strip()

                # 공통 필드 추출 (id/jira_key 를 제외한 나머지)
                # jira_key 컬럼이 있으면 JIRA 이슈 기준으로 upsert,
                # jira_key 가 비어 있으면 "로컬 전용 이슈" 로 신규 생성한다.
                jira_key_idx = col_idx.get("jira_key")
                jira_key: str | None
                if jira_key_idx is None or jira_key_idx < 0 or jira_key_idx >= len(row):
                    jira_key = None
                else:
                    raw_key = row[jira_key_idx]
                    jira_key = str(raw_key) if raw_key not in (None, "") else None

                fields: Dict[str, Any] = {}
                for name in [
                    "issue_type",
                    "summary",
                    "description",
                    "status",
                    "priority",
                    "assignee",
                    "reporter",
                    "labels",
                    "components",
                    "security_level",
                    "fix_versions",
                    "affects_versions",
                    "rtm_environment",
                    "due_date",
                    "created",
                    "updated",
                    "attachments",
                    "epic_link",
                    "sprint",
                ]:
                    idx = col_idx.get(name)
                    if idx is None or idx < 0 or idx >= len(row):
                        continue
                    val = row[idx]
                    if val is not None:
                        fields[name] = str(val)

                # 폴더 경로 (optional): folder_path 컬럼이 있으면 이를 기반으로 folder_id 를 계산
                folder_id = None
                folder_idx = col_idx.get("folder_path")
                if folder_idx is not None and 0 <= folder_idx < len(row):
                    raw_path = row[folder_idx]
                    if raw_path not in (None, ""):
                        issue_type_for_folder = fields.get("issue_type")
                        folder_id = ensure_folder_path(
                            conn,
                            project_id=project_id,
                            path=str(raw_path),
                            issue_type=issue_type_for_folder,
                        )
                        if folder_id:
                            fields["folder_id"] = folder_id

                # 0-1) 이 행에서 실제로 업데이트/생성한 이슈의 id (엑셀 ref 매핑에 사용)
                target_issue_id: int | None = None

                # 0-2) local_id 로 업데이트 가능한 경우: 해당 행은 "수정"으로만 처리
                if local_id is not None:
                    issue = get_issue_by_id(conn, local_id)
                    if issue and issue.get("project_id") == project_id:
                        update_issue_fields(conn, local_id, fields)
                        target_issue_id = local_id
                        # 아래 jira_key / 신규 생성 로직은 건너뛴다.
                        pass
                    else:
                        # ID 가 없거나 다른 프로젝트인 경우에는 아래 jira_key / 신규 생성 로직으로 넘어간다.
                        target_issue_id = None

                # local_id 기반 업데이트가 안 된 경우에만 jira_key / 신규 생성 로직 수행
                if target_issue_id is None:
                    if jira_key:
                        # JIRA Key 가 있는 경우: 기존 이슈 update 또는 신규 JIRA-연동 이슈 insert
                        issue = get_issue_by_jira_key(conn, project_id, jira_key)
                        if issue:
                            update_issue_fields(conn, issue["id"], fields)
                            target_issue_id = int(issue["id"])
                        else:
                            cur.execute(
                                """
                                INSERT INTO issues (
                                    project_id, folder_id, jira_key, issue_type, summary,
                                    status, priority, assignee, reporter, labels, components,
                                    security_level, fix_versions, affects_versions,
                                    rtm_environment, due_date, created, updated, is_deleted
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'), 0)
                                """,
                                (
                                    project_id,
                                    folder_id,
                                    jira_key,
                                    fields.get("issue_type") or "",
                                    fields.get("summary") or "",
                                    fields.get("status") or "",
                                    fields.get("priority") or "",
                                    fields.get("assignee") or "",
                                    fields.get("reporter") or "",
                                    fields.get("labels") or "",
                                    fields.get("components") or "",
                                    fields.get("security_level") or "",
                                    fields.get("fix_versions") or "",
                                    fields.get("affects_versions") or "",
                                    fields.get("rtm_environment") or "",
                                    fields.get("due_date") or "",
                                ),
                            )
                            conn.commit()
                            try:
                                target_issue_id = int(cur.lastrowid)
                            except Exception:
                                # lastrowid 를 얻지 못하는 경우, jira_key 로 재조회
                                issue = get_issue_by_jira_key(conn, project_id, jira_key)
                                if issue:
                                    target_issue_id = int(issue["id"])
                    else:
                        # jira_key が 비어 있는 경우:
                        #   - 엑셀에 local ID 가 없어도, issue_type / summary 정보만 있으면
                        #     로컬 전용 이슈를 자동으로 생성한다.
                        issue_type = fields.get("issue_type")
                        summary = fields.get("summary", "")
                        if not issue_type:
                            # issue_type 이 없는 행은 스킵 (스키마상 NOT NULL)
                            continue

                        new_issue_id = create_local_issue(
                            conn,
                            project_id=project_id,
                            issue_type=issue_type,
                            folder_id=folder_id,
                            summary=summary,
                        )
                        # create_local_issue 로 기본 값은 들어갔으므로,
                        # 나머지 필드가 있다면 추가로 업데이트한다.
                        # (issue_type / summary 는 동일 값으로 한 번 더 세팅해도 무방)
                        update_issue_fields(conn, new_issue_id, fields)
                        target_issue_id = new_issue_id

                # 이 행에 excel_key 가 있다면, 해당 키로 issue_id 를 매핑해 둔다.
                if excel_ref and target_issue_id is not None:
                    excel_issue_ref[excel_ref] = target_issue_id

    # --- TestcaseSteps 시트 처리
    if "Testcase_Steps" in wb.sheetnames:
        ws = wb["Testcase_Steps"]
    elif "TestcaseSteps" in wb.sheetnames:
        ws = wb["TestcaseSteps"]
    else:
        ws = None

    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"TestcaseSteps 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("TestcaseSteps", header)
            # sheet 간 매핑을 위해 issue_id 를 우선 사용, 없으면 issue_jira_key 로 fallback
            steps_by_issue_id: Dict[int, List[Dict[str, Any]]] = {}
            preconditions_by_issue_id: Dict[int, str] = {}
            for row in rows[1:]:
                if not any(row):
                    continue

                issue = None
                issue_id_val: int | None = None

                # 1) issue_id 로 우선 매핑 (Issues 시트의 id 와 동일)
                id_idx = col_idx.get("issue_id")
                if id_idx is not None and 0 <= id_idx < len(row):
                    raw_id = row[id_idx]
                    if raw_id not in (None, ""):
                        try:
                            issue_id_val = int(str(raw_id).split(".")[0])
                        except Exception:
                            issue_id_val = None
                        if issue_id_val is not None:
                            issue = get_issue_by_id(conn, issue_id_val)
                            if issue and issue.get("project_id") != project_id:
                                issue = None

                # 2) excel_key 로 매핑 (Issues 시트에서 excel_issue_ref 로 기록된 경우)
                if issue is None:
                    excel_ref_idx = col_idx.get("excel_key")
                    if excel_ref_idx is not None and 0 <= excel_ref_idx < len(row):
                        raw_ref = row[excel_ref_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                issue = get_issue_by_id(conn, mapped_id)
                                if issue and issue.get("project_id") != project_id:
                                    issue = None

                # 3) fallback: issue_jira_key 로 매핑 (기존 파일 호환)
                if issue is None:
                    key_idx = col_idx.get("issue_jira_key")
                    if key_idx is not None and 0 <= key_idx < len(row):
                        jira_key = row[key_idx]
                        if jira_key:
                            jira_key = str(jira_key)
                            issue = get_issue_by_jira_key(conn, project_id, jira_key)

                if not issue:
                    continue

                issue_id = int(issue["id"])

                order_no = None
                action = None
                inp = None
                expected = None
                group_no = None

                # Preconditions (Test Case 단위, 마지막 값이 우선)
                idx_pre = col_idx.get("preconditions")
                if idx_pre is not None and 0 <= idx_pre < len(row):
                    pre_val = row[idx_pre]
                    if pre_val not in (None, ""):
                        preconditions_by_issue_id[issue_id] = str(pre_val)

                idx_group = col_idx.get("group_no")
                if idx_group is not None and 0 <= idx_group < len(row):
                    group_no = row[idx_group]

                idx_order = col_idx.get("order_no")
                if idx_order is not None and 0 <= idx_order < len(row):
                    order_no = row[idx_order]
                idx_action = col_idx.get("action")
                if idx_action is not None and 0 <= idx_action < len(row):
                    action = row[idx_action]
                idx_input = col_idx.get("input")
                if idx_input is not None and 0 <= idx_input < len(row):
                    inp = row[idx_input]
                idx_expected = col_idx.get("expected")
                if idx_expected is not None and 0 <= idx_expected < len(row):
                    expected = row[idx_expected]

                steps_by_issue_id.setdefault(issue_id, []).append(
                    {
                        "group_no": int(group_no) if group_no is not None else 1,
                        "order_no": int(order_no) if order_no is not None else 0,
                        "action": str(action) if action is not None else "",
                        "input": str(inp) if inp is not None else "",
                        "expected": str(expected) if expected is not None else "",
                    }
                )

            for issue_id, steps in steps_by_issue_id.items():
                replace_steps_for_issue(conn, issue_id, steps)
                pre = preconditions_by_issue_id.get(issue_id)
                if pre is not None:
                    update_issue_fields(conn, issue_id, {"preconditions": pre})

    # --- Relations 시트 처리
    if "Relations" in wb.sheetnames:
        ws = wb["Relations"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"Relations 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("Relations", header)
            # src_issue_id 기준으로 relations 를 교체한다.
            rels_by_src_id: Dict[int, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue

                # 출발 이슈 찾기: jira_key → excel_key 순서로 매핑
                src_issue = None
                src_jira_idx = col_idx.get("src_jira_key")
                if src_jira_idx is not None and 0 <= src_jira_idx < len(row):
                    raw = row[src_jira_idx]
                    if raw not in (None, ""):
                        src_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not src_issue:
                    src_excel_idx = col_idx.get("src_excel_key")
                    if src_excel_idx is not None and 0 <= src_excel_idx < len(row):
                        raw_ref = row[src_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                src_issue = get_issue_by_id(conn, mapped_id)
                                if src_issue and src_issue.get("project_id") != project_id:
                                    src_issue = None
                if not src_issue:
                    continue

                # 도착 이슈 찾기: jira_key → excel_key 순서로 매핑
                dst_issue = None
                dst_jira_idx = col_idx.get("dst_jira_key")
                if dst_jira_idx is not None and 0 <= dst_jira_idx < len(row):
                    raw = row[dst_jira_idx]
                    if raw not in (None, ""):
                        dst_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not dst_issue:
                    dst_excel_idx = col_idx.get("dst_excel_key")
                    if dst_excel_idx is not None and 0 <= dst_excel_idx < len(row):
                        raw_ref = row[dst_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                dst_issue = get_issue_by_id(conn, mapped_id)
                                if dst_issue and dst_issue.get("project_id") != project_id:
                                    dst_issue = None
                if not dst_issue:
                    continue

                rel_type = ""
                type_idx = col_idx.get("relation_type")
                if type_idx is not None and 0 <= type_idx < len(row):
                    val = row[type_idx]
                    if val is not None:
                        rel_type = str(val)

                src_id = int(src_issue["id"])
                rels_by_src_id.setdefault(src_id, []).append(
                    {
                        "dst_issue_id": int(dst_issue["id"]),
                        "relation_type": rel_type or "",
                    }
                )

            for src_id, rel_list in rels_by_src_id.items():
                if rel_list:
                    replace_relations_for_issue(conn, src_id, rel_list)

    # --- TestPlanTestcases 시트 처리
    if "TestPlan_Testcases" in wb.sheetnames:
        ws = wb["TestPlan_Testcases"]
    elif "TestPlanTestcases" in wb.sheetnames:
        ws = wb["TestPlanTestcases"]
    else:
        ws = None
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"TestPlanTestcases 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("TestPlanTestcases", header)
            by_tp_id: Dict[int, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue

                # Test Plan 이슈 찾기 (jira_key → excel_key)
                tp_issue = None
                tp_jira_idx = col_idx.get("testplan_jira_key")
                if tp_jira_idx is not None and 0 <= tp_jira_idx < len(row):
                    raw = row[tp_jira_idx]
                    if raw not in (None, ""):
                        tp_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not tp_issue:
                    tp_excel_idx = col_idx.get("testplan_excel_key")
                    if tp_excel_idx is not None and 0 <= tp_excel_idx < len(row):
                        raw_ref = row[tp_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                tp_issue = get_issue_by_id(conn, mapped_id)
                                if tp_issue and tp_issue.get("project_id") != project_id:
                                    tp_issue = None
                if not tp_issue:
                    continue

                # Test Case 이슈 찾기 (jira_key → excel_key)
                tc_issue = None
                tc_jira_idx = col_idx.get("testcase_jira_key")
                if tc_jira_idx is not None and 0 <= tc_jira_idx < len(row):
                    raw = row[tc_jira_idx]
                    if raw not in (None, ""):
                        tc_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not tc_issue:
                    tc_excel_idx = col_idx.get("testcase_excel_key")
                    if tc_excel_idx is not None and 0 <= tc_excel_idx < len(row):
                        raw_ref = row[tc_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                tc_issue = get_issue_by_id(conn, mapped_id)
                                if tc_issue and tc_issue.get("project_id") != project_id:
                                    tc_issue = None
                if not tc_issue:
                    continue

                order_idx = col_idx.get("order_no")
                order_no = 0
                if order_idx is not None and 0 <= order_idx < len(row):
                    val = row[order_idx]
                    if val is not None:
                        try:
                            order_no = int(val)
                        except Exception:
                            order_no = 0

                tp_id = int(tp_issue["id"])
                by_tp_id.setdefault(tp_id, []).append(
                    {"testcase_id": int(tc_issue["id"]), "order_no": order_no}
                )

            for tp_id, mappings in by_tp_id.items():
                if not mappings:
                    continue
                records: List[Dict[str, Any]] = []
                for m in mappings:
                    records.append(
                        {
                            "testplan_id": tp_id,
                            "testcase_id": m["testcase_id"],
                            "order_no": m["order_no"],
                        }
                    )
                if records:
                    replace_testplan_testcases(conn, tp_id, records)

    # --- TestExecutions 시트 처리
    if "Test_Executions" in wb.sheetnames:
        ws = wb["Test_Executions"]
    elif "TestExecutions" in wb.sheetnames:
        ws = wb["TestExecutions"]
    else:
        ws = None
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"TestExecutions 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("TestExecutions", header)
            for row in rows[1:]:
                if not any(row):
                    continue

                # Test Execution 이슈 찾기 (jira_key → excel_key)
                te_issue = None
                key_idx = col_idx.get("testexecution_jira_key")
                if key_idx is not None and 0 <= key_idx < len(row):
                    raw = row[key_idx]
                    if raw not in (None, ""):
                        te_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not te_issue:
                    excel_idx = col_idx.get("testexecution_excel_key")
                    if excel_idx is not None and 0 <= excel_idx < len(row):
                        raw_ref = row[excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                te_issue = get_issue_by_id(conn, mapped_id)
                                if te_issue and te_issue.get("project_id") != project_id:
                                    te_issue = None
                if not te_issue:
                    continue

                env = None
                start_date = None
                end_date = None
                result_val = None
                executed_by = None

                idx_env = col_idx.get("environment")
                if idx_env is not None and 0 <= idx_env < len(row):
                    env = row[idx_env]
                idx_start = col_idx.get("start_date")
                if idx_start is not None and 0 <= idx_start < len(row):
                    start_date = row[idx_start]
                idx_end = col_idx.get("end_date")
                if idx_end is not None and 0 <= idx_end < len(row):
                    end_date = row[idx_end]
                idx_result = col_idx.get("result")
                if idx_result is not None and 0 <= idx_result < len(row):
                    result_val = row[idx_result]
                idx_exec = col_idx.get("executed_by")
                if idx_exec is not None and 0 <= idx_exec < len(row):
                    executed_by = row[idx_exec]

                te_row = get_or_create_testexecution_for_issue(conn, te_issue["id"])
                meta = {}
                if env is not None:
                    meta["environment"] = str(env)
                if start_date is not None:
                    meta["start_date"] = str(start_date)
                if end_date is not None:
                    meta["end_date"] = str(end_date)
                if result_val is not None:
                    meta["result"] = str(result_val)
                if executed_by is not None:
                    meta["executed_by"] = str(executed_by)
                if meta:
                    update_testexecution_for_issue(conn, te_issue["id"], meta)

    # --- TestcaseExecutions 시트 처리
    if "Testcase_Executions" in wb.sheetnames:
        ws = wb["Testcase_Executions"]
    elif "TestcaseExecutions" in wb.sheetnames:
        ws = wb["TestcaseExecutions"]
    else:
        ws = None
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"TestcaseExecutions 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("TestcaseExecutions", header)
            by_te_id: Dict[int, List[Dict[str, Any]]] = {}
            for row in rows[1:]:
                if not any(row):
                    continue

                # TE 이슈 찾기 (jira_key → excel_key)
                te_issue = None
                te_jira_idx = col_idx.get("testexecution_jira_key")
                if te_jira_idx is not None and 0 <= te_jira_idx < len(row):
                    raw = row[te_jira_idx]
                    if raw not in (None, ""):
                        te_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not te_issue:
                    te_excel_idx = col_idx.get("testexecution_excel_key")
                    if te_excel_idx is not None and 0 <= te_excel_idx < len(row):
                        raw_ref = row[te_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                te_issue = get_issue_by_id(conn, mapped_id)
                                if te_issue and te_issue.get("project_id") != project_id:
                                    te_issue = None
                if not te_issue:
                    continue

                # TC 이슈 찾기 (jira_key → excel_key)
                tc_issue = None
                tc_jira_idx = col_idx.get("testcase_jira_key")
                if tc_jira_idx is not None and 0 <= tc_jira_idx < len(row):
                    raw = row[tc_jira_idx]
                    if raw not in (None, ""):
                        tc_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not tc_issue:
                    tc_excel_idx = col_idx.get("testcase_excel_key")
                    if tc_excel_idx is not None and 0 <= tc_excel_idx < len(row):
                        raw_ref = row[tc_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                tc_issue = get_issue_by_id(conn, mapped_id)
                                if tc_issue and tc_issue.get("project_id") != project_id:
                                    tc_issue = None
                if not tc_issue:
                    continue

                order_no = 0
                assignee = ""
                result_val = ""
                env = ""
                time_val = None
                defects = ""
                tce_test_key = ""

                idx_order = col_idx.get("order_no")
                if idx_order is not None and 0 <= idx_order < len(row):
                    val = row[idx_order]
                    if val is not None:
                        try:
                            order_no = int(val)
                        except Exception:
                            order_no = 0
                idx_assignee = col_idx.get("assignee")
                if idx_assignee is not None and 0 <= idx_assignee < len(row):
                    val = row[idx_assignee]
                    if val is not None:
                        assignee = str(val)
                idx_res = col_idx.get("result")
                if idx_res is not None and 0 <= idx_res < len(row):
                    val = row[idx_res]
                    if val is not None:
                        result_val = str(val)
                idx_env = col_idx.get("rtm_environment")
                if idx_env is not None and 0 <= idx_env < len(row):
                    val = row[idx_env]
                    if val is not None:
                        env = str(val)
                idx_time = col_idx.get("actual_time")
                if idx_time is not None and 0 <= idx_time < len(row):
                    val = row[idx_time]
                    if val not in (None, ""):
                        try:
                            time_val = int(str(val).split(".")[0])
                        except Exception:
                            time_val = None
                idx_def = col_idx.get("defects")
                if idx_def is not None and 0 <= idx_def < len(row):
                    val = row[idx_def]
                    if val is not None:
                        defects = str(val)
                idx_tce_key = col_idx.get("tce_test_key")
                if idx_tce_key is not None and 0 <= idx_tce_key < len(row):
                    val = row[idx_tce_key]
                    if val is not None:
                        tce_test_key = str(val)

                te_id = int(te_issue["id"])
                tc_id = int(tc_issue["id"])
                by_te_id.setdefault(te_id, []).append(
                    {
                        "testcase_id": tc_id,
                        "order_no": order_no,
                        "assignee": assignee,
                        "result": result_val,
                        "actual_time": time_val,
                        "rtm_environment": env,
                        "defects": defects,
                        "tce_test_key": tce_test_key,
                    }
                )

            for te_id, items in by_te_id.items():
                te_row = get_or_create_testexecution_for_issue(conn, te_id)
                records: List[Dict[str, Any]] = []
                for item in items:
                    rec: Dict[str, Any] = {
                        "testcase_id": item["testcase_id"],
                        "order_no": item["order_no"],
                        "assignee": item["assignee"],
                        "result": item["result"],
                        "rtm_environment": item["rtm_environment"],
                        "defects": item["defects"],
                    }
                    if item.get("actual_time") is not None:
                        rec["actual_time"] = item["actual_time"]
                    if item.get("tce_test_key"):
                        rec["tce_test_key"] = item["tce_test_key"]
                    records.append(rec)
                if records:
                    replace_testcase_executions(conn, te_row["id"], records)

    # --- TestcaseStepExecutions 시트 처리 (옵션)
    if "TestcaseStep_Executions" in wb.sheetnames:
        ws = wb["TestcaseStep_Executions"]
    elif "TestcaseStepExecutions" in wb.sheetnames:
        ws = wb["TestcaseStepExecutions"]
    else:
        ws = None
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            step_index += 1
            _progress(f"TestcaseStepExecutions 시트 처리 중 ({len(rows) - 1} data rows)", step_index, total_steps)
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = _get_col_index("TestcaseStepExecutions", header)

            # 캐시: (testcase_id) -> {(group_no, order_no): step_id}
            step_cache: Dict[int, Dict[tuple[int, int], int]] = {}
            # 캐시: (te_issue_id, tc_issue_id) -> testcase_execution_id
            tce_cache: Dict[tuple[int, int], int] = {}
            records_by_tce: Dict[int, List[Dict[str, Any]]] = {}

            cur = conn.cursor()

            for row in rows[1:]:
                if not any(row):
                    continue
                g_idx = col_idx.get("group_no")
                o_idx = col_idx.get("order_no")
                if g_idx is None or o_idx is None:
                    continue

                # TE 이슈 찾기 (jira_key → excel_key)
                te_issue = None
                te_jira_idx = col_idx.get("testexecution_jira_key")
                if te_jira_idx is not None and 0 <= te_jira_idx < len(row):
                    raw = row[te_jira_idx]
                    if raw not in (None, ""):
                        te_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not te_issue:
                    te_excel_idx = col_idx.get("testexecution_excel_key")
                    if te_excel_idx is not None and 0 <= te_excel_idx < len(row):
                        raw_ref = row[te_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                te_issue = get_issue_by_id(conn, mapped_id)
                                if te_issue and te_issue.get("project_id") != project_id:
                                    te_issue = None

                # TC 이슈 찾기 (jira_key → excel_key)
                tc_issue = None
                tc_jira_idx = col_idx.get("testcase_jira_key")
                if tc_jira_idx is not None and 0 <= tc_jira_idx < len(row):
                    raw = row[tc_jira_idx]
                    if raw not in (None, ""):
                        tc_issue = get_issue_by_jira_key(conn, project_id, str(raw))
                if not tc_issue:
                    tc_excel_idx = col_idx.get("testcase_excel_key")
                    if tc_excel_idx is not None and 0 <= tc_excel_idx < len(row):
                        raw_ref = row[tc_excel_idx]
                        if raw_ref not in (None, ""):
                            ref = str(raw_ref).strip()
                            mapped_id = excel_issue_ref.get(ref)
                            if mapped_id is not None:
                                tc_issue = get_issue_by_id(conn, mapped_id)
                                if tc_issue and tc_issue.get("project_id") != project_id:
                                    tc_issue = None

                if not te_issue or not tc_issue:
                    continue

                try:
                    group_no = int(row[g_idx]) if row[g_idx] not in (None, "") else 1
                except Exception:
                    group_no = 1
                try:
                    order_no = int(row[o_idx]) if row[o_idx] not in (None, "") else 1
                except Exception:
                    order_no = 1

                # 상태/결과/증빙
                status = ""
                actual_result = ""
                evidence = ""
                s_idx = col_idx.get("status")
                if s_idx is not None and 0 <= s_idx < len(row):
                    val = row[s_idx]
                    if val is not None:
                        status = str(val)
                ar_idx = col_idx.get("actual_result")
                if ar_idx is not None and 0 <= ar_idx < len(row):
                    val = row[ar_idx]
                    if val is not None:
                        actual_result = str(val)
                ev_idx = col_idx.get("evidence")
                if ev_idx is not None and 0 <= ev_idx < len(row):
                    val = row[ev_idx]
                    if val is not None:
                        evidence = str(val)

                # TE/TC 이슈 → TCE 찾기
                te_row = get_or_create_testexecution_for_issue(conn, te_issue["id"])
                te_id = te_row["id"]
                tc_id = tc_issue["id"]

                # TestcaseExecution id 찾기/캐시
                pair = (te_id, tc_id)
                tce_id = tce_cache.get(pair)
                if tce_id is None:
                    cur.execute(
                        """
                        SELECT id FROM testcase_executions
                         WHERE testexecution_id = ? AND testcase_id = ?
                         ORDER BY order_no ASC, id ASC
                        """,
                        (te_id, tc_id),
                    )
                    r = cur.fetchone()
                    if not r:
                        continue
                    tce_id = int(r[0])
                    tce_cache[pair] = tce_id

                # testcase_step_id 찾기/캐시
                step_map = step_cache.get(tc_id)
                if step_map is None:
                    step_map = {}
                    steps = get_steps_for_issue(conn, tc_id)
                    for s in steps:
                        try:
                            g = int(s.get("group_no", 1) or 1)
                            o = int(s.get("order_no", 1) or 1)
                            sid = int(s.get("id"))
                        except Exception:
                            continue
                        step_map[(g, o)] = sid
                    step_cache[tc_id] = step_map
                step_id = step_map.get((group_no, order_no))
                if not step_id:
                    continue

                records_by_tce.setdefault(tce_id, []).append(
                    {
                        "testcase_step_id": step_id,
                        "status": status,
                        "actual_result": actual_result,
                        "evidence": evidence,
                    }
                )

            # TCE 별로 Step 실행 상태 저장
            for tce_id, recs in records_by_tce.items():
                replace_step_executions_for_tce(conn, tce_id, recs)
