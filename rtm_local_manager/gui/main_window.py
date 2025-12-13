
"""
gui/main_window.py - PySide6 GUI for RTM Local Manager.

This file provides:
- MainWindow: QMainWindow with left(Local)/right(JIRA) split view
- PanelWidget: reusable panel containing tree + issue tabs
- IssueTabWidget: RTM-style tabbed issue view (wireframe)
- Simple binding from DB folder/issue tree to QTreeView using QStandardItemModel
"""

from __future__ import annotations

import sys
import json
from typing import Dict, Any, List



from PySide6.QtCore import Qt, QEvent, QItemSelectionModel
from PySide6.QtGui import QStandardItemModel, QStandardItem, QAction, QKeySequence, QShortcut, QColor
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QSplitter,
    QTreeView,
    QTabWidget,
    QTabBar,
    QLabel,
    QPushButton,
    QToolBar,
    QStatusBar,
    QFileDialog,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QAbstractItemView,
    QStyle,
    QLineEdit,
    QDialog,
    QDialogButtonBox,
    QListWidget,
    QListWidgetItem,
    QRadioButton,
    QButtonGroup,
    QComboBox,
    QFormLayout,
    QInputDialog,
    QToolButton,
)



from backend.db import (
    get_connection,
    init_db,
    get_or_create_project,
    fetch_folder_tree,
    Project,
    get_issue_by_id,
    get_issue_by_jira_key,
    update_issue_fields,
    get_steps_for_issue,
    replace_steps_for_issue,
    get_relations_for_issue,
    replace_relations_for_issue,
    get_testplan_testcases,
    replace_testplan_testcases,
    get_or_create_testexecution_for_issue,
    update_testexecution_for_issue,
    get_testcase_executions,
    replace_testcase_executions,
    get_step_executions_for_tce,
    replace_step_executions_for_tce,
    get_testcase_execution_by_id,
    create_local_issue,
    soft_delete_issue,
    create_folder_node,
    delete_folder_if_empty,
    move_issue_to_folder,
    move_folder,
)
from backend import jira_mapping, excel_io
from backend.field_presets import load_presets, save_presets
from backend.local_settings import load_local_settings, save_local_settings
from backend.excel_mapping import load_mapping as load_excel_mapping, save_mapping as save_excel_mapping

from backend.jira_api import (
    load_config_from_file,
    save_config_to_file,
    JiraRTMClient,
    JiraConfig,
    DEFAULT_ENDPOINTS,
    DEFAULT_ENDPOINT_PARAMS,
)
from backend.logger import get_logger
from backend.sync import sync_tree, map_rtm_type_to_local



class IssueTabWidget(QTabWidget):
    """
    RTM 스타일 이슈 탭(Details / Steps / Requirements / Relations / Test Cases / Executions / Defects).

    - Details 탭: Summary, Status, Priority, Assignee, Reporter, Labels, Components,
                  RTM Environment, Due Date, Description 등을 편집 가능하도록 구성
    - 나머지 탭은 현재 와이어프레임 수준(라벨만)이며, 점진적으로 구현 예정
    """

    def __init__(self, parent=None):
        super().__init__(parent)

        self._current_issue: Dict[str, Any] | None = None
        self._current_details_columns: int = 2
        self._detail_label_widgets: list[QLabel] = []
        # Activity 영역에서 사용할 JIRA 댓글 원본 목록 캐시
        self._activity_comments: list[Dict[str, Any]] = []

        self.details_tab = QWidget()
        self._init_details_tab()

        self.steps_tab = QWidget()
        self._init_steps_tab()

        self.requirements_tab = QWidget()
        self._init_requirements_tab()

        self.relations_tab = QWidget()
        self._init_relations_tab()

        self.testcases_tab = QWidget()
        self._init_testcases_tab()

        self.executions_tab = QWidget()
        self._init_executions_tab()

        self.defects_tab = QWidget()
        self._init_defects_tab()

        self.addTab(self.details_tab, "Details")
        self.addTab(self.steps_tab, "Steps")
        self.addTab(self.requirements_tab, "Requirements")
        self.addTab(self.relations_tab, "Relations")
        self.addTab(self.testcases_tab, "Test Cases")
        self.addTab(self.executions_tab, "Executions")
        self.addTab(self.defects_tab, "Defects")

    # ------------------------------------------------------------------ Details

    def _init_details_tab(self):
        from PySide6.QtWidgets import (
            QGridLayout,
            QLineEdit,
            QTextEdit,
            QVBoxLayout,
            QHBoxLayout,
            QGroupBox,
            QPushButton,
            QLabel,
            QListWidget,
            QComboBox,
        )

        main_layout = QVBoxLayout()
        # 필드 로딩 중에는 dirty 플래그가 올라가지 않도록 제어하기 위한 플래그
        self._suppress_dirty: bool = False
        self._details_grid = QGridLayout()

        # issues 테이블 스키마 기반 필드들
        self.ed_local_id = QLineEdit()
        self.ed_jira_key = QLineEdit()
        self.ed_issue_type = QLineEdit()
        self.ed_summary = QLineEdit()
        # 일부 필드는 JIRA 메타데이터(상태/우선순위 등)에서 선택할 수 있도록 콤보박스로 구성
        self.ed_status = QComboBox()
        self.ed_status.setEditable(True)
        self.ed_priority = QComboBox()
        self.ed_priority.setEditable(True)
        self.ed_assignee = QLineEdit()
        self.ed_reporter = QLineEdit()
        self.ed_labels = QLineEdit()
        self.ed_components = QComboBox()
        self.ed_components.setEditable(True)
        self.ed_security_level = QLineEdit()
        self.ed_fix_versions = QComboBox()
        self.ed_fix_versions.setEditable(True)
        self.ed_affects_versions = QComboBox()
        self.ed_affects_versions.setEditable(True)
        self.ed_epic_link = QLineEdit()
        self.ed_sprint = QLineEdit()
        # RTM Environment 도 사전 정의 목록에서 선택할 수 있도록 콤보박스로 구성
        self.ed_rtm_env = QComboBox()
        self.ed_rtm_env.setEditable(True)
        self.ed_due_date = QLineEdit()
        self.ed_created = QLineEdit()
        self.ed_updated = QLineEdit()
        self.ed_attachments = QLineEdit()
        self.txt_description = QTextEdit()

        # 생성/수정 일시는 보통 읽기 전용
        self.ed_local_id.setReadOnly(True)
        self.ed_jira_key.setReadOnly(True)
        self.ed_issue_type.setReadOnly(True)
        self.ed_created.setReadOnly(True)
        self.ed_updated.setReadOnly(True)

        # 동적 다단 레이아웃을 위한 필드 정의 (Description 제외)
        # JIRA 의 기본 요약 필드(summary)도 독립 입력 필드로 노출한다.
        self._details_fields: list[tuple[str, QWidget]] = [
            ("Local ID", self.ed_local_id),
            ("JIRA Key", self.ed_jira_key),
            ("Issue Type", self.ed_issue_type),
            ("Summary", self.ed_summary),
            ("Status", self.ed_status),
            ("Priority", self.ed_priority),
            ("Assignee", self.ed_assignee),
            ("Reporter", self.ed_reporter),
            ("RTM Environment", self.ed_rtm_env),
            ("Components", self.ed_components),
            ("Labels", self.ed_labels),
            ("Fix Versions", self.ed_fix_versions),
            ("Affects Versions", self.ed_affects_versions),
            ("Epic Link", self.ed_epic_link),
            ("Sprint", self.ed_sprint),
            ("Security Level", self.ed_security_level),
            ("Due Date", self.ed_due_date),
            ("Created", self.ed_created),
            ("Updated", self.ed_updated),
            ("Attachments", self.ed_attachments),
        ]

        # 초기에는 2단 레이아웃으로 구성
        self._current_details_columns = 2
        self._rebuild_details_grid(self._current_details_columns)
        main_layout.addLayout(self._details_grid)

        # Description 은 하단에 단일 열로 크게 배치
        main_layout.addWidget(QLabel("Description"))
        main_layout.addWidget(self.txt_description)

        # Activity (local or JIRA comments / history)
        activity_group = QGroupBox("Activity")
        activity_layout = QVBoxLayout(activity_group)
        activity_btn_row = QHBoxLayout()
        self.btn_refresh_activity = QPushButton("Load Activity from JIRA")
        self.btn_add_comment = QPushButton("Add Comment...")
        self.btn_edit_comment = QPushButton("Edit Last Comment...")
        self.btn_delete_comment = QPushButton("Delete Last Comment")
        activity_btn_row.addWidget(self.btn_refresh_activity)
        activity_btn_row.addWidget(self.btn_add_comment)
        activity_btn_row.addWidget(self.btn_edit_comment)
        activity_btn_row.addWidget(self.btn_delete_comment)
        activity_btn_row.addStretch()
        self.txt_activity = QTextEdit()
        self.txt_activity.setReadOnly(False)
        self.txt_activity.setPlaceholderText("이슈별 Activity(로컬 메모 또는 JIRA Comments / History)를 표시합니다.")
        activity_layout.addLayout(activity_btn_row)
        activity_layout.addWidget(self.txt_activity)

        main_layout.addWidget(activity_group)

        # Attachments
        attachments_group = QGroupBox("Attachments")
        attachments_layout = QVBoxLayout(attachments_group)
        attachments_btn_row = QHBoxLayout()
        self.btn_upload_attachment = QPushButton("Upload...")
        self.btn_delete_attachment = QPushButton("Delete")
        self.btn_open_attachment = QPushButton("Open")
        attachments_btn_row.addWidget(self.btn_upload_attachment)
        attachments_btn_row.addWidget(self.btn_delete_attachment)
        attachments_btn_row.addWidget(self.btn_open_attachment)
        attachments_btn_row.addStretch()
        attachments_layout.addLayout(attachments_btn_row)

        self.attachments_list = QListWidget()
        attachments_layout.addWidget(self.attachments_list)

        main_layout.addWidget(attachments_group)

        self.details_tab.setLayout(main_layout)

        # Details 탭 필드 변경 시 현재 이슈를 "편집됨(unsaved)" 상태로 표시하기 위한 핸들러
        # (JIRA 사용 가능 여부와 무관하게 항상 연결한다.)
        def _mark_dirty_from_details():
            # set_issue() 등에서 프로그램적으로 값을 채우는 동안에는 더티로 보지 않는다.
            if getattr(self, "_suppress_dirty", False):
                return
            main_win = self.window()
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()

        # 읽기 전용 필드를 제외한 주요 편집 필드에 변경 감지 연결
        # QLineEdit 기반 필드들
        editable_edits = [
            self.ed_summary,
            self.ed_assignee,
            self.ed_reporter,
            self.ed_labels,
            self.ed_security_level,
            self.ed_epic_link,
            self.ed_sprint,
            self.ed_due_date,
            self.ed_attachments,
        ]
        for ed in editable_edits:
            ed.textEdited.connect(_mark_dirty_from_details)
        # 콤보박스 및 Description 변경 감지
        self.ed_status.currentTextChanged.connect(_mark_dirty_from_details)
        self.ed_priority.currentTextChanged.connect(_mark_dirty_from_details)
        self.ed_rtm_env.currentTextChanged.connect(_mark_dirty_from_details)
        self.ed_components.currentTextChanged.connect(_mark_dirty_from_details)
        self.ed_fix_versions.currentTextChanged.connect(_mark_dirty_from_details)
        self.ed_affects_versions.currentTextChanged.connect(_mark_dirty_from_details)
        self.txt_description.textChanged.connect(_mark_dirty_from_details)

    def apply_field_options(self, options: Dict[str, List[str]]) -> None:
        """
        JIRA 메타데이터(상태/우선순위 등)에서 가져온 필드 옵션을
        Details 탭 콤보박스(예: Status / Priority)에 적용한다.
        (dirty 플래그 연결은 __init__ 에서 한 번만 설정한다.)
        """
        # Status
        status_list = options.get("status") or []
        self.ed_status.blockSignals(True)
        current_status = self.ed_status.currentText()
        self.ed_status.clear()
        self.ed_status.addItem("")  # 빈 값 허용
        for name in status_list:
            self.ed_status.addItem(name)
        # 기존 값이 있으면 그대로 유지
        if current_status:
            self.ed_status.setCurrentText(current_status)
        self.ed_status.blockSignals(False)

        # Priority
        pri_list = options.get("priority") or []
        self.ed_priority.blockSignals(True)
        current_pri = self.ed_priority.currentText()
        self.ed_priority.clear()
        self.ed_priority.addItem("")
        for name in pri_list:
            self.ed_priority.addItem(name)
        if current_pri:
            self.ed_priority.setCurrentText(current_pri)
        self.ed_priority.blockSignals(False)

        # RTM Environment (콤보박스로 구성)
        env_list = options.get("rtm_environment") or []
        self.ed_rtm_env.blockSignals(True)
        current_env = self.ed_rtm_env.currentText()
        self.ed_rtm_env.clear()
        self.ed_rtm_env.addItem("")
        for name in env_list:
            self.ed_rtm_env.addItem(name)
        if current_env:
            self.ed_rtm_env.setCurrentText(current_env)
        self.ed_rtm_env.blockSignals(False)

        # Components (콤보박스, 다중 값은 쉼표로 직접 입력)
        comp_list = options.get("components") or []
        self.ed_components.blockSignals(True)
        current_comp = self.ed_components.currentText()
        self.ed_components.clear()
        self.ed_components.addItem("")
        for name in comp_list:
            self.ed_components.addItem(name)
        if current_comp:
            self.ed_components.setCurrentText(current_comp)
        self.ed_components.blockSignals(False)

        # Versions (Fix / Affects)
        ver_list = options.get("versions") or []

        self.ed_fix_versions.blockSignals(True)
        cur_fix = self.ed_fix_versions.currentText()
        self.ed_fix_versions.clear()
        self.ed_fix_versions.addItem("")
        for name in ver_list:
            self.ed_fix_versions.addItem(name)
        if cur_fix:
            self.ed_fix_versions.setCurrentText(cur_fix)
        self.ed_fix_versions.blockSignals(False)

        self.ed_affects_versions.blockSignals(True)
        cur_aff = self.ed_affects_versions.currentText()
        self.ed_affects_versions.clear()
        self.ed_affects_versions.addItem("")
        for name in ver_list:
            self.ed_affects_versions.addItem(name)
        if cur_aff:
            self.ed_affects_versions.setCurrentText(cur_aff)
        self.ed_affects_versions.blockSignals(False)

    # ------------------------------------------------------------------ Details layout helpers (responsive columns)

    def _rebuild_details_grid(self, columns: int) -> None:
        """
        Details 상단 메타 필드를 주어진 columns 수(1~4)로 재배치한다.
        - 각 필드는 (Label, Editor) 쌍으로, 한 column 당 Label+Editor 를 차지한다.
        """
        if columns < 1:
            columns = 1
        if columns > 4:
            columns = 4

        grid = self._details_grid

        # 이전에 생성된 레이블 위젯은 완전히 제거 (레이아웃에서도, 부모 관계에서도)
        for lbl in getattr(self, "_detail_label_widgets", []):
            try:
                lbl.setParent(None)
                lbl.deleteLater()
            except Exception:
                pass
        self._detail_label_widgets = []

        # 기존 레이아웃 아이템 제거 (에디터는 나중에 다시 addWidget 으로 배치)
        while grid.count():
            grid.takeAt(0)

        # 필드 재배치
        row = 0
        col_count = columns * 2  # Label / Editor 쌍

        for idx, (label_text, editor) in enumerate(self._details_fields):
            col_group = idx % columns
            row = idx // columns
            label_col = col_group * 2
            editor_col = label_col + 1

            lbl = QLabel(label_text)
            lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid.addWidget(lbl, row, label_col)

            # 필드 입력란의 가로 폭이 지나치게 넓어지지 않도록,
            # 현재 탭 폭과 columns 를 기준으로 대략 2/3 수준으로 제한한다.
            try:
                # 한 column 이 차지하는 이론상 폭: self.width() / columns
                # 그 중 2/3 만 최대 폭으로 사용
                approx_col_width = max(80, int(self.width() / max(columns, 1)))
                max_editor_width = max(120, int(approx_col_width * 2 / 3))
                editor.setMaximumWidth(max_editor_width)
            except Exception:
                # width 계산 실패 시에는 기본 동작 유지
                pass

            grid.addWidget(editor, row, editor_col)
            self._detail_label_widgets.append(lbl)

        # 컬럼 폭 비율: label 는 좁게, field 는 넓게
        for c in range(col_count):
            if c % 2 == 0:  # label
                grid.setColumnStretch(c, 0)
            else:           # editor
                grid.setColumnStretch(c, 1)

        self._current_details_columns = columns

    def resizeEvent(self, event) -> None:
        """
        탭 위젯(각 내부 창)의 실제 폭에 따라 Details 상단 필드의 단 수(1~4단)를 동적으로 조정한다.
        """
        super().resizeEvent(event)

        # 이 IssueTabWidget 이 속한 내부 창(좌/우 패널)의 실제 폭을 기준으로 columns 결정
        details_width = self.width()

        # 폭 기준 임계값 (pixel)
        if details_width < 550:
            cols = 1
        elif details_width < 900:
            cols = 2
        elif details_width < 1300:
            cols = 3
        else:
            cols = 4

        if cols != getattr(self, "_current_details_columns", 2):
            self._rebuild_details_grid(cols)

    # ------------------------------------------------------------------ Other tabs (wireframe for now)

    def _init_steps_tab(self):
        from PySide6.QtWidgets import (
            QVBoxLayout,
            QHBoxLayout,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QLabel,
            QTextEdit,
            QSplitter,
            QWidget,
        )

        layout = QVBoxLayout()

        # 상단 설명
        layout.addWidget(QLabel("Preconditions & Test Case Steps (Group, Order, Action, Input, Expected)"))

        # 상단(Preconditions) / 하단(Steps 버튼 + 테이블) 을 분리해서 높이를 조절할 수 있도록 QSplitter 사용
        splitter = QSplitter(Qt.Vertical)

        # 상단: Preconditions 영역
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addWidget(QLabel("Preconditions"))
        self.txt_preconditions = QTextEdit()
        self.txt_preconditions.setPlaceholderText("Enter preconditions for this test case...")
        top_layout.addWidget(self.txt_preconditions)

        # 하단: 버튼 + Steps 테이블
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        btn_layout = QHBoxLayout()
        self.btn_add_group = QPushButton("Add Group")
        self.btn_delete_group = QPushButton("Delete Selected Group")
        self.btn_add_step = QPushButton("Add Step")
        self.btn_delete_step = QPushButton("Delete Selected Step")
        btn_layout.addWidget(self.btn_add_group)
        btn_layout.addWidget(self.btn_delete_group)
        btn_layout.addWidget(self.btn_add_step)
        btn_layout.addWidget(self.btn_delete_step)
        btn_layout.addStretch()
        bottom_layout.addLayout(btn_layout)

        # 테이블: Group / Order / Action / Input / Expected
        self.steps_table = QTableWidget()
        self.steps_table.setColumnCount(5)
        self.steps_table.setHorizontalHeaderLabels(["Group", "Order", "Action", "Input", "Expected"])
        self.steps_table.horizontalHeader().setStretchLastSection(True)
        bottom_layout.addWidget(self.steps_table)

        splitter.addWidget(top_widget)
        splitter.addWidget(bottom_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)

        layout.addWidget(splitter)

        # 시그널
        self.btn_add_group.clicked.connect(self._on_add_group_clicked)
        self.btn_delete_group.clicked.connect(self._on_delete_group_clicked)
        self.btn_add_step.clicked.connect(self._on_add_step_clicked)
        self.btn_delete_step.clicked.connect(self._on_delete_step_clicked)
        # 스텝 테이블 직접 편집도 더티로 인식
        self.steps_table.itemChanged.connect(self._on_steps_item_changed)

        # Preconditions 텍스트 변경도 더티로 인식
        self.txt_preconditions.textChanged.connect(self._on_preconditions_changed)

        self.steps_tab.setLayout(layout)

    # --------------------------- Steps UI helpers ---------------------------

    def _current_group_no(self) -> int:
        """현재 선택된 행의 group_no, 없으면 마지막 그룹 번호(또는 1)를 돌려준다."""
        row = self.steps_table.currentRow()
        if row >= 0:
            item = self.steps_table.item(row, 0)
            if item and item.text().strip():
                try:
                    return int(item.text())
                except ValueError:
                    pass
        # 선택이 없으면 현재 테이블에서 가장 큰 group_no 를 사용 (없으면 1)
        max_group = 0
        for r in range(self.steps_table.rowCount()):
            it = self.steps_table.item(r, 0)
            if not it or not it.text().strip():
                continue
            try:
                g = int(it.text())
                if g > max_group:
                    max_group = g
            except ValueError:
                continue
        return max_group if max_group > 0 else 1

    def _next_group_no(self) -> int:
        """새 그룹을 만들 때 사용할 group_no (현재 최대값 + 1, 없으면 1)."""
        max_group = 0
        for r in range(self.steps_table.rowCount()):
            it = self.steps_table.item(r, 0)
            if not it or not it.text().strip():
                continue
            try:
                g = int(it.text())
                if g > max_group:
                    max_group = g
            except ValueError:
                continue
        return max_group + 1 if max_group > 0 else 1

    def _renumber_orders_per_group(self) -> None:
        """각 group_no 별로 order_no 를 1부터 다시 매겨준다."""
        # group_no, row 인덱스를 모아서 그룹별로 정렬 후 order 재부여
        groups: Dict[int, List[int]] = {}
        for r in range(self.steps_table.rowCount()):
            item_group = self.steps_table.item(r, 0)
            try:
                g = int(item_group.text()) if item_group and item_group.text().strip() else 1
            except ValueError:
                g = 1
            groups.setdefault(g, []).append(r)
        for g, rows in groups.items():
            for idx, r in enumerate(rows, start=1):
                item_order = self.steps_table.item(r, 1)
                if item_order is None:
                    item_order = QTableWidgetItem(str(idx))
                    self.steps_table.setItem(r, 1, item_order)
                else:
                    item_order.setText(str(idx))

    def _on_add_group_clicked(self):
        """새 그룹을 추가하고 첫 스텝 한 줄을 만든다."""
        from PySide6.QtWidgets import QTableWidgetItem

        # 항상 새로운 그룹 번호 사용
        new_group = self._next_group_no()
        row = self.steps_table.rowCount()
        self.steps_table.insertRow(row)
        self.steps_table.setItem(row, 0, QTableWidgetItem(str(new_group)))
        self.steps_table.setItem(row, 1, QTableWidgetItem("1"))
        self._renumber_orders_per_group()

    def _on_delete_group_clicked(self):
        """선택된 행의 group_no 에 해당하는 모든 스텝을 삭제."""
        row = self.steps_table.currentRow()
        if row < 0:
            return
        item_group = self.steps_table.item(row, 0)
        if not item_group or not item_group.text().strip():
            return
        try:
            target_group = int(item_group.text())
        except ValueError:
            return

        rows_to_delete = [r for r in range(self.steps_table.rowCount()) if
                          (self.steps_table.item(r, 0) and self.steps_table.item(r, 0).text().strip() == str(target_group))]
        for r in reversed(rows_to_delete):
            self.steps_table.removeRow(r)
        self._renumber_orders_per_group()

    def _on_add_step_clicked(self):
        """현재 그룹에 새 Step 한 줄 추가."""
        from PySide6.QtWidgets import QTableWidgetItem

        group_no = self._current_group_no()
        # 현재 그룹의 마지막 행 뒤에 삽입
        insert_row = self.steps_table.rowCount()
        for r in range(self.steps_table.rowCount()):
            item_group = self.steps_table.item(r, 0)
            if not item_group or not item_group.text().strip():
                continue
            try:
                g = int(item_group.text())
            except ValueError:
                continue
            if g > group_no and r < insert_row:
                insert_row = r
                break

        self.steps_table.insertRow(insert_row)
        self.steps_table.setItem(insert_row, 0, QTableWidgetItem(str(group_no)))
        # order_no 는 나중에 _renumber_orders_per_group 에서 재계산
        self._renumber_orders_per_group()

    def _on_delete_step_clicked(self):
        # 선택된 행 삭제 (복수 선택 시 모두 삭제).
        selected = self.steps_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected}, reverse=True)
        for r in rows:
            self.steps_table.removeRow(r)
        self._renumber_orders_per_group()

    def _on_steps_item_changed(self, item):
        """Steps 테이블 내용이 변경되면 해당 Test Case 를 더티로 표시."""
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_preconditions_changed(self):
        """Preconditions 텍스트 변경 시 현재 Test Case 를 더티로 표시."""
        # set_issue() 에서 preconditions 를 로드하는 동안에는 _suppress_dirty 로 제어
        if getattr(self, "_suppress_dirty", False):
            return
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    # --------------------------- Steps binding helpers ---------------------------

    def load_steps(self, steps: List[Dict[str, Any]]):
        """DB 등에서 읽어온 steps 리스트를 테이블에 로드."""
        from PySide6.QtWidgets import QTableWidgetItem
        # 로딩 중에는 itemChanged 핸들러가 더티 플래그를 올리지 않도록 시그널 일시 차단
        self.steps_table.blockSignals(True)
        self.steps_table.setRowCount(0)
        for s in steps:
            row = self.steps_table.rowCount()
            self.steps_table.insertRow(row)
            group_val = str(s.get("group_no", 1))
            order_val = str(s.get("order_no", 1))
            self.steps_table.setItem(row, 0, QTableWidgetItem(group_val))
            self.steps_table.setItem(row, 1, QTableWidgetItem(order_val))
            self.steps_table.setItem(row, 2, QTableWidgetItem(s.get("action") or ""))
            self.steps_table.setItem(row, 3, QTableWidgetItem(s.get("input") or ""))
            self.steps_table.setItem(row, 4, QTableWidgetItem(s.get("expected") or ""))
        self.steps_table.blockSignals(False)
        self._renumber_orders_per_group()

    def collect_steps(self) -> List[Dict[str, Any]]:
        """테이블의 내용을 읽어 steps 리스트로 반환."""
        steps: List[Dict[str, Any]] = []
        rows = self.steps_table.rowCount()
        for r in range(rows):
            group_item = self.steps_table.item(r, 0)
            order_item = self.steps_table.item(r, 1)
            action_item = self.steps_table.item(r, 2)
            input_item = self.steps_table.item(r, 3)
            expected_item = self.steps_table.item(r, 4)
            try:
                group_no = int(group_item.text()) if group_item and group_item.text().strip() else 1
            except ValueError:
                group_no = 1
            try:
                order_no = int(order_item.text()) if order_item and order_item.text().strip() else 1
            except ValueError:
                order_no = 1
            steps.append(
                {
                    "group_no": group_no,
                    "order_no": order_no,
                    "action": action_item.text().strip() if action_item else "",
                    "input": input_item.text().strip() if input_item else "",
                    "expected": expected_item.text().strip() if expected_item else "",
                }
            )
        return steps

    def set_preconditions_text(self, text: str) -> None:
        if hasattr(self, "txt_preconditions"):
            self.txt_preconditions.setPlainText(text or "")

    def get_preconditions_text(self) -> str:
        if hasattr(self, "txt_preconditions"):
            return self.txt_preconditions.toPlainText().strip()
        return ""

    def _init_requirements_tab(self):
        from PySide6.QtWidgets import QVBoxLayout, QTableWidget, QTableWidgetItem, QLabel

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Linked Requirements (read-only, derived from relations)"))

        self.requirements_table = QTableWidget()
        self.requirements_table.setColumnCount(3)
        self.requirements_table.setHorizontalHeaderLabels(["Req ID", "Jira Key", "Summary"])
        self.requirements_table.horizontalHeader().setStretchLastSection(True)
        self.requirements_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.requirements_table)

        self.requirements_tab.setLayout(layout)

    def load_requirements(self, requirements: List[Dict[str, Any]]):
        """Relations 정보에서 필터링한 Requirements 리스트를 테이블에 로드."""
        self.requirements_table.setRowCount(0)
        for req in requirements:
            row = self.requirements_table.rowCount()
            self.requirements_table.insertRow(row)
            self.requirements_table.setItem(row, 0, QTableWidgetItem(str(req.get("dst_issue_id"))))
            self.requirements_table.setItem(row, 1, QTableWidgetItem(req.get("dst_jira_key") or ""))
            self.requirements_table.setItem(row, 2, QTableWidgetItem(req.get("dst_summary") or ""))

    def _init_relations_tab(self):
        from PySide6.QtWidgets import QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QLabel

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Relations (src = current issue)"))

        btn_layout = QHBoxLayout()
        self.btn_add_relation = QPushButton("Add Relation")
        self.btn_delete_relation = QPushButton("Delete Selected Relation")
        btn_layout.addWidget(self.btn_add_relation)
        btn_layout.addWidget(self.btn_delete_relation)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.relations_table = QTableWidget()
        self.relations_table.setColumnCount(4)
        self.relations_table.setHorizontalHeaderLabels(["Relation Type", "Dst Issue ID", "Dst Jira Key", "Dst Summary"])
        self.relations_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.relations_table)

        self.btn_add_relation.clicked.connect(self._on_add_relation_clicked)
        self.btn_delete_relation.clicked.connect(self._on_delete_relation_clicked)

        # Relations 테이블 직접 편집(예: Dst Issue ID, Type 수정) 도 더티로 인식
        self.relations_table.itemChanged.connect(self._on_relations_item_changed)

        self.relations_tab.setLayout(layout)

    def _on_add_relation_clicked(self):
        """
        Relations 추가 팝업:
        - 기존에 등록된 이슈 목록에서 선택 (Dst Issue)
        - Relation Type 선택
        - 또는 Web Link(URL + Text) 를 입력하여 Relations 테이블에 추가
        """
        # 메인 윈도우와 DB/프로젝트 정보 가져오기
        main_win = self.window()
        issues: List[Dict[str, Any]] = []
        if hasattr(main_win, "conn") and getattr(main_win, "project", None) is not None:
            try:
                cur = main_win.conn.cursor()
                cur.execute(
                    """
                    SELECT id, issue_type, jira_key, summary
                      FROM issues
                     WHERE project_id = ? AND is_deleted = 0
                     ORDER BY issue_type, jira_key, summary
                    """,
                    (main_win.project.id,),
                )
                rows = cur.fetchall()
                issues = [dict(row) for row in rows]
            except Exception as e:
                print(f"[WARN] Failed to load issues for relations dialog: {e}")

        dlg = QDialog(self)
        dlg.setWindowTitle("Add Relation")

        layout = QVBoxLayout(dlg)

        # 모드 선택: 이슈 / Web Link
        mode_group = QButtonGroup(dlg)
        rb_issue = QRadioButton("Link to existing issue")
        rb_weblink = QRadioButton("Web link")
        rb_issue.setChecked(True)
        mode_group.addButton(rb_issue)
        mode_group.addButton(rb_weblink)

        mode_layout = QHBoxLayout()
        mode_layout.addWidget(rb_issue)
        mode_layout.addWidget(rb_weblink)
        mode_layout.addStretch()
        layout.addLayout(mode_layout)

        # 이슈 목록
        issue_list = QListWidget()
        for iss in issues:
            key = iss.get("jira_key") or f"ID={iss.get('id')}"
            itype = (iss.get("issue_type") or "").upper()
            summary = iss.get("summary") or ""
            text = f"[{itype}] {key}"
            if summary:
                text += f" - {summary}"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, iss.get("id"))          # dst_issue_id
            item.setData(Qt.UserRole + 1, iss.get("jira_key") or "")
            item.setData(Qt.UserRole + 2, summary)
            issue_list.addItem(item)
        layout.addWidget(issue_list)

        # Web link 입력 영역
        url_layout = QFormLayout()
        url_edit = QLineEdit()
        text_edit = QLineEdit()
        url_layout.addRow("URL", url_edit)
        url_layout.addRow("Link Text", text_edit)
        layout.addLayout(url_layout)

        # Link type 선택 (JIRA issue link type 목록 사용)
        rel_type_combo = QComboBox()
        rel_type_combo.setEditable(True)
        link_type_names: List[str] = []
        if getattr(main_win, "jira_available", False) and getattr(main_win, "jira_client", None):
            try:
                link_types_json = main_win.jira_client.get_issue_link_types()
                for lt in link_types_json.get("issueLinkTypes", []):
                    name = lt.get("name")
                    if name:
                        link_type_names.append(str(name))
            except Exception as e_lt:
                print(f"[WARN] Failed to load JIRA issue link types: {e_lt}")
        # 중복 제거 및 정렬
        link_type_names = sorted(set(link_type_names))
        if not link_type_names:
            # JIRA에서 가져오지 못한 경우 기본값
            link_type_names = ["Relates", "Blocks", "Cloners", "Duplicates"]
        rel_type_combo.addItems(link_type_names + ["Web Link"])
        rel_layout = QHBoxLayout()
        # JIRA 용어에 맞게 Link Type 으로 표기
        rel_layout.addWidget(QLabel("Link Type:"))
        rel_layout.addWidget(rel_type_combo)
        layout.addLayout(rel_layout)

        # 버튼 박스
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(btn_box)

        def on_accept():
            row = self.relations_table.rowCount()
            self.relations_table.insertRow(row)

            rel_type = rel_type_combo.currentText().strip() or ""

            if rb_issue.isChecked():
                item = issue_list.currentItem()
                if not item:
                    # 선택된 이슈가 없으면 행을 추가하지 않고 종료
                    self.relations_table.removeRow(row)
                    dlg.reject()
                    return
                dst_id = item.data(Qt.UserRole)
                dst_key = item.data(Qt.UserRole + 1) or ""
                dst_summary = item.data(Qt.UserRole + 2) or ""

                self.relations_table.setItem(row, 0, QTableWidgetItem(rel_type))
                self.relations_table.setItem(row, 1, QTableWidgetItem(str(dst_id)))
                self.relations_table.setItem(row, 2, QTableWidgetItem(dst_key))
                self.relations_table.setItem(row, 3, QTableWidgetItem(dst_summary))
            else:
                # Web link 모드: 현재 DB 스키마상 dst_issue_id 기반으로만 저장되므로,
                # 우선 Relations UI 에서만 관리하고, dst_issue_id 는 비워 둔다.
                url = url_edit.text().strip()
                link_text = text_edit.text().strip() or url

                if not url:
                    self.relations_table.removeRow(row)
                    dlg.reject()
                    return

                if not rel_type:
                    rel_type = "Web Link"

                self.relations_table.setItem(row, 0, QTableWidgetItem(rel_type))
                self.relations_table.setItem(row, 1, QTableWidgetItem(""))  # dst_issue_id 비움
                self.relations_table.setItem(row, 2, QTableWidgetItem(url))
                self.relations_table.setItem(row, 3, QTableWidgetItem(link_text))

            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

        # Relations 편집 이후에는 현재 이슈를 더티로 표시
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_delete_relation_clicked(self):
        """선택된 relation 행 삭제."""
        selected = self.relations_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected}, reverse=True)
        for r in rows:
            self.relations_table.removeRow(r)

        # Relations 변경도 더티로 인식
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    # --------------------------- Relations binding helpers ---------------------------

    def load_relations(self, relations: List[Dict[str, Any]], link_types: List[str] | None = None):
        """DB에서 읽어온 relations 리스트를 테이블에 로드.

        link_types 가 주어지면 0열(relation_type)을 콤보박스로 구성하여 사용자가 목록에서 선택할 수 있게 한다.
        """
        from PySide6.QtWidgets import QComboBox

        # 로딩 중에는 itemChanged 핸들러가 더티 플래그를 올리지 않도록 시그널 일시 차단
        self.relations_table.blockSignals(True)
        self.relations_table.setRowCount(0)
        use_combo = bool(link_types)
        for rel in relations:
            row = self.relations_table.rowCount()
            self.relations_table.insertRow(row)

            rel_type = rel.get("relation_type") or ""
            if use_combo:
                cmb = QComboBox()
                cmb.setEditable(True)
                cmb.addItem("")
                for name in link_types or []:
                    cmb.addItem(name)
                cmb.setCurrentText(rel_type)
                self.relations_table.setCellWidget(row, 0, cmb)
            else:
                self.relations_table.setItem(row, 0, QTableWidgetItem(rel_type))

            self.relations_table.setItem(row, 1, QTableWidgetItem(str(rel.get("dst_issue_id"))))
            self.relations_table.setItem(row, 2, QTableWidgetItem(rel.get("dst_jira_key") or ""))
            self.relations_table.setItem(row, 3, QTableWidgetItem(rel.get("dst_summary") or ""))
        self.relations_table.blockSignals(False)

    def collect_relations(self) -> List[Dict[str, Any]]:
        """Relations 테이블에서 relation 리스트를 수집 (dst_issue_id, relation_type)."""
        rels: List[Dict[str, Any]] = []
        rows = self.relations_table.rowCount()
        for r in range(rows):
            w = self.relations_table.cellWidget(r, 0)
            if w is not None:
                # 콤보박스가 있는 경우
                from PySide6.QtWidgets import QComboBox

                if isinstance(w, QComboBox):
                    rel_type = w.currentText().strip()
                else:
                    rel_type = ""
            else:
                type_item = self.relations_table.item(r, 0)
                rel_type = type_item.text().strip() if type_item else ""

            dst_id_item = self.relations_table.item(r, 1)
            dst_id_text = dst_id_item.text().strip() if dst_id_item else ""
            if not dst_id_text:
                continue
            try:
                dst_id = int(dst_id_text)
            except ValueError:
                continue
            rels.append(
                {
                    "relation_type": rel_type,
                    "dst_issue_id": dst_id,
                }
            )
        return rels

    def _on_relations_item_changed(self, item):
        """Relations 테이블 셀 편집 시 현재 이슈를 더티로 표시."""
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _init_testcases_tab(self):
        from PySide6.QtWidgets import (
            QVBoxLayout,
            QHBoxLayout,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QLabel,
        )

        layout = QVBoxLayout()
        # 헤더 라벨은 이슈 타입에 따라 동적으로 변경된다.
        self.lbl_testcases_header = QLabel("")
        layout.addWidget(self.lbl_testcases_header)

        btn_layout = QHBoxLayout()
        # REQUIREMENT / DEFECT 등에서 사용하는 버튼
        self.btn_cover_by_tc = QPushButton("Cover by TC...")
        self.btn_create_tc = QPushButton("Create New Test Case...")
        # TEST_PLAN 에서 사용하는 버튼
        self.btn_add_tp_tc = QPushButton("Add Test Case")
        self.btn_del_tp_tc = QPushButton("Delete Selected")
        self.btn_edit_tp_order = QPushButton("Edit order")
        # TEST_CASE 탭에서 사용하는 버튼 (기존에는 Panel 헤더에 있었던 것)
        self.btn_execute_tc = QPushButton("Execute")
        self.btn_add_to_testplan = QPushButton("Add to Test Plan")
        self.btn_link_requirement = QPushButton("Link to Requirement")

        btn_layout.addWidget(self.btn_cover_by_tc)
        btn_layout.addWidget(self.btn_create_tc)
        btn_layout.addWidget(self.btn_add_tp_tc)
        btn_layout.addWidget(self.btn_del_tp_tc)
        btn_layout.addWidget(self.btn_edit_tp_order)
        btn_layout.addWidget(self.btn_execute_tc)
        btn_layout.addWidget(self.btn_add_to_testplan)
        btn_layout.addWidget(self.btn_link_requirement)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 단일 테이블을 이슈 타입에 따라 다르게 사용한다.
        self.testplan_tc_table = QTableWidget()
        self.testplan_tc_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.testplan_tc_table)

        # 시그널 연결
        self.btn_cover_by_tc.clicked.connect(self._on_cover_by_testcase_clicked)
        self.btn_create_tc.clicked.connect(self._on_create_testcase_clicked)
        self.btn_add_tp_tc.clicked.connect(self._on_add_tp_tc_clicked)
        self.btn_del_tp_tc.clicked.connect(self._on_del_tp_tc_clicked)
        self.btn_edit_tp_order.clicked.connect(self._on_toggle_tp_order_edit)

        # Test Plan order 편집 모드 상태
        self._tp_order_edit_mode: bool = False

        self.testcases_tab.setLayout(layout)

    def _on_add_tp_tc_clicked(self):
        """
        Test Plan - Test Case 매핑에 기존 Test Case 들을 추가한다.
        (Test Plan 이슈의 Test Cases 탭에서 'Add Test Case' 버튼)
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QDialogButtonBox,
            QTableWidget,
            QTableWidgetItem,
            QLabel,
        )

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return

        if getattr(main_win, "current_issue_id", None) is None:
            main_win.status_bar.showMessage("No Test Plan selected.")
            return

        issue = get_issue_by_id(main_win.conn, main_win.current_issue_id)
        if not issue or (issue.get("issue_type") or "").upper() != "TEST_PLAN":
            main_win.status_bar.showMessage("Current issue is not a TEST_PLAN.")
            return

        # 프로젝트 내 TEST_CASE 이슈 목록 조회
        cur = main_win.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary, priority, assignee, components, rtm_environment
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'TEST_CASE'
             ORDER BY jira_key, summary
            """,
            (main_win.project.id,),
        )
        rows = cur.fetchall()
        testcases = [dict(r) for r in rows]
        if not testcases:
            main_win.status_bar.showMessage("No TEST_CASE issues found in this project.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Add Test Case to Test Plan")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select Test Cases to add to this Test Plan:"))

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(
            ["Jira Key", "Summary", "Priority", "Assignee", "Components", "RTM Env"]
        )
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.MultiSelection)

        for tc in testcases:
            row_idx = table.rowCount()
            table.insertRow(row_idx)
            table.setItem(row_idx, 0, QTableWidgetItem(tc.get("jira_key") or ""))
            table.setItem(row_idx, 1, QTableWidgetItem(tc.get("summary") or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(tc.get("priority") or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(tc.get("assignee") or ""))
            table.setItem(row_idx, 4, QTableWidgetItem(tc.get("components") or ""))
            table.setItem(row_idx, 5, QTableWidgetItem(tc.get("rtm_environment") or ""))
            table.item(row_idx, 0).setData(Qt.UserRole, tc.get("id"))

        vbox.addWidget(table)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            selected_rows = {idx.row() for idx in table.selectedIndexes()}
            if not selected_rows:
                dlg.reject()
                return

            tp_id = int(main_win.current_issue_id)
            # 기존 매핑 읽기
            existing = get_testplan_testcases(main_win.conn, tp_id)
            existing_tc_ids = {
                int(r.get("testcase_id")) for r in existing if r.get("testcase_id")
            }
            records = [
                {
                    "testcase_id": int(r["testcase_id"]),
                    "order_no": int(r.get("order_no", 0) or 0),
                }
                for r in existing
            ]

            for r_idx in selected_rows:
                item = table.item(r_idx, 0)
                if not item:
                    continue
                tc_id = item.data(Qt.UserRole)
                if not tc_id or int(tc_id) in existing_tc_ids:
                    continue
                records.append({"testcase_id": int(tc_id), "order_no": 0})

            # order_no 재정렬
            for i, rec in enumerate(records, start=1):
                rec["order_no"] = i

            replace_testplan_testcases(main_win.conn, tp_id, records)

            # UI 갱신
            tp_records = get_testplan_testcases(main_win.conn, tp_id)
            if hasattr(main_win.left_panel.issue_tabs, "load_testplan_testcases"):
                main_win.left_panel.issue_tabs.load_testplan_testcases(tp_records)

            main_win.status_bar.showMessage(
                f"Added {len(selected_rows)} Test Case(s) to Test Plan."
            )
            # Test Plan 이 변경되었으므로 더티로 표시
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def _on_del_tp_tc_clicked(self):
        """선택된 매핑 행 삭제 및 DB 반영."""
        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return
        if getattr(main_win, "current_issue_id", None) is None:
            main_win.status_bar.showMessage("No Test Plan selected.")
            return

        selected = self.testplan_tc_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected}, reverse=True)
        for r in rows:
            self.testplan_tc_table.removeRow(r)

        # 남은 행 기준으로 order_no 재계산 후 DB 갱신
        try:
            tp_records = self.collect_testplan_testcases()
            # 순서 재정렬
            for i, rec in enumerate(tp_records, start=1):
                rec["order_no"] = i
            replace_testplan_testcases(main_win.conn, int(main_win.current_issue_id), tp_records)

            # UI 재로드
            new_records = get_testplan_testcases(
                main_win.conn, int(main_win.current_issue_id)
            )
            if hasattr(main_win.left_panel.issue_tabs, "load_testplan_testcases"):
                main_win.left_panel.issue_tabs.load_testplan_testcases(new_records)
        except Exception as e:
            print(f"[WARN] Failed to delete test plan testcases: {e}")

        # Test Plan 의 Test Case 매핑이 바뀌었으므로 더티로 표시
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_toggle_tp_order_edit(self):
        """
        Test Plan Test Cases 탭의 'Edit order' / 'Accept order' 토글 버튼.
        - 사용자는 Order 컬럼을 수정한 뒤, 'Accept order' 를 눌러 순서를 확정한다.
        """
        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return
        if getattr(main_win, "current_issue_id", None) is None:
            main_win.status_bar.showMessage("No Test Plan selected.")
            return

        # 첫 클릭: 편집 모드 진입
        if not self._tp_order_edit_mode:
            self._tp_order_edit_mode = True
            self.btn_edit_tp_order.setText("Accept order")
            main_win.status_bar.showMessage(
                "Edit the 'Order' column, then click 'Accept order' to save."
            )
            return

        # 두 번째 클릭: 현재 Order 값을 기준으로 순서 확정 및 DB 저장
        try:
            tp_id = int(main_win.current_issue_id)
            records = self.collect_testplan_testcases()
            # 사용자가 입력한 order_no 기준으로 정렬
            records.sort(key=lambda r: int(r.get("order_no", 0) or 0))
            # 1..N 으로 재번호 부여
            for i, rec in enumerate(records, start=1):
                rec["order_no"] = i

            replace_testplan_testcases(main_win.conn, tp_id, records)

            # UI 재로드
            new_records = get_testplan_testcases(main_win.conn, tp_id)
            if hasattr(main_win.left_panel.issue_tabs, "load_testplan_testcases"):
                main_win.left_panel.issue_tabs.load_testplan_testcases(new_records)

            main_win.status_bar.showMessage("Test Plan Test Cases order has been saved.")
        except Exception as e:
            print(f"[WARN] Failed to accept Test Plan order: {e}")
            main_win.status_bar.showMessage(f"Failed to save order: {e}")
        finally:
            self._tp_order_edit_mode = False
            self.btn_edit_tp_order.setText("Edit order")

        # 순서가 저장되었으므로, 현재 Test Plan 이 변경되었다고 표시
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    # --------------------------- Requirement: Cover by / Create Test Case --------

    def _on_cover_by_testcase_clicked(self):
        """
        REQUIREMENT 이슈에서 'Cover by Test Case...' 버튼을 눌렀을 때:
        - 현재 프로젝트의 TEST_CASE 이슈 목록을 보여주고
        - 선택된 Test Case 들을 현재 Requirement 와 relations 로 연결한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QDialogButtonBox,
            QTableWidgetItem,
            QLabel,
        )

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return

        if getattr(main_win, "current_issue_id", None) is None:
            main_win.status_bar.showMessage("No issue selected; cannot cover by test case.")
            return

        # 현재 이슈가 REQUIREMENT 인지 확인
        src_issue = get_issue_by_id(main_win.conn, main_win.current_issue_id)
        if not src_issue or (src_issue.get("issue_type") or "").upper() != "REQUIREMENT":
            main_win.status_bar.showMessage("Current issue is not a REQUIREMENT; 'Cover by Test Case' is disabled.")
            return

        # 프로젝트 내 TEST_CASE 이슈 목록 조회
        cur = main_win.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary, priority, assignee, components, rtm_environment
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'TEST_CASE'
             ORDER BY jira_key, summary
            """,
            (main_win.project.id,),
        )
        rows = cur.fetchall()
        testcases = [dict(r) for r in rows]

        if not testcases:
            main_win.status_bar.showMessage("No TEST_CASE issues found in this project.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Cover by Test Case")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select Test Cases to cover this Requirement:"))

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(
            ["Jira Key", "Summary", "Priority", "Assignee", "Components", "RTM Env"]
        )
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.MultiSelection)

        for tc in testcases:
            row_idx = table.rowCount()
            table.insertRow(row_idx)
            table.setItem(row_idx, 0, QTableWidgetItem(tc.get("jira_key") or ""))
            table.setItem(row_idx, 1, QTableWidgetItem(tc.get("summary") or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(tc.get("priority") or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(tc.get("assignee") or ""))
            table.setItem(row_idx, 4, QTableWidgetItem(tc.get("components") or ""))
            table.setItem(row_idx, 5, QTableWidgetItem(tc.get("rtm_environment") or ""))
            # 내부적으로 issue_id 저장
            table.item(row_idx, 0).setData(Qt.UserRole, tc.get("id"))

        vbox.addWidget(table)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            selected_rows = {idx.row() for idx in table.selectedIndexes()}
            if not selected_rows:
                dlg.reject()
                return

            # 기존 relations 를 가져와서, 선택된 Test Case 들을 추가한 새 리스트 구성
            existing_rels = get_relations_for_issue(main_win.conn, main_win.current_issue_id)
            existing_dst_ids = {int(r.get("dst_issue_id")) for r in existing_rels if r.get("dst_issue_id")}
            new_rels = [
                {"dst_issue_id": r.get("dst_issue_id"), "relation_type": r.get("relation_type") or ""}
                for r in existing_rels
            ]

            for r in selected_rows:
                item = table.item(r, 0)
                if not item:
                    continue
                tc_id = item.data(Qt.UserRole)
                if not tc_id or tc_id in existing_dst_ids:
                    continue
                new_rels.append({"dst_issue_id": int(tc_id), "relation_type": "Tests"})

            replace_relations_for_issue(main_win.conn, main_win.current_issue_id, new_rels)

            # UI 갱신: Relations / Requirements / Linked Test Cases
            rels = get_relations_for_issue(main_win.conn, main_win.current_issue_id)
            if hasattr(main_win.left_panel.issue_tabs, "load_relations"):
                link_types = main_win.jira_field_options.get("relation_types", [])
                main_win.left_panel.issue_tabs.load_relations(rels, link_types)
            if hasattr(main_win.left_panel.issue_tabs, "load_requirements"):
                reqs = [
                    r for r in rels if (r.get("dst_issue_type") or "").upper() == "REQUIREMENT"
                ]
                main_win.left_panel.issue_tabs.load_requirements(reqs)
            if hasattr(main_win.left_panel.issue_tabs, "load_linked_testcases"):
                tcs = [
                    r for r in rels if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"
                ]
                main_win.left_panel.issue_tabs.load_linked_testcases(tcs)

            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def _on_create_testcase_clicked(self):
        """
        REQUIREMENT 이슈에서 'Create New Test Case...' 버튼:
        - 간단한 팝업에서 Summary 등을 입력 받아
        - 동일 프로젝트 / 동일 폴더에 TEST_CASE 로컬 이슈를 생성하고
        - 현재 Requirement 와 relation 으로 연결한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QFormLayout,
            QLineEdit,
            QDialogButtonBox,
            QLabel,
        )

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return
        if getattr(main_win, "current_issue_id", None) is None:
            QMessageBox.warning(
                self,
                "Create Test Case",
                "No Requirement is selected.\n\n"
                "Please select a Requirement in the left tree first.",
            )
            return

        src_issue = get_issue_by_id(main_win.conn, main_win.current_issue_id)
        if not src_issue or (src_issue.get("issue_type") or "").upper() != "REQUIREMENT":
            QMessageBox.warning(
                self,
                "Create Test Case",
                "Current issue is not a REQUIREMENT.\n\n"
                "Select a Requirement in the tree to create a Test Case linked to it.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Create New Test Case")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("New Test Case will be created as local-only issue linked to this Requirement."))

        form = QFormLayout()
        ed_summary = QLineEdit()
        ed_summary.setPlaceholderText("Summary")
        ed_priority = QLineEdit()
        ed_assignee = QLineEdit()
        form.addRow("Summary*", ed_summary)
        form.addRow("Priority", ed_priority)
        form.addRow("Assignee", ed_assignee)
        vbox.addLayout(form)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            summary = ed_summary.text().strip()
            if not summary:
                main_win.status_bar.showMessage("Summary is required to create a Test Case.")
                return

            folder_id = src_issue.get("folder_id")
            from backend.db import create_local_issue, update_issue_fields
            new_tc_id = create_local_issue(
                main_win.conn,
                project_id=main_win.project.id,
                issue_type="TEST_CASE",
                folder_id=folder_id,
                summary=summary,
            )
            fields: Dict[str, Any] = {}
            if ed_priority.text().strip():
                fields["priority"] = ed_priority.text().strip()
            if ed_assignee.text().strip():
                fields["assignee"] = ed_assignee.text().strip()
            if fields:
                update_issue_fields(main_win.conn, new_tc_id, fields)

            # Requirement 와 새 Test Case 를 relation 으로 연결
            existing_rels = get_relations_for_issue(main_win.conn, main_win.current_issue_id)
            new_rels = [
                {"dst_issue_id": r.get("dst_issue_id"), "relation_type": r.get("relation_type") or ""}
                for r in existing_rels
            ]
            new_rels.append({"dst_issue_id": new_tc_id, "relation_type": "Tests"})
            replace_relations_for_issue(main_win.conn, main_win.current_issue_id, new_rels)

            # UI 갱신
            rels = get_relations_for_issue(main_win.conn, main_win.current_issue_id)
            if hasattr(main_win.left_panel.issue_tabs, "load_relations"):
                main_win.left_panel.issue_tabs.load_relations(rels)
            if hasattr(main_win.left_panel.issue_tabs, "load_requirements"):
                reqs = [
                    r for r in rels if (r.get("dst_issue_type") or "").upper() == "REQUIREMENT"
                ]
                main_win.left_panel.issue_tabs.load_requirements(reqs)
            if hasattr(main_win.left_panel.issue_tabs, "load_linked_testcases"):
                tcs = [
                    r for r in rels if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"
                ]
                main_win.left_panel.issue_tabs.load_linked_testcases(tcs)

            main_win.reload_local_tree()
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    # ------------------------------------------------------------------ Test Cases top-level actions

    def _selected_local_testcase_ids(self) -> list[int]:
        """
        좌측(Local) 트리에서 선택된 TEST_CASE 이슈들의 로컬 ID 리스트를 반환.
        현재 상단 탭이 Test Cases 일 때 Add to Test Plan / Link to Requirement 에서 사용한다.
        """
        ids: list[int] = []
        view = self.left_panel.tree_view
        model = view.model()
        if model is None:
            return ids
        for idx in view.selectedIndexes():
            if idx.column() != 0:
                continue
            item = model.itemFromIndex(idx)
            if not item:
                continue
            node_type = item.data(Qt.UserRole)
            if node_type != "ISSUE":
                continue
            issue_id = item.data(Qt.UserRole + 1)
            issue_type = (item.data(Qt.UserRole + 3) or "").upper()
            if not issue_id or issue_type != "TEST_CASE":
                continue
            try:
                ids.append(int(issue_id))
            except (TypeError, ValueError):
                continue
        return ids

    def on_add_testcases_to_testplan_clicked(self):
        """
        상단 Test Cases 탭에서 'Add to Test Plan' 버튼:
        - 트리에서 선택된 TEST_CASE 이슈들을 하나의 Test Plan 에 추가한다.
        """
        tc_ids = self._selected_local_testcase_ids()
        if not tc_ids:
            self.status_bar.showMessage("No Test Cases selected in the tree.")
            return

        # 프로젝트 내 TEST_PLAN 목록 조회
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'TEST_PLAN'
             ORDER BY jira_key, summary
            """,
            (self.project.id,),
        )
        rows = cur.fetchall()
        plans = [dict(r) for r in rows]
        if not plans:
            self.status_bar.showMessage("No TEST_PLAN issues found; cannot add Test Cases.")
            return

        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QDialogButtonBox,
            QListWidget,
            QListWidgetItem,
            QLabel,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Add to Test Plan")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select a Test Plan to add the selected Test Cases:"))

        lst = QListWidget()
        for p in plans:
            key = p.get("jira_key") or f"ID={p.get('id')}"
            text = key
            if p.get("summary"):
                text += f" - {p.get('summary')}"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, p.get("id"))
            lst.addItem(item)
        vbox.addWidget(lst)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            item = lst.currentItem()
            if not item:
                self.status_bar.showMessage("No Test Plan selected.")
                return
            tp_id = item.data(Qt.UserRole)
            if not tp_id:
                return

            # 기존 매핑을 읽고 선택된 Test Case 들을 추가
            existing = get_testplan_testcases(self.conn, int(tp_id))
            existing_tc_ids = {int(r.get("testcase_id")) for r in existing if r.get("testcase_id")}
            records = [
                {"testcase_id": int(r["testcase_id"]), "order_no": int(r.get("order_no", 0) or 0)}
                for r in existing
            ]

            for tc_id in tc_ids:
                if tc_id in existing_tc_ids:
                    continue
                records.append({"testcase_id": tc_id, "order_no": 0})

            # order_no 재정렬
            for idx, rec in enumerate(records, start=1):
                rec["order_no"] = idx

            replace_testplan_testcases(self.conn, int(tp_id), records)
            self.status_bar.showMessage(
                f"Added {len(tc_ids)} Test Case(s) to Test Plan (id={tp_id})."
            )
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def on_link_testcases_to_requirement_clicked(self):
        """
        상단 Test Cases 탭에서 'Link to Requirement' 버튼:
        - 트리에서 선택된 TEST_CASE 이슈들을 선택한 Requirement 와 relation 으로 연결한다.
        """
        tc_ids = self._selected_local_testcase_ids()
        if not tc_ids:
            self.status_bar.showMessage("No Test Cases selected in the tree.")
            return

        # 프로젝트 내 REQUIREMENT 목록 조회
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'REQUIREMENT'
             ORDER BY jira_key, summary
            """,
            (self.project.id,),
        )
        rows = cur.fetchall()
        reqs = [dict(r) for r in rows]
        if not reqs:
            self.status_bar.showMessage("No REQUIREMENT issues found; cannot link.")
            return

        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QDialogButtonBox,
            QListWidget,
            QListWidgetItem,
            QLabel,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Link to Requirement")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select a Requirement to link the selected Test Cases:"))

        lst = QListWidget()
        for r in reqs:
            key = r.get("jira_key") or f"ID={r.get('id')}"
            text = key
            if r.get("summary"):
                text += f" - {r.get('summary')}"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, r.get("id"))
            lst.addItem(item)
        vbox.addWidget(lst)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            item = lst.currentItem()
            if not item:
                self.status_bar.showMessage("No Requirement selected.")
                return
            req_id = item.data(Qt.UserRole)
            if not req_id:
                return

            # Requirement 를 src 로 하는 relations 를 읽고, 선택된 Test Case 들을 추가
            existing_rels = get_relations_for_issue(self.conn, int(req_id))
            existing_dst_ids = {
                int(r.get("dst_issue_id")) for r in existing_rels if r.get("dst_issue_id")
            }
            new_rels = [
                {"dst_issue_id": r.get("dst_issue_id"), "relation_type": r.get("relation_type") or ""}
                for r in existing_rels
            ]
            for tc_id in tc_ids:
                if tc_id in existing_dst_ids:
                    continue
                new_rels.append({"dst_issue_id": tc_id, "relation_type": "Tests"})

            replace_relations_for_issue(self.conn, int(req_id), new_rels)

            # 현재 선택 이슈/탭이 이 Requirement 이거나 관련 있을 경우, UI 갱신
            if self.current_issue_id == int(req_id):
                rels = get_relations_for_issue(self.conn, int(req_id))
                if hasattr(self.left_panel.issue_tabs, "load_relations"):
                    link_types = self.jira_field_options.get("relation_types", [])
                    self.left_panel.issue_tabs.load_relations(rels, link_types)
                if hasattr(self.left_panel.issue_tabs, "load_requirements"):
                    reqs2 = [
                        r
                        for r in rels
                        if (r.get("dst_issue_type") or "").upper() == "REQUIREMENT"
                    ]
                    self.left_panel.issue_tabs.load_requirements(reqs2)
                if hasattr(self.left_panel.issue_tabs, "load_linked_testcases"):
                    tcs2 = [
                        r
                        for r in rels
                        if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"
                    ]
                    self.left_panel.issue_tabs.load_linked_testcases(tcs2)

            self.status_bar.showMessage(
                f"Linked {len(tc_ids)} Test Case(s) to Requirement (id={req_id})."
            )
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    # --------------------------- Test Plan / Test Cases binding ------------------

    def load_testplan_testcases(self, records: List[Dict[str, Any]]):
        """DB에서 읽어온 Test Plan - Test Case 매핑을 테이블에 로드."""
        # TEST_PLAN 용 컬럼 구성 보장
        self.testplan_tc_table.setColumnCount(4)
        self.testplan_tc_table.setHorizontalHeaderLabels(
            ["Order", "Test Case ID", "Jira Key", "Summary"]
        )
        self.testplan_tc_table.setRowCount(0)
        for rec in records:
            row = self.testplan_tc_table.rowCount()
            self.testplan_tc_table.insertRow(row)
            self.testplan_tc_table.setItem(row, 0, QTableWidgetItem(str(rec.get("order_no", row + 1))))
            self.testplan_tc_table.setItem(row, 1, QTableWidgetItem(str(rec.get("testcase_id"))))
            self.testplan_tc_table.setItem(row, 2, QTableWidgetItem(rec.get("jira_key") or ""))
            self.testplan_tc_table.setItem(row, 3, QTableWidgetItem(rec.get("summary") or ""))

    def collect_testplan_testcases(self) -> List[Dict[str, Any]]:
        """테이블에서 Test Plan - Test Case 매핑 정보를 수집."""
        records: List[Dict[str, Any]] = []
        rows = self.testplan_tc_table.rowCount()
        for r in range(rows):
            order_item = self.testplan_tc_table.item(r, 0)
            tcid_item = self.testplan_tc_table.item(r, 1)
            try:
                order_no = int(order_item.text()) if order_item and order_item.text().strip() else r + 1
            except ValueError:
                order_no = r + 1
            if not tcid_item or not tcid_item.text().strip():
                continue
            try:
                testcase_id = int(tcid_item.text().strip())
            except ValueError:
                continue
            records.append(
                {
                    "order_no": order_no,
                    "testcase_id": testcase_id,
                }
            )
        return records

    def _init_executions_tab(self):
        from PySide6.QtWidgets import (
            QVBoxLayout,
            QHBoxLayout,
            QFormLayout,
            QLabel,
            QLineEdit,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QComboBox,
            QSplitter,
            QWidget,
        )

        layout = QVBoxLayout()
        # 실행/실행결과 탭 헤더는 이슈 타입에 따라 동적으로 변경
        self.lbl_exec_header = QLabel("Executions")
        layout.addWidget(self.lbl_exec_header)

        splitter = QSplitter(Qt.Vertical)

        # 상단: 대시보드 + 필터 + Test Execution 메타 정보 + 버튼들
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # 상단 대시보드: TE executed 요약
        dash_layout = QHBoxLayout()
        self.lbl_te_executed = QLabel("TE executed: 0/0 (0%)")
        dash_layout.addWidget(self.lbl_te_executed)
        dash_layout.addStretch()
        top_layout.addLayout(dash_layout)

        # 상단 필터바: Assignee / Result / RTM Env 기준 필터
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Assignee:"))
        self.cmb_te_filter_assignee = QComboBox()
        self.cmb_te_filter_assignee.setEditable(False)
        filter_layout.addWidget(self.cmb_te_filter_assignee)

        filter_layout.addWidget(QLabel("Result:"))
        self.cmb_te_filter_result = QComboBox()
        self.cmb_te_filter_result.setEditable(False)
        filter_layout.addWidget(self.cmb_te_filter_result)

        filter_layout.addWidget(QLabel("RTM Env:"))
        self.cmb_te_filter_env = QComboBox()
        self.cmb_te_filter_env.setEditable(False)
        filter_layout.addWidget(self.cmb_te_filter_env)

        filter_layout.addStretch()
        top_layout.addLayout(filter_layout)

        # 상단: Test Execution 메타 정보 (Environment, Start/End, Result, Executed By)
        form_layout = QFormLayout()
        self.ed_te_env = QLineEdit()
        self.ed_te_start = QLineEdit()
        self.ed_te_end = QLineEdit()
        self.cmb_te_result = QComboBox()
        self.cmb_te_result.setEditable(False)
        self.cmb_te_result.addItems(
            [
                "",
                "In progress",
                "Pass",
                "Fail",
                "Blocked",
                "Passed with restrictions",
            ]
        )
        self.ed_te_executed_by = QLineEdit()
        form_layout.addRow("Environment", self.ed_te_env)
        form_layout.addRow("Start Date", self.ed_te_start)
        form_layout.addRow("End Date", self.ed_te_end)
        form_layout.addRow("Result", self.cmb_te_result)
        form_layout.addRow("Executed By", self.ed_te_executed_by)
        top_layout.addLayout(form_layout)

        # 중간: 버튼
        btn_layout = QHBoxLayout()
        # TEST_PLAN 용: Execute Test Plan
        self.btn_execute_plan = QPushButton("Execute Test Plan")
        # TEST_EXECUTION 용: Test Case Execution 행 추가/삭제/일괄 편집
        self.btn_add_tc_exec = QPushButton("Add Test Case Execution")
        self.btn_del_tc_exec = QPushButton("Delete Selected")
        self.btn_edit_tc_exec = QPushButton("Edit Selected...")
        self.btn_create_defect = QPushButton("Create Defect...")
        self.btn_link_defect = QPushButton("Link Defect...")
        self.btn_clear_defects = QPushButton("Clear Defects")
        btn_layout.addWidget(self.btn_execute_plan)
        btn_layout.addWidget(self.btn_add_tc_exec)
        btn_layout.addWidget(self.btn_del_tc_exec)
        btn_layout.addWidget(self.btn_edit_tc_exec)
        btn_layout.addWidget(self.btn_create_defect)
        btn_layout.addWidget(self.btn_link_defect)
        btn_layout.addWidget(self.btn_clear_defects)
        btn_layout.addStretch()
        top_layout.addLayout(btn_layout)

        # 하단: Test Case Executions 테이블
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        self.tc_exec_table = QTableWidget()
        self.tc_exec_table.setColumnCount(9)
        self.tc_exec_table.setHorizontalHeaderLabels(
            [
                "Order",
                "Test Case ID",
                "Jira Key",
                "Summary",
                "Assignee",
                "Result",
                "RTM Env",
                "Actual Time (min)",
                "Defects",
            ]
        )
        self.tc_exec_table.horizontalHeader().setStretchLastSection(True)
        bottom_layout.addWidget(self.tc_exec_table)

        splitter.addWidget(top_widget)
        splitter.addWidget(bottom_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)

        layout.addWidget(splitter)

        self.btn_execute_plan.clicked.connect(self._on_execute_test_plan_clicked)
        self.btn_add_tc_exec.clicked.connect(self._on_add_tc_exec_clicked)
        self.btn_del_tc_exec.clicked.connect(self._on_del_tc_exec_clicked)
        self.btn_edit_tc_exec.clicked.connect(self._on_edit_tc_exec_clicked)
        self.btn_link_defect.clicked.connect(self._on_link_defect_to_tces_clicked)
        self.btn_create_defect.clicked.connect(self._on_create_defect_for_tces_clicked)
        self.btn_clear_defects.clicked.connect(self._on_clear_defects_from_tces_clicked)
        # TCE 테이블 직접 편집도 더티로 인식
        self.tc_exec_table.itemChanged.connect(self._on_tc_exec_item_changed)

        # 더블클릭 시 선택된 Test Case Execution 의 Step 실행 상세를 여는 핸들러 연결
        self.tc_exec_table.cellDoubleClicked.connect(self._on_tc_exec_double_clicked)

        # 필터 변경 시 실시간으로 테이블 필터링
        self.cmb_te_filter_assignee.currentIndexChanged.connect(self._apply_tc_exec_filters)
        self.cmb_te_filter_result.currentIndexChanged.connect(self._apply_tc_exec_filters)
        self.cmb_te_filter_env.currentIndexChanged.connect(self._apply_tc_exec_filters)

        self.executions_tab.setLayout(layout)

    # --------------------------- Linked Test Cases (Requirement / Defect 등) -----

    def load_linked_testcases(self, records: List[Dict[str, Any]]):
        """
        Relations 정보 등에서 필터링한 Test Case 리스트를 Test Cases 탭에 로드.
        REQUIREMENT / DEFECT 등의 이슈 타입에서 사용한다.
        """
        # REQUIREMENT / DEFECT 용 컬럼 구성
        self.testplan_tc_table.setColumnCount(6)
        self.testplan_tc_table.setHorizontalHeaderLabels(
            ["Jira Key", "Summary", "Priority", "Assignee", "Components", "RTM Env"]
        )
        self.testplan_tc_table.setRowCount(0)
        from PySide6.QtWidgets import QTableWidgetItem

        for rec in records:
            row = self.testplan_tc_table.rowCount()
            self.testplan_tc_table.insertRow(row)
            self.testplan_tc_table.setItem(row, 0, QTableWidgetItem(rec.get("dst_jira_key") or ""))
            self.testplan_tc_table.setItem(row, 1, QTableWidgetItem(rec.get("dst_summary") or ""))
            # 아래 필드들은 relations 조회 결과에 아직 없을 수 있으므로 빈 값으로 둔다.
            # 향후 jira_mapping.extract_relations_from_jira / DB 스키마 확장으로 보강 가능.
            self.testplan_tc_table.setItem(
                row, 2, QTableWidgetItem(rec.get("dst_priority") or "")
            )
            self.testplan_tc_table.setItem(
                row, 3, QTableWidgetItem(rec.get("dst_assignee") or "")
            )
            self.testplan_tc_table.setItem(
                row, 4, QTableWidgetItem(rec.get("dst_components") or "")
            )
            self.testplan_tc_table.setItem(
                row, 5, QTableWidgetItem(rec.get("dst_rtm_environment") or "")
            )

    def _on_add_tc_exec_clicked(self):
        """Test Case Execution 행 추가 (수동으로 한 건 더 추가)."""
        row = self.tc_exec_table.rowCount()
        self.tc_exec_table.insertRow(row)
        self.tc_exec_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
        # 새로운 TCE 추가 → 현재 Test Execution 이슈를 더티로 표시
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_edit_tc_exec_clicked(self):
        """
        Test Execution 화면에서 선택된 Test Case Executions 에 대해
        Assignee / Result / RTM Env / Defects 를 일괄 수정하는 팝업.
        (실제 DB 반영은 Save Local Issue 시점에 collect_testcase_executions()를 통해 이뤄진다.)
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QFormLayout,
            QLineEdit,
            QDialogButtonBox,
            QLabel,
        )

        selected = self.tc_exec_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected})

        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Selected Test Case Executions")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(
            QLabel("Leave a field empty to keep current value. Non-empty fields overwrite all selected rows.")
        )

        form = QFormLayout()
        ed_assignee = QLineEdit()
        ed_result = QLineEdit()
        ed_env = QLineEdit()
        ed_actual_time = QLineEdit()
        ed_defects = QLineEdit()
        form.addRow("Assignee", ed_assignee)
        form.addRow("Result", ed_result)
        form.addRow("RTM Env", ed_env)
        form.addRow("Actual Time (min)", ed_actual_time)
        form.addRow("Defects", ed_defects)
        vbox.addLayout(form)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            assignee = ed_assignee.text().strip()
            result = ed_result.text().strip()
            env = ed_env.text().strip()
            at_text = ed_actual_time.text().strip()
            defects = ed_defects.text().strip()

            if not any([assignee, result, env, at_text, defects]):
                dlg.reject()
                return

            from PySide6.QtWidgets import QTableWidgetItem

            for r in rows:
                if assignee:
                    item = self.tc_exec_table.item(r, 4) or QTableWidgetItem()
                    item.setText(assignee)
                    self.tc_exec_table.setItem(r, 4, item)
                if result:
                    item = self.tc_exec_table.item(r, 5) or QTableWidgetItem()
                    item.setText(result)
                    self.tc_exec_table.setItem(r, 5, item)
                if env:
                    item = self.tc_exec_table.item(r, 6) or QTableWidgetItem()
                    item.setText(env)
                    self.tc_exec_table.setItem(r, 6, item)
                if at_text:
                    item = self.tc_exec_table.item(r, 7) or QTableWidgetItem()
                    item.setText(at_text)
                    self.tc_exec_table.setItem(r, 7, item)
                if defects:
                    item = self.tc_exec_table.item(r, 8) or QTableWidgetItem()
                    item.setText(defects)
                    self.tc_exec_table.setItem(r, 8, item)

            # 일괄 수정 완료 → 현재 Test Execution 이슈를 더티로 표시
            main_win = self.window()
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def _on_link_defect_to_tces_clicked(self):
        """
        Executions 탭에서 선택된 TCE 들에 대해 기존 Defect 를 링크한다.
        - 로컬 DB 내 DEFECT 이슈 목록을 보여주고, 선택된 Defect 의 Jira Key 를
          TCE 의 'Defects' 컬럼 문자열에 추가한다.
        - 이후 on_save_issue_clicked() / on_push_issue_clicked() 를 통해
          RTM Test Execution TestCases payload 의 defects 필드로 반영된다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QLabel,
            QListWidget,
            QListWidgetItem,
            QDialogButtonBox,
        )

        selected = self.tc_exec_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected})

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return

        # 프로젝트 내 DEFECT 이슈 목록 조회
        cur = main_win.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary, status, priority
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'DEFECT'
             ORDER BY jira_key, summary
            """,
            (main_win.project.id,),
        )
        rows_def = cur.fetchall()
        if not rows_def:
            main_win.status_bar.showMessage("No Defect issues found in this project.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Link Existing Defect")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select a Defect to link to the selected Test Case Executions:"))

        lst = QListWidget()
        for r in rows_def:
            key = r["jira_key"] or f"ID={r['id']}"
            text = f"{key} - {r['summary'] or ''} [{r['status'] or ''}/{r['priority'] or ''}]"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, r["jira_key"])
            lst.addItem(item)
        vbox.addWidget(lst)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            it = lst.currentItem()
            if not it:
                dlg.reject()
                return
            defect_key = it.data(Qt.UserRole)
            if not defect_key:
                dlg.reject()
                return

            from PySide6.QtWidgets import QTableWidgetItem

            for r in rows:
                item = self.tc_exec_table.item(r, 8)
                existing = item.text().strip() if item else ""
                keys = [x.strip() for x in existing.split(",") if x.strip()] if existing else []
                if defect_key in keys:
                    continue
                keys.append(defect_key)
                new_val = ", ".join(keys)
                if not item:
                    item = QTableWidgetItem()
                item.setText(new_val)
                self.tc_exec_table.setItem(r, 8, item)

                # RTM TCE Defect 링크 (tce_test_key 가 있고 JIRA 사용 가능할 때만)
                if (
                    getattr(main_win, "jira_available", False)
                    and getattr(main_win, "jira_client", None)
                ):
                    order_item = self.tc_exec_table.item(r, 0)
                    tce_test_key = order_item.data(Qt.UserRole + 1) if order_item else None
                    if tce_test_key:
                        try:
                            # issueId 는 생략하고 defect_test_key 만 전달 (RTM 문서상 둘 다 허용)
                            main_win.jira_client.link_tce_defect(str(tce_test_key), str(defect_key))
                        except Exception as e_link:
                            print(
                                f"[WARN] Failed to link defect {defect_key} to TCE {tce_test_key} in RTM: {e_link}"
                            )

            # Defects 링크 변경 → 현재 Test Execution 이슈를 더티로 표시
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)
        dlg.exec()

    def _on_create_defect_for_tces_clicked(self):
        """
        Executions 탭에서 선택된 TCE 들을 기준으로 새 Defect 를 생성하고, 해당 TCE 들에 링크한다.

        흐름:
          1) 요약(Summary) / 설명(Description) 입력 다이얼로그 표시
          2) 로컬 DB 에 DEFECT 이슈 한 건 생성
          3) RTM / JIRA 에 DEFECT 엔티티 생성 (create_entity)
          4) 생성된 Defect 의 jira_key 를 TCE Defects 컬럼에 추가
          5) RTM TCE Defect 링크 API 호출 (tce_test_key 가 있는 경우)
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QFormLayout,
            QLineEdit,
            QTextEdit,
            QDialogButtonBox,
            QLabel,
        )

        selected = self.tc_exec_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected})

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return
        if not getattr(main_win, "jira_available", False) or not getattr(main_win, "jira_client", None):
            # JIRA/RTM 이 설정되지 않은 경우에는 현재 단계에서는 Defect 생성 기능을 제공하지 않는다.
            return

        # 1) 요약/설명 입력 다이얼로그
        dlg = QDialog(self)
        dlg.setWindowTitle("Create Defect for Selected Executions")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Enter Defect summary and (optional) description:"))

        form = QFormLayout()
        ed_summary = QLineEdit()
        ed_description = QTextEdit()
        form.addRow("Summary", ed_summary)
        form.addRow("Description", ed_description)
        vbox.addLayout(form)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            summary = ed_summary.text().strip()
            description = ed_description.toPlainText().strip()
            if not summary:
                dlg.reject()
                return

            from backend.db import create_local_issue, update_issue_fields, get_issue_by_id
            import backend.jira_mapping as jira_mapping
            from PySide6.QtWidgets import QTableWidgetItem

            # 2) 로컬 DEFECT 이슈 생성
            try:
                new_issue_id = create_local_issue(
                    main_win.conn,
                    project_id=main_win.project.id,
                    issue_type="DEFECT",
                    folder_id=None,
                    summary=summary,
                )
            except Exception as e_local:
                print(f"[ERROR] Failed to create local DEFECT issue: {e_local}")
                dlg.reject()
                return

            # 생성 직후 issue 레코드를 읽어 payload 생성
            issue = get_issue_by_id(main_win.conn, new_issue_id)
            if not issue:
                print("[ERROR] Newly created DEFECT issue not found in DB.")
                dlg.reject()
                return

            # description / local_only 필드 보정
            try:
                update_issue_fields(
                    main_win.conn,
                    new_issue_id,
                    {
                        "description": description,
                        "local_only": 0,
                    },
                )
                issue = get_issue_by_id(main_win.conn, new_issue_id)
            except Exception:
                # description 업데이트 실패는 치명적이지 않으므로 무시
                issue = get_issue_by_id(main_win.conn, new_issue_id)

            # 3) RTM / JIRA 에 DEFECT 엔티티 생성
            defect_key = None
            try:
                payload = jira_mapping.build_jira_create_payload("DEFECT", issue)
                resp = main_win.jira_client.create_entity("DEFECT", payload)
                if isinstance(resp, dict):
                    defect_key = resp.get("key") or resp.get("jiraKey") or resp.get("issueKey")
            except Exception as e_create:
                print(f"[ERROR] Failed to create DEFECT in JIRA/RTM: {e_create}")
                defect_key = None

            if not defect_key:
                # 원격 생성 실패 시 로컬 이슈는 남기되, 이번 동작은 취소
                dlg.reject()
                return

            # 생성된 jira_key 를 로컬 DB 에 업데이트
            try:
                update_issue_fields(main_win.conn, new_issue_id, {"jira_key": defect_key})
            except Exception as e_upd:
                print(f"[WARN] Failed to update local DEFECT jira_key: {e_upd}")

            # 4) 선택된 TCE 행들의 Defects 컬럼에 jira_key 추가 + 5) RTM 링크
            for r in rows:
                item = self.tc_exec_table.item(r, 8)
                existing = item.text().strip() if item else ""
                keys = [x.strip() for x in existing.split(",") if x.strip()] if existing else []
                if defect_key not in keys:
                    keys.append(defect_key)
                new_val = ", ".join(keys)
                if not item:
                    item = QTableWidgetItem()
                item.setText(new_val)
                self.tc_exec_table.setItem(r, 8, item)

                # RTM TCE Defect 링크
                order_item = self.tc_exec_table.item(r, 0)
                tce_test_key = order_item.data(Qt.UserRole + 1) if order_item else None
                if tce_test_key:
                    try:
                        main_win.jira_client.link_tce_defect(str(tce_test_key), str(defect_key))
                    except Exception as e_link:
                        print(
                            f"[WARN] Failed to link new defect {defect_key} to TCE {tce_test_key} in RTM: {e_link}"
                        )

            # 좌측 Defects 탭 갱신 (현재 이슈가 TEST_EXECUTION 인 경우)
            try:
                if hasattr(main_win, "_refresh_defects_tab_for_current_issue"):
                    main_win._refresh_defects_tab_for_current_issue()
            except Exception:
                pass

            # 새 Defect 생성 및 링크까지 완료 → 현재 Test Execution 더티 표시
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)
        dlg.exec()

    def _on_clear_defects_from_tces_clicked(self):
        """
        Executions 탭에서 선택된 TCE 들의 Defects 컬럼을 비운다.
        """
        from PySide6.QtWidgets import QTableWidgetItem

        selected = self.tc_exec_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected})
        main_win = self.window()
        for r in rows:
            # 기존 Defects 문자열
            item = self.tc_exec_table.item(r, 8) or QTableWidgetItem()
            existing = item.text().strip()

            # RTM TCE Defect 언링크 (tce_test_key 가 있고 JIRA 사용 가능할 때만)
            if existing and getattr(main_win, "jira_available", False) and getattr(
                main_win, "jira_client", None
            ):
                order_item = self.tc_exec_table.item(r, 0)
                tce_test_key = order_item.data(Qt.UserRole + 1) if order_item else None
                if tce_test_key:
                    keys = [x.strip() for x in existing.split(",") if x.strip()]
                    for defect_key in keys:
                        try:
                            main_win.jira_client.unlink_tce_defect(str(tce_test_key), str(defect_key))
                        except Exception as e_unlink:
                            print(
                                f"[WARN] Failed to unlink defect {defect_key} from TCE {tce_test_key} in RTM: {e_unlink}"
                            )

            # 로컬 Defects 컬럼 비우기
            item.setText("")
            self.tc_exec_table.setItem(r, 8, item)

        # Defects 컬럼 변경 → 현재 Test Execution 이슈를 더티로 표시
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_tc_exec_item_changed(self, item):
        """TCE 테이블 셀 편집 시 현재 Test Execution 을 더티로 표시."""
        main_win = self.window()
        if hasattr(main_win, "mark_current_issue_dirty"):
            main_win.mark_current_issue_dirty()

    def _on_tc_exec_double_clicked(self, row: int, column: int):
        """
        Test Case Executions 테이블에서 행을 더블클릭했을 때:
        - 해당 Test Case 의 설계 Steps + 실행 상태(testcase_step_executions)를 합쳐
          Step 실행 상세 팝업을 띄운다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QTableWidget,
            QTableWidgetItem,
            QDialogButtonBox,
            QLabel,
            QComboBox,
        )

        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return

        order_item = self.tc_exec_table.item(row, 0)
        tcid_item = self.tc_exec_table.item(row, 1)
        if not order_item or not tcid_item:
            return

        tce_id = order_item.data(Qt.UserRole)
        if not tce_id:
            return
        try:
            tce_id = int(tce_id)
        except (TypeError, ValueError):
            return

        try:
            testcase_id = int(tcid_item.text().strip())
        except (TypeError, ValueError):
            return

        # 설계 단계 Steps 및 기존 Step Execution 상태 로딩
        steps = get_steps_for_issue(main_win.conn, testcase_id)
        existing_execs = get_step_executions_for_tce(main_win.conn, tce_id)
        exec_map = {
            int(rec["testcase_step_id"]): rec for rec in existing_execs if rec.get("testcase_step_id") is not None
        }

        # RTM Test Case Execution testKey (있으면 Step 상태를 RTM에도 반영 가능)
        tce_row = get_testcase_execution_by_id(main_win.conn, tce_id)
        tce_test_key = tce_row.get("tce_test_key") if tce_row else None

        dlg = QDialog(self)
        dlg.setWindowTitle("Execute Test Case - Steps")
        vbox = QVBoxLayout(dlg)

        tc_label = self.tc_exec_table.item(row, 3)  # Summary
        summary_text = tc_label.text() if tc_label else ""
        vbox.addWidget(QLabel(f"Test Case ID: {testcase_id}  Summary: {summary_text}"))

        # 상단: 일괄 상태 변경 콤보
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("Change steps status:"))
        cmb_status = QComboBox()
        cmb_status.addItems(
            [
                "",
                "Not run",
                "In progress",
                "Pass",
                "Fail",
                "Blocked",
                "Passed with restrictions",
            ]
        )
        btn_apply_all = QPushButton("Apply to all")
        top_layout.addWidget(cmb_status)
        top_layout.addWidget(btn_apply_all)
        top_layout.addStretch()
        vbox.addLayout(top_layout)

        # Steps 테이블: Group / Order / Action / Input / Expected / Status / Actual Result / Evidence
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(
            ["Group", "Order", "Action", "Input", "Expected", "Status", "Actual Result", "Evidence"]
        )
        table.horizontalHeader().setStretchLastSection(True)

        for s in steps:
            step_id = s.get("id")
            row_idx = table.rowCount()
            table.insertRow(row_idx)

            item_group = QTableWidgetItem(str(s.get("group_no", 1)))
            item_group.setFlags(item_group.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 0, item_group)

            item_order = QTableWidgetItem(str(s.get("order_no", 1)))
            item_order.setFlags(item_order.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 1, item_order)

            item_action = QTableWidgetItem(s.get("action") or "")
            item_action.setFlags(item_action.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 2, item_action)

            item_input = QTableWidgetItem(s.get("input") or "")
            item_input.setFlags(item_input.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 3, item_input)

            item_expected = QTableWidgetItem(s.get("expected") or "")
            item_expected.setFlags(item_expected.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_idx, 4, item_expected)

            # 기존 실행 상태
            exec_rec = exec_map.get(int(step_id)) if step_id is not None else None
            status_val = exec_rec.get("status") if exec_rec else ""
            actual_val = exec_rec.get("actual_result") if exec_rec else ""
            evidence_val = exec_rec.get("evidence") if exec_rec else ""

            item_status = QTableWidgetItem(status_val or "")
            table.setItem(row_idx, 5, item_status)
            item_actual = QTableWidgetItem(actual_val or "")
            table.setItem(row_idx, 6, item_actual)
            item_evidence = QTableWidgetItem(evidence_val or "")
            table.setItem(row_idx, 7, item_evidence)

            # 숨은 데이터로 testcase_step_id 저장
            item_group.setData(Qt.UserRole, step_id)

        vbox.addWidget(table)

        # 일괄 상태 적용 버튼 로직
        def apply_all_status():
            val = cmb_status.currentText().strip()
            if not val:
                return
            for r in range(table.rowCount()):
                item = table.item(r, 5)
                if item is None:
                    item = QTableWidgetItem()
                    table.setItem(r, 5, item)
                item.setText(val)

        btn_apply_all.clicked.connect(apply_all_status)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            records: list[Dict[str, Any]] = []
            for r in range(table.rowCount()):
                item_group = table.item(r, 0)
                step_id = item_group.data(Qt.UserRole) if item_group else None
                if not step_id:
                    continue
                status_item = table.item(r, 5)
                actual_item = table.item(r, 6)
                evidence_item = table.item(r, 7)
                records.append(
                    {
                        "row_index": r,
                        "testcase_step_id": int(step_id),
                        "status": status_item.text().strip() if status_item else "",
                        "actual_result": actual_item.text().strip() if actual_item else "",
                        "evidence": evidence_item.text().strip() if evidence_item else "",
                    }
                )

            # 1) 로컬 Step Execution 저장
            replace_step_executions_for_tce(main_win.conn, tce_id, records)

            # 2) RTM Test Case Execution Step API 연계 (tce_test_key 가 있는 경우에만)
            if (
                tce_test_key
                and getattr(main_win, "jira_available", False)
                and getattr(main_win, "jira_client", None)
            ):
                try:
                    for rec in records:
                        step_index = rec["row_index"] + 1  # 현재 UI 순서를 RTM stepIndex 로 사용
                        status_text = rec["status"]
                        actual_text = rec["actual_result"]

                        # Status 업데이트
                        if status_text:
                            status_payload = {
                                "statusName": status_text,
                                "name": status_text,
                            }
                            try:
                                main_win.jira_client.set_tce_step_status(
                                    tce_test_key, step_index, status_payload
                                )
                            except Exception as e_stat:
                                print(
                                    f"[WARN] Failed to push TCE step status to RTM (tce={tce_test_key}, step={step_index}): {e_stat}"
                                )

                        # Comment(여기서는 Actual Result 를 코멘트로 사용) 업데이트
                        if actual_text:
                            try:
                                main_win.jira_client.set_tce_step_comment(
                                    tce_test_key, step_index, actual_text
                                )
                            except Exception as e_cmt:
                                print(
                                    f"[WARN] Failed to push TCE step comment to RTM (tce={tce_test_key}, step={step_index}): {e_cmt}"
                                )
                except Exception as e_rtm:
                    print(f"[WARN] Failed to sync step executions to RTM: {e_rtm}")

            # Step 실행 결과가 변경되었으므로, 현재 Test Execution 이슈를 더티로 표시
            if hasattr(main_win, "mark_current_issue_dirty"):
                main_win.mark_current_issue_dirty()

            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def _on_execute_test_plan_clicked(self):
        """
        TEST_PLAN 이슈의 Executions 탭에서 'Execute Test Plan' 버튼:
        - 현재 Test Plan 을 기반으로 새로운 TEST_EXECUTION 이슈를 생성하고,
        - 해당 플랜에 연결된 모든 Test Case 에 대한 Test Case Execution 행을 초기화한다.
        """
        main_win = self.window()
        if not hasattr(main_win, "conn") or getattr(main_win, "project", None) is None:
            return
        if getattr(main_win, "current_issue_id", None) is None:
            main_win.status_bar.showMessage("No Test Plan selected.")
            return

        # 현재 이슈가 TEST_PLAN 인지 확인
        plan_issue = get_issue_by_id(main_win.conn, main_win.current_issue_id)
        if not plan_issue or (plan_issue.get("issue_type") or "").upper() != "TEST_PLAN":
            main_win.status_bar.showMessage("Current issue is not a TEST_PLAN.")
            return

        # 이 Test Plan 에 연결된 Test Case 목록 (로컬 기준)
        tp_records = get_testplan_testcases(main_win.conn, int(main_win.current_issue_id))
        if not tp_records:
            main_win.status_bar.showMessage("This Test Plan has no Test Cases mapped.")
            return

        # 동기화 영향 안내 및 확인:
        # - RTM 사용 가능 + jira_key 가 있을 경우: RTM 에 실제 Test Execution 이 생성되고
        #   그 결과가 로컬 DB 에 동기화된다.
        # - 그렇지 않은 경우: 로컬 전용 Test Execution 이 생성된다.
        msg = (
            "이 Test Plan 을 실행하여 새로운 Test Execution 을 생성합니다.\n\n"
            "- JIRA/RTM 이 구성되어 있고, 이 플랜에 JIRA Key 가 있으면:\n"
            "  RTM `/test-execution/execute/{TestPlanKey}` 가 호출되어 서버에도 Test Execution 이 생성되고,\n"
            "  그 결과가 로컬 DB (issues/testexecutions/testcase_executions)에 저장됩니다.\n"
            "- 그렇지 않으면 로컬 DB 에만 Test Execution 이 생성됩니다.\n\n"
            "이미 존재하는 서버측 Test Execution/TCE 와 충돌할 수 있으므로,\n"
            "필요 시 JIRA/RTM 상태를 먼저 확인한 뒤 진행하는 것이 좋습니다.\n\n"
            "계속 진행하시겠습니까?"
        )
        ret = QMessageBox.question(
            main_win,
            "Execute Test Plan",
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        # ------------------------------------------------------------------
        # 1) JIRA/RTM 이 사용 가능한 경우: RTM Test Plan Execute API 호출
        # ------------------------------------------------------------------
        te_issue_id: int | None = None
        plan_jira_key = plan_issue.get("jira_key")
        if getattr(main_win, "jira_available", False) and getattr(main_win, "jira_client", None) and plan_jira_key:
            try:
                main_win.status_bar.showMessage(f"Executing Test Plan in RTM for {plan_jira_key}...")
                QApplication.setOverrideCursor(Qt.WaitCursor)

                # RTM Test Execution 생성
                te_resp = main_win.jira_client.execute_test_plan(plan_jira_key)
                te_key = None
                if isinstance(te_resp, dict):
                    te_key = (
                        te_resp.get("testKey")
                        or te_resp.get("key")
                        or te_resp.get("jiraKey")
                    )

                if te_key:
                    # 로컬 TEST_EXECUTION 이슈 생성 (RTM 에 생성된 TE 와 매핑)
                    folder_id = plan_issue.get("folder_id")
                    summary = te_resp.get("summary") if isinstance(te_resp, dict) else None
                    if not summary:
                        summary = f"{plan_issue.get('summary') or 'Test Plan'} (execution)"

                    te_issue_id = create_local_issue(
                        main_win.conn,
                        project_id=main_win.project.id,
                        issue_type="TEST_EXECUTION",
                        folder_id=folder_id,
                        summary=summary,
                    )

                    # Details 필드: 플랜에서 일부 복사 + jira_key / parent_issue_id 설정
                    fields: Dict[str, Any] = {}
                    for key in ["rtm_environment", "priority", "assignee", "labels", "components"]:
                        val = plan_issue.get(key)
                        if val:
                            fields[key] = val
                    fields["parent_issue_id"] = int(main_win.current_issue_id)
                    fields["jira_key"] = te_key
                    update_issue_fields(main_win.conn, te_issue_id, fields)

                    # RTM Test Execution 상세 및 TCE 목록을 로드하여 로컬 testexecutions / testcase_executions 에 반영
                    te_json = main_win.jira_client.get_testexecution_details(te_key)
                    te_meta = jira_mapping.map_jira_testexecution_meta_to_local(te_json)
                    te_row = get_or_create_testexecution_for_issue(main_win.conn, te_issue_id)
                    if te_meta:
                        update_testexecution_for_issue(main_win.conn, te_issue_id, te_meta)

                    tce_json = main_win.jira_client.get_testexecution_testcases(te_key)
                    tce_items = jira_mapping.map_jira_testexecution_testcases_to_local(tce_json)
                    if tce_items:
                        from backend.db import get_issue_by_jira_key

                        tce_records: list[Dict[str, Any]] = []
                        for item in tce_items:
                            tc_key = item.get("testcase_key")
                            if not tc_key:
                                continue
                            tc_issue = get_issue_by_jira_key(main_win.conn, main_win.project.id, tc_key)
                            if not tc_issue:
                                continue
                            tce_records.append(
                                {
                                    "order_no": item.get("order_no") or 0,
                                    "testcase_id": tc_issue["id"],
                                    "assignee": item.get("assignee") or "",
                                    "result": item.get("result") or "",
                                    "rtm_environment": item.get("rtm_environment") or "",
                                    "defects": item.get("defects") or "",
                                    "actual_time": item.get("actual_time"),
                                }
                            )
                        if tce_records:
                            replace_testcase_executions(main_win.conn, te_row["id"], tce_records)

                    # 트리/상태 갱신
                    main_win.reload_local_tree()
                    main_win.status_bar.showMessage(
                        f"Created new Test Execution (RTM) {te_key} for this Test Plan."
                    )
                    QApplication.restoreOverrideCursor()
                    return
                else:
                    main_win.status_bar.showMessage(
                        "RTM execute API did not return a test execution key; falling back to local-only creation."
                    )
            except Exception as e_rtm:
                QApplication.restoreOverrideCursor()
                print(f"[WARN] Failed to execute Test Plan in RTM: {e_rtm}")
                main_win.status_bar.showMessage(
                    f"Failed to execute Test Plan in RTM ({e_rtm}); falling back to local-only creation."
                )

        # ------------------------------------------------------------------
        # 2) RTM 사용 불가이거나 실패 시: 기존 로컬 전용 Test Execution 생성 플로우
        # ------------------------------------------------------------------

        folder_id = plan_issue.get("folder_id")
        summary = f"{plan_issue.get('summary') or 'Test Plan'} (execution)"
        te_issue_id = create_local_issue(
            main_win.conn,
            project_id=main_win.project.id,
            issue_type="TEST_EXECUTION",
            folder_id=folder_id,
            summary=summary,
        )

        # Details 필드 일부를 플랜에서 복사 (환경/우선순위/라벨 등)
        fields_local: Dict[str, Any] = {}
        for key in ["rtm_environment", "priority", "assignee", "labels", "components"]:
            val = plan_issue.get(key)
            if val:
                fields_local[key] = val
        fields_local["parent_issue_id"] = int(main_win.current_issue_id)
        update_issue_fields(main_win.conn, te_issue_id, fields_local)

        # testexecutions 메타 행 초기화
        te_row_local = get_or_create_testexecution_for_issue(main_win.conn, te_issue_id)
        te_meta_local: Dict[str, Any] = {}
        if plan_issue.get("rtm_environment"):
            te_meta_local["environment"] = plan_issue.get("rtm_environment") or ""
        update_testexecution_for_issue(main_win.conn, te_issue_id, te_meta_local)

        # testcase_executions 행 생성 (각 Test Case 에 대해 기본 행)
        records: list[Dict[str, Any]] = []
        for idx, rec in enumerate(tp_records, start=1):
            tc_id = rec.get("testcase_id")
            if not tc_id:
                continue
            records.append(
                {
                    "testcase_id": int(tc_id),
                    "order_no": int(rec.get("order_no", idx) or idx),
                    "assignee": "",
                    "result": "",
                    "rtm_environment": plan_issue.get("rtm_environment") or "",
                    "defects": "",
                }
            )
        replace_testcase_executions(main_win.conn, te_row_local["id"], records)

        # 트리/상태 갱신: 새 Test Execution 은 상단 'Test Executions' 탭에서 확인 가능
        main_win.reload_local_tree()
        main_win.status_bar.showMessage(
            f"Created new Test Execution (local only, id={te_issue_id}) for this Test Plan."
        )

    def _on_del_tc_exec_clicked(self):
        """선택된 Test Case Execution 행 삭제 및 order 재정렬."""
        selected = self.tc_exec_table.selectedIndexes()
        if not selected:
            return
        rows = sorted({idx.row() for idx in selected}, reverse=True)
        for r in rows:
            self.tc_exec_table.removeRow(r)
        # order 재정렬
        for i in range(self.tc_exec_table.rowCount()):
            item = self.tc_exec_table.item(i, 0)
            if item is None:
                item = QTableWidgetItem(str(i + 1))
                self.tc_exec_table.setItem(i, 0, item)
            else:
                item.setText(str(i + 1))

    # --------------------------- Test Execution binding helpers ------------------

    def load_testexecution(self, meta: Dict[str, Any] | None, tc_execs: List[Dict[str, Any]]):
        """Test Execution 메타 & Test Case Executions를 탭에 로드."""
        if not meta:
            self.ed_te_env.setText("")
            self.ed_te_start.setText("")
            self.ed_te_end.setText("")
            self.cmb_te_result.setCurrentIndex(0)
            self.ed_te_executed_by.setText("")
        else:
            self.ed_te_env.setText(meta.get("environment") or "")
            self.ed_te_start.setText(meta.get("start_date") or "")
            self.ed_te_end.setText(meta.get("end_date") or "")
            result_val = meta.get("result") or ""
            # 콤보박스에서 동일 텍스트를 찾고, 없으면 첫 번째 항목("")로 설정
            idx = self.cmb_te_result.findText(result_val)
            if idx < 0:
                idx = 0
            self.cmb_te_result.setCurrentIndex(idx)
            self.ed_te_executed_by.setText(meta.get("executed_by") or "")

        self.tc_exec_table.setRowCount(0)
        assignees: set[str] = set()
        results: set[str] = set()
        envs: set[str] = set()

        for rec in tc_execs:
            row = self.tc_exec_table.rowCount()
            self.tc_exec_table.insertRow(row)
            item_order = QTableWidgetItem(str(rec.get("order_no", row + 1)))
            # UserRole 에 testcase_executions.id 저장 (Step 실행 상세에 사용)
            item_order.setData(Qt.UserRole, rec.get("id"))
            # UserRole+1 에 RTM Test Case Execution testKey 저장 (존재하는 경우)
            if rec.get("tce_test_key"):
                item_order.setData(Qt.UserRole + 1, rec.get("tce_test_key"))
            self.tc_exec_table.setItem(row, 0, item_order)
            self.tc_exec_table.setItem(row, 1, QTableWidgetItem(str(rec.get("testcase_id"))))
            self.tc_exec_table.setItem(row, 2, QTableWidgetItem(rec.get("jira_key") or ""))
            self.tc_exec_table.setItem(row, 3, QTableWidgetItem(rec.get("summary") or ""))
            assignee_val = rec.get("assignee") or ""
            result_val = rec.get("result") or ""
            env_val = rec.get("rtm_environment") or ""
            self.tc_exec_table.setItem(row, 4, QTableWidgetItem(assignee_val))
            self.tc_exec_table.setItem(row, 5, QTableWidgetItem(result_val))
            self.tc_exec_table.setItem(row, 6, QTableWidgetItem(env_val))
            at = rec.get("actual_time")
            at_text = "" if at in (None, "") else str(at)
            self.tc_exec_table.setItem(row, 7, QTableWidgetItem(at_text))
            self.tc_exec_table.setItem(row, 8, QTableWidgetItem(rec.get("defects") or ""))

            if assignee_val:
                assignees.add(assignee_val)
            if result_val:
                results.add(result_val)
            if env_val:
                envs.add(env_val)

        # 대시보드: 실행된 Test Case 수/전체/퍼센트 표시
        total = len(tc_execs)
        executed = 0
        for rec in tc_execs:
            if (rec.get("result") or "").strip():
                executed += 1
        if total > 0:
            pct = executed * 100.0 / total
            self.lbl_te_executed.setText(f"TE executed: {executed}/{total} ({pct:.1f}%)")
        else:
            self.lbl_te_executed.setText("TE executed: 0/0 (0%)")

        # 필터 콤보박스 갱신 (첫 항목은 항상 'All')
        def _fill_filter_combo(combo: QComboBox, values: set[str]):
            current = combo.currentText() if combo.count() > 0 else "All"
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("All")
            for v in sorted(values):
                combo.addItem(v)
            # 이전 선택이 여전히 유효하면 그대로 유지
            idx = combo.findText(current)
            if idx < 0:
                idx = 0
            combo.setCurrentIndex(idx)
            combo.blockSignals(False)

        _fill_filter_combo(self.cmb_te_filter_assignee, assignees)
        _fill_filter_combo(self.cmb_te_filter_result, results)
        _fill_filter_combo(self.cmb_te_filter_env, envs)

        # 필터 적용
        self._apply_tc_exec_filters()

    def collect_testexecution_meta(self) -> Dict[str, Any]:
        """현재 탭의 Test Execution 메타 정보 수집."""
        return {
            "environment": self.ed_te_env.text().strip(),
            "start_date": self.ed_te_start.text().strip(),
            "end_date": self.ed_te_end.text().strip(),
            "result": self.cmb_te_result.currentText().strip(),
            "executed_by": self.ed_te_executed_by.text().strip(),
        }

    def collect_testcase_executions(self) -> List[Dict[str, Any]]:
        """현재 탭의 Test Case Execution 리스트 수집."""
        records: List[Dict[str, Any]] = []
        rows = self.tc_exec_table.rowCount()
        for r in range(rows):
            order_item = self.tc_exec_table.item(r, 0)
            tcid_item = self.tc_exec_table.item(r, 1)
            assignee_item = self.tc_exec_table.item(r, 4)
            result_item = self.tc_exec_table.item(r, 5)
            env_item = self.tc_exec_table.item(r, 6)
            at_item = self.tc_exec_table.item(r, 7)
            defects_item = self.tc_exec_table.item(r, 8)
            try:
                order_no = int(order_item.text()) if order_item and order_item.text().strip() else r + 1
            except ValueError:
                order_no = r + 1
            if not tcid_item or not tcid_item.text().strip():
                continue
            try:
                testcase_id = int(tcid_item.text().strip())
            except ValueError:
                continue
            actual_time = 0
            if at_item and at_item.text().strip():
                try:
                    actual_time = int(at_item.text().strip())
                except ValueError:
                    actual_time = 0

            # RTM Test Case Execution testKey (있으면 유지)
            tce_test_key = None
            if order_item is not None:
                val = order_item.data(Qt.UserRole + 1)
                if val:
                    tce_test_key = str(val)
            rec: Dict[str, Any] = {
                    "order_no": order_no,
                    "testcase_id": testcase_id,
                    "assignee": assignee_item.text().strip() if assignee_item else "",
                    "result": result_item.text().strip() if result_item else "",
                "actual_time": actual_time,
                    "rtm_environment": env_item.text().strip() if env_item else "",
                    "defects": defects_item.text().strip() if defects_item else "",
                }
            if tce_test_key:
                rec["tce_test_key"] = tce_test_key
            records.append(rec)
        return records

    def _apply_tc_exec_filters(self):
        """
        상단 필터(Assignee / Result / RTM Env)에 따라 Test Case Executions 테이블을 필터링한다.
        """
        if not hasattr(self, "tc_exec_table"):
            return

        assignee = (
            self.cmb_te_filter_assignee.currentText()
            if hasattr(self, "cmb_te_filter_assignee")
            else "All"
        )
        result = (
            self.cmb_te_filter_result.currentText()
            if hasattr(self, "cmb_te_filter_result")
            else "All"
        )
        env = (
            self.cmb_te_filter_env.currentText()
            if hasattr(self, "cmb_te_filter_env")
            else "All"
        )

        assignee = assignee if assignee != "All" else ""
        result = result if result != "All" else ""
        env = env if env != "All" else ""

        rows = self.tc_exec_table.rowCount()
        for r in range(rows):
            a_item = self.tc_exec_table.item(r, 4)
            res_item = self.tc_exec_table.item(r, 5)
            env_item = self.tc_exec_table.item(r, 6)
            a_txt = a_item.text().strip() if a_item else ""
            res_txt = res_item.text().strip() if res_item else ""
            env_txt = env_item.text().strip() if env_item else ""

            visible = True
            if assignee and a_txt != assignee:
                visible = False
            if result and res_txt != result:
                visible = False
            if env and env_txt != env:
                visible = False

            self.tc_exec_table.setRowHidden(r, not visible)

    def _init_defects_tab(self):
        from PySide6.QtWidgets import (
            QVBoxLayout,
            QHBoxLayout,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QLabel,
        )

        layout = QVBoxLayout()
        header = QLabel("Defects linked to this issue / its executions")
        layout.addWidget(header)

        # 상단 버튼: 간단한 Defect 링크/생성 액션 (로컬/RTM 연동은 단계적으로 확장)
        btn_layout = QHBoxLayout()
        self.btn_defect_refresh = QPushButton("Refresh Defects")
        self.btn_defect_open = QPushButton("Open in Details")
        btn_layout.addWidget(self.btn_defect_refresh)
        btn_layout.addWidget(self.btn_defect_open)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Defects 테이블
        self.defects_table = QTableWidget()
        # Local + JIRA 메타 + 어디에서 링크되었는지
        self.defects_table.setColumnCount(7)
        self.defects_table.setHorizontalHeaderLabels(
            ["Local ID", "Jira Key", "Summary", "Status", "Priority", "Assigned To", "Linked From"]
        )
        self.defects_table.horizontalHeader().setStretchLastSection(True)
        self.defects_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 더블클릭 시 해당 Defect 를 Details 에서 열기
        self.defects_table.cellDoubleClicked.connect(
            lambda row, col: self.window().on_open_selected_defect_clicked()
        )
        layout.addWidget(self.defects_table)

        # 시그널은 MainWindow 쪽에서 연결

        self.defects_tab.setLayout(layout)

    # --------------------------- Defects tab helpers ---------------------------

    def load_defects_for_issue(self, defects: List[Dict[str, Any]]) -> None:
        """
        Defects 탭에 표시할 결함 리스트를 로드한다.
        defects 형식:
          - id, jira_key, summary, status, priority, linked_from (문자열)
        """
        from PySide6.QtWidgets import QTableWidgetItem

        self.defects_table.setRowCount(0)
        for rec in defects:
            row = self.defects_table.rowCount()
            self.defects_table.insertRow(row)
            self.defects_table.setItem(row, 0, QTableWidgetItem(str(rec.get("id") or "")))
            self.defects_table.setItem(row, 1, QTableWidgetItem(rec.get("jira_key") or ""))
            self.defects_table.setItem(row, 2, QTableWidgetItem(rec.get("summary") or ""))
            self.defects_table.setItem(row, 3, QTableWidgetItem(rec.get("status") or ""))
            self.defects_table.setItem(row, 4, QTableWidgetItem(rec.get("priority") or ""))
            self.defects_table.setItem(row, 5, QTableWidgetItem(rec.get("assignee") or ""))
            self.defects_table.setItem(row, 6, QTableWidgetItem(rec.get("linked_from") or ""))

    # ------------------------------------------------------------------ Binding helpers

    def _set_tab_visible(self, widget: QWidget, visible: bool) -> None:
        """
        PySide6(Qt6) 의 setTabVisible 지원 여부에 따라 탭 가시성/활성화를 제어.
        """
        idx = self.indexOf(widget)
        if idx < 0:
            return
        if hasattr(self, "setTabVisible"):
            self.setTabVisible(idx, visible)
        else:
            self.setTabEnabled(idx, visible)

    def update_tabs_for_issue_type(self, issue_type: str | None) -> None:
        """
        RTM 이슈 타입별로 필요한 탭만 보이도록 구성.

        SW 사양서 기준:
        - REQUIREMENT : Details / Test Cases / Relations
        - TEST_CASE   : Details / Steps / Requirements / Relations
        - TEST_PLAN   : Details / Test Cases / Executions / Relations
        - TEST_EXECUTION : Details / Executions / Relations
        - DEFECT      : Details / Test Cases / Relations
        - 기타/None   : 모든 탭 표시
        """
        # 1) 기본값: 모두 보이게
        all_widgets = [
            self.details_tab,
            self.steps_tab,
            self.requirements_tab,
            self.relations_tab,
            self.testcases_tab,
            self.executions_tab,
            self.defects_tab,
        ]
        for w in all_widgets:
            self._set_tab_visible(w, True)

        if not issue_type:
            return

        it = issue_type.upper()

        visible_map = {
            "REQUIREMENT": {"details", "testcases", "relations"},
            "TEST_CASE": {"details", "steps", "requirements", "relations"},
            "TEST_PLAN": {"details", "testcases", "executions", "relations"},
            "TEST_EXECUTION": {"details", "executions", "relations"},
            "DEFECT": {"details", "testcases", "relations"},
        }

        visible = visible_map.get(it)
        if visible is None:
            # 정의되지 않은 타입이면 모든 탭 유지
            return

        name_map = {
            "details": self.details_tab,
            "steps": self.steps_tab,
            "requirements": self.requirements_tab,
            "relations": self.relations_tab,
            "testcases": self.testcases_tab,
            "executions": self.executions_tab,
            "defects": self.defects_tab,
        }

        for name, widget in name_map.items():
            self._set_tab_visible(widget, name in visible)

        # 항상 Details 탭부터 보이도록 설정 (추적성/입력 흐름의 기본 진입점)
        if "details" in visible:
            self.setCurrentWidget(self.details_tab)

        # Test Cases 탭은 이슈 타입에 따라 헤더/버튼/컬럼 구성을 달리한다.
        if "testcases" in visible and hasattr(self, "configure_testcases_tab_for_issue_type"):
            self.configure_testcases_tab_for_issue_type(it)

        # Executions 탭도 TEST_PLAN / TEST_EXECUTION 에 따라 다르게 구성한다.
        if "executions" in visible and hasattr(self, "configure_executions_tab_for_issue_type"):
            self.configure_executions_tab_for_issue_type(it)

    def configure_testcases_tab_for_issue_type(self, issue_type: str) -> None:
        """
        Test Cases 탭을 이슈 타입에 맞게 구성한다.
        - REQUIREMENT / DEFECT: Linked Test Cases 뷰 (Cover by Test Case / Create New Test Case)
        - TEST_PLAN: Test Plan - Test Case 매핑 테이블
        - TEST_CASE: 현재 Test Case 상세 (헤더만 간단히)
        """
        it = issue_type.upper()

        if it == "TEST_PLAN":
            # 헤더 텍스트
            self.lbl_testcases_header.setText("Test Plan - Test Cases")
            # 버튼 표시 (Add / Create / Delete / Edit order)
            self.btn_cover_by_tc.setVisible(False)
            self.btn_create_tc.setVisible(True)
            self.btn_add_tp_tc.setVisible(True)
            self.btn_del_tp_tc.setVisible(True)
            self.btn_edit_tp_order.setVisible(True)
            # 컬럼 구성은 load_testplan_testcases 에서 보장
        elif it == "TEST_CASE":
            # 순수 Test Case 이슈에서는 상단 헤더만 간단히 두고,
            # 실제 생성/관리는 좌측 패널의 'Create Test Case' 버튼을 사용하도록 한다.
            self.lbl_testcases_header.setText("Test Cases")
            self.btn_cover_by_tc.setVisible(False)
            self.btn_create_tc.setVisible(False)
            self.btn_add_tp_tc.setVisible(False)
            self.btn_del_tp_tc.setVisible(False)
            self.btn_edit_tp_order.setVisible(False)
        else:
            # REQUIREMENT / DEFECT 등: Linked Test Cases
            self.lbl_testcases_header.setText("Linked Test Cases (covered by / related test cases)")
            self.btn_cover_by_tc.setVisible(True)
            self.btn_create_tc.setVisible(True)
            self.btn_add_tp_tc.setVisible(False)
            self.btn_del_tp_tc.setVisible(False)
            self.btn_edit_tp_order.setVisible(False)
            # 컬럼 구성은 load_linked_testcases 에서 보장

    def configure_executions_tab_for_issue_type(self, issue_type: str) -> None:
        """
        Executions 탭을 이슈 타입에 맞게 구성한다.
        - TEST_PLAN      : 'Execute Test Plan' 버튼 중심의 뷰 (이 플랜에서 생성된 실행들을 관리)
        - TEST_EXECUTION : Test Case Executions 편집 뷰
        """
        it = issue_type.upper()

        if it == "TEST_PLAN":
            # 헤더/버튼 구성
            self.lbl_exec_header.setText("Executions for this Test Plan")
            # Test Plan 에서는 실제 실행(Execute Test Plan) 버튼만 보이게 하고
            # 개별 Test Case Execution 편집용 UI는 숨긴다.
            self.btn_execute_plan.setVisible(True)
            self.btn_add_tc_exec.setVisible(False)
            self.btn_del_tc_exec.setVisible(False)
            self.btn_edit_tc_exec.setVisible(False)
            self.tc_exec_table.setVisible(False)
        elif it == "TEST_EXECUTION":
            # Test Execution 상세/TC Executions 편집
            self.lbl_exec_header.setText("Test Case Executions")
            self.btn_execute_plan.setVisible(False)
            self.btn_add_tc_exec.setVisible(True)
            self.btn_del_tc_exec.setVisible(True)
            self.btn_edit_tc_exec.setVisible(True)
            self.tc_exec_table.setVisible(True)
        else:
            # 기타 타입에서는 Executions 탭을 단순 비활성화 수준으로 둔다.
            self.lbl_exec_header.setText("Executions")
            self.btn_execute_plan.setVisible(False)
            self.btn_add_tc_exec.setVisible(False)
            self.btn_del_tc_exec.setVisible(False)
            self.btn_edit_tc_exec.setVisible(False)
            self.tc_exec_table.setVisible(False)

    def set_issue(self, issue: Dict[str, Any] | None) -> None:
        """
        현재 선택된 이슈의 필드를 Details 탭에 로드하고,
        RTM 이슈 타입별 탭 구성을 적용한다.

        issues 테이블 스키마 필드 매핑:
          - summary, description, status, priority, assignee, reporter
          - labels, components, security_level, fix_versions, affects_versions
          - rtm_environment, due_date, created, updated, attachments
        """
        # 이슈 로딩 중에는 더티 플래그가 올라가지 않도록 suppress 플래그 설정
        self._suppress_dirty = True
        try:
            self._current_issue = issue

            issue_type = None
            if issue is not None:
                issue_type = issue.get("issue_type")
            self.update_tabs_for_issue_type(issue_type)

            # 이슈가 없는 경우: 모든 필드를 초기화하고 종료
            if not issue:
                self.ed_local_id.setText("")
                self.ed_jira_key.setText("")
                self.ed_issue_type.setText("")
                self.ed_summary.setText("")
                # 콤보박스 / 라인에디트 필드 초기화
                self.ed_status.setCurrentText("")
                self.ed_priority.setCurrentText("")
                self.ed_assignee.setText("")
                self.ed_reporter.setText("")
                self.ed_labels.setText("")
                self.ed_components.setText("")
                self.ed_security_level.setText("")
                self.ed_fix_versions.setText("")
                self.ed_affects_versions.setText("")
                self.ed_epic_link.setText("")
                self.ed_sprint.setText("")
                self.ed_rtm_env.setCurrentText("")
                self.ed_due_date.setText("")
                self.ed_created.setText("")
                self.ed_updated.setText("")
                self.ed_attachments.setText("")
                if hasattr(self, "attachments_list"):
                    self.attachments_list.clear()
                self.txt_description.setPlainText("")
                # Activity / Preconditions 초기화
                if hasattr(self, "set_activity_text"):
                    self.set_activity_text("")
                if hasattr(self, "set_preconditions_text"):
                    self.set_preconditions_text("")
                return

            # ------------------------------ 이슈가 있는 경우: Details 필드 채우기
            issue_type_str = (issue.get("issue_type") or "").upper()

            # 기본 메타 필드 매핑 (Local ID / JIRA Key / Issue Type / Summary)
            local_id = issue.get("id")
            self.ed_local_id.setText(str(local_id) if local_id is not None else "")
            self.ed_jira_key.setText(issue.get("jira_key") or "")
            self.ed_issue_type.setText(issue_type_str)
            self.ed_summary.setText(issue.get("summary") or "")

            # 콤보박스/텍스트 필드 설정
            self.ed_status.setCurrentText(issue.get("status") or "")
            self.ed_priority.setCurrentText(issue.get("priority") or "")
            self.ed_assignee.setText(issue.get("assignee") or "")
            self.ed_reporter.setText(issue.get("reporter") or "")
            self.ed_labels.setText(issue.get("labels") or "")
            self.ed_components.setText(issue.get("components") or "")
            self.ed_security_level.setText(issue.get("security_level") or "")
            self.ed_fix_versions.setText(issue.get("fix_versions") or "")
            self.ed_affects_versions.setText(issue.get("affects_versions") or "")
            self.ed_epic_link.setText(issue.get("epic_link") or "")
            self.ed_sprint.setText(issue.get("sprint") or "")
            self.ed_rtm_env.setCurrentText(issue.get("rtm_environment") or "")
            self.ed_due_date.setText(issue.get("due_date") or "")
            self.ed_created.setText(issue.get("created") or "")
            self.ed_updated.setText(issue.get("updated") or "")

            attachments_raw = issue.get("attachments")
            self.ed_attachments.setText(attachments_raw or "")

            # Description
            self.txt_description.setPlainText(issue.get("description") or "")

            # Preconditions (TEST_CASE)
            pre = issue.get("preconditions") or ""
            if hasattr(self, "set_preconditions_text"):
                self.set_preconditions_text(pre)

            # 로컬 Activity 텍스트가 있으면 Activity 영역에 표시
            local_activity = issue.get("local_activity")
            if local_activity is not None and hasattr(self, "set_activity_text"):
                self.set_activity_text(local_activity or "")
                # JIRA 댓글 캐시는 비운다 (로컬 Activity 모드)
                if hasattr(self, "set_activity_comments"):
                    self.set_activity_comments([])

            # Attachments 리스트 표시 (attachments 컬럼은 JIRA JSON 문자열 또는 리스트)
            if hasattr(self, "attachments_list"):
                self._load_attachments_list(attachments_raw)
        finally:
            # 로딩 완료 후에는 다시 사용자 입력을 dirty 로 인식
            self._suppress_dirty = False

    def set_activity_text(self, text: str) -> None:
        """Activity 영역 텍스트를 설정한다 (로컬 메모 또는 JIRA Comments/History)."""
        self.txt_activity.setPlainText(text or "")

    def set_activity_comments(self, comments: list[Dict[str, Any]] | None) -> None:
        """
        Activity 영역에서 사용할 JIRA 댓글 원본 목록을 캐시한다.
        - on_edit_comment / on_delete_comment 에서 마지막 댓글을 찾을 때 사용.
        """
        self._activity_comments = list(comments or [])

    def get_activity_comments(self) -> list[Dict[str, Any]]:
        """현재 캐시된 JIRA 댓글 목록을 반환."""
        return list(self._activity_comments or [])

    def _load_attachments_list(self, attachments_raw: Any) -> None:
        """attachments 컬럼(JSON 문자열 또는 list)을 파싱하여 리스트 위젯에 표시."""
        from PySide6.QtWidgets import QListWidgetItem

        self.attachments_list.clear()

        if not attachments_raw:
            return

        data = None
        try:
            if isinstance(attachments_raw, str):
                data = json.loads(attachments_raw)
            elif isinstance(attachments_raw, list):
                data = attachments_raw
        except Exception:
            # 파싱 실패 시에는 원본 문자열만 한 줄로 보여준다.
            item = QListWidgetItem(str(attachments_raw))
            self.attachments_list.addItem(item)
            return

        if not isinstance(data, list):
            return

        for att in data:
            if not isinstance(att, dict):
                continue
            name = att.get("filename") or att.get("fileName") or att.get("name") or ""
            size = att.get("size") or att.get("filesize")
            att_id = att.get("id") or att.get("attachmentId")
            content_url = att.get("content") or att.get("contentUrl") or att.get("self")
            local_path = att.get("local_path")

            if not name and not (att_id or local_path):
                continue

            text = name
            if size is not None:
                text += f" ({size} bytes)"
            if local_path:
                text += f"  [{local_path}]"

            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, att_id)
            item.setData(Qt.UserRole + 1, content_url)
            # Qt.UserRole + 2 에 로컬 파일 경로(attachments 루트 기준 상대 경로) 저장
            if local_path:
                item.setData(Qt.UserRole + 2, local_path)
            self.attachments_list.addItem(item)

    def get_issue_updates(self) -> Dict[str, Any]:
        """
        Details 탭에서 사용자가 수정한 필드 값을 dict로 반환.
        DB update 및 JIRA sync 시 활용.
        """
        return {
            "summary": self.ed_summary.text().strip(),
            "status": self.ed_status.text().strip(),
            "priority": self.ed_priority.text().strip(),
            "assignee": self.ed_assignee.text().strip(),
            "reporter": self.ed_reporter.text().strip(),
            "labels": self.ed_labels.text().strip(),
            "components": self.ed_components.text().strip(),
            "security_level": self.ed_security_level.text().strip(),
            "fix_versions": self.ed_fix_versions.text().strip(),
            "affects_versions": self.ed_affects_versions.text().strip(),
            "rtm_environment": self.ed_rtm_env.text().strip(),
            "due_date": self.ed_due_date.text().strip(),
            "attachments": self.ed_attachments.text().strip(),
            "description": self.txt_description.toPlainText().strip(),
        }
class PanelWidget(QWidget):
    """
    좌/우 패널 공통 레이아웃:
    - 상단 헤더(라벨 + 버튼)
    - 중간 트리뷰
    - 하단 이슈 탭
    """

    def __init__(self, title: str, is_online: bool = False, parent=None):
        super().__init__(parent)

        self.is_online = is_online

        main_layout = QVBoxLayout(self)

        # 상단 헤더
        header_layout = QHBoxLayout()
        title_label = QLabel(title)
        title_label.setStyleSheet("font-weight: bold;")

        header_layout.addWidget(title_label)
        header_layout.addStretch()

        if is_online:
            # 온라인(JIRA) 패널: 트리 갱신 / Local 로 동기화 / JIRA 엔티티 생성/삭제
            self.btn_refresh = QPushButton("Refresh")
            self.btn_sync_down = QPushButton("Sync → Local")
            self.btn_create_jira = QPushButton("Create in JIRA")
            self.btn_delete_jira = QPushButton("Delete in JIRA")
            header_layout.addWidget(self.btn_refresh)
            header_layout.addWidget(self.btn_sync_down)
            header_layout.addWidget(self.btn_create_jira)
            header_layout.addWidget(self.btn_delete_jira)
        else:
            # 로컬 패널: 상단 헤더에는 제목만 두고, 트리 바로 위 아이콘 툴바에서
            # 폴더/이슈 생성/삭제 및 저장/동기화 아이콘을 제공한다.
            self.btn_add_folder = QToolButton()
            self.btn_delete_folder = QToolButton()
            self.btn_new_issue = QToolButton()
            self.btn_save_issue = QToolButton()
            self.btn_delete_issue = QToolButton()
            self.btn_sync_up = QToolButton()

            style = self.style()
            self.btn_add_folder.setIcon(style.standardIcon(QStyle.SP_FileDialogNewFolder))
            self.btn_delete_folder.setIcon(style.standardIcon(QStyle.SP_TrashIcon))
            self.btn_new_issue.setIcon(style.standardIcon(QStyle.SP_FileIcon))
            self.btn_save_issue.setIcon(style.standardIcon(QStyle.SP_DialogSaveButton))
            self.btn_delete_issue.setIcon(style.standardIcon(QStyle.SP_TrashIcon))
            self.btn_sync_up.setIcon(style.standardIcon(QStyle.SP_ArrowRight))

            self.btn_add_folder.setToolTip("Add Folder")
            self.btn_delete_folder.setToolTip("Delete Folder")
            self.btn_new_issue.setToolTip("New Issue")
            self.btn_save_issue.setToolTip("Save Local Issue")
            self.btn_delete_issue.setToolTip("Delete Issue")
            self.btn_sync_up.setToolTip("Sync selected issue to JIRA/RTM (Push)")

            for btn in [
                self.btn_add_folder,
                self.btn_delete_folder,
                self.btn_new_issue,
                self.btn_save_issue,
                self.btn_delete_issue,
                self.btn_sync_up,
            ]:
                btn.setAutoRaise(True)

            # 헤더에는 우측에 Sync → JIRA 만 간단히 배치
            header_layout.addWidget(self.btn_sync_up)

        main_layout.addLayout(header_layout)

        # 모듈(최상위) 탭바: Dashboard / Requirements / Test Cases / Test Plans / Test Executions / Defects
        from PySide6.QtWidgets import QTabBar

        self.module_tab_bar = QTabBar()
        self.module_tab_bar.addTab("Dashboard")
        self.module_tab_bar.addTab("Requirements")
        self.module_tab_bar.addTab("Test Cases")
        self.module_tab_bar.addTab("Test Plans")
        self.module_tab_bar.addTab("Test Executions")
        self.module_tab_bar.addTab("Defects")
        self.module_tab_bar.setExpanding(False)
        main_layout.addWidget(self.module_tab_bar)

        # 모듈 탭 바로 아래, 트리 바로 위에 아이콘 기반 툴바 배치
        if not is_online:
            tree_toolbar = QHBoxLayout()
            tree_toolbar.setContentsMargins(0, 0, 0, 0)
            tree_toolbar.setSpacing(4)
            tree_toolbar.addWidget(self.btn_add_folder)
            tree_toolbar.addWidget(self.btn_delete_folder)
            tree_toolbar.addSpacing(8)
            tree_toolbar.addWidget(self.btn_new_issue)
            tree_toolbar.addWidget(self.btn_save_issue)
            tree_toolbar.addWidget(self.btn_delete_issue)
            tree_toolbar.addSpacing(8)
            tree_toolbar.addWidget(self.btn_sync_up)

            # Test Case 전용 상단 액션 (모듈 탭 바로 아래, 트리 위에 위치)
            self.btn_execute_tc = QPushButton("Execute")
            self.btn_add_to_testplan = QPushButton("Add to Test Plan")
            self.btn_link_requirement = QPushButton("Link to Requirement")
            self.btn_execute_tc.setVisible(False)
            self.btn_add_to_testplan.setVisible(False)
            self.btn_link_requirement.setVisible(False)

            tree_toolbar.addSpacing(8)
            tree_toolbar.addWidget(self.btn_execute_tc)
            tree_toolbar.addWidget(self.btn_add_to_testplan)
            tree_toolbar.addWidget(self.btn_link_requirement)

            tree_toolbar.addStretch()
            main_layout.addLayout(tree_toolbar)
        else:
            # 온라인 패널용 트리 툴바: 폴더/이슈 생성/삭제, 동기화 관련 아이콘 제공
            tree_toolbar = QHBoxLayout()
            tree_toolbar.setContentsMargins(0, 0, 0, 0)
            tree_toolbar.setSpacing(4)

            self.btn_online_refresh = QToolButton()
            self.btn_online_sync_down = QToolButton()
            self.btn_online_create = QToolButton()
            self.btn_online_delete = QToolButton()

            style = self.style()
            self.btn_online_refresh.setIcon(style.standardIcon(QStyle.SP_BrowserReload))
            self.btn_online_sync_down.setIcon(style.standardIcon(QStyle.SP_ArrowLeft))
            self.btn_online_create.setIcon(style.standardIcon(QStyle.SP_FileIcon))
            self.btn_online_delete.setIcon(style.standardIcon(QStyle.SP_TrashIcon))

            self.btn_online_refresh.setToolTip("Refresh Online Tree")
            self.btn_online_sync_down.setToolTip("Sync from JIRA → Local")
            self.btn_online_create.setToolTip("Create in JIRA/RTM")
            self.btn_online_delete.setToolTip("Delete in JIRA/RTM")

            for btn in [
                self.btn_online_refresh,
                self.btn_online_sync_down,
                self.btn_online_create,
                self.btn_online_delete,
            ]:
                btn.setAutoRaise(True)

            tree_toolbar.addWidget(self.btn_online_refresh)
            tree_toolbar.addWidget(self.btn_online_sync_down)
            tree_toolbar.addSpacing(8)
            tree_toolbar.addWidget(self.btn_online_create)
            tree_toolbar.addWidget(self.btn_online_delete)
            tree_toolbar.addStretch()
            main_layout.addLayout(tree_toolbar)

        # 트리 + 탭을 좌/우로 배치
        tree_and_tabs = QSplitter(Qt.Horizontal)
        tree_and_tabs.setChildrenCollapsible(False)

        # 트리 (좌측/우측 패널 공통)
        self.tree_view = QTreeView()
        self.tree_view.setHeaderHidden(True)
        # 다중 선택은 Ctrl/Shift 기반으로만 허용하고,
        # 마우스 드래그는 Drag & Drop 이동에 사용하도록 한다.
        self.tree_view.setSelectionMode(QTreeView.ExtendedSelection)
        if not is_online:
            # 로컬 트리: 드래그로 폴더/이슈를 이동할 수 있도록 Drag&Drop 활성화
            self.tree_view.setDragEnabled(True)
            self.tree_view.setAcceptDrops(True)
            self.tree_view.setDropIndicatorShown(True)
            self.tree_view.setDragDropMode(QAbstractItemView.DragDrop)
        # 양쪽 패널 모두 동일하게 최소 폭을 두고, 세부 내용 영역이 우선적으로 변동되도록 한다.
        self.tree_view.setMinimumWidth(260)
        tree_and_tabs.addWidget(self.tree_view)

        # 이슈 탭 (우측)
        self.issue_tabs = IssueTabWidget()
        tree_and_tabs.addWidget(self.issue_tabs)

        # 창 전체 폭이 줄어들 때는 우측(세부내용) 영역이 주로 줄어들고,
        # 좌측 트리 폭은 최대한 유지되도록 스트레치 비율을 조정
        tree_and_tabs.setStretchFactor(0, 0)
        tree_and_tabs.setStretchFactor(1, 1)

        main_layout.addWidget(tree_and_tabs)


class MainWindow(QMainWindow):
    def __init__(self, db_path=None, config_path: str = "jira_config.json", mode: str = "both"):
        """
        mode:
          - "local"  : 로컬 관리 전용 윈도우 (온라인 패널/툴바 요소 숨김)
          - "online" : 온라인(JIRA RTM) 전용 윈도우 (로컬 패널/툴바 요소 숨김)
          - "both"   : 기존처럼 좌/우 패널을 한 창에 모두 배치
        """
        super().__init__()
        self.logger = get_logger(__name__)
        self.mode = mode.lower()
        # 동기화 전략: "server" = 서버 우선(Pull 중심), "local" = 로컬 우선(Push 중심), "merge" = 최대 항목 병합
        self.sync_mode: str = "server"
        # 현재/과거에 GUI 상에서 수정되었으나 저장되지 않은 이슈 id 집합
        self.dirty_issue_ids: set[int] = set()
        # 현재 선택된 이슈에 대해 편의를 위해 별도 플래그도 유지
        self.current_issue_dirty: bool = False

        # 설정 파일 경로를 보관하여 Settings 다이얼로그에서 사용
        self.config_path = config_path

        self.logger.info(
            "MainWindow init: db_path=%s, config_path=%s, mode=%s",
            db_path,
            config_path,
            self.mode,
        )
        title_suffix = {
            "local": " - Local",
            "online": " - Online",
            "both": "",
        }.get(self.mode, "")
        self.setWindowTitle(f"RTM Local Manager (JIRA RTM Sync Tool){title_suffix}")
        self.resize(1600, 900)

        # DB, Jira client 초기화
        self.conn = get_connection(db_path)
        init_db(self.conn)

        try:
            self.jira_config = load_config_from_file(config_path)
            # 프로젝트 레코드는 JIRA 사용 여부와 무관하게 생성/로드한다.
            self.project = get_or_create_project(
                self.conn,
                project_key=self.jira_config.project_key,
                project_id=self.jira_config.project_id,
                name=self.jira_config.project_key,
                base_url=self.jira_config.base_url,
            )

            base_url = str(self.jira_config.base_url or "").strip()
            if "your-jira-server.example.com" in base_url:
                # 샘플/플레이스홀더 URL 그대로인 경우: JIRA 연동 비활성화 (오프라인 모드)
                self.jira_client = None
                self.jira_available = False
                self.logger.warning(
                    "Jira config loaded but base_url is a placeholder "
                    "'https://your-jira-server.example.com'; JIRA integration disabled."
                )
            else:
                self.jira_client = JiraRTMClient(self.jira_config)
                self.jira_available = True
        except Exception as e:
            # 설정 파일이 없거나 잘못된 경우: 오프라인 전용(Local Only) 프로젝트로 시작
            self.jira_config = None
            self.project = get_or_create_project(
                self.conn,
                project_key="LOCAL",
                project_id=0,
                name="Local Only",
                base_url=None,
            )
            self.jira_client = None
            self.jira_available = False
            # jira_config.json 이 없는 것은 정상적인 최초 실행 시나리오이므로,
            # 불필요한 전체 스택트레이스를 로그에 남기지 않는다.
            if isinstance(e, FileNotFoundError):
                self.logger.info("Jira config file not found; starting in offline (LOCAL) mode.")
            else:
                self.logger.warning("Jira config not loaded: %s", e, exc_info=True)

        # currently selected issue id (local DB)
        self.current_issue_id: int | None = None
        self.current_issue_type: str | None = None
        self.current_testexecution_id: int | None = None

        # 로컬 트리 Cut/Copy/Paste 용 클립보드
        # 예: {"mode": "cut"|"copy", "items": [{"kind": "FOLDER"/"ISSUE", "id": "..."}]}
        self._tree_clipboard: Dict[str, Any] | None = None

        # 로컬 / 온라인 트리의 이슈 타입 필터
        # - Dashboard 탭: None (필터 없음, 모든 이슈 타입 표시)
        # - 나머지 탭: REQUIREMENT / TEST_CASE / TEST_PLAN / TEST_EXECUTION / DEFECT
        self.local_issue_type_filter: str | None = None
        self.online_issue_type_filter: str | None = None

        # JIRA 필드 옵션 캐시 (statuses, priorities 등)
        self.jira_field_options: Dict[str, List[str]] = {}
        # 사용자 정의 필드 프리셋 (예: RTM Environment)
        self.field_presets: Dict[str, List[str]] = load_presets()
        # 로컬 Activity / Attachments 동작 설정
        self.local_settings: Dict[str, Any] = load_local_settings()
        # Excel 컬럼 매핑 (시트/논리필드 ↔ 엑셀 헤더 이름)
        self.excel_column_mapping: Dict[str, Dict[str, str]] = load_excel_mapping()

        # 메뉴바 생성 (File / Local / JIRA / Help)
        self._create_menu_bar()

        # 메인 리본/툴바 (카테고리별 그룹화된 버튼 모음)
        toolbar = QToolBar("Ribbon")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)

        style = self.style()

        from PySide6.QtWidgets import QGroupBox

        ribbon = QWidget()
        ribbon_layout = QHBoxLayout(ribbon)
        ribbon_layout.setContentsMargins(4, 2, 4, 2)
        ribbon_layout.setSpacing(8)

        # ------------------------------------------------------------------
        # Local
        # ------------------------------------------------------------------
        grp_local = QGroupBox("Local")
        gl = QVBoxLayout(grp_local)

        # Excel Import/Export
        row_local_excel = QHBoxLayout()
        self.btn_import_excel = QPushButton("Import Excel")
        self.btn_export_excel = QPushButton("Export Excel")
        self.btn_import_excel.setIcon(style.standardIcon(QStyle.SP_DialogOpenButton))
        self.btn_import_excel.setToolTip("Import project data from Excel → Local SQLite")
        self.btn_export_excel.setIcon(style.standardIcon(QStyle.SP_DialogSaveButton))
        self.btn_export_excel.setToolTip("Export current project from Local SQLite → Excel file")
        row_local_excel.addWidget(self.btn_import_excel)
        row_local_excel.addWidget(self.btn_export_excel)
        gl.addLayout(row_local_excel)

        # Local Issues / Folders
        row_local_issue = QHBoxLayout()
        self.btn_ribbon_save_issue = QPushButton("Save Issue")
        self.btn_ribbon_new_issue = QPushButton("New Issue")
        self.btn_ribbon_delete_issue = QPushButton("Delete Issue")
        row_local_issue.addWidget(self.btn_ribbon_save_issue)
        row_local_issue.addWidget(self.btn_ribbon_new_issue)
        row_local_issue.addWidget(self.btn_ribbon_delete_issue)
        gl.addLayout(row_local_issue)

        row_local_folder = QHBoxLayout()
        self.btn_ribbon_add_folder = QPushButton("Add Folder")
        self.btn_ribbon_delete_folder = QPushButton("Delete Folder")
        row_local_folder.addWidget(self.btn_ribbon_add_folder)
        row_local_folder.addWidget(self.btn_ribbon_delete_folder)
        gl.addLayout(row_local_folder)

        ribbon_layout.addWidget(grp_local)

        # ------------------------------------------------------------------
        # JIRA / RTM (Online only)
        # ------------------------------------------------------------------
        grp_jira = QGroupBox("JIRA / RTM (Online only)")
        gj = QVBoxLayout(grp_jira)

        row_jira_tree = QHBoxLayout()
        self.btn_ribbon_refresh_online = QPushButton("Refresh Online Tree")
        self.btn_ribbon_delete_in_jira = QPushButton("Delete in JIRA")
        row_jira_tree.addWidget(self.btn_ribbon_refresh_online)
        row_jira_tree.addWidget(self.btn_ribbon_delete_in_jira)
        gj.addLayout(row_jira_tree)

        row_jira_issue = QHBoxLayout()
        self.btn_ribbon_create_in_jira = QPushButton("Create in JIRA")
        row_jira_issue.addWidget(self.btn_ribbon_create_in_jira)
        gj.addLayout(row_jira_issue)

        row_jira_search = QHBoxLayout()
        self.jira_filter_edit = QLineEdit()
        self.jira_filter_edit.setPlaceholderText("JQL or JIRA key (e.g. project = KVHSICCU)")
        self.btn_ribbon_search_jira = QPushButton("Search")
        row_jira_search.addWidget(self.jira_filter_edit)
        row_jira_search.addWidget(self.btn_ribbon_search_jira)
        gj.addLayout(row_jira_search)

        ribbon_layout.addWidget(grp_jira)

        # ------------------------------------------------------------------
        # Sync (JIRA ↔ Local)
        # ------------------------------------------------------------------
        grp_sync = QGroupBox("Sync (JIRA ↔ Local)")
        gs = QVBoxLayout(grp_sync)

        # 1행: 버튼들
        row_sync = QHBoxLayout()
        self.btn_full_sync = QPushButton("Full Sync (Tree)")
        self.btn_full_sync.setIcon(style.standardIcon(QStyle.SP_BrowserReload))
        self.btn_full_sync.setToolTip("Full RTM tree sync: JIRA → Local SQLite")
        self.btn_ribbon_pull = QPushButton("Pull Issue")
        self.btn_ribbon_push = QPushButton("Push Issue")
        row_sync.addWidget(self.btn_full_sync)
        row_sync.addWidget(self.btn_ribbon_pull)
        row_sync.addWidget(self.btn_ribbon_push)
        gs.addLayout(row_sync)

        # 2행: 동기화 모드 선택
        from PySide6.QtWidgets import QComboBox, QLabel

        row_sync_mode = QHBoxLayout()
        lbl_sync_mode = QLabel("Sync Mode:")
        self.cmb_sync_mode = QComboBox()
        self.cmb_sync_mode.addItems(
            [
                "서버 우선 (Server → Local)",
                "로컬 우선 (Local → Server)",
                "병합 (Max / Merge)",
            ]
        )
        # 기본 값: 서버 우선
        self.cmb_sync_mode.setCurrentIndex(0)
        row_sync_mode.addWidget(lbl_sync_mode)
        row_sync_mode.addWidget(self.cmb_sync_mode)
        row_sync_mode.addStretch()
        gs.addLayout(row_sync_mode)

        ribbon_layout.addWidget(grp_sync)

        # ------------------------------------------------------------------
        # Execution / Defects
        # ------------------------------------------------------------------
        grp_exec = QGroupBox("Execution / Defects")
        ge = QVBoxLayout(grp_exec)

        row_exec_main = QHBoxLayout()
        self.btn_ribbon_execute_plan = QPushButton("Execute Test Plan")
        self.btn_ribbon_link_defect = QPushButton("Link Defect")
        self.btn_ribbon_clear_defects = QPushButton("Clear Defects")
        row_exec_main.addWidget(self.btn_ribbon_execute_plan)
        row_exec_main.addWidget(self.btn_ribbon_link_defect)
        row_exec_main.addWidget(self.btn_ribbon_clear_defects)
        ge.addLayout(row_exec_main)

        row_exec_def = QHBoxLayout()
        self.btn_ribbon_refresh_defects = QPushButton("Refresh Defects")
        self.btn_ribbon_export_te_report = QPushButton("Export TE Report")
        row_exec_def.addWidget(self.btn_ribbon_refresh_defects)
        row_exec_def.addWidget(self.btn_ribbon_export_te_report)
        ge.addLayout(row_exec_def)

        ribbon_layout.addWidget(grp_exec)

        # ------------------------------------------------------------------
        # View / Layout
        # ------------------------------------------------------------------
        grp_view = QGroupBox("View / Layout")
        gv = QVBoxLayout(grp_view)

        row_view_panels = QHBoxLayout()
        self.btn_ribbon_show_local = QPushButton("Local Panel")
        self.btn_ribbon_show_online = QPushButton("Online Panel")
        self.btn_ribbon_show_local.setCheckable(True)
        self.btn_ribbon_show_online.setCheckable(True)
        # 실제 표시 상태(mode)에 맞게 초기 체크 상태를 아래에서 다시 설정한다.
        self.btn_ribbon_show_local.setChecked(True)
        self.btn_ribbon_show_online.setChecked(True)
        row_view_panels.addWidget(self.btn_ribbon_show_local)
        row_view_panels.addWidget(self.btn_ribbon_show_online)
        gv.addLayout(row_view_panels)

        row_view_layout = QHBoxLayout()
        self.btn_ribbon_layout_h = QPushButton("Left / Right")
        self.btn_ribbon_layout_v = QPushButton("Top / Bottom")
        row_view_layout.addWidget(self.btn_ribbon_layout_h)
        row_view_layout.addWidget(self.btn_ribbon_layout_v)
        gv.addLayout(row_view_layout)

        ribbon_layout.addWidget(grp_view)

        ribbon_layout.addStretch(1)
        toolbar.addWidget(ribbon)

        # 중앙 위젯: 좌/우(기본) 또는 상/하(옵션) 배치를 위한 스플리터
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        self.left_panel = PanelWidget("Local", is_online=False)
        self.right_panel = PanelWidget("JIRA RTM (Online)", is_online=True)

        # 스플리터를 속성으로 보관 (레이아웃 전환 시 재사용)
        self.main_splitter = splitter

        # 기본 레이아웃 모드: 좌(로컬) / 우(온라인)
        self.layout_mode: str = "horizontal"  # "horizontal" or "vertical"

        splitter.addWidget(self.left_panel)
        splitter.addWidget(self.right_panel)

        # mode 에 따라 기본 표시 여부/체크 상태 제어
        if self.mode == "local":
            # 온라인 패널은 기본적으로 숨김
            self.right_panel.setVisible(False)
            if hasattr(self, "act_view_online"):
                self.act_view_online.setChecked(False)
            # 리본 버튼 상태도 View 메뉴와 동일하게 맞춘다.
            if hasattr(self, "btn_ribbon_show_online"):
                self.btn_ribbon_show_online.setChecked(False)
            if hasattr(self, "btn_ribbon_show_local"):
                self.btn_ribbon_show_local.setChecked(True)
        elif self.mode == "online":
            # 로컬 패널은 숨기고 온라인만 표시
            self.left_panel.setVisible(False)
            if hasattr(self, "act_view_local"):
                self.act_view_local.setChecked(False)
            if hasattr(self, "btn_ribbon_show_local"):
                self.btn_ribbon_show_local.setChecked(False)
            if hasattr(self, "btn_ribbon_show_online"):
                self.btn_ribbon_show_online.setChecked(True)
        else:
            # both 모드인 경우, 두 패널/버튼 모두 활성
            if hasattr(self, "act_view_local"):
                self.act_view_local.setChecked(True)
            if hasattr(self, "act_view_online"):
                self.act_view_online.setChecked(True)
            if hasattr(self, "btn_ribbon_show_local"):
                self.btn_ribbon_show_local.setChecked(True)
            if hasattr(self, "btn_ribbon_show_online"):
                self.btn_ribbon_show_online.setChecked(True)

        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 3)

        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.addWidget(splitter)

        self.setCentralWidget(container)

        # 상태바
        status = QStatusBar()
        self.setStatusBar(status)
        self.status_bar = status

        # JIRA 연결 상태 표시 라벨 (항상 보이는 인디케이터)
        from PySide6.QtWidgets import QLabel as _QLabel  # 지역 import 로 순환 방지
        self.jira_status_label = _QLabel()
        self.status_bar.addPermanentWidget(self.jira_status_label)
        self._update_jira_status_label()

        # 좌/우 패널의 모듈 탭바를 MainWindow 핸들러에 연결
        self.left_panel.module_tab_bar.currentChanged.connect(self._on_local_issue_type_tab_changed)
        self.right_panel.module_tab_bar.currentChanged.connect(self._on_online_issue_type_tab_changed)

        # 시그널 연결
        self._connect_signals()

        # JIRA 메타데이터(상태/우선순위 등)를 불러와 Details 탭 콤보박스에 적용
        if self.jira_available and self.jira_client:
            self._load_jira_field_options()
            try:
                self.left_panel.issue_tabs.apply_field_options(self.jira_field_options)
            except Exception as e_meta:
                self.logger.warning("Failed to apply JIRA field options to Details tab: %s", e_meta)

    def _load_jira_field_options(self) -> None:
        """
        JIRA REST API 를 사용하여 상태(Status) / 우선순위(Priority) 등의
        사전 정의 값을 조회해 Details 탭 콤보박스에서 선택할 수 있게 한다.
        """
        options: Dict[str, List[str]] = {}

        # Status 목록
        try:
            res = self.jira_client.get_statuses()
            statuses: List[str] = []
            if isinstance(res, list):
                for s in res:
                    if isinstance(s, dict):
                        name = s.get("name")
                        if name:
                            statuses.append(str(name))
            options["status"] = statuses
        except Exception as e_stat:
            self.logger.warning("Failed to load JIRA statuses: %s", e_stat)

        # Priority 목록
        try:
            res = self.jira_client.get_priorities()
            priorities: List[str] = []
            if isinstance(res, list):
                for p in res:
                    if isinstance(p, dict):
                        name = p.get("name")
                        if name:
                            priorities.append(str(name))
            options["priority"] = priorities
        except Exception as e_pri:
            self.logger.warning("Failed to load JIRA priorities: %s", e_pri)

        # Components / Versions (프로젝트 메타)
        try:
            proj_meta = self.jira_client.get_project_metadata()
            comps: List[str] = []
            vers: List[str] = []
            if isinstance(proj_meta, dict):
                comp_items = proj_meta.get("components")
                if isinstance(comp_items, list):
                    for c in comp_items:
                        if isinstance(c, dict):
                            name = c.get("name")
                            if name:
                                comps.append(str(name))
                ver_items = proj_meta.get("versions")
                if isinstance(ver_items, list):
                    for v in ver_items:
                        if isinstance(v, dict):
                            name = v.get("name")
                            if name:
                                vers.append(str(name))
            # JIRA 메타 + 사용자 프리셋 병합 (중복 제거)
            preset_comps = self.field_presets.get("components", [])
            preset_vers = self.field_presets.get("versions", [])
            comp_all = list(dict.fromkeys([*preset_comps, *comps]))
            ver_all = list(dict.fromkeys([*preset_vers, *vers]))
            options["components"] = comp_all
            options["versions"] = ver_all
        except Exception as e_proj:
            self.logger.warning("Failed to load JIRA project metadata (components/versions): %s", e_proj)

        # Issue Link Types (Relations 탭에서 relation_type 선택용)
        try:
            link_raw = self.jira_client.get_issue_link_types()
            link_types: List[str] = []
            if isinstance(link_raw, dict):
                items = link_raw.get("issueLinkTypes")
                if isinstance(items, list):
                    for it in items:
                        if isinstance(it, dict):
                            name = it.get("name")
                            if name:
                                link_types.append(str(name))
            preset_rel = self.field_presets.get("relation_types", [])
            rel_all = list(dict.fromkeys([*preset_rel, *link_types]))
            options["relation_types"] = rel_all
        except Exception as e_lt:
            self.logger.warning("Failed to load JIRA issue link types: %s", e_lt)

        # RTM Environment 는 JIRA 에서 직접 제공되지 않으므로, 로컬 프리셋 기반으로 설정
        options["rtm_environment"] = self.field_presets.get("rtm_environment", [])

        self.jira_field_options = options

    def on_refresh_jira_metadata_clicked(self) -> None:
        """
        Settings > Refresh JIRA Metadata 메뉴:
        - JIRA 에서 상태/우선순위 메타데이터를 다시 읽어와 Details 탭 콤보박스에 반영한다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is offline; cannot refresh metadata.")
            return
        self._load_jira_field_options()
        try:
            self.left_panel.issue_tabs.apply_field_options(self.jira_field_options)
            self.status_bar.showMessage("Refreshed JIRA metadata (statuses/priorities).")
        except Exception as e_meta:
            self.status_bar.showMessage(f"Failed to apply JIRA metadata: {e_meta}")

    def on_edit_field_presets_clicked(self) -> None:
        """
        Settings > Field Presets... 메뉴:
        - 콤보박스로 구성된 필드(Status / Priority / RTM Environment / Components / Versions / Relation Types)의
          프리셋 목록을 탭 형태로 편집할 수 있는 UI 를 제공한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QListWidget,
            QListWidgetItem,
            QLineEdit,
            QPushButton,
            QDialogButtonBox,
            QLabel,
            QTabWidget,
            QWidget,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Field Presets")
        vbox = QVBoxLayout(dlg)

        tabs = QTabWidget()
        vbox.addWidget(tabs)

        # 편집 대상 필드와 라벨 텍스트 정의
        preset_defs: list[tuple[str, str, str]] = [
            ("status", "Status", "예: To Do, In Progress, Done"),
            ("priority", "Priority", "예: Highest, High, Medium, Low"),
            ("rtm_environment", "RTM Environment", "예: DEV, QA, STAGE, PROD"),
            ("components", "Components", "예: Backend, API, UI"),
            ("versions", "Versions", "예: 1.0.0, 1.0.1"),
            ("relation_types", "Relation Types", "예: Tests, Is tested by, Relates, Blocks"),
        ]

        tab_lists: dict[str, QListWidget] = {}
        tab_edits: dict[str, QLineEdit] = {}

        for key, title, placeholder in preset_defs:
            page = QWidget()
            pv = QVBoxLayout(page)
            pv.addWidget(QLabel(f"{title} presets (one per line):"))

            lst = QListWidget()
            for val in self.field_presets.get(key, []):
                if not val:
                    continue
                lst.addItem(QListWidgetItem(str(val)))
            pv.addWidget(lst)

            btn_row = QHBoxLayout()
            ed_new = QLineEdit()
            ed_new.setPlaceholderText(placeholder)
            btn_add = QPushButton("Add/Update")
            btn_del = QPushButton("Delete Selected")
            btn_row.addWidget(ed_new)
            btn_row.addWidget(btn_add)
            btn_row.addWidget(btn_del)
            pv.addLayout(btn_row)

            def make_add(lst_ref: QListWidget, ed_ref: QLineEdit):
                def _on_add() -> None:
                    name = ed_ref.text().strip()
                    if not name:
                        return
                    for i in range(lst_ref.count()):
                        item = lst_ref.item(i)
                        if item.text().strip().lower() == name.lower():
                            item.setText(name)
                            ed_ref.clear()
                            return
                    lst_ref.addItem(QListWidgetItem(name))
                    ed_ref.clear()

                return _on_add

            def make_del(lst_ref: QListWidget):
                def _on_del() -> None:
                    selected = lst_ref.selectedIndexes()
                    if not selected:
                        return
                    rows = sorted({idx.row() for idx in selected}, reverse=True)
                    for r in rows:
                        lst_ref.takeItem(r)

                return _on_del

            btn_add.clicked.connect(make_add(lst, ed_new))
            btn_del.clicked.connect(make_del(lst))

            tab_lists[key] = lst
            tab_edits[key] = ed_new
            tabs.addTab(page, title)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept() -> None:
            # 각 탭의 리스트 내용을 field_presets 에 반영
            for key, _, _ in preset_defs:
                lst = tab_lists.get(key)
                if not lst:
                    continue
                values: List[str] = []
                for i in range(lst.count()):
                    txt = lst.item(i).text().strip()
                    if txt and txt not in values:
                        values.append(txt)
                self.field_presets[key] = values

            save_presets(self.field_presets)

            # 프리셋을 다시 옵션에 반영
            self._load_jira_field_options()
            try:
                self.left_panel.issue_tabs.apply_field_options(self.jira_field_options)
            except Exception as e_meta:
                self.logger.warning("Failed to apply updated presets to Details/Relations tabs: %s", e_meta)

            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def on_edit_api_endpoints_clicked(self) -> None:
        """
        Settings > REST API Endpoint Settings... :
        - jira_config.json 의 endpoints 맵을 표 형태로 편집할 수 있는 다이얼로그.
        - JIRA REST / RTM REST 의 모든 주요 엔드포인트를 key/path 로 노출한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QLabel,
            QTableWidget,
            QTableWidgetItem,
            QHeaderView,
            QDialogButtonBox,
            QMessageBox,
        )
        import re

        dlg = QDialog(self)
        dlg.setWindowTitle("REST API Endpoint Settings")

        vbox = QVBoxLayout(dlg)
        vbox.addWidget(
            QLabel(
                "각 엔드포인트의 논리 이름(key)과 경로 템플릿(path)을 수정할 수 있습니다.\n"
                "예: /rest/rtm/1.0/api/test-case/{testKey}\n"
                "{...} 안의 플레이스홀더 이름은 코드에서 사용하는 것과 동일해야 합니다.\n\n"
                "여러 버전의 엔드포인트를 우선순위대로 시도하려면 세미콜론(;)으로 구분해서 입력하세요.\n"
                "예: /rest/rtm/1.0/api/tree/{projectId}/{treeType}; /rest/rtm/1.0/api/v2/tree/{projectId}/{treeType}"
            )
        )

        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Key", "Path Template", "Allowed Placeholders"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)

        # 현재 설정 값 + 기본값 병합
        current_endpoints = getattr(self.jira_config, "endpoints", None) or DEFAULT_ENDPOINTS
        merged = dict(DEFAULT_ENDPOINTS)
        merged.update(current_endpoints)

        for key in sorted(merged.keys()):
            row = table.rowCount()
            table.insertRow(row)
            item_key = QTableWidgetItem(key)
            item_key.setFlags(item_key.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 0, item_key)

            item_path = QTableWidgetItem(merged[key])
            table.setItem(row, 1, item_path)

            allowed = DEFAULT_ENDPOINT_PARAMS.get(key, [])
            item_allowed = QTableWidgetItem(", ".join(allowed) if allowed else "(none)")
            item_allowed.setFlags(item_allowed.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 2, item_allowed)

        vbox.addWidget(table)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept() -> None:
            # 테이블에서 endpoints dict 재구성
            new_eps: Dict[str, str] = {}
            invalid_msgs: list[str] = []

            for r in range(table.rowCount()):
                key_item = table.item(r, 0)
                path_item = table.item(r, 1)
                if not key_item:
                    continue
                k = key_item.text().strip()
                if not k:
                    continue
                v = (path_item.text().strip() if path_item else "") or DEFAULT_ENDPOINTS.get(k, "")

                # 플레이스홀더 유효성 검사: {name} 패턴을 추출하여, 정의된 리스트 내인지 확인
                used = set(re.findall(r"\{([^{}]+)\}", v))
                allowed_names = set(DEFAULT_ENDPOINT_PARAMS.get(k, []))
                unknown = used - allowed_names
                if unknown:
                    invalid_msgs.append(
                        f"- {k}: 허용되지 않은 플레이스홀더 {', '.join(sorted(unknown))}"
                    )

                new_eps[k] = v

            if invalid_msgs:
                QMessageBox.warning(
                    dlg,
                    "Invalid placeholders",
                    "다음 엔드포인트의 Path Template 에 허용되지 않은 플레이스홀더가 포함되어 있습니다.\n\n"
                    + "\n".join(invalid_msgs)
                    + "\n\n"
                    + "각 Key 별로 'Allowed Placeholders' 열에 표시된 이름만 사용할 수 있습니다.",
                )
                return

            try:
                # jira_config 가 없을 수도 있으므로 기본값으로 생성
                cfg = getattr(self, "jira_config", None)
                if not isinstance(cfg, JiraConfig):
                    if not self.config_path:
                        QMessageBox.warning(dlg, "Config Error", "config_path 가 설정되지 않았습니다.")
                        return
                    # 최소한의 기본값으로 새 JiraConfig 생성
                    cfg = JiraConfig(
                        base_url="",
                        username="",
                        api_token="",
                        project_key="LOCAL",
                        project_id=0,
                        endpoints=new_eps,
                    )
                else:
                    cfg.endpoints = new_eps

                save_config_to_file(self.config_path, cfg)
                self.jira_config = cfg
            except Exception as e:
                QMessageBox.warning(
                    dlg,
                    "Save Failed",
                    f"jira_config.json 저장 중 오류가 발생했습니다.\n\n{e}",
                )
                return

            self.status_bar.showMessage("REST API 엔드포인트 설정을 저장했습니다.", 5000)
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def on_edit_api_settings_clicked(self) -> None:
        """
        Settings > REST API & Auth Settings... 메뉴:
        - jira_config.json 의 base_url / username / api_token / project_key / project_id 를
          GUI 에서 수정할 수 있도록 한다.
        - Test Connection 버튼으로 설정 값이 유효한지 간단히 검증한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QFormLayout,
            QLineEdit,
            QDialogButtonBox,
            QLabel,
            QHBoxLayout,
            QPushButton,
            QMessageBox,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("REST API & Auth Settings")

        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("JIRA / RTM 서버 연결 설정을 입력하세요."))

        form = QFormLayout()

        ed_base_url = QLineEdit()
        ed_username = QLineEdit()
        ed_token = QLineEdit()
        ed_token.setEchoMode(QLineEdit.Password)
        ed_project_key = QLineEdit()
        ed_project_id = QLineEdit()

        # 현재 설정 값 로드 (가능한 경우)
        cfg = getattr(self, "jira_config", None)
        if isinstance(cfg, JiraConfig):
            ed_base_url.setText(cfg.base_url or "")
            ed_username.setText(cfg.username or "")
            ed_token.setText(cfg.api_token or "")
            ed_project_key.setText(cfg.project_key or "")
            ed_project_id.setText(str(cfg.project_id))

        form.addRow("Base URL", ed_base_url)
        form.addRow("Username", ed_username)
        form.addRow("API Token / Password", ed_token)
        form.addRow("Project Key", ed_project_key)
        form.addRow("Project ID", ed_project_id)

        vbox.addLayout(form)

        # Test Connection 버튼
        btn_row = QHBoxLayout()
        btn_test = QPushButton("Test Connection")
        btn_row.addWidget(btn_test)
        btn_row.addStretch()
        vbox.addLayout(btn_row)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def build_config_from_inputs() -> JiraConfig | None:
            base_url = ed_base_url.text().strip()
            username = ed_username.text().strip()
            token = ed_token.text().strip()
            project_key = ed_project_key.text().strip() or "LOCAL"
            project_id_text = ed_project_id.text().strip()
            try:
                project_id_val = int(project_id_text) if project_id_text else 0
            except ValueError:
                QMessageBox.warning(
                    dlg,
                    "Invalid Project ID",
                    "Project ID 는 정수여야 합니다.",
                )
                return None
            if not base_url:
                QMessageBox.warning(dlg, "Invalid Base URL", "Base URL 을 입력하세요.")
                return None
            if not username or not token:
                QMessageBox.warning(
                    dlg,
                    "Invalid Auth",
                    "Username 과 API Token/Password 를 모두 입력하세요.",
                )
                return None
            return JiraConfig(
                base_url=base_url,
                username=username,
                api_token=token,
                project_key=project_key,
                project_id=project_id_val,
            )

        def on_test_clicked() -> None:
            cfg_obj = build_config_from_inputs()
            if cfg_obj is None:
                return
            client = JiraRTMClient(cfg_obj)
            try:
                # 간단한 메타데이터 호출로 연결 확인
                client.get_priorities()
                QMessageBox.information(
                    dlg,
                    "Test Connection",
                    "JIRA/RTM REST API 호출에 성공했습니다.",
                )
            except Exception as e:
                QMessageBox.warning(
                    dlg,
                    "Test Connection Failed",
                    f"JIRA/RTM REST API 호출에 실패했습니다.\n\n{e}",
                )

        def on_accept_clicked() -> None:
            cfg_obj = build_config_from_inputs()
            if cfg_obj is None:
                return
            try:
                save_config_to_file(self.config_path, cfg_obj)
            except Exception as e:
                QMessageBox.warning(
                    dlg,
                    "Save Failed",
                    f"jira_config.json 저장 중 오류가 발생했습니다.\n\n{e}",
                )
                return

            # 저장이 성공했으면 메모리 상의 설정과 Jira 클라이언트도 갱신
            self.jira_config = cfg_obj
            try:
                self.project = get_or_create_project(
                    self.conn,
                    project_key=cfg_obj.project_key,
                    project_id=cfg_obj.project_id,
                    name=cfg_obj.project_key,
                    base_url=cfg_obj.base_url,
                )
            except Exception:
                # project 초기화 실패는 치명적이지 않으므로 로그만 남기고 진행
                self.logger.exception("Failed to update local project info from API settings.")

            try:
                self.jira_client = JiraRTMClient(cfg_obj)
                self.jira_available = True
            except Exception:
                self.jira_client = None
                self.jira_available = False
                self.logger.exception("Failed to initialize JiraRTMClient with new settings.")

            self.status_bar.showMessage("REST API & Auth 설정을 저장했습니다.", 5000)
            dlg.accept()

        btn_test.clicked.connect(on_test_clicked)
        btn_box.accepted.connect(on_accept_clicked)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def on_open_api_tester_clicked(self) -> None:
        """
        Settings > REST API Tester... :
        - 사용자가 임의의 메서드/경로/바디를 입력해서 REST 호출을 시험할 수 있는 다이얼로그.
        - 현재 jira_config 의 base_url / auth 를 사용한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QLabel,
            QComboBox,
            QLineEdit,
            QPlainTextEdit,
            QPushButton,
            QDialogButtonBox,
            QMessageBox,
        )
        from requests.auth import HTTPBasicAuth
        import requests

        cfg: JiraConfig | None = getattr(self, "jira_config", None)
        if not isinstance(cfg, JiraConfig):
            QMessageBox.warning(
                self,
                "Config not loaded",
                "JIRA 설정이 로드되지 않았습니다. 먼저 REST API & Auth Settings 에서 설정을 저장하세요.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("REST API Tester")

        vbox = QVBoxLayout(dlg)

        row_base = QHBoxLayout()
        row_base.addWidget(QLabel("Base URL:"))
        ed_base_url = QLineEdit()
        ed_base_url.setText(cfg.base_url or "")
        row_base.addWidget(ed_base_url)
        vbox.addLayout(row_base)

        # Method + Path
        row_top = QHBoxLayout()
        cmb_method = QComboBox()
        cmb_method.addItems(["GET", "POST", "PUT", "DELETE"])
        ed_path = QLineEdit()
        ed_path.setPlaceholderText("예: /rest/api/2/priority 또는 /rest/rtm/1.0/api/test-execution/...")
        row_top.addWidget(QLabel("Method:"))
        row_top.addWidget(cmb_method)
        row_top.addWidget(QLabel("Path:"))
        row_top.addWidget(ed_path)
        vbox.addLayout(row_top)

        # Request body
        vbox.addWidget(QLabel("Request Body (JSON/text, 선택 사항):"))
        ed_body = QPlainTextEdit()
        vbox.addWidget(ed_body)

        # Send 버튼
        btn_send = QPushButton("Send Request")
        vbox.addWidget(btn_send)

        # Response 표시
        lbl_status = QLabel("Status: -")
        vbox.addWidget(lbl_status)
        vbox.addWidget(QLabel("Response Body:"))
        ed_resp = QPlainTextEdit()
        ed_resp.setReadOnly(True)
        vbox.addWidget(ed_resp)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        vbox.addWidget(btn_box)

        def on_send() -> None:
            method = cmb_method.currentText().strip().upper()
            path = ed_path.text().strip()
            if not path:
                QMessageBox.warning(dlg, "Invalid Path", "Path 를 입력하세요.")
                return
            base_url = ed_base_url.text().strip() or cfg.base_url
            if not base_url:
                QMessageBox.warning(dlg, "Invalid Base URL", "Base URL 을 입력하세요.")
                return
            url = base_url.rstrip("/") + path

            text_body = ed_body.toPlainText().strip()
            json_body = None
            data_body = None
            headers = {"Accept": "application/json"}
            if text_body:
                # JSON 시도, 실패하면 그냥 text 로 전송
                try:
                    import json as _json

                    json_body = _json.loads(text_body)
                    headers["Content-Type"] = "application/json"
                except Exception:
                    data_body = text_body

            try:
                resp = requests.request(
                    method,
                    url,
                    headers=headers,
                    auth=HTTPBasicAuth(cfg.username, cfg.api_token),
                    json=json_body,
                    data=data_body,
                )
            except Exception as e:
                lbl_status.setText(f"Status: ERROR ({e})")
                ed_resp.setPlainText(str(e))
                return

            lbl_status.setText(f"Status: {resp.status_code}")
            text = resp.text
            # JSON 응답이면 보기 좋게 포맷팅
            try:
                import json as _json

                parsed = _json.loads(text)
                text = _json.dumps(parsed, ensure_ascii=False, indent=2)
            except Exception:
                pass
            ed_resp.setPlainText(text)

        btn_send.clicked.connect(on_send)
        btn_box.rejected.connect(dlg.reject)

        dlg.resize(900, 700)
        dlg.exec()

    def on_edit_excel_mapping_clicked(self) -> None:
        """
        Settings > Excel Column Mapping... 메뉴:
        - 각 시트별로 'DB 논리 필드 이름' 과 '엑셀 헤더 텍스트' 를 매핑할 수 있는 UI.
        - Import 시 열 순서와 무관하게, 이 매핑 정보를 기반으로 컬럼을 찾는다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QTabWidget,
            QWidget,
            QTableWidget,
            QTableWidgetItem,
            QLabel,
            QDialogButtonBox,
        )
        from PySide6.QtCore import Qt

        # 시트별로 import 에서 사용하는 논리 필드 목록 정의
        SHEET_FIELDS: Dict[str, List[str]] = {
            "Issues": [
                "id",
                "excel_key",
                "jira_key",
                "issue_type",
                "folder_path",
                "summary",
                "description",
                "status",
                "priority",
                "assignee",
                "reporter",
                "labels",
                "components",
                "security_level",
                "fix_versions",
                "affects_versions",
                "rtm_environment",
                "due_date",
                "created",
                "updated",
                "attachments",
                "epic_link",
                "sprint",
            ],
            "TestcaseSteps": [
                "issue_id",
                "excel_key",
                "issue_jira_key",
                "preconditions",
                "group_no",
                "order_no",
                "action",
                "input",
                "expected",
            ],
            "Relations": [
                "src_jira_key",
                "src_excel_key",
                "dst_jira_key",
                "dst_excel_key",
                "relation_type",
            ],
            "TestPlanTestcases": [
                "testplan_jira_key",
                "testplan_excel_key",
                "testcase_jira_key",
                "testcase_excel_key",
                "order_no",
            ],
            "TestExecutions": [
                "testexecution_jira_key",
                "testexecution_excel_key",
                "environment",
                "start_date",
                "end_date",
                "result",
                "executed_by",
            ],
            "TestcaseExecutions": [
                "testexecution_jira_key",
                "testexecution_excel_key",
                "testcase_jira_key",
                "testcase_excel_key",
                "order_no",
                "assignee",
                "result",
                "actual_time",
                "rtm_environment",
                "defects",
                "tce_test_key",
            ],
            "TestcaseStepExecutions": [
                "testexecution_jira_key",
                "testexecution_excel_key",
                "testcase_jira_key",
                "testcase_excel_key",
                "group_no",
                "order_no",
                "status",
                "actual_result",
                "evidence",
            ],
        }

        dlg = QDialog(self)
        dlg.setWindowTitle("Excel Column Mapping")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(
            QLabel(
                "각 시트별로 DB 논리 필드 이름과 실제 엑셀 헤더 텍스트를 매핑합니다.\n"
                "- 기본값: 논리 필드 이름과 동일한 헤더를 가진다고 가정합니다.\n"
                "- 엑셀에서 헤더 이름을 바꾼 경우, 오른쪽 열에 실제 헤더 이름을 입력해 주세요."
            )
        )

        tabs = QTabWidget()
        vbox.addWidget(tabs)

        tables: Dict[str, QTableWidget] = {}

        # 현재 매핑 로드 (메모리 캐시가 아닌 파일 기준으로 한 번 더 읽는다)
        current_mapping = load_excel_mapping()

        for sheet_name, fields in SHEET_FIELDS.items():
            page = QWidget()
            pv = QVBoxLayout(page)

            table = QTableWidget(len(fields), 2, parent=page)
            table.setHorizontalHeaderLabels(["DB Field", "Excel Column Header"])
            table.horizontalHeader().setStretchLastSection(True)

            sheet_map = current_mapping.get(sheet_name, {})

            for row, logical_name in enumerate(fields):
                # 왼쪽: 논리 필드 이름 (읽기 전용)
                item_field = QTableWidgetItem(logical_name)
                item_field.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                table.setItem(row, 0, item_field)

                # 오른쪽: 실제 엑셀 헤더 이름 (기본은 논리 이름과 동일)
                excel_header = sheet_map.get(logical_name, logical_name)
                item_header = QTableWidgetItem(excel_header)
                table.setItem(row, 1, item_header)

            pv.addWidget(table)
            tabs.addTab(page, sheet_name)
            tables[sheet_name] = table

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept() -> None:
            new_mapping: Dict[str, Dict[str, str]] = {}
            for sheet_name, table in tables.items():
                sheet_map: Dict[str, str] = {}
                for row in range(table.rowCount()):
                    field_item = table.item(row, 0)
                    header_item = table.item(row, 1)
                    if field_item is None or header_item is None:
                        continue
                    logical_name = field_item.text().strip()
                    excel_header = header_item.text().strip()
                    # 기본값(논리 이름과 동일)인 경우는 저장하지 않아도 된다.
                    if not logical_name or not excel_header:
                        continue
                    if excel_header == logical_name:
                        continue
                    sheet_map[logical_name] = excel_header
                if sheet_map:
                    new_mapping[sheet_name] = sheet_map

            # 메모리/파일 모두 갱신
            self.excel_column_mapping = new_mapping
            save_excel_mapping(new_mapping)
            self.status_bar.showMessage("Saved Excel column mapping.", 5000)
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

        # 최초 로드: 초기 필터는 Dashboard(None)로 두고 양쪽 트리 모두 로드
        self._on_local_issue_type_tab_changed(self.left_panel.module_tab_bar.currentIndex())
        self._on_online_issue_type_tab_changed(self.right_panel.module_tab_bar.currentIndex())

        # JIRA가 사용 가능한 경우에만 오른쪽 패널 활성화
        if not self.jira_available:
            self.right_panel.setEnabled(False)
            self.status_bar.showMessage("Offline mode (jira_config.json not loaded)")

    # --------------------------------------------------------------------- Dirty state helpers

    def mark_current_issue_dirty(self) -> None:
        """
        현재 선택된 이슈에 대해 '저장되지 않은 변경 있음' 상태를 기록하고
        로컬 트리에서 주황색으로 표시되도록 한다.

        - Details 탭의 필드 변경, Description 수정 등에서 호출된다.
        """
        if self.current_issue_id is None:
            return
        try:
            issue_id = int(self.current_issue_id)
        except Exception:
            return

        # 전체 dirty 세트에 추가
        self.dirty_issue_ids.add(issue_id)
        # 현재 이슈 플래그도 true 로
        self.current_issue_dirty = True

        # 트리 색상/툴팁 갱신
        try:
            self.reload_local_tree()
        except Exception:
            # 트리 갱신 실패가 전체 동작을 막지 않도록 예외는 무시
            pass

    # --------------------------------------------------------------------- GUI helpers

    def _update_jira_status_label(self) -> None:
        """
        상태바 우측에 JIRA 연결 상태를 'JIRA: Online / Offline' 으로 표시.
        """
        if getattr(self, "jira_available", False):
            text = "JIRA: Online"
            color = "#007700"
        else:
            text = "JIRA: Offline"
            color = "#AA0000"
        self.jira_status_label.setText(text)
        self.jira_status_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    # ------------------------------------------------------------------ Top-level module tabs (Local / Online)

    def _issue_type_from_index(self, index: int) -> str | None:
        """
        모듈 탭 인덱스를 이슈 타입 문자열로 변환한다.
        - 0: Dashboard (None 반환, 필터 없음)
        - 1: REQUIREMENT
        - 2: TEST_CASE
        - 3: TEST_PLAN
        - 4: TEST_EXECUTION
        - 5: DEFECT
        """
        mapping = {
            1: "REQUIREMENT",
            2: "TEST_CASE",
            3: "TEST_PLAN",
            4: "TEST_EXECUTION",
            5: "DEFECT",
        }
        return mapping.get(index)

    def _tree_type_from_issue_type(self, issue_type: str | None) -> str:
        """
        RTM REST API 의 treeType 문자열로 변환한다.
        - requirements
        - test-cases
        - test-plans
        - test-executions
        - defects
        """
        mapping = {
            "REQUIREMENT": "requirements",
            "TEST_CASE": "test-cases",
            "TEST_PLAN": "test-plans",
            "TEST_EXECUTION": "test-executions",
            "DEFECT": "defects",
            None: "requirements",
        }
        return mapping.get(issue_type, "requirements")

    def _on_local_issue_type_tab_changed(self, index: int) -> None:
        """
        좌측(로컬) 패널의 모듈 탭 변경 시 호출.
        로컬 트리 필터와 상단 로컬 전용 버튼 상태를 업데이트한다.
        """
        self.local_issue_type_filter = self._issue_type_from_index(index)

        # 로컬 패널 헤더의 버튼 텍스트/표시를 이슈 타입에 맞게 조정
        if self.left_panel and hasattr(self.left_panel, "btn_new_issue"):
            it = self.local_issue_type_filter or ""
            text_map = {
                "REQUIREMENT": "Create Requirement",
                "TEST_CASE": "Create Test Case",
                "TEST_PLAN": "Create Test Plan",
                "TEST_EXECUTION": "Create Test Execution",
                "DEFECT": "Create Defect",
            }
            self.left_panel.btn_new_issue.setText(text_map.get(it, "New Issue"))

            # Test Case 전용 상단 액션(Execute / Add to Test Plan / Link to Requirement)은
            # TEST_CASE 모듈 탭에서만 표시
            is_tc = it == "TEST_CASE"
            if hasattr(self.left_panel, "btn_execute_tc"):
                self.left_panel.btn_execute_tc.setVisible(is_tc)
            if hasattr(self.left_panel, "btn_add_to_testplan"):
                self.left_panel.btn_add_to_testplan.setVisible(is_tc)
            if hasattr(self.left_panel, "btn_link_requirement"):
                self.left_panel.btn_link_requirement.setVisible(is_tc)

        # 로컬 트리 재로드
        self.reload_local_tree()

    def _on_online_issue_type_tab_changed(self, index: int) -> None:
        """
        우측(온라인) 패널의 모듈 탭 변경 시 호출.
        온라인 RTM 트리 필터를 업데이트한다.
        """
        self.online_issue_type_filter = self._issue_type_from_index(index)

        if self.jira_available:
            try:
                self.on_refresh_online_tree()
            except Exception:
                # 온라인 트리 로딩 실패는 치명적이지 않으므로 무시
                pass

    # ------------------------------------------------------------------ Test Cases top-level actions (MainWindow)

    def _selected_local_testcase_ids(self) -> list[int]:
        """
        좌측(Local) 트리에서 선택된 TEST_CASE 이슈들의 로컬 ID 리스트를 반환.
        상단 탭이 Test Cases 일 때 Add to Test Plan / Link to Requirement 에서 사용한다.
        """
        ids: list[int] = []
        view = self.left_panel.tree_view
        model = view.model()
        if model is None:
            return ids
        for idx in view.selectedIndexes():
            if idx.column() != 0:
                continue
            item = model.itemFromIndex(idx)
            if not item:
                continue
            node_type = item.data(Qt.UserRole)
            if node_type != "ISSUE":
                continue
            issue_id = item.data(Qt.UserRole + 1)
            issue_type = (item.data(Qt.UserRole + 3) or "").upper()
            if not issue_id or issue_type != "TEST_CASE":
                continue
            try:
                ids.append(int(issue_id))
            except (TypeError, ValueError):
                continue
        return ids

    def on_add_testcases_to_testplan_clicked(self) -> None:
        """
        상단 Test Cases 탭에서 'Add to Test Plan' 버튼:
        - 트리에서 선택된 TEST_CASE 이슈들을 하나의 Test Plan 에 추가한다.
        """
        tc_ids = self._selected_local_testcase_ids()
        if not tc_ids:
            self.status_bar.showMessage("No Test Cases selected in the tree.")
            return

        # 프로젝트 내 TEST_PLAN 목록 조회
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'TEST_PLAN'
             ORDER BY jira_key, summary
            """,
            (self.project.id,),
        )
        rows = cur.fetchall()
        plans = [dict(r) for r in rows]
        if not plans:
            self.status_bar.showMessage("No TEST_PLAN issues found; cannot add Test Cases.")
            return

        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QDialogButtonBox,
            QListWidget,
            QListWidgetItem,
            QLabel,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Add to Test Plan")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select a Test Plan to add the selected Test Cases:"))

        lst = QListWidget()
        for p in plans:
            key = p.get("jira_key") or f"ID={p.get('id')}"
            text = key
            if p.get("summary"):
                text += f" - {p.get('summary')}"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, p.get("id"))
            lst.addItem(item)
        vbox.addWidget(lst)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            item = lst.currentItem()
            if not item:
                self.status_bar.showMessage("No Test Plan selected.")
                return
            tp_id = item.data(Qt.UserRole)
            if not tp_id:
                return

            # 기존 매핑을 읽고 선택된 Test Case 들을 추가
            existing = get_testplan_testcases(self.conn, int(tp_id))
            existing_tc_ids = {
                int(r.get("testcase_id")) for r in existing if r.get("testcase_id")
            }
            records = [
                {"testcase_id": int(r["testcase_id"]), "order_no": int(r.get("order_no", 0) or 0)}
                for r in existing
            ]

            for tc_id in tc_ids:
                if tc_id in existing_tc_ids:
                    continue
                records.append({"testcase_id": tc_id, "order_no": 0})

            # order_no 재정렬
            for idx, rec in enumerate(records, start=1):
                rec["order_no"] = idx

            replace_testplan_testcases(self.conn, int(tp_id), records)
            self.status_bar.showMessage(
                f"Added {len(tc_ids)} Test Case(s) to Test Plan (id={tp_id})."
            )
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def on_link_testcases_to_requirement_clicked(self) -> None:
        """
        상단 Test Cases 탭에서 'Link to Requirement' 버튼:
        - 트리에서 선택된 TEST_CASE 이슈들을 선택한 Requirement 와 relation 으로 연결한다.
        """
        tc_ids = self._selected_local_testcase_ids()
        if not tc_ids:
            self.status_bar.showMessage("No Test Cases selected in the tree.")
            return

        # 프로젝트 내 REQUIREMENT 목록 조회
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, jira_key, summary
              FROM issues
             WHERE project_id = ? AND is_deleted = 0 AND UPPER(issue_type) = 'REQUIREMENT'
             ORDER BY jira_key, summary
            """,
            (self.project.id,),
        )
        rows = cur.fetchall()
        reqs = [dict(r) for r in rows]
        if not reqs:
            self.status_bar.showMessage("No REQUIREMENT issues found; cannot link.")
            return

        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QDialogButtonBox,
            QListWidget,
            QListWidgetItem,
            QLabel,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Link to Requirement")
        vbox = QVBoxLayout(dlg)
        vbox.addWidget(QLabel("Select a Requirement to link the selected Test Cases:"))

        lst = QListWidget()
        for r in reqs:
            key = r.get("jira_key") or f"ID={r.get('id')}"
            text = key
            if r.get("summary"):
                text += f" - {r.get('summary')}"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, r.get("id"))
            lst.addItem(item)
        vbox.addWidget(lst)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def on_accept():
            item = lst.currentItem()
            if not item:
                self.status_bar.showMessage("No Requirement selected.")
                return
            req_id = item.data(Qt.UserRole)
            if not req_id:
                return

            # Requirement 를 src 로 하는 relations 를 읽고, 선택된 Test Case 들을 추가
            existing_rels = get_relations_for_issue(self.conn, int(req_id))
            existing_dst_ids = {
                int(r.get("dst_issue_id")) for r in existing_rels if r.get("dst_issue_id")
            }
            new_rels = [
                {"dst_issue_id": r.get("dst_issue_id"), "relation_type": r.get("relation_type") or ""}
                for r in existing_rels
            ]
            for tc_id in tc_ids:
                if tc_id in existing_dst_ids:
                    continue
                new_rels.append({"dst_issue_id": tc_id, "relation_type": "Tests"})

            replace_relations_for_issue(self.conn, int(req_id), new_rels)

            # 현재 선택 이슈가 이 Requirement 인 경우, Relations / Requirements / Linked Test Cases 갱신
            if self.current_issue_id == int(req_id):
                rels = get_relations_for_issue(self.conn, int(req_id))
                if hasattr(self.left_panel.issue_tabs, "load_relations"):
                    link_types = self.jira_field_options.get("relation_types", [])
                    self.left_panel.issue_tabs.load_relations(rels, link_types)
                if hasattr(self.left_panel.issue_tabs, "load_requirements"):
                    reqs2 = [
                        r
                        for r in rels
                        if (r.get("dst_issue_type") or "").upper() == "REQUIREMENT"
                    ]
                    self.left_panel.issue_tabs.load_requirements(reqs2)
                if hasattr(self.left_panel.issue_tabs, "load_linked_testcases"):
                    tcs2 = [
                        r
                        for r in rels
                        if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"
                    ]
                    self.left_panel.issue_tabs.load_linked_testcases(tcs2)

            self.status_bar.showMessage(
                f"Linked {len(tc_ids)} Test Case(s) to Requirement (id={req_id})."
            )
            dlg.accept()

        btn_box.accepted.connect(on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

        # 로컬/온라인 트리를 현재 필터에 맞게 다시 로드
        self.reload_local_tree()
        if self.jira_available:
            try:
                self.on_refresh_online_tree()
            except Exception:
                # 온라인 트리 로딩 실패는 치명적이지 않으므로 무시
                pass

    def on_jira_filter_search(self) -> None:
        """
        리본 메뉴 우측 JIRA 필터 입력창에서 엔터를 치면,
        입력된 문자열을 JQL 로 간주하고 Jira 검색 REST API
        (예: GET /rest/api/2/search, [Jira REST API 9.12.0](https://docs.atlassian.com/software/jira/docs/api/REST/9.12.0/))
        를 호출하여 Online 패널 트리에 검색 결과를 표시한다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is not configured; cannot search.")
            return

        text = self.jira_filter_edit.text().strip()
        if not text:
            return

        try:
            # JQL 검색 수행
            res = self.jira_client.search_issues(text, max_results=100)
            issues = res.get("issues") or []

            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["JIRA Search Results"])
            root_item = model.invisibleRootItem()

            for issue in issues:
                key = issue.get("key") or ""
                fields = issue.get("fields") or {}
                summary = fields.get("summary") or ""
                issue_type_name = ""
                issuetype = fields.get("issuetype")
                if isinstance(issuetype, dict):
                    issue_type_name = issuetype.get("name") or ""
                # RTM 타입 이름을 로컬 타입(REQUIREMENT/TEST_CASE/...) 으로 매핑
                local_type = map_rtm_type_to_local(issue_type_name) if issue_type_name else ""

                label = f"{key} - {summary}" if summary else key
                item = QStandardItem(label)
                item.setEditable(False)
                # node_type(UserRole) 에 이슈 타입, UserRole+1 에 jira_key 저장 (온라인 트리와 동일 규약)
                item.setData(local_type or issue_type_name.upper(), Qt.UserRole)
                item.setData(key, Qt.UserRole + 1)
                root_item.appendRow(item)

            self.right_panel.tree_view.setModel(model)
            self.right_panel.tree_view.expandAll()
            self.right_panel.tree_view.setSelectionMode(QTreeView.ExtendedSelection)

            # 검색 결과가 있으면 첫 번째 이슈를 자동 선택하여 상세 정보 로드
            if issues:
                first_index = model.index(0, 0)
                if first_index.isValid():
                    self.right_panel.tree_view.setCurrentIndex(first_index)
                    # selectionModel 시그널이 이미 on_online_tree_selection_changed 에 연결되어 있으므로
                    # 여기서 별도 호출은 필요 없음.

            self.status_bar.showMessage(f"JIRA search finished: {len(issues)} issue(s) found.")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to search JIRA: {e}")
            print(f"[ERROR] JIRA JQL search failed for query='{text}': {e}")

    # --------------------------------------------------------------------- Tree + selection handling

    def reload_local_tree(self):
        """
        현재 project 의 folders/issues 를 SQLite 에서 읽어와
        왼쪽(Local) 패널의 QTreeView 에 바인딩한다.
        """
        if not self.project:
            return

        tree_data = fetch_folder_tree(self.conn, self.project.id)

        type_filter = self.local_issue_type_filter

        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["Name"])

        root_item = model.invisibleRootItem()

        # 아이콘 준비: 폴더 / 이슈 (이슈 타입별로 다른 아이콘 사용)
        style = self.left_panel.style()
        folder_icon = style.standardIcon(QStyle.SP_DirIcon)
        default_issue_icon = style.standardIcon(QStyle.SP_FileIcon)
        issue_icon_map = {
            "REQUIREMENT": style.standardIcon(QStyle.SP_FileDialogEnd),
            "TEST_CASE": style.standardIcon(QStyle.SP_FileDialogDetailedView),
            "TEST_PLAN": style.standardIcon(QStyle.SP_FileDialogContentsView),
            "TEST_EXECUTION": style.standardIcon(QStyle.SP_CommandLink),
            "DEFECT": style.standardIcon(QStyle.SP_MessageBoxWarning),
        }

        def add_node(parent_item: QStandardItem, node: Dict[str, Any]) -> bool:
            """
            현재 이슈 타입 필터에 맞는 이슈가 하나도 없는 폴더 서브트리는 숨기고,
            각 이슈 타입 탭마다 '자신에게 해당하는' 폴더 트리만 보이도록 구성한다.

            RETURNS: 이 노드(폴더/이슈)가 실제로 트리에 추가되었는지 여부.
            """
            node_type = node.get("node_type") or "FOLDER"

            # 이슈 노드
            if node_type == "ISSUE":
                issue_type = (node.get("issue_type") or "").upper()
                if type_filter and issue_type != type_filter:
                    return False

                jira_key = node.get("jira_key") or ""
                label = node.get("summary") or jira_key or f"ISSUE {node.get('id')}"
                item = QStandardItem(label)
                icon = issue_icon_map.get(issue_type, default_issue_icon)
                item.setIcon(icon)
                item.setEditable(False)
                item.setData("ISSUE", Qt.UserRole)
                item.setData(node.get("id"), Qt.UserRole + 1)
                item.setData(jira_key, Qt.UserRole + 2)
                item.setData(node.get("issue_type") or "", Qt.UserRole + 3)

                # 트리에서 로컬 전용 / JIRA 연동 / 편집 중(unsaved) 이슈 구분:
                # - jira_key 없음: 로컬 전용 → 녹색 텍스트
                # - jira_key 있음: JIRA/RTM 연동 이슈 → 기본색
                # - dirty_issue_ids 에 포함된 이슈 → 주황색 텍스트로 "편집됨" 표시
                issue_id_val = int(node.get("id")) if node.get("id") is not None else None
                is_dirty = issue_id_val is not None and issue_id_val in getattr(self, "dirty_issue_ids", set())
                if is_dirty:
                    item.setForeground(QColor("#CC6600"))  # Unsaved edits
                    base_tip = (
                        "Local only issue (no JIRA key)"
                        if not jira_key
                        else f"Linked to JIRA/RTM issue: {jira_key}"
                    )
                    item.setToolTip(base_tip + "\n(Unsaved local changes)")
                else:
                    if not jira_key:
                        item.setForeground(QColor("#007700"))  # Local only
                        item.setToolTip("Local only issue (no JIRA key)")
                    else:
                        # 기본 색으로 두고, Jira Key 정보만 툴팁에 표시
                        item.setToolTip(f"Linked to JIRA/RTM issue: {jira_key}")

                parent_item.appendRow(item)
                return True

            # 폴더 노드
            label = node.get("name") or f"Folder {node.get('id')}"
            folder_id = str(node.get("id") or "")
            item = QStandardItem(label)
            item.setEditable(False)
            item.setData("FOLDER", Qt.UserRole)
            item.setData(folder_id, Qt.UserRole + 1)
            item.setIcon(folder_icon)

            has_visible_child = False
            for child in node.get("children", []):
                if add_node(item, child):
                    has_visible_child = True

            # 사용자가 로컬에서 생성한 폴더: LOCAL-<TYPE>-<uuid> 형태
            if folder_id.startswith("LOCAL-"):
                parts = folder_id.split("-", 2)
                local_type = parts[1].upper() if len(parts) >= 3 else ""

                # 구버전(타입 정보 없는 LOCAL-xxxx)은 어떤 탭에도 표시하지 않음
                if not local_type:
                    return False

                # 현재 탭 타입과 다른 로컬 폴더는 숨김
                if type_filter and local_type != type_filter:
                    return False

                # 자신의 타입 탭에서는, 자식이 없어도 항상 표시
                parent_item.appendRow(item)
                return True

            # RTM 에서 내려온 폴더는, 현재 타입 이슈/폴더가 하나도 없으면 숨긴다.
            if not has_visible_child:
                return False

            parent_item.appendRow(item)
            return True

        for root in tree_data.get("roots", []):
            add_node(root_item, root)

        self.left_panel.tree_view.setModel(model)
        self.left_panel.tree_view.expandAll()
        self.left_panel.tree_view.setSelectionMode(QTreeView.ExtendedSelection)

        # selectionModel 이 새로 생성되므로, selectionChanged 시그널을 다시 연결한다.
        try:
            self.left_panel.tree_view.selectionModel().selectionChanged.connect(
                self.on_local_tree_selection_changed
            )
        except Exception:
            pass

    # --------------------------------------------------------------------- Local issue create/delete

    def on_new_local_issue_clicked(self):
        """
        현재 선택된 폴더/이슈를 기준으로 새 로컬 이슈를 생성한다.
        - 좌측 모듈 탭(local_issue_type_filter)에 따라 issue_type 결정.
        - 선택이 폴더면 해당 folder_id 하위에 생성.
        - 선택이 이슈면 그 이슈의 folder_id 하위에 생성.
        - 선택이 없으면 folder_id 없이 프로젝트 루트에 생성.
        """
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot create issue.")
            return

        issue_type = (self.local_issue_type_filter or "REQUIREMENT").upper()

        folder_id: str | None = None
        model = self.left_panel.tree_view.model()
        selection_model = self.left_panel.tree_view.selectionModel()
        if model is not None and selection_model is not None:
            indexes = selection_model.selectedIndexes()
            if indexes:
                item = model.itemFromIndex(indexes[0])
                if item:
                    kind = item.data(Qt.UserRole)
                    if kind == "FOLDER":
                        folder_id = item.data(Qt.UserRole + 1)
                    elif kind == "ISSUE":
                        issue_id = item.data(Qt.UserRole + 1)
                        if issue_id:
                            issue = get_issue_by_id(self.conn, int(issue_id))
                            if issue:
                                folder_id = issue.get("folder_id")

        summary = f"New {issue_type}"
        new_id = create_local_issue(self.conn, self.project.id, issue_type=issue_type, folder_id=folder_id, summary=summary)
        self.logger.info("Created new local issue id=%s, type=%s, folder_id=%s", new_id, issue_type, folder_id)
        self.status_bar.showMessage(f"Created new local issue (type={issue_type}).")
        self.reload_local_tree()

    def on_delete_local_issue_clicked(self):
        """
        현재 선택된 로컬 이슈(복수 선택 포함)를 soft delete 한다(is_deleted=1).
        폴더는 삭제하지 않으며, 이슈가 선택되지 않은 경우 아무 것도 하지 않는다.
        """
        model = self.left_panel.tree_view.model()
        selection_model = self.left_panel.tree_view.selectionModel()
        if model is None or selection_model is None:
            self.status_bar.showMessage("No local tree to delete from.")
            return

        indexes = selection_model.selectedIndexes()
        issue_ids = []
        for idx in indexes:
            item = model.itemFromIndex(idx)
            if item and item.data(Qt.UserRole) == "ISSUE":
                issue_id = item.data(Qt.UserRole + 1)
                if issue_id is not None:
                    issue_ids.append(int(issue_id))

        if not issue_ids:
            self.status_bar.showMessage("No issue selected to delete.")
            return

        # 사용자 확인 (복수 선택 요약)
        count = len(issue_ids)
        ret = QMessageBox.question(
            self,
            "Delete Local Issue",
            f"Delete {count} local issue(s)?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        for iid in issue_ids:
            soft_delete_issue(self.conn, iid)

        self.status_bar.showMessage(f"Deleted {count} local issue(s).")
        self.reload_local_tree()

    # --------------------------------------------------------------------- Local folder create/delete

    def on_add_local_folder_clicked(self):
        """
        현재 선택된 폴더/이슈를 기준으로 새 폴더를 생성한다.
        - 선택이 폴더면 그 하위에
        - 선택이 이슈면 해당 이슈의 folder_id 하위에
        - 선택이 없으면 루트에
        """
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot create folder.")
            return

        from PySide6.QtWidgets import QInputDialog

        name, ok = QInputDialog.getText(self, "New Folder", "Folder name:")
        if not ok or not name.strip():
            return
        name = name.strip()

        folder_parent_id: str | None = None
        model = self.left_panel.tree_view.model()
        selection_model = self.left_panel.tree_view.selectionModel()
        if model is not None and selection_model is not None:
            indexes = selection_model.selectedIndexes()
            if indexes:
                item = model.itemFromIndex(indexes[0])
                if item:
                    kind = item.data(Qt.UserRole)
                    if kind == "FOLDER":
                        folder_parent_id = item.data(Qt.UserRole + 1)
                    elif kind == "ISSUE":
                        issue_id = item.data(Qt.UserRole + 1)
                        if issue_id:
                            issue = get_issue_by_id(self.conn, int(issue_id))
                            if issue:
                                folder_parent_id = issue.get("folder_id")

        issue_type = (self.local_issue_type_filter or "REQUIREMENT").upper()
        new_folder_id = create_folder_node(
            self.conn,
            project_id=self.project.id,
            name=name,
            parent_id=folder_parent_id,
            issue_type=issue_type,
        )
        self.logger.info("Created new folder id=%s under parent=%s", new_folder_id, folder_parent_id)

        # RTM Tree API 와의 정합성: JIRA 가 사용 가능한 경우, 서버 트리에도 폴더 생성 시도
        if self.jira_available and self.jira_client:
            try:
                # parent_id 가 특정 RTM testKey 와 직접적으로 1:1 맵핑되는 구조는 아니므로,
                # 여기서는 프로젝트 루트 하위에 생성하는 간단한 예시로 둔다.
                self.jira_client.create_tree_folder(
                    project_id=self.project.project_id,
                    name=name,
                    parent_test_key=None,
                    issue_type=issue_type,
                )
            except Exception as e_tree:
                self.logger.warning("Failed to create folder in RTM tree: %s", e_tree)

        self.status_bar.showMessage(f"Created new folder '{name}'.")
        self.reload_local_tree()

    def on_delete_local_folder_clicked(self):
        """
        현재 선택된 폴더들을 삭제한다(복수 선택 포함).
        - 하위에 폴더/이슈가 있으면 해당 폴더는 삭제하지 않고 경고 메시지를 표시.
        """
        model = self.left_panel.tree_view.model()
        selection_model = self.left_panel.tree_view.selectionModel()
        if model is None or selection_model is None:
            self.status_bar.showMessage("No local tree to delete from.")
            return

        indexes = selection_model.selectedIndexes()
        folder_ids: list[str] = []
        for idx in indexes:
            item = model.itemFromIndex(idx)
            if item and item.data(Qt.UserRole) == "FOLDER":
                fid = item.data(Qt.UserRole + 1)
                if fid is not None and fid not in folder_ids:
                    folder_ids.append(str(fid))

        if not folder_ids:
            self.status_bar.showMessage("No folder selected to delete.")
            return

        # 사용자 확인 (복수 선택 요약)
        count = len(folder_ids)
        ret = QMessageBox.question(
            self,
            "Delete Folder",
            f"Delete {count} selected folder(s)?\n\n(Only empty folders can be deleted.)",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        deleted = 0
        failed = 0
        for folder_id in folder_ids:
            ok = delete_folder_if_empty(self.conn, str(folder_id))
            if ok:
                deleted += 1
            else:
                failed += 1

        # TODO: RTM Tree API 와 동기화 (트리 구조와 폴더-이슈 매핑 정책이 확정되면
        #       delete_tree_folder(testKey) 호출을 통해 서버 트리에서도 삭제 처리)
        # 현재는 로컬 전용 삭제로 두고, 향후 RTM tree 의 folder <-> local folder_id 매핑을
        # 프로젝트 설정이나 별도 테이블로 관리하도록 확장할 수 있다.

        if deleted:
            self.status_bar.showMessage(
                f"Deleted {deleted} folder(s)."
                + (" Some folders are not empty and were not deleted." if failed else ""),
            )
        elif failed:
            self.status_bar.showMessage(
                "Folder(s) are not empty; cannot delete. Remove child folders/issues first."
            )

        self.reload_local_tree()

    def on_local_tree_selection_changed(self, selected, deselected):
        """
        왼쪽(Local) 트리에서 선택된 이슈가 변경되었을 때 Details 탭에 로드하고,
        current_issue_id / current_issue_type 을 갱신한다.
        복수 선택이더라도 Details 탭은 '첫 번째 이슈' 기준으로 표시한다.
        """
        model = self.left_panel.tree_view.model()
        if model is None:
            self.current_issue_id = None
            self.current_issue_type = None
            self.left_panel.issue_tabs.set_issue(None)
            return

        selected_indexes = self.left_panel.tree_view.selectionModel().selectedIndexes()
        issue_id = None
        issue_type = None

        for idx in selected_indexes:
            item = model.itemFromIndex(idx)
            if item and item.data(Qt.UserRole) == "ISSUE":
                issue_id = item.data(Qt.UserRole + 1)
                issue_type = item.data(Qt.UserRole + 3)
                break

        self.current_issue_id = issue_id
        self.current_issue_type = issue_type

        if issue_id is None:
            self.left_panel.issue_tabs.set_issue(None)
            return

        issue = get_issue_by_id(self.conn, issue_id)
        if not issue:
            self.left_panel.issue_tabs.set_issue(None)
            return

        tabs = self.left_panel.issue_tabs
        tabs.set_issue(issue)

        # ------------------------------------------------------------------
        # 하위 탭 데이터 로드 (로컬 DB 기준)
        issue_type = (issue.get("issue_type") or "").upper()

        # Relations (모든 이슈 공통)
        try:
            rels = get_relations_for_issue(self.conn, issue_id)
            if hasattr(tabs, "load_relations"):
                tabs.load_relations(rels)
            # Requirements 탭: relations 중 REQUIREMENT 타입만 필터링
            if hasattr(tabs, "load_requirements"):
                reqs = [r for r in rels if (r.get("dst_issue_type") or "").upper() == "REQUIREMENT"]
                tabs.load_requirements(reqs)
            # Test Cases 탭: relations 중 TEST_CASE 타입만 필터링 (REQUIREMENT / DEFECT 등에서 사용)
            if hasattr(tabs, "load_linked_testcases"):
                tcs = [r for r in rels if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"]
                tabs.load_linked_testcases(tcs)
        except Exception as e_rel:
            print(f"[WARN] Failed to load local relations: {e_rel}")

        # Test Case: Steps 탭 데이터
        if issue_type == "TEST_CASE":
            try:
                steps = get_steps_for_issue(self.conn, issue_id)
                if hasattr(tabs, "load_steps"):
                    tabs.load_steps(steps)
            except Exception as e_steps:
                print(f"[WARN] Failed to load local steps: {e_steps}")

        # Test Plan: Test Cases 탭 (Plan - Test Case 매핑)
        if issue_type == "TEST_PLAN":
            try:
                tp_records = get_testplan_testcases(self.conn, issue_id)
                if hasattr(tabs, "load_testplan_testcases"):
                    tabs.load_testplan_testcases(tp_records)
            except Exception as e_tp:
                print(f"[WARN] Failed to load local test plan testcases: {e_tp}")

        # Test Execution: Executions 탭 (메타 + Test Case Executions)
        if issue_type == "TEST_EXECUTION":
            try:
                te_row = get_or_create_testexecution_for_issue(self.conn, issue_id)
                from backend.db import get_testcase_executions  # local import to avoid circular

                tc_execs = get_testcase_executions(self.conn, te_row["id"])
                if hasattr(tabs, "load_testexecution"):
                    tabs.load_testexecution(te_row, tc_execs)
            except Exception as e_te:
                print(f"[WARN] Failed to load local test execution data: {e_te}")

        # Defects 탭: 현재 이슈와 관련된 모든 Defect 이슈를 간단히 로드
        try:
            self._refresh_defects_tab_for_current_issue()
        except Exception as e_def:
            print(f"[WARN] Failed to load defects for current issue: {e_def}")

    # --------------------------------------------------------------------- Online tree selection (JIRA RTM)

    def on_online_tree_selection_changed(self, selected, deselected):
        """
        오른쪽(JIRA RTM Online) 트리에서 이슈를 선택했을 때,
        해당 이슈의 상세 정보를 JIRA REST/RTM API 로 조회하여 우측 이슈 탭에 표시한다.
        """
        if not self.jira_available or not self.jira_client:
            return

        model = self.right_panel.tree_view.model()
        if model is None:
            self.right_panel.issue_tabs.set_issue(None)
            return

        selection_model = self.right_panel.tree_view.selectionModel()
        if selection_model is None:
            self.right_panel.issue_tabs.set_issue(None)
            return

        selected_indexes = selection_model.selectedIndexes()
        if not selected_indexes:
            self.right_panel.issue_tabs.set_issue(None)
            return

        item = model.itemFromIndex(selected_indexes[0])
        if not item:
            self.right_panel.issue_tabs.set_issue(None)
            return

        node_type = (item.data(Qt.UserRole) or "").upper()
        jira_key = item.data(Qt.UserRole + 1) or ""

        # 폴더는 무시
        if node_type == "FOLDER" or not jira_key:
            self.right_panel.issue_tabs.set_issue(None)
            return

        issue_type = node_type  # RTM Tree 에서 오는 type 을 그대로 사용

        tabs = self.right_panel.issue_tabs

        try:
            # 1) 기본 필드: Jira 표준 REST API 로 개별 이슈 조회
            #    (fields 구조를 가진 JSON -> jira_mapping.map_jira_to_local 로 변환)
            jira_issue_json = self.jira_client.get_jira_issue(jira_key)
            updates = jira_mapping.map_jira_to_local(issue_type, jira_issue_json)
            issue_like: Dict[str, Any] = {
                "issue_type": issue_type,
                "jira_key": jira_key,
                **updates,
            }
            tabs.set_issue(issue_like)

            # 2) Relations (Jira issue links)
            try:
                rel_entries = jira_mapping.extract_relations_from_jira(jira_issue_json)
                if hasattr(tabs, "load_relations"):
                    # Online 뷰에서는 dst_issue_id 가 없으므로 None 으로 두고 표시만 한다.
                    rels_for_ui = []
                    for r in rel_entries:
                        rels_for_ui.append(
                            {
                                "relation_type": r.get("relation_type") or "",
                                "dst_issue_id": None,
                                "dst_jira_key": r.get("dst_jira_key") or "",
                                "dst_summary": r.get("dst_summary") or "",
                            }
                        )
                    tabs.load_relations(rels_for_ui)
            except Exception as e_rel:
                print(f"[WARN] Failed to load online relations: {e_rel}")

            # 3) RTM 전용 정보: 예를 들어 Test Case Steps 등은 RTM v1 REST API 사용
            if issue_type == "TEST_CASE":
                try:
                    steps_json = self.jira_client.get_testcase_steps(jira_key)
                    local_steps = jira_mapping.map_jira_testcase_steps_to_local(steps_json)
                    if hasattr(tabs, "load_steps"):
                        tabs.load_steps(local_steps)
                except Exception as e_steps:
                    print(f"[WARN] Failed to load online Test Case steps: {e_steps}")

        except Exception as e:
            self.status_bar.showMessage(f"Failed to load online issue {jira_key}: {e}")
            print(f"[ERROR] Failed to load online issue {jira_key}: {e}")

    # --------------------------------------------------------------------- Full sync / online tree

    def on_full_sync_clicked(self):
        """
        JIRA RTM Tree 전체를 내려 받아(Local DB 동기화) 왼쪽 트리를 재구성한다.
        JIRA 사용이 불가능한 경우에는 동작하지 않는다.
        """
        if not self.jira_available or not self.jira_client or not self.project:
            self.status_bar.showMessage("Cannot full sync: Jira RTM not configured.")
            return

        # 동기화 영향 설명 및 사용자 확인
        ret = QMessageBox.question(
            self,
            "Full Sync – JIRA → Local",
            (
                "RTM Tree 전체를 JIRA 서버에서 다시 읽어 와 로컬 DB와 트리를 재구성합니다.\n\n"
                "- JIRA 상의 최신 트리 구조가 기준이 됩니다.\n"
                "- 로컬에서만 추가/이동한 폴더/이슈 정보는 덮어써질 수 있습니다.\n\n"
                "계속 진행하시겠습니까?"
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        try:
            self.status_bar.showMessage("Syncing RTM tree from JIRA to local DB...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            sync_tree(self.project, self.jira_client, self.conn)
            self.reload_local_tree()
            # 온라인 트리도 함께 갱신
            self.on_refresh_online_tree()

            self.status_bar.showMessage("Full tree sync completed.")
        except Exception as e:
            self.status_bar.showMessage(f"Full sync failed: {e}")
            print(f"[ERROR] Full sync failed: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def on_refresh_online_tree(self):
        """
        오른쪽(JIRA RTM Online) 패널 트리를 서버에서 직접 조회하여 표시한다.
        - /rest/rtm/1.0/api/tree/{projectId}/{treeType} 응답 구조를 사용
        - treeType 은 현재 온라인 패널의 모듈 탭(Requirements / Test Cases / ...)에 따라
          requirements / test-cases / test-plans / test-executions / defects 로 설정된다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot refresh online tree: Jira RTM not configured.")
            return

        try:
            self.status_bar.showMessage("Loading online RTM tree from JIRA...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            tree_type = self._tree_type_from_issue_type(self.online_issue_type_filter)
            tree = self.jira_client.get_tree(tree_type=tree_type)

            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["JIRA RTM Tree"])

            root_item = model.invisibleRootItem()

            # 아이콘 준비: 폴더 / 이슈 (온라인 트리)
            style = self.right_panel.style()
            folder_icon = style.standardIcon(QStyle.SP_DirIcon)
            issue_icon = style.standardIcon(QStyle.SP_FileIcon)

            def add_online_node(parent_item: QStandardItem, node: Dict[str, Any]) -> bool:
                """
                RTM 온라인 트리에서도 현재 이슈 타입에 해당하는 서브트리만 보여준다.
                RETURNS: 이 노드가 실제로 추가되었는지 여부.
                
                실제 응답 구조 (requirements, test-cases, test-plans, test-executions, defects):
                - 루트 노드: id, testKey (예: "F-KVHSICCU-RQ"), folderName (예: "All"), children 배열
                - 폴더 노드: testKey (예: "F-KVHSICCU-RQ-6"), folderName (예: "ICCU"), children 배열
                - 이슈 노드: testKey (예: "KVHSICCU-73"), issueId (예: 3142769)
                """
                # 폴더인지 이슈인지 판단: folderName이 있으면 폴더, issueId가 있으면 이슈
                folder_name = node.get("folderName")
                issue_id = node.get("issueId")
                test_key = node.get("testKey") or ""
                
                # 폴더 노드 처리
                if folder_name is not None:
                    label = f"[Folder] {folder_name}"
                    item = QStandardItem(label)
                    item.setEditable(False)
                    item.setData("FOLDER", Qt.UserRole)
                    item.setData(test_key, Qt.UserRole + 1)
                    item.setIcon(folder_icon)

                    has_visible_child = False
                    children = node.get("children", [])
                    for child in children:
                        if add_online_node(item, child):
                            has_visible_child = True

                    if not has_visible_child:
                        return False

                    parent_item.appendRow(item)
                    return True
                
                # 이슈 노드 처리
                if issue_id is not None:
                    # testKey를 키로 사용, 이름은 testKey 사용
                    # 이슈 타입 필터링은 현재 필터가 없거나, 노드의 type 필드와 일치할 때만 표시
                    node_type_from_field = (node.get("type") or "").upper()
                    type_filter = self.online_issue_type_filter
                    
                    # type 필터가 있고, 노드의 type이 필터와 일치하지 않으면 스킵
                    # (실제 응답에 type 필드가 없을 수 있으므로, 필터가 있으면 현재 트리 타입을 기준으로 판단)
                    if type_filter and node_type_from_field and node_type_from_field != type_filter:
                        return False
                    
                    # 이슈 표시: testKey를 기본으로 사용
                    label = test_key if test_key else f"(no key) - {issue_id}"
                    
                    # 현재 트리 타입에 맞는 이슈 타입 설정
                    default_issue_type = type_filter or "REQUIREMENT"
                    if tree_type == "test-cases":
                        default_issue_type = "TEST_CASE"
                    elif tree_type == "test-plans":
                        default_issue_type = "TEST_PLAN"
                    elif tree_type == "test-executions":
                        default_issue_type = "TEST_EXECUTION"
                    elif tree_type == "defects":
                        default_issue_type = "DEFECT"
                    
                    item = QStandardItem(label)
                    item.setEditable(False)
                    item.setData(node_type_from_field or default_issue_type, Qt.UserRole)
                    item.setData(test_key, Qt.UserRole + 1)
                    item.setIcon(issue_icon)
                    parent_item.appendRow(item)
                    return True
                
                # 폴더도 이슈도 아닌 경우 (예: 루트 노드 또는 children만 있는 노드)
                # 루트 노드는 children만 처리
                children = node.get("children", [])
                has_visible_child = False
                for child in children:
                    if add_online_node(parent_item, child):
                        has_visible_child = True
                
                return has_visible_child

            # 실제 응답 구조에 맞게 루트 노드 처리
            # 응답이 단일 객체인 경우 (id, testKey, folderName, children을 가진 루트 객체)
            # 또는 배열인 경우, 또는 dict에 roots/children 키가 있는 경우를 모두 처리
            if isinstance(tree, list):
                # 배열인 경우: 각 요소를 루트로 처리
                for r in tree:
                    add_online_node(root_item, r)
            elif isinstance(tree, dict):
                # 딕셔너리인 경우
                if "roots" in tree:
                    # roots 키가 있으면 그것을 사용
                    roots = tree.get("roots", [])
                    for r in roots:
                        add_online_node(root_item, r)
                elif "children" in tree:
                    # children 키가 있으면 루트 객체 자체를 처리 (루트가 폴더일 수 있음)
                    add_online_node(root_item, tree)
                else:
                    # 그 외의 경우: 루트 객체 자체를 처리
                    add_online_node(root_item, tree)

            self.right_panel.tree_view.setModel(model)
            self.right_panel.tree_view.expandAll()
            self.right_panel.tree_view.setSelectionMode(QTreeView.ExtendedSelection)

            self.status_bar.showMessage("Online RTM tree refreshed.")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to refresh online tree: {e}")
            print(f"[ERROR] Failed to refresh online tree: {e}")
        finally:
            QApplication.restoreOverrideCursor()



    # ------------------------------------------------------------------ 메뉴바 구성

    def _create_menu_bar(self) -> None:
        """상단 메뉴바에서 주요 기능을 카테고리별로 제공한다."""
        menubar = self.menuBar()

        # File 메뉴: 엑셀 입/출력, 종료
        file_menu = menubar.addMenu("File")
        act_import = QAction("Import from Excel...", self)
        act_import.triggered.connect(self.on_import_excel_clicked)
        act_import.setShortcut(QKeySequence("Ctrl+I"))
        file_menu.addAction(act_import)

        act_export = QAction("Export to Excel...", self)
        act_export.triggered.connect(self.on_export_excel_clicked)
        # 선택 단축키: Ctrl+E
        act_export.setShortcut(QKeySequence("Ctrl+E"))
        file_menu.addAction(act_export)

        file_menu.addSeparator()
        act_exit = QAction("Exit", self)
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        # Local 메뉴: 로컬 이슈 저장, 트리 갱신
        local_menu = menubar.addMenu("Local")
        act_save = QAction("Save Current Issue", self)
        act_save.triggered.connect(self.on_save_issue_clicked)
        act_save.setShortcut(QKeySequence("Ctrl+S"))
        local_menu.addAction(act_save)

        act_refresh_tree = QAction("Refresh Local Tree", self)
        act_refresh_tree.triggered.connect(self.on_full_sync_clicked)
        act_refresh_tree.setShortcut(QKeySequence("F5"))
        local_menu.addAction(act_refresh_tree)

        # JIRA / RTM 메뉴: 온라인 연동 기능
        jira_menu = menubar.addMenu("JIRA / RTM")
        act_pull = QAction("Pull from JIRA (Selected)", self)
        act_pull.triggered.connect(self.on_pull_issue_clicked)
        jira_menu.addAction(act_pull)

        act_push = QAction("Push to JIRA (Selected)", self)
        act_push.triggered.connect(self.on_push_issue_clicked)
        jira_menu.addAction(act_push)

        jira_menu.addSeparator()
        act_create = QAction("Create New Issue in JIRA", self)
        act_create.triggered.connect(self.on_create_in_jira_clicked)
        jira_menu.addAction(act_create)

        act_delete = QAction("Delete Issue from JIRA", self)
        act_delete.triggered.connect(self.on_delete_in_jira_clicked)
        jira_menu.addAction(act_delete)

        # View 메뉴: 좌/우(로컬/온라인) 윈도우 표시 및 레이아웃 제어
        view_menu = menubar.addMenu("View")
        self.act_view_local = QAction("Show Local Window", self, checkable=True)
        # 기본값: 로컬 창은 항상 표시
        self.act_view_local.setChecked(True)
        self.act_view_online = QAction("Show Online Window", self, checkable=True)
        # 기본값: 온라인 창은 숨김 (필요 시 View 메뉴에서 켜도록 함)
        # mode 가 "online" 인 경우에는 반대로 설정
        if getattr(self, "mode", "both") == "online":
            self.act_view_online.setChecked(True)
            self.act_view_local.setChecked(False)
        elif getattr(self, "mode", "both") == "local":
            self.act_view_online.setChecked(False)
        else:
            self.act_view_online.setChecked(True)

        self.act_view_local.triggered.connect(self._on_toggle_local_window)
        self.act_view_online.triggered.connect(self._on_toggle_online_window)

        view_menu.addAction(self.act_view_local)
        view_menu.addAction(self.act_view_online)

        # 레이아웃(좌/우 vs 상/하) 전환
        layout_menu = view_menu.addMenu("Layout")
        self.act_layout_horizontal = QAction("Left / Right (Local | Online)", self, checkable=True)
        self.act_layout_vertical = QAction("Top / Bottom (Online / Local)", self, checkable=True)
        self.act_layout_horizontal.setChecked(True)

        self.act_layout_horizontal.triggered.connect(self._on_set_layout_horizontal)
        self.act_layout_vertical.triggered.connect(self._on_set_layout_vertical)

        layout_menu.addAction(self.act_layout_horizontal)
        layout_menu.addAction(self.act_layout_vertical)

        # Help 메뉴
        help_menu = menubar.addMenu("Help")
        act_help = QAction("Help...", self)
        act_help.triggered.connect(self._show_help_dialog)
        help_menu.addAction(act_help)
        act_about = QAction("About RTM Local Manager", self)
        act_about.triggered.connect(self._show_about_dialog)
        help_menu.addAction(act_about)

        # Settings 메뉴: 필드 프리셋 편집, JIRA 메타데이터 새로고침, Excel 컬럼 매핑
        settings_menu = menubar.addMenu("Settings")
        act_presets = QAction("Field Presets...", self)
        act_presets.triggered.connect(self.on_edit_field_presets_clicked)
        settings_menu.addAction(act_presets)

        act_refresh_meta = QAction("Refresh JIRA Metadata", self)
        act_refresh_meta.triggered.connect(self.on_refresh_jira_metadata_clicked)
        settings_menu.addAction(act_refresh_meta)

        act_api_settings = QAction("REST API && Auth Settings...", self)
        act_api_settings.triggered.connect(self.on_edit_api_settings_clicked)
        settings_menu.addAction(act_api_settings)

        act_api_endpoints = QAction("REST API Endpoint Settings...", self)
        act_api_endpoints.triggered.connect(self.on_edit_api_endpoints_clicked)
        settings_menu.addAction(act_api_endpoints)

        act_api_tester = QAction("REST API Tester...", self)
        act_api_tester.triggered.connect(self.on_open_api_tester_clicked)
        settings_menu.addAction(act_api_tester)

        act_excel_mapping = QAction("Excel Column Mapping...", self)
        act_excel_mapping.triggered.connect(self.on_edit_excel_mapping_clicked)
        settings_menu.addAction(act_excel_mapping)

        act_local_settings = QAction("Local Activity && Attachments...", self)
        act_local_settings.triggered.connect(self.on_edit_local_settings_clicked)
        settings_menu.addAction(act_local_settings)

    def on_edit_local_settings_clicked(self) -> None:
        """
        Settings > Local Activity && Attachments...
        - 로컬 Activity 타임스탬프 / 첨부 루트 디렉터리 / 자동 업·다운로드 여부를 설정한다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QCheckBox,
            QPushButton,
            QFileDialog,
            QDialogButtonBox,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Local Activity && Attachments Settings")

        vbox = QVBoxLayout(dlg)

        cfg = self.local_settings or {}
        act_cfg = cfg.get("activity", {}) or {}
        att_cfg = cfg.get("attachments", {}) or {}

        # Activity 설정
        vbox.addWidget(QLabel("<b>Local Activity</b>"))
        row_ts = QHBoxLayout()
        chk_ts = QCheckBox("새 Activity 추가 시 타임스탬프 자동 추가")
        chk_ts.setChecked(bool(act_cfg.get("append_timestamp_on_add", True)))
        row_ts.addWidget(chk_ts)
        row_ts.addStretch()
        vbox.addLayout(row_ts)

        row_fmt = QHBoxLayout()
        row_fmt.addWidget(QLabel("타임스탬프 포맷 (datetime.strftime):"))
        edt_fmt = QLineEdit()
        edt_fmt.setText(str(act_cfg.get("timestamp_format") or "%Y-%m-%d %H:%M"))
        row_fmt.addWidget(edt_fmt)
        vbox.addLayout(row_fmt)

        # Attachments 설정
        vbox.addSpacing(12)
        vbox.addWidget(QLabel("<b>Local Attachments</b>"))

        row_root = QHBoxLayout()
        row_root.addWidget(QLabel("첨부 루트 디렉터리:"))
        edt_root = QLineEdit()
        edt_root.setText(str(att_cfg.get("root_dir") or ""))
        btn_browse = QPushButton("Browse...")

        def _browse_root():
            base = edt_root.text().strip() or ""
            directory = QFileDialog.getExistingDirectory(
                self,
                "Select attachments root directory",
                base or "",
            )
            if directory:
                edt_root.setText(directory)

        btn_browse.clicked.connect(_browse_root)
        row_root.addWidget(edt_root)
        row_root.addWidget(btn_browse)
        vbox.addLayout(row_root)

        chk_auto_dl = QCheckBox("Pull from JIRA 시 서버 첨부 자동 다운로드")
        chk_auto_dl.setChecked(bool(att_cfg.get("auto_download_on_pull", True)))
        vbox.addWidget(chk_auto_dl)

        chk_auto_ul = QCheckBox("Push to JIRA 시 로컬 첨부 자동 업로드")
        chk_auto_ul.setChecked(bool(att_cfg.get("auto_upload_on_push", True)))
        vbox.addWidget(chk_auto_ul)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        vbox.addWidget(btn_box)

        def _on_accept():
            # Activity 설정 반영
            new_act = {
                "append_timestamp_on_add": chk_ts.isChecked(),
                "timestamp_format": edt_fmt.text().strip() or "%Y-%m-%d %H:%M",
            }
            # Attachments 설정 반영
            new_att = {
                "root_dir": edt_root.text().strip(),
                "auto_download_on_pull": chk_auto_dl.isChecked(),
                "auto_upload_on_push": chk_auto_ul.isChecked(),
            }

            self.local_settings["activity"] = new_act
            self.local_settings["attachments"] = new_att
            save_local_settings(self.local_settings)
            dlg.accept()

        btn_box.accepted.connect(_on_accept)
        btn_box.rejected.connect(dlg.reject)

        dlg.exec()

    def _show_about_dialog(self) -> None:
        QMessageBox.information(
            self,
            "About RTM Local Manager",
            "RTM Local Manager\n\n"
            "JIRA RTM 요구사항/테스트/결함을 로컬 SQLite DB와 동기화하여\n"
            "오프라인에서도 편리하게 관리하기 위한 도구입니다.",
        )

    def _show_help_dialog(self) -> None:
        """
        Help > Help... 메뉴:
        - 좌측에는 도움말 항목 트리, 우측에는 상세 설명을 표시하는 간단한 뷰어.
        - 상단 콤보박스로 한국어/영어를 전환할 수 있다.
        """
        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QHBoxLayout,
            QLabel,
            QComboBox,
            QTreeWidget,
            QTreeWidgetItem,
            QTextBrowser,
            QDialogButtonBox,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("RTM Local Manager – Help")

        vbox = QVBoxLayout(dlg)

        # 언어 선택
        row_lang = QHBoxLayout()
        row_lang.addWidget(QLabel("Language / 언어:"))
        cmb_lang = QComboBox()
        cmb_lang.addItem("한국어", userData="ko")
        cmb_lang.addItem("English", userData="en")
        row_lang.addWidget(cmb_lang)
        row_lang.addStretch()
        vbox.addLayout(row_lang)

        # 좌: 토픽 트리, 우: 내용
        hbox = QHBoxLayout()

        tree = QTreeWidget()
        tree.setHeaderHidden(True)

        txt = QTextBrowser()
        hbox.addWidget(tree, 1)
        hbox.addWidget(txt, 3)

        vbox.addLayout(hbox)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        vbox.addWidget(btn_box)

        # 도움말 토픽 정의 (ID 를 key 로 사용)
        topics = [
            ("overview", "개요 / Overview"),
            ("local_tree", "로컬 트리 / Local Tree"),
            ("online_tree", "온라인 트리 / Online Tree"),
            ("excel", "엑셀 Import/Export"),
            ("sync", "동기화 / Sync JIRA ↔ Local"),
            ("shortcuts", "단축키 / Shortcuts"),
        ]

        help_text_ko = {
            "overview": (
                "<h3>개요</h3>"
                "<p>RTM Local Manager 는 JIRA + Deviniti RTM 의 요구사항 / 테스트 / 결함 데이터를 "
                "로컬 SQLite DB 와 엑셀 파일로 관리할 수 있게 해주는 데스크톱 도구입니다.</p>"
                "<ul>"
                "<li>좌측(Local) 패널: 로컬 DB 트리 및 이슈 편집</li>"
                "<li>우측(Online) 패널: JIRA RTM 트리 조회 및 비교 (옵션)</li>"
                "<li>상단 Ribbon: Excel Import/Export, Sync, JIRA 연동 기능</li>"
                "</ul>"
            ),
            "local_tree": (
                "<h3>로컬 트리 / Local Tree</h3>"
                "<p>좌측 트리에서 Requirement / Test Case / Test Plan / Test Execution / Defect 를 "
                "폴더 구조로 관리합니다.</p>"
                "<ul>"
                "<li>상단 탭에서 이슈 타입(Requirements, Test Cases 등)을 선택하면 해당 타입만 필터링됩니다.</li>"
                "<li>폴더를 선택하면 우측 탭에서 하위 이슈 목록을, 이슈를 선택하면 상세 탭을 볼 수 있습니다.</li>"
                "<li>이슈를 편집하면 트리에서 주황색(더티)으로 표시되고, Ctrl+S 또는 Save Issue 로 저장합니다.</li>"
                "</ul>"
            ),
            "online_tree": (
                "<h3>온라인 트리 / Online Tree</h3>"
                "<p>우측 트리는 JIRA RTM 의 Tree Structure 를 그대로 보여줍니다.</p>"
                "<ul>"
                "<li>우측 상단 모듈 탭(Requirements / Test Cases / Test Plans / Test Executions / Defects)에 따라 "
                "RTM REST API 의 treeType (requirements, test-cases 등)을 자동으로 선택합니다.</li>"
                "<li>\"Refresh Online Tree\" 버튼으로 서버 트리를 다시 불러올 수 있습니다.</li>"
                "<li>JIRA 설정이 올바르지 않으면 Settings &gt; REST API &amp; Auth Settings 에서 수정합니다.</li>"
                "</ul>"
            ),
            "excel": (
                "<h3>엑셀 Import / Export</h3>"
                "<ul>"
                "<li>Ribbon 의 Local 섹션에서 Import Excel / Export Excel 버튼으로 로컬 DB 와 엑셀 파일 간에 데이터를 교환합니다.</li>"
                "<li>Export 한 엑셀에는 Issues, TestcaseSteps, Relations, TestPlanTestcases, "
                "TestExecutions, TestcaseExecutions, TestcaseStepExecutions 시트가 포함됩니다.</li>"
                "<li>각 시트의 컬럼 헤더에는 작성 방법과 예시가 Comment 로 달려 있습니다.</li>"
                "<li>Settings &gt; Excel Column Mapping... 에서 엑셀 헤더와 DB 필드를 매핑할 수 있습니다.</li>"
                "</ul>"
            ),
            "sync": (
                "<h3>동기화 / Sync</h3>"
                "<ul>"
                "<li>Full Sync (Tree): JIRA RTM 의 Tree Structure (requirements, test-cases, test-plans, "
                "test-executions, defects) 를 순차적으로 읽어 로컬 DB 와 트리를 재구성합니다.</li>"
                "<li>Pull Issue / Push Issue: 현재 선택된 이슈에 대해 JIRA &lt;-&gt; Local 단방향 동기화를 수행합니다.</li>"
                "<li>REST API 엔드포인트와 인증 정보는 Settings &gt; REST API &amp; Auth Settings / "
                "REST API Endpoint Settings 에서 수정 가능합니다.</li>"
                "</ul>"
            ),
            "shortcuts": (
                "<h3>단축키 / Shortcuts</h3>"
                "<ul>"
                "<li>Ctrl+S: 현재 선택된 로컬 이슈 저장</li>"
                "<li>Ctrl+C / Ctrl+X / Ctrl+V (계획): 트리 항목 복사/잘라내기/붙여넣기</li>"
                "<li>기타 단축키는 추후 문서에서 확장 예정입니다.</li>"
                "</ul>"
            ),
        }

        help_text_en = {
            "overview": (
                "<h3>Overview</h3>"
                "<p>RTM Local Manager is a desktop tool for managing Requirements / Test Cases / "
                "Test Plans / Test Executions / Defects from JIRA + Deviniti RTM in a local SQLite DB and Excel files.</p>"
                "<ul>"
                "<li>Left (Local) panel: local DB tree and issue editor</li>"
                "<li>Right (Online) panel: JIRA RTM tree view (optional)</li>"
                "<li>Top ribbon: Excel import/export, sync, and JIRA integration actions</li>"
                "</ul>"
            ),
            "local_tree": (
                "<h3>Local Tree</h3>"
                "<p>The left tree shows local Requirements, Test Cases, Test Plans, Test Executions and Defects "
                "organized into folders.</p>"
                "<ul>"
                "<li>Use the top module tabs (Requirements, Test Cases, ...) to filter by issue type.</li>"
                "<li>Select a folder to see its children; select an issue to see its details on the right.</li>"
                "<li>When an issue has unsaved local changes, it is highlighted in orange in the tree; "
                "press Ctrl+S or click Save Issue to persist it.</li>"
                "</ul>"
            ),
            "online_tree": (
                "<h3>Online Tree</h3>"
                "<p>The right-hand tree shows the RTM Tree Structure from the server.</p>"
                "<ul>"
                "<li>The current online module tab determines RTM treeType "
                "(requirements, test-cases, test-plans, test-executions, defects).</li>"
                "<li>Use the \"Refresh Online Tree\" button to reload the tree from RTM.</li>"
                "<li>If REST calls fail, verify your configuration in "
                "Settings &gt; REST API &amp; Auth Settings.</li>"
                "</ul>"
            ),
            "excel": (
                "<h3>Excel Import / Export</h3>"
                "<ul>"
                "<li>Use Import Excel / Export Excel buttons in the Local ribbon group to move data "
                "between the local DB and Excel files.</li>"
                "<li>The exported workbook contains Issues, TestcaseSteps, Relations, TestPlanTestcases, "
                "TestExecutions, TestcaseExecutions and TestcaseStepExecutions sheets.</li>"
                "<li>Each header cell has a comment describing how to fill the column and showing examples.</li>"
                "<li>You can customize header-to-field mapping via Settings &gt; Excel Column Mapping...</li>"
                "</ul>"
            ),
            "sync": (
                "<h3>Synchronization</h3>"
                "<ul>"
                "<li>Full Sync (Tree): sequentially fetch RTM tree structures for "
                "requirements / test-cases / test-plans / test-executions / defects and merge them into the local DB.</li>"
                "<li>Pull Issue / Push Issue: one-way sync between JIRA and the current local issue.</li>"
                "<li>REST API endpoints and authentication can be adjusted via Settings &gt; "
                "REST API &amp; Auth Settings / REST API Endpoint Settings.</li>"
                "</ul>"
            ),
            "shortcuts": (
                "<h3>Shortcuts</h3>"
                "<ul>"
                "<li>Ctrl+S: Save currently selected local issue</li>"
                "<li>Ctrl+C / Ctrl+X / Ctrl+V (planned): Copy / Cut / Paste tree items</li>"
                "<li>Additional shortcuts will be documented later.</li>"
                "</ul>"
            ),
        }

        # 트리 채우기
        for topic_id, label in topics:
            item = QTreeWidgetItem([label])
            item.setData(0, Qt.UserRole, topic_id)
            tree.addTopLevelItem(item)

        # 기본 선택: 개요
        if tree.topLevelItemCount() > 0:
            tree.setCurrentItem(tree.topLevelItem(0))

        def update_text() -> None:
            lang = cmb_lang.currentData() or "ko"
            item = tree.currentItem()
            if not item:
                return
            topic_id = item.data(0, Qt.UserRole)
            if lang == "en":
                text = help_text_en.get(topic_id, "")
            else:
                text = help_text_ko.get(topic_id, "")
            txt.setHtml(text)

        tree.currentItemChanged.connect(lambda *_: update_text())
        cmb_lang.currentIndexChanged.connect(lambda *_: update_text())

        update_text()

        btn_box.rejected.connect(dlg.reject)
        dlg.resize(900, 700)
        dlg.exec()

    # ------------------------------------------------------------------ View menu handlers

    def _on_toggle_local_window(self, checked: bool) -> None:
        """
        View > Show Local Window 체크박스에 따라 좌측(로컬) 패널 표시/숨김.
        """
        if self.left_panel is None:
            return
        self.left_panel.setVisible(checked)

    def _on_toggle_online_window(self, checked: bool) -> None:
        """
        View > Show Online Window 체크박스에 따라 우측(온라인) 패널 표시/숨김.
        """
        if self.right_panel is None:
            return
        self.right_panel.setVisible(checked)

    # ------------------------------------------------------------------ Layout mode handlers (horizontal / vertical)

    def _set_layout_mode(self, mode: str) -> None:
        """
        중앙 스플리터의 배치를 좌/우(horiz) 또는 상/하(vert, Online/Local)로 전환한다.
        - "horizontal":  Local | Online
        - "vertical"  :  Online / Local
        """
        mode = mode.lower()
        if mode not in ("horizontal", "vertical"):
            return
        if getattr(self, "layout_mode", None) == mode:
            return

        self.layout_mode = mode

        splitter: QSplitter = self.main_splitter

        if mode == "horizontal":
            splitter.setOrientation(Qt.Horizontal)
            # Left | Right
            splitter.insertWidget(0, self.left_panel)
            splitter.insertWidget(1, self.right_panel)
        else:
            splitter.setOrientation(Qt.Vertical)
            # Top(Online) / Bottom(Local)
            splitter.insertWidget(0, self.right_panel)
            splitter.insertWidget(1, self.left_panel)

        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 3)

    # ------------------------------------------------------------------ Helpers for local attachments root

    def _get_attachments_root(self):
        """
        Settings > Local Activity && Attachments... 에서 지정한 첨부 루트 디렉터리를 사용한다.
        - settings.attachments.root_dir 가 비어 있으면 기본값(rtm_local_manager/attachments)을 사용.
        """
        from pathlib import Path
        from backend.attachments_fs import get_attachments_root

        try:
            cfg = (self.local_settings or {}).get("attachments", {}) if hasattr(self, "local_settings") else {}
            root_dir = str(cfg.get("root_dir") or "").strip()
            if root_dir:
                p = Path(root_dir).expanduser()
                p.mkdir(parents=True, exist_ok=True)
                return p
        except Exception:
            pass
        return get_attachments_root()

    def _on_set_layout_horizontal(self, checked: bool) -> None:
        """
        View > Layout > Left / Right 선택 시 호출.
        """
        if not checked:
            # 다른 액션에 의해 해제되는 경우는 무시
            return
        self.act_layout_vertical.setChecked(False)
        self.act_layout_horizontal.setChecked(True)
        self._set_layout_mode("horizontal")

    def _on_set_layout_vertical(self, checked: bool) -> None:
        """
        View > Layout > Top / Bottom 선택 시 호출.
        """
        if not checked:
            return
        self.act_layout_horizontal.setChecked(False)
        self.act_layout_vertical.setChecked(True)
        self._set_layout_mode("vertical")

    # ------------------------------------------------------------------ Tree context menus (right-click)

    def _create_issue_from_tree_context(self, issue_type: str) -> None:
        """
        좌측(Local) 트리 컨텍스트 메뉴에서 선택한 이슈 타입으로 새 로컬 이슈를 생성한다.
        - 폴더/이슈 선택 규칙은 on_new_local_issue_clicked 와 동일.
        """
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot create issue.")
            return

        resolved_type = (issue_type or "REQUIREMENT").upper()

        folder_id: str | None = None
        model = self.left_panel.tree_view.model()
        selection_model = self.left_panel.tree_view.selectionModel()
        if model is not None and selection_model is not None:
            indexes = selection_model.selectedIndexes()
            if indexes:
                item = model.itemFromIndex(indexes[0])
                if item:
                    kind = item.data(Qt.UserRole)
                    if kind == "FOLDER":
                        folder_id = item.data(Qt.UserRole + 1)
                    elif kind == "ISSUE":
                        issue_id = item.data(Qt.UserRole + 1)
                        if issue_id:
                            issue = get_issue_by_id(self.conn, int(issue_id))
                            if issue:
                                folder_id = issue.get("folder_id")

        summary = f"New {resolved_type}"
        new_id = create_local_issue(
            self.conn,
            self.project.id,
            issue_type=resolved_type,
            folder_id=folder_id,
            summary=summary,
        )
        self.logger.info(
            "Created new local issue (context menu) id=%s, type=%s, folder_id=%s",
            new_id,
            resolved_type,
            folder_id,
        )
        self.status_bar.showMessage(f"Created new local issue (type={resolved_type}).")
        self.reload_local_tree()

    def _on_local_tree_context_menu(self, pos):
        """
        좌측(Local) 트리 우클릭 컨텍스트 메뉴.
        - 복수 선택 지원: 선택된 이슈/폴더에 대해 삭제 메뉴 제공.
        - 트리 전용 클립보드 액션(Select All / Copy / Cut / Paste) 제공.
        """
        from PySide6.QtWidgets import QMenu

        view = self.left_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        indexes = selection_model.selectedIndexes()
        has_issue = False
        has_folder = False
        for idx in indexes:
            item = model.itemFromIndex(idx)
            if not item:
                continue
            kind = item.data(Qt.UserRole)
            if kind == "ISSUE":
                has_issue = True
            elif kind == "FOLDER":
                has_folder = True

        menu = QMenu(view)

        # New 서브메뉴: 폴더 및 각 이슈 타입 생성
        new_menu = menu.addMenu("New")
        act_new_folder = new_menu.addAction("Folder")
        new_menu.addSeparator()
        act_new_req = new_menu.addAction("Requirement")
        act_new_tc = new_menu.addAction("Test Case")
        act_new_tp = new_menu.addAction("Test Plan")
        act_new_te = new_menu.addAction("Test Execution")
        act_new_def = new_menu.addAction("Defect")

        menu.addSeparator()

        # 공통 트리 클립보드 액션들
        act_select_all = menu.addAction("Select All")
        act_copy = menu.addAction("Copy")
        act_cut = menu.addAction("Cut")
        act_paste = menu.addAction("Paste")
        # Paste 는 내부 트리 클립보드가 있을 때만 의미가 있으므로 상태에 따라 enable 제어
        act_paste.setEnabled(bool(getattr(self, "_tree_clipboard", None)))

        # 선택이 없는 경우에도 Select All / Paste 는 유효하지만,
        # Copy / Cut 은 selection 이 없으면 내부적으로 no-op + 상태바 메시지만 출력된다.
        menu.addSeparator()

        act_del_issue = act_del_folder = None
        if has_issue:
            act_del_issue = menu.addAction("Delete Selected Issue(s)")
        if has_folder:
            act_del_folder = menu.addAction("Delete Selected Folder(s)")

        action = menu.exec(view.viewport().mapToGlobal(pos))
        if action is None:
            return

        # New 서브메뉴 우선 처리
        if action == act_new_folder:
            self.on_add_local_folder_clicked()
            return
        if action == act_new_req:
            self._create_issue_from_tree_context("REQUIREMENT")
            return
        if action == act_new_tc:
            self._create_issue_from_tree_context("TEST_CASE")
            return
        if action == act_new_tp:
            self._create_issue_from_tree_context("TEST_PLAN")
            return
        if action == act_new_te:
            self._create_issue_from_tree_context("TEST_EXECUTION")
            return
        if action == act_new_def:
            self._create_issue_from_tree_context("DEFECT")
            return

        # 트리 클립보드 관련 액션
        if action == act_select_all:
            self._on_local_tree_select_all()
            return
        if action == act_copy:
            self._on_local_tree_copy()
            return
        if action == act_cut:
            self._on_local_tree_cut()
            return
        if action == act_paste:
            self._on_local_tree_paste()
            return

        # 삭제 액션 처리
        if act_del_issue is not None and action == act_del_issue:
            self.on_delete_local_issue_clicked()
        elif act_del_folder is not None and action == act_del_folder:
            self.on_delete_local_folder_clicked()

    def _on_online_tree_context_menu(self, pos):
        """
        우측(온라인 JIRA RTM) 트리 우클릭 컨텍스트 메뉴.
        - 복수 선택된 이슈들에 대해 'Delete in JIRA' 를 호출한다.
        """
        from PySide6.QtWidgets import QMenu

        if not self.jira_available:
            return

        view = self.right_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        indexes = selection_model.selectedIndexes()
        has_issue = False
        for idx in indexes:
            item = model.itemFromIndex(idx)
            if item and item.data(Qt.UserRole) == "ISSUE":
                has_issue = True
                break

        if not has_issue:
            return

        menu = QMenu(view)
        act_del = menu.addAction("Delete Selected Issue(s) in JIRA")
        action = menu.exec(view.viewport().mapToGlobal(pos))
        if action is None or action != act_del:
            return

        self._delete_online_issues_in_jira()

    # ------------------------------------------------------------------ Local tree clipboard handlers (Cut/Copy/Paste/Select All)

    def _on_local_tree_select_all(self) -> None:
        """로컬 트리의 모든 항목을 선택한다."""
        if not self.left_panel or not self.left_panel.tree_view:
            return
        self.left_panel.tree_view.selectAll()

    def _collect_local_tree_selection(self, mode: str) -> None:
        """
        로컬 트리에서 선택된 폴더/이슈들을 클립보드에 저장한다.
        mode: "cut" 또는 "copy"
        """
        if not self.left_panel or not self.left_panel.tree_view:
            return
        view = self.left_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        items: list[Dict[str, Any]] = []
        for idx in selection_model.selectedIndexes():
            if idx.column() != 0:
                continue
            item = model.itemFromIndex(idx)
            if not item:
                continue
            kind = item.data(Qt.UserRole)
            if kind not in ("FOLDER", "ISSUE"):
                continue
            item_id = item.data(Qt.UserRole + 1)
            if not item_id:
                continue
            # summary/name 은 트리 표시용 텍스트를 그대로 사용한다.
            label_text = item.text() if hasattr(item, "text") else str(item.data(Qt.DisplayRole) or "")
            items.append({"kind": kind, "id": item_id, "summary": label_text})

        if not items:
            self.status_bar.showMessage("No items selected in local tree.")
            return

        self._tree_clipboard = {"mode": mode, "items": items}
        self.status_bar.showMessage(
            f"{len(items)} item(s) copied from local tree (mode={mode}).", 3000
        )

    def _on_local_tree_copy(self) -> None:
        """로컬 트리 선택 항목을 'copy' 모드로 클립보드에 저장."""
        self._collect_local_tree_selection("copy")

    def _on_local_tree_cut(self) -> None:
        """로컬 트리 선택 항목을 'cut' 모드로 클립보드에 저장."""
        self._collect_local_tree_selection("cut")

    def _on_local_tree_paste(self) -> None:
        """
        로컬 트리에서 Paste:
        - 현재 선택된 폴더/이슈 위치를 기준으로, 클립보드에 있는 폴더/이슈를 이동시킨다.
        - 현재 구현에서는 'copy' 와 'cut' 모두 이동(move) 로 동작한다.
        """
        if not self._tree_clipboard or not self.left_panel or not self.left_panel.tree_view:
            return

        mode = self._tree_clipboard.get("mode", "cut")
        items = self._tree_clipboard.get("items") or []
        if not items:
            return

        if mode == "cut":
            # 잘라내기: 실제 이동
            self._move_items_in_local_tree(items, mode="cut")
        else:
            # 복사: 원본은 그대로 두고, 동일한 구조를 새로 생성
            self._copy_items_in_local_tree(items)

    def _copy_items_in_local_tree(self, items: list[Dict[str, Any]]) -> None:
        """
        로컬 트리에서 선택된 폴더/이슈 리스트(items)를
        현재 선택된 위치를 기준으로 '복사(duplicate)' 한다.

        - FOLDER: 하위 폴더/이슈까지 재귀적으로 모두 복사.
        - ISSUE : 해당 이슈를 동일한 메타데이터로 새 로컬 이슈로 복제.
        """
        if not self.left_panel or not self.left_panel.tree_view:
            return

        view = self.left_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        # 붙여넣기 대상 폴더 결정 (move 와 동일한 규칙)
        target_folder_id: str | None = None
        indexes = selection_model.selectedIndexes()
        if indexes:
            item = model.itemFromIndex(indexes[0])
            if item:
                kind = item.data(Qt.UserRole)
                if kind == "FOLDER":
                    target_folder_id = item.data(Qt.UserRole + 1)
                elif kind == "ISSUE":
                    issue_id = item.data(Qt.UserRole + 1)
                    try:
                        issue = get_issue_by_id(self.conn, int(issue_id))
                        target_folder_id = issue.get("folder_id")
                    except Exception:
                        target_folder_id = None

        total_copied = 0
        for entry in items:
            kind = entry.get("kind")
            item_id = entry.get("id")
            if not item_id:
                continue
            try:
                if kind == "ISSUE":
                    src_issue = get_issue_by_id(self.conn, int(item_id))
                    if not src_issue:
                        continue
                    self._duplicate_issue_from_row(src_issue, target_folder_id)
                    total_copied += 1
                elif kind == "FOLDER":
                    total_copied += self._duplicate_folder_subtree(
                        str(item_id), target_folder_id
                    )
            except Exception as e:
                self.logger.warning("Failed to copy %s %s: %s", kind, item_id, e)

        if total_copied:
            self.reload_local_tree()
            self.status_bar.showMessage(
                f"Copied {total_copied} item(s) in local tree.", 3000
            )

    def _move_items_in_local_tree(self, items: list[Dict[str, Any]], mode: str = "cut") -> None:
        """
        공통 로직: 로컬 트리에서 선택된 폴더/이슈 리스트(items)를
        현재 선택된 위치를 기준으로 이동시킨다.

        - Ctrl+V (Paste) 뿐 아니라, 마우스 드래그&드롭에서도 재사용한다.
        """
        view = self.left_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        # 붙여넣기 대상 폴더 결정
        target_folder_id: str | None = None
        indexes = selection_model.selectedIndexes()
        if indexes:
            item = model.itemFromIndex(indexes[0])
            if item:
                kind = item.data(Qt.UserRole)
                if kind == "FOLDER":
                    target_folder_id = item.data(Qt.UserRole + 1)
                elif kind == "ISSUE":
                    issue_id = item.data(Qt.UserRole + 1)
                    try:
                        issue = get_issue_by_id(self.conn, int(issue_id))
                        target_folder_id = issue.get("folder_id")
                    except Exception:
                        target_folder_id = None

        # target_folder_id 가 None 이면 루트로 이동
        moved_count = 0
        for entry in items:
            kind = entry.get("kind")
            item_id = entry.get("id")
            if not item_id:
                continue
            try:
                if kind == "ISSUE":
                    move_issue_to_folder(self.conn, int(item_id), target_folder_id)
                    moved_count += 1
                elif kind == "FOLDER":
                    move_folder(self.conn, str(item_id), target_folder_id)
                    moved_count += 1
            except Exception as e:
                self.logger.warning("Failed to move %s %s: %s", kind, item_id, e)

        if moved_count:
            self.reload_local_tree()
            self.status_bar.showMessage(f"Moved {moved_count} item(s) in local tree.", 3000)

        # cut/copy 모드 모두에서 클립보드를 비울지 여부는 취향 차이이나,
        # 여기서는 cut 인 경우에만 비운다.
        if mode == "cut":
            self._tree_clipboard = None

    def _duplicate_issue_from_row(
        self, issue: Dict[str, Any], target_folder_id: str | None
    ) -> int:
        """
        단일 이슈 row(dict)를 기반으로, 동일한 메타데이터를 가진 새 로컬 이슈를 생성한다.

        - jira_key / jira_id 는 복사하지 않고, 로컬 전용 이슈로 만든다.
        - created / updated 는 새로 생성된 시각으로 세팅된다(create_local_issue 사용).
        """
        project_id = int(issue.get("project_id") or 0)
        issue_type = (issue.get("issue_type") or "REQUIREMENT").upper()
        summary = issue.get("summary") or f"Copy of {issue.get('id')}"

        # 1차: 최소 필드로 로컬 이슈 생성
        new_issue_id = create_local_issue(
            self.conn,
            project_id=project_id,
            issue_type=issue_type,
            folder_id=target_folder_id,
            summary=summary,
        )

        # 2차: 나머지 메타 필드 복사 (jira_key/jira_id/is_deleted/local_only/dirty 제외)
        fields: Dict[str, Any] = {}
        for key in [
            "description",
            "status",
            "priority",
            "assignee",
            "reporter",
            "labels",
            "components",
            "security_level",
            "fix_versions",
            "affects_versions",
            "rtm_environment",
            "due_date",
            "attachments",
            "epic_link",
            "sprint",
            "preconditions",
            "local_activity",
        ]:
            if key in issue:
                fields[key] = issue.get(key)

        if fields:
            update_issue_fields(self.conn, new_issue_id, fields)

        return new_issue_id

    def _duplicate_folder_subtree(
        self, src_folder_id: str, new_parent_id: str | None
    ) -> int:
        """
        주어진 src_folder_id 를 루트로 하는 폴더/이슈 서브트리를
        new_parent_id 하위에 재귀적으로 복사한다.

        RETURNS: 생성된 폴더/이슈 개수 (루트 폴더 포함).
        """
        cur = self.conn.cursor()

        # 원본 폴더 정보 조회
        cur.execute("SELECT * FROM folders WHERE id = ?", (src_folder_id,))
        row = cur.fetchone()
        if not row:
            return 0
        src = dict(row)

        project_id = int(src.get("project_id") or (self.project.id if self.project else 0))
        name = src.get("name") or ""

        # LOCAL-<TYPE>-uuid 형태인 경우, TYPE 을 추출하여 동일 타입의 로컬 폴더로 복사
        issue_type_hint: str | None = None
        fid = str(src.get("id") or "")
        if fid.startswith("LOCAL-"):
            parts = fid.split("-", 2)
            if len(parts) >= 3 and parts[1]:
                issue_type_hint = parts[1].upper()

        # 현재 로컬 탭 타입이 있으면 그것도 힌트로 사용
        if not issue_type_hint and getattr(self, "local_issue_type_filter", None):
            issue_type_hint = str(self.local_issue_type_filter).upper()

        # 새 폴더 생성 (RTM 에서 내려온 폴더라도, 복사본은 로컬 전용 폴더로 만든다)
        new_folder_id = create_folder_node(
            self.conn,
            project_id=project_id,
            name=name,
            parent_id=new_parent_id,
            issue_type=issue_type_hint,
        )

        created_count = 1  # 새 폴더 1개

        # 이 폴더에 속한 이슈들 복제
        cur.execute(
            "SELECT * FROM issues WHERE folder_id = ? AND is_deleted = 0",
            (src_folder_id,),
        )
        for issue_row in cur.fetchall():
            issue = dict(issue_row)
            self._duplicate_issue_from_row(issue, new_folder_id)
            created_count += 1

        # 하위 폴더들 재귀 복제
        cur.execute("SELECT id FROM folders WHERE parent_id = ?", (src_folder_id,))
        for child_row in cur.fetchall():
            child_id = child_row["id"]
            created_count += self._duplicate_folder_subtree(child_id, new_folder_id)

        return created_count

    # ------------------------------------------------------------------
    # Local tree drag & drop (마우스 드래그로 폴더/이슈 이동)
    # ------------------------------------------------------------------

    def eventFilter(self, obj, event):
        """
        트리 드래그&드롭 처리를 위해, 좌측(Local) 트리의 Drop 이벤트를 가로채서
        DB 기반 move_* 함수로 폴더/이슈를 이동시킨다.
        """
        try:
            view = getattr(self, "left_panel", None) and getattr(self.left_panel, "tree_view", None)
            if view is not None and event.type() == QEvent.Drop and obj in (view, view.viewport()):
                self._handle_local_tree_drop(event)
                return True
        except Exception as e:
            # 드래그 처리 중 예외가 발생해도 전체 GUI 가 죽지 않도록 방어
            if hasattr(self, "logger"):
                self.logger.warning("Error handling local tree drop: %s", e)

        return super().eventFilter(obj, event)

    def _handle_local_tree_drop(self, event) -> None:
        """
        좌측(Local) 트리에서의 Drop 이벤트 처리:
        - 현재 선택된 폴더/이슈들을 Drag source 로 간주
        - 드롭 위치에 따라 대상 폴더를 계산한 뒤, _move_items_in_local_tree 로 이동 처리
        """
        if not self.left_panel or not self.left_panel.tree_view:
            return

        view = self.left_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            return

        # 드래그된 항목: 현재 선택된 폴더/이슈들
        items: list[Dict[str, Any]] = []
        for idx in selection_model.selectedIndexes():
            if idx.column() != 0:
                continue
            item = model.itemFromIndex(idx)
            if not item:
                continue
            kind = item.data(Qt.UserRole)
            if kind not in ("FOLDER", "ISSUE"):
                continue
            item_id = item.data(Qt.UserRole + 1)
            if not item_id:
                continue
            items.append({"kind": kind, "id": item_id})

        if not items:
            return

        # 드롭 위치를 기준으로 "현재 선택" 을 강제로 설정하여,
        # _move_items_in_local_tree 의 로직을 재사용한다.
        pos = getattr(event, "position", None)
        if callable(pos):
            drop_pos = pos().toPoint()
        else:
            drop_pos = event.pos()

        target_index = view.indexAt(drop_pos)
        selection_model.clearSelection()
        if target_index.isValid():
            selection_model.select(
                target_index,
                QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Current,
            )

        # 드래그는 "이동" 개념에 더 가깝기 때문에 mode="cut" 으로 취급
        self._move_items_in_local_tree(items, mode="cut")

        event.accept()

    def _connect_signals(self):
        # ------------------------------------------------------------------
        # Ribbon: Local (SQLite only)
        # ------------------------------------------------------------------
        # Full sync 버튼: JIRA 트리 → Local DB → Local Tree reload
        self.btn_full_sync.clicked.connect(self.on_full_sync_clicked)

        # Excel Import/Export
        self.btn_import_excel.clicked.connect(self.on_import_excel_clicked)
        self.btn_export_excel.clicked.connect(self.on_export_excel_clicked)

        # Local Issues / Folders (리본)
        self.btn_ribbon_save_issue.clicked.connect(self.on_save_issue_clicked)
        self.btn_ribbon_new_issue.clicked.connect(self.on_new_local_issue_clicked)
        self.btn_ribbon_delete_issue.clicked.connect(self.on_delete_local_issue_clicked)
        self.btn_ribbon_add_folder.clicked.connect(self.on_add_local_folder_clicked)
        self.btn_ribbon_delete_folder.clicked.connect(self.on_delete_local_folder_clicked)

        # Test Case 상단 액션들 (로컬 패널 트리 툴바)
        if hasattr(self.left_panel, "btn_add_to_testplan"):
            self.left_panel.btn_add_to_testplan.clicked.connect(self.on_add_testcases_to_testplan_clicked)
        if hasattr(self.left_panel, "btn_link_requirement"):
            self.left_panel.btn_link_requirement.clicked.connect(self.on_link_testcases_to_requirement_clicked)
        # Execute 버튼은 향후 Test Execution 흐름과 연계 (현재는 별도 동작 없음)

        # ------------------------------------------------------------------
        # Local tree: keyboard shortcuts (Ctrl+A/C/X/V)
        # ------------------------------------------------------------------
        if self.left_panel and hasattr(self.left_panel, "tree_view"):
            view = self.left_panel.tree_view
            # 드래그&드롭 이동 처리를 위해 MainWindow 가 Drop 이벤트를 가로챈다.
            # Drop 이벤트는 보통 viewport 에서 발생하므로 viewport 에 필터를 설치한다.
            view.viewport().installEventFilter(self)
            # 기본 드롭 액션을 Move 로 설정하여, Qt 가 Copy 중심으로 동작하지 않도록 한다.
            view.setDefaultDropAction(Qt.MoveAction)
            QShortcut(QKeySequence("Ctrl+A"), view, activated=self._on_local_tree_select_all)
            QShortcut(QKeySequence("Ctrl+C"), view, activated=self._on_local_tree_copy)
            QShortcut(QKeySequence("Ctrl+X"), view, activated=self._on_local_tree_cut)
            QShortcut(QKeySequence("Ctrl+V"), view, activated=self._on_local_tree_paste)

        # ------------------------------------------------------------------
        # Ribbon: JIRA / RTM (Online only)
        # ------------------------------------------------------------------
        if self.jira_available:
            self.btn_ribbon_refresh_online.clicked.connect(self.on_refresh_online_tree)
            self.btn_ribbon_delete_in_jira.clicked.connect(self.on_delete_in_jira_clicked)
            self.btn_ribbon_create_in_jira.clicked.connect(self.on_create_in_jira_clicked)
            self.btn_ribbon_search_jira.clicked.connect(self.on_jira_filter_search)

        # ------------------------------------------------------------------
        # Ribbon: Sync (JIRA ↔ Local)
        # ------------------------------------------------------------------
        if self.jira_available:
            self.btn_ribbon_pull.clicked.connect(self.on_pull_issue_clicked)
            self.btn_ribbon_push.clicked.connect(self.on_push_issue_clicked)

        # Sync 모드 콤보박스 변경 시 내부 sync_mode 상태 업데이트
        if hasattr(self, "cmb_sync_mode"):
            def _on_sync_mode_changed(index: int):
                text = self.cmb_sync_mode.currentText()
                if "서버 우선" in text or "Server" in text:
                    self.sync_mode = "server"
                elif "로컬 우선" in text or "Local" in text:
                    self.sync_mode = "local"
                else:
                    self.sync_mode = "merge"
                # 상태바에 현재 모드 표시 (간단한 피드백)
                if hasattr(self, "status_bar"):
                    self.status_bar.showMessage(f"Sync mode set to: {self.sync_mode}")

            self.cmb_sync_mode.currentIndexChanged.connect(_on_sync_mode_changed)

        # ------------------------------------------------------------------
        # Ribbon: Execution / Defects
        # ------------------------------------------------------------------
        # 리본에서 실행/Defect 버튼은 각 탭의 버튼을 대신 눌러주는 형태로 연결
        self.btn_ribbon_execute_plan.clicked.connect(
            lambda: getattr(self.left_panel.issue_tabs, "btn_execute_plan", None)
            and self.left_panel.issue_tabs.btn_execute_plan.click()
        )
        self.btn_ribbon_link_defect.clicked.connect(
            lambda: getattr(self.left_panel.issue_tabs, "btn_link_defect", None)
            and self.left_panel.issue_tabs.btn_link_defect.click()
        )
        self.btn_ribbon_clear_defects.clicked.connect(
            lambda: getattr(self.left_panel.issue_tabs, "btn_clear_defects", None)
            and self.left_panel.issue_tabs.btn_clear_defects.click()
        )
        self.btn_ribbon_refresh_defects.clicked.connect(self.on_refresh_defects_tab_clicked)
        self.btn_ribbon_export_te_report.clicked.connect(self.on_export_testexecution_report_clicked)

        # ------------------------------------------------------------------
        # Ribbon: View / Layout
        # ------------------------------------------------------------------
        self.btn_ribbon_show_local.toggled.connect(self._on_toggle_local_window)
        self.btn_ribbon_show_online.toggled.connect(self._on_toggle_online_window)
        self.btn_ribbon_layout_h.clicked.connect(lambda: self._on_set_layout_horizontal(True))
        self.btn_ribbon_layout_v.clicked.connect(lambda: self._on_set_layout_vertical(True))

        # Right panel buttons (온라인 관련)
        if self.jira_available:
            self.right_panel.btn_refresh.clicked.connect(self.on_refresh_online_tree)
            self.right_panel.btn_sync_down.clicked.connect(self.on_full_sync_clicked)
            # JIRA create/delete (온라인 패널 헤더에 위치)
            self.right_panel.btn_create_jira.clicked.connect(self.on_create_in_jira_clicked)
            self.right_panel.btn_delete_jira.clicked.connect(self.on_delete_in_jira_clicked)
            # 온라인 패널 트리 selection
            r_selection = self.right_panel.tree_view.selectionModel()
            if r_selection is not None:
                r_selection.selectionChanged.connect(self.on_online_tree_selection_changed)

        # Local panel 트리 selection
        selection_model = self.left_panel.tree_view.selectionModel()
        if selection_model is not None:
            selection_model.selectionChanged.connect(self.on_local_tree_selection_changed)
        else:
            self.logger.debug("Local tree_view selectionModel is None at _connect_signals; will be connected after model set.")

        # Local panel 폴더/이슈 생성/삭제/저장 버튼
        self.left_panel.btn_add_folder.clicked.connect(self.on_add_local_folder_clicked)
        self.left_panel.btn_delete_folder.clicked.connect(self.on_delete_local_folder_clicked)
        self.left_panel.btn_new_issue.clicked.connect(self.on_new_local_issue_clicked)
        self.left_panel.btn_delete_issue.clicked.connect(self.on_delete_local_issue_clicked)
        self.left_panel.btn_save_issue.clicked.connect(self.on_save_issue_clicked)

        # Local / Online 트리 컨텍스트 메뉴 (우클릭)
        self.left_panel.tree_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.left_panel.tree_view.customContextMenuRequested.connect(self._on_local_tree_context_menu)
        if self.jira_available:
            self.right_panel.tree_view.setContextMenuPolicy(Qt.CustomContextMenu)
            self.right_panel.tree_view.customContextMenuRequested.connect(self._on_online_tree_context_menu)

        # Details 탭 Activity / Comment / Attachments 버튼
        try:
            # 왼쪽 패널(로컬)은 로컬 Activity/Attachments 를 관리
            self.left_panel.issue_tabs.btn_refresh_activity.clicked.connect(
                self.on_refresh_local_activity_clicked
            )
            self.left_panel.issue_tabs.btn_add_comment.clicked.connect(
                self.on_add_local_activity_clicked
            )
            self.left_panel.issue_tabs.btn_edit_comment.clicked.connect(
                self.on_edit_local_activity_clicked
            )
            self.left_panel.issue_tabs.btn_delete_comment.clicked.connect(
                self.on_delete_local_activity_clicked
            )
            self.left_panel.issue_tabs.btn_upload_attachment.clicked.connect(
                self.on_upload_attachment_clicked
            )
            self.left_panel.issue_tabs.btn_delete_attachment.clicked.connect(
                self.on_delete_attachment_clicked
            )
            self.left_panel.issue_tabs.btn_open_attachment.clicked.connect(
                self.on_open_attachment_clicked
            )
            # 오른쪽 패널(온라인)은 JIRA Activity 조회/조작용으로 사용
            if hasattr(self, "right_panel") and hasattr(self.right_panel, "issue_tabs"):
                self.right_panel.issue_tabs.btn_refresh_activity.clicked.connect(
                    self.on_refresh_activity_clicked
                )
                self.right_panel.issue_tabs.btn_add_comment.clicked.connect(
                    self.on_add_comment_clicked
                )
                self.right_panel.issue_tabs.btn_edit_comment.clicked.connect(
                    self.on_edit_comment_clicked
                )
                self.right_panel.issue_tabs.btn_delete_comment.clicked.connect(
                    self.on_delete_comment_clicked
                )
            # Defects 탭 버튼
            self.left_panel.issue_tabs.btn_defect_refresh.clicked.connect(
                self.on_refresh_defects_tab_clicked
            )
            self.left_panel.issue_tabs.btn_defect_open.clicked.connect(
                self.on_open_selected_defect_clicked
            )
        except Exception:
            # 방어적: 위젯 초기화 순서 문제가 있을 경우를 대비
            pass


        # --------------------------------------------------------------------- Excel import/export

    def on_export_excel_clicked(self):
        """
        현재 Project 전체를 Excel 파일(.xlsx)로 내보낸다.
        - Issues, TestcaseSteps, Relations, TestPlanTestcases, TestExecutions, TestcaseExecutions 시트 생성
        """
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file_path:
            return

        try:
            excel_io.export_project_to_excel(self.conn, self.project.id, file_path)
            self.status_bar.showMessage(f"Exported project to Excel: {file_path}")
        except Exception as e:
            self.status_bar.showMessage(f"Excel export failed: {e}")
            print(f"[ERROR] Excel export failed: {e}")

    def on_import_excel_clicked(self):
        """
        Excel 파일(.xlsx)에서 데이터를 읽어와 로컬 DB에 병합(import)한다.
        - Issues / TestcaseSteps / Relations 시트를 1차 대상으로 처리
        - Import 후 로컬 트리를 새로고침한다.
        """
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot import.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import from Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file_path:
            return

        from PySide6.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QLabel,
            QProgressBar,
            QTextEdit,
            QDialogButtonBox,
            QApplication,
        )
        from PySide6.QtCore import Qt

        # 진행율/현황 표시용 팝업 다이얼로그
        dlg = QDialog(self)
        dlg.setWindowTitle("Import from Excel – Progress")
        dlg.setWindowModality(Qt.ApplicationModal)
        layout = QVBoxLayout(dlg)

        lbl_status = QLabel("Excel 파일을 읽는 중입니다...")
        progress = QProgressBar()
        progress.setRange(0, 100)
        progress.setValue(0)
        txt_log = QTextEdit()
        txt_log.setReadOnly(True)

        layout.addWidget(lbl_status)
        layout.addWidget(progress)
        layout.addWidget(txt_log)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_close = btn_box.button(QDialogButtonBox.Close)
        btn_close.setEnabled(False)
        layout.addWidget(btn_box)

        def on_close() -> None:
            dlg.close()

        btn_box.rejected.connect(on_close)

        dlg.resize(600, 400)
        dlg.show()
        QApplication.processEvents()

        # import_project_from_excel 에 전달할 진행 콜백
        def _progress_cb(message: str, current: int, total: int) -> None:
            if total <= 0:
                total = 1
            value = max(0, min(100, int(current * 100 / total)))
            progress.setValue(value)
            lbl_status.setText(message)
            txt_log.append(message)
            QApplication.processEvents()

        try:
            excel_io.import_project_from_excel(self.conn, self.project.id, file_path, progress_cb=_progress_cb)
            self.reload_local_tree()
            done_msg = f"Imported data from Excel: {file_path}"
            self.status_bar.showMessage(done_msg)
            lbl_status.setText("Import completed successfully.")
            txt_log.append(done_msg)
            progress.setValue(100)
        except Exception as e:
            err_msg = f"Excel import failed: {e}"
            self.status_bar.showMessage(err_msg)
            txt_log.append(err_msg)
            lbl_status.setText("Import failed.")
            print(f"[ERROR] Excel import failed: {e}")
        finally:
            btn_close.setEnabled(True)
            # 사용자가 진행 내역을 확인할 수 있도록 다이얼로그를 닫을 때까지 대기
            dlg.exec()

    def on_export_testexecution_report_clicked(self):
        """
        현재 선택된 TEST_EXECUTION 이슈를 기준으로 간단한 실행 결과 리포트를 Excel(.xlsx) 로 내보낸다.

        구성:
          - Summary 시트: TE 메타(Env/Start/End/Result/ExecutedBy) + TCE 결과 집계(PASS/FAIL 등 카운트)
          - TCEs 시트: 각 Test Case Execution 행 (Order, Test Case Key, Summary, Assignee, Result, Env, Actual Time, Defects)
        """
        # 현재 이슈 확인
        if self.current_issue_id is None:
            self.status_bar.showMessage("No issue selected; cannot export Test Execution report.")
            return

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage("Current issue not found in DB; cannot export report.")
            return

        if (issue.get("issue_type") or "").upper() != "TEST_EXECUTION":
            self.status_bar.showMessage("Current issue is not a TEST_EXECUTION; report export is only for Test Executions.")
            return

        jira_key = issue.get("jira_key") or ""

        from backend.db import get_or_create_testexecution_for_issue, get_testcase_executions

        # TE 메타 + TCE 목록 조회
        te_row = get_or_create_testexecution_for_issue(self.conn, self.current_issue_id)
        te_id = te_row["id"]
        tce_rows = get_testcase_executions(self.conn, te_id)

        # 파일 저장 위치 선택
        default_name = f"TestExecutionReport_{jira_key or 'LOCAL'}"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Test Execution Report",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file_path:
            return

        try:
            openpyxl = excel_io._ensure_openpyxl()  # type: ignore[attr-defined]
        except Exception as e:
            self.status_bar.showMessage(str(e))
            return

        from openpyxl.utils import get_column_letter  # type: ignore

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Summary 시트: TE 메타
        ws_summary.append(["Field", "Value"])
        ws_summary.append(["Test Execution Key", jira_key])
        ws_summary.append(["Environment", te_row.get("environment") or ""])
        ws_summary.append(["Start Date", te_row.get("start_date") or ""])
        ws_summary.append(["End Date", te_row.get("end_date") or ""])
        ws_summary.append(["Result", te_row.get("result") or ""])
        ws_summary.append(["Executed By", te_row.get("executed_by") or ""])
        ws_summary.append([])

        # TCE 결과 집계
        total = len(tce_rows)
        result_counts: dict[str, int] = {}
        for r in tce_rows:
            res = (r.get("result") or "").strip() or "(empty)"
            result_counts[res] = result_counts.get(res, 0) + 1

        ws_summary.append(["Total Test Cases", total])
        ws_summary.append([])
        ws_summary.append(["Result", "Count"])
        for res, cnt in result_counts.items():
            ws_summary.append([res, cnt])

        # 두 번째 시트: TCE 목록
        ws_tces = wb.create_sheet("TCEs")
        ws_tces.append(
            [
                "Order",
                "Test Case Key",
                "Summary",
                "Assignee",
                "Result",
                "RTM Env",
                "Actual Time (min)",
                "Defects",
            ]
        )

        # Test Case summary 조회를 위한 헬퍼
        cur = self.conn.cursor()
        for r in tce_rows:
            tc_id = r.get("testcase_id")
            jira_tc = ""
            summary_tc = ""
            if tc_id:
                cur.execute(
                    "SELECT jira_key, summary FROM issues WHERE id = ? AND is_deleted = 0",
                    (tc_id,),
                )
                row_tc = cur.fetchone()
                if row_tc:
                    jira_tc = row_tc["jira_key"] or ""
                    summary_tc = row_tc["summary"] or ""

            ws_tces.append(
                [
                    r.get("order_no"),
                    jira_tc,
                    summary_tc,
                    r.get("assignee") or "",
                    r.get("result") or "",
                    r.get("rtm_environment") or "",
                    r.get("actual_time") or "",
                    r.get("defects") or "",
                ]
            )

        # 간단한 자동 너비 조정
        for ws in (ws_summary, ws_tces):
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                    except Exception:
                        val = ""
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

        try:
            wb.save(file_path)
            self.status_bar.showMessage(f"Exported Test Execution report to: {file_path}")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to save Test Execution report: {e}")

    def on_refresh_activity_clicked(self):
        """
        오른쪽 온라인 패널에서 선택된 이슈의 JIRA Comments / History 를
        조회하여 해당 패널의 Activity 영역에 표시한다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is offline; cannot load activity.")
            return

        # 온라인 패널 IssueTabWidget 의 JIRA Key 기준으로 Activity 조회
        tabs = self.right_panel.issue_tabs
        jira_key = tabs.ed_jira_key.text().strip()
        if not jira_key:
            self.status_bar.showMessage("Selected online issue has no JIRA Key; no remote activity.")
            return

        try:
            self.status_bar.showMessage(f"Loading activity from JIRA for {jira_key}...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            data = self.jira_client.get_jira_issue(jira_key, expand="comments,changelog")

            # Comments
            activity_lines: list[str] = []
            fields = data.get("fields", {})
            comments = None
            comment_container = fields.get("comment")
            if isinstance(comment_container, dict):
                comments = comment_container.get("comments")
            if isinstance(comments, list):
                # IssueTabWidget 에 원본 댓글 목록 캐시 (온라인 패널 기준)
                tabs.set_activity_comments(comments)

                activity_lines.append("=== Comments ===")
                for c in comments:
                    author = (
                        c.get("author", {}).get("displayName")
                        if isinstance(c.get("author"), dict)
                        else ""
                    )
                    created = c.get("created", "")
                    body = c.get("body", "")
                    activity_lines.append(f"- [{created}] {author}:")
                    activity_lines.append(str(body))
                    activity_lines.append("")
            else:
                # 댓글이 없으면 캐시도 비운다.
                tabs.set_activity_comments([])

            # Changelog (History)
            changelog = data.get("changelog", {})
            histories = changelog.get("histories") if isinstance(changelog, dict) else None
            if isinstance(histories, list) and histories:
                activity_lines.append("=== History ===")
                for h in histories:
                    author = (
                        h.get("author", {}).get("displayName")
                        if isinstance(h.get("author"), dict)
                        else ""
                    )
                    created = h.get("created", "")
                    items = h.get("items") or []
                    activity_lines.append(f"- [{created}] {author}")
                    for it in items:
                        field = it.get("field", "")
                        from_val = it.get("fromString", "")
                        to_val = it.get("toString", "")
                        activity_lines.append(f"    {field}: '{from_val}' → '{to_val}'")
                    activity_lines.append("")

            if not activity_lines:
                activity_text = "No comments or history found for this issue."
            else:
                activity_text = "\n".join(activity_lines)

            tabs.set_activity_text(activity_text)
            self.status_bar.showMessage(f"Loaded activity from JIRA for {jira_key}")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to load activity: {e}")
            print(f"[ERROR] Failed to load activity for {jira_key}: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def on_add_comment_clicked(self):
        """
        현재 선택된 이슈에 대해 JIRA Issue Comment 를 추가한다.
        - 간단한 멀티라인 입력 다이얼로그를 띄운 뒤, add_issue_comment 호출
        - 성공 시 Activity 를 다시 로드하여 반영
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is offline; cannot add comment.")
            return

        tabs = self.right_panel.issue_tabs
        jira_key = tabs.ed_jira_key.text().strip()
        if not jira_key:
            self.status_bar.showMessage("Selected online issue has no JIRA Key; cannot add comment.")
            return

        text, ok = QInputDialog.getMultiLineText(
            self,
            "Add Comment",
            f"Add comment to {jira_key}:",
            "",
        )
        if not ok or not text.strip():
            return

        try:
            self.status_bar.showMessage(f"Adding comment to {jira_key}...")
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self.jira_client.add_issue_comment(jira_key, text.strip())
            # 댓글 추가 후 Activity 새로고침
            self.on_refresh_activity_clicked()
        except Exception as e:
            self.status_bar.showMessage(f"Failed to add comment: {e}")
            print(f"[ERROR] Failed to add comment for {jira_key}: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def on_edit_comment_clicked(self):
        """
        Details 탭 Activity 에서 마지막 JIRA 댓글을 수정한다.
        - on_refresh_activity_clicked 로 캐시된 _activity_comments 의 마지막 항목을 대상으로 함.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is offline; cannot edit comment.")
            return

        tabs = self.right_panel.issue_tabs
        jira_key = tabs.ed_jira_key.text().strip()
        if not jira_key:
            self.status_bar.showMessage("Selected online issue has no JIRA Key; cannot edit comment.")
            return

        comments = tabs.get_activity_comments()
        if not comments:
            self.status_bar.showMessage("No comments to edit for this issue.")
            return

        last = comments[-1]
        comment_id = last.get("id")
        body = last.get("body") or ""
        if not comment_id:
            self.status_bar.showMessage("Last comment has no id; cannot edit.")
            return

        text, ok = QInputDialog.getMultiLineText(
            self,
            "Edit Last Comment",
            f"Edit last comment on {jira_key}:",
            str(body),
        )
        if not ok:
            return
        new_text = text.strip()
        if not new_text or new_text == str(body).strip():
            return

        try:
            self.status_bar.showMessage(f"Updating last comment on {jira_key}...")
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self.jira_client.update_issue_comment(jira_key, comment_id, new_text)
            self.on_refresh_activity_clicked()
        except Exception as e:
            self.status_bar.showMessage(f"Failed to update comment: {e}")
            print(f"[ERROR] Failed to update comment for {jira_key}: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def on_delete_comment_clicked(self):
        """
        Details 탭 Activity 에서 마지막 JIRA 댓글을 삭제한다.
        - on_refresh_activity_clicked 로 캐시된 _activity_comments 의 마지막 항목을 대상으로 함.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("JIRA is offline; cannot delete comment.")
            return

        tabs = self.right_panel.issue_tabs
        jira_key = tabs.ed_jira_key.text().strip()
        if not jira_key:
            self.status_bar.showMessage("Selected online issue has no JIRA Key; cannot delete comment.")
            return

        comments = tabs.get_activity_comments()
        if not comments:
            self.status_bar.showMessage("No comments to delete for this issue.")
            return

        last = comments[-1]
        comment_id = last.get("id")
        body = last.get("body") or ""
        if not comment_id:
            self.status_bar.showMessage("Last comment has no id; cannot delete.")
            return

        preview = str(body)
        if len(preview) > 200:
            preview = preview[:200] + "..."

        confirm = QMessageBox.question(
            self,
            "Delete Last Comment",
            f"Delete the last comment on {jira_key}?\n\n{preview}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        try:
            self.status_bar.showMessage(f"Deleting last comment on {jira_key}...")
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self.jira_client.delete_issue_comment(jira_key, comment_id)
            self.on_refresh_activity_clicked()
        except Exception as e:
            self.status_bar.showMessage(f"Failed to delete comment: {e}")
            print(f"[ERROR] Failed to delete comment for {jira_key}: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    # --------------------------------------------------------------------- Local Activity (left panel)

    def on_refresh_local_activity_clicked(self):
        """
        현재 선택된 로컬 이슈의 local_activity 텍스트를 Activity 영역에 표시한다.
        JIRA 와는 무관한 순수 로컬 메모 영역이다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot load local activity.")
            return

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage("Issue not found in local DB; cannot load local activity.")
            return

        text = issue.get("local_activity") or ""
        self.left_panel.issue_tabs.set_activity_text(text)
        self.status_bar.showMessage("Loaded local activity.")

    def on_add_local_activity_clicked(self):
        """
        현재 선택된 로컬 이슈의 local_activity 에 새 메모를 추가한다.
        - 기존 텍스트 뒤에 공백 줄을 추가하고, 새 입력 내용을 이어붙인다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot add activity.")
            return

        tabs = self.left_panel.issue_tabs
        current_text = tabs.txt_activity.toPlainText()

        text, ok = QInputDialog.getMultiLineText(
            self,
            "Add Local Activity",
            "현재 이슈에 추가할 메모를 입력하세요:",
            "",
        )
        if not ok or not text.strip():
            return

        new_block = text.strip()

        # 설정에 따라 타임스탬프를 자동으로 앞에 붙인다.
        try:
            from datetime import datetime

            cfg = (self.local_settings or {}).get("activity", {}) if hasattr(self, "local_settings") else {}
            if cfg.get("append_timestamp_on_add", False):
                fmt = cfg.get("timestamp_format") or "%Y-%m-%d %H:%M"
                ts = datetime.now().strftime(fmt)
                new_block = f"[{ts}] {new_block}"
        except Exception:
            # 타임스탬프 생성 실패 시에는 그냥 원본 텍스트만 사용
            pass

        if current_text.strip():
            combined = current_text.rstrip() + "\n\n" + new_block
        else:
            combined = new_block

        from backend.db import update_issue_fields

        update_issue_fields(self.conn, self.current_issue_id, {"local_activity": combined})
        tabs.set_activity_text(combined)
        self.status_bar.showMessage("Added local activity.")

    def on_edit_local_activity_clicked(self):
        """
        현재 선택된 로컬 이슈의 local_activity 전체 텍스트를 편집한다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot edit activity.")
            return

        tabs = self.left_panel.issue_tabs
        current_text = tabs.txt_activity.toPlainText()

        text, ok = QInputDialog.getMultiLineText(
            self,
            "Edit Local Activity",
            "현재 이슈의 Activity 내용을 편집합니다:",
            current_text,
        )
        if not ok:
            return

        new_text = text.strip()
        from backend.db import update_issue_fields

        update_issue_fields(self.conn, self.current_issue_id, {"local_activity": new_text})
        tabs.set_activity_text(new_text)
        self.status_bar.showMessage("Updated local activity.")

    def on_delete_local_activity_clicked(self):
        """
        현재 선택된 로컬 이슈의 local_activity 를 모두 삭제한다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot delete activity.")
            return

        tabs = self.left_panel.issue_tabs
        current_text = tabs.txt_activity.toPlainText().strip()
        if not current_text:
            self.status_bar.showMessage("No local activity to delete.")
            return

        confirm = QMessageBox.question(
            self,
            "Delete Local Activity",
            "현재 이슈의 Activity 내용을 모두 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        from backend.db import update_issue_fields

        update_issue_fields(self.conn, self.current_issue_id, {"local_activity": ""})
        tabs.set_activity_text("")
        self.status_bar.showMessage("Deleted local activity.")

    # --------------------------------------------------------------------- Attachments (Details tab)

    def _reload_issue_from_jira_standard(self, issue_type: str, jira_key: str) -> None:
        """
        JIRA 표준 Issue REST (/rest/api/2/issue/{key}) 를 사용해
        현재 선택된 이슈의 필드를 다시 가져와 로컬 DB와 UI를 갱신한다.

        특히 attachments 필드(JSON 배열)를 동기화할 때 사용한다.
        """
        try:
            jira_issue_json = self.jira_client.get_jira_issue(jira_key)
            updates = jira_mapping.map_jira_to_local(issue_type, jira_issue_json)
            if updates:
                update_issue_fields(self.conn, self.current_issue_id, updates)
            # DB 에 반영된 최신 이슈를 다시 로딩하여 UI 갱신
            refreshed = get_issue_by_id(self.conn, self.current_issue_id)
            if refreshed:
                self.left_panel.issue_tabs.set_issue(refreshed)
        except Exception as e:
            self.status_bar.showMessage(f"Failed to reload issue from JIRA: {e}")
            print(f"[ERROR] Failed to reload issue {jira_key} from JIRA: {e}")

    def on_upload_attachment_clicked(self):
        """
        현재 선택된 로컬 이슈에 대해 로컬 첨부파일을 추가한다.
        - 파일 선택 다이얼로그에서 선택한 파일을 attachments/<ISSUE_TYPE>/<ISSUE_ID>/ 아래에 복사
        - issues.attachments 컬럼에 JSON 메타(파일명, 크기, local_path 등)를 저장하고 리스트를 갱신
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot add attachment.")
            return

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage("Issue not found in local DB; cannot add attachment.")
            return

        issue_type = issue.get("issue_type") or "UNKNOWN"

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Add attachment to local issue (type={issue_type})",
            "",
            "All Files (*)",
        )
        if not file_path:
            return

        from pathlib import Path
        import shutil
        import os
        import json
        from backend.attachments_fs import get_issue_attachments_dir
        from backend.db import update_issue_fields

        try:
            src = Path(file_path)
            # 설정에 따라 첨부 루트 디렉터리를 결정
            root = self._get_attachments_root()
            dst_dir = get_issue_attachments_dir(issue_type, self.current_issue_id, root=root)
            dst_dir.mkdir(parents=True, exist_ok=True)
            dst = dst_dir / src.name
            shutil.copy2(src, dst)

            attachments_root = self._get_attachments_root()
            rel_path = str(dst.relative_to(attachments_root))
            size = dst.stat().st_size if dst.exists() else None

            # 기존 메타 로드
            raw = issue.get("attachments")
            items = []
            if raw:
                try:
                    if isinstance(raw, str):
                        items = json.loads(raw)
                    elif isinstance(raw, list):
                        items = raw
                except Exception:
                    items = []

            if not isinstance(items, list):
                items = []

            items.append(
                {
                    "filename": src.name,
                    "size": size,
                    "id": None,
                    "content": None,
                    "local_path": rel_path,
                }
            )

            json_text = json.dumps(items, ensure_ascii=False)
            update_issue_fields(self.conn, self.current_issue_id, {"attachments": json_text})

            # UI 갱신
            self.left_panel.issue_tabs._load_attachments_list(json_text)
            self.status_bar.showMessage("Added local attachment.")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to add local attachment: {e}")
            print(f"[ERROR] Failed to add local attachment: {e}")

    def on_delete_attachment_clicked(self):
        """
        Details 탭의 첨부 리스트에서 선택된 로컬 첨부파일을 삭제한다.
        - attachments/<ISSUE_TYPE>/<ISSUE_ID>/ 아래의 파일을 삭제하고
          issues.attachments 메타에서도 제거한다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No local issue selected; cannot delete attachment.")
            return

        tabs = self.left_panel.issue_tabs
        current_item = tabs.attachments_list.currentItem()
        if current_item is None:
            self.status_bar.showMessage("No attachment selected.")
            return

        local_path = current_item.data(Qt.UserRole + 2)
        if not local_path:
            self.status_bar.showMessage("Selected attachment has no local path; cannot delete locally.")
            return

        confirm = QMessageBox.question(
            self,
            "Delete Local Attachment",
            f"Delete selected local attachment?\n\n{local_path}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        from pathlib import Path
        import json
        from backend.db import update_issue_fields

        try:
            root = self._get_attachments_root()
            full_path = root / local_path
            try:
                if full_path.exists():
                    full_path.unlink()
            except Exception:
                # 파일 삭제 실패는 메타만 정리하고 지나간다.
                pass

            issue = get_issue_by_id(self.conn, self.current_issue_id)
            raw = issue.get("attachments")
            items = []
            if raw:
                try:
                    if isinstance(raw, str):
                        items = json.loads(raw)
                    elif isinstance(raw, list):
                        items = raw
                except Exception:
                    items = []
            if not isinstance(items, list):
                items = []

            new_items = []
            for att in items:
                if not isinstance(att, dict):
                    continue
                if att.get("local_path") == local_path:
                    continue
                new_items.append(att)

            json_text = json.dumps(new_items, ensure_ascii=False)
            update_issue_fields(self.conn, self.current_issue_id, {"attachments": json_text})
            tabs._load_attachments_list(json_text)
            self.status_bar.showMessage("Deleted local attachment.")
        except Exception as e:
            self.status_bar.showMessage(f"Failed to delete local attachment: {e}")
            print(f"[ERROR] Failed to delete local attachment: {e}")

    def on_open_attachment_clicked(self):
        """
        Details 탭의 첨부 리스트에서 선택된 항목을 연다.
        - 로컬 경로(local_path)가 있으면 해당 파일을 OS 기본 핸들러로 연다.
        - 로컬 경로가 없고 JIRA URL 이 있으면 브라우저로 연다.
        """
        import webbrowser
        from pathlib import Path

        tabs = self.left_panel.issue_tabs
        current_item = tabs.attachments_list.currentItem()
        if current_item is None:
            self.status_bar.showMessage("No attachment selected.")
            return

        local_path = current_item.data(Qt.UserRole + 2)
        if local_path:
            root = self._get_attachments_root()
            full_path = root / local_path
            if not full_path.exists():
                self.status_bar.showMessage(f"Local attachment not found: {full_path}")
                return
            webbrowser.open(full_path.as_uri())
            return

        url = current_item.data(Qt.UserRole + 1)
        if not url:
            self.status_bar.showMessage("Selected attachment has no URL or local path.")
            return

        webbrowser.open(url)

    # --------------------------------------------------------------------- Defects tab (local view)

    def _collect_defects_for_issue(self, issue_id: int) -> List[Dict[str, Any]]:
        """
        현재 이슈 및 관련 Test Execution / Test Case Execution 으로부터
        Defect 로컬 이슈들을 수집한다.

        1) relations 테이블에서 dst_issue_type = 'DEFECT' 인 항목
        2) testcase_executions.defects 문자열(TC 키 목록) 에 해당하는 이슈들
        """
        defects: Dict[int, Dict[str, Any]] = {}

        # 1) Relations 기반 Defects
        rels = get_relations_for_issue(self.conn, issue_id)
        for r in rels:
            if (r.get("dst_issue_type") or "").upper() != "DEFECT":
                continue
            dst_id = r.get("dst_issue_id")
            if not dst_id:
                continue
            try:
                dst_id_int = int(dst_id)
            except (TypeError, ValueError):
                continue
            if dst_id_int in defects:
                # 링크 출처 정보만 추가
                lf = defects[dst_id_int].setdefault("linked_from", "")
                lf2 = f"Relation({issue_id})"
                if lf:
                    if lf2 not in lf.split(", "):
                        defects[dst_id_int]["linked_from"] = lf + ", " + lf2
                else:
                    defects[dst_id_int]["linked_from"] = lf2
                continue

            iss = get_issue_by_id(self.conn, dst_id_int)
            if not iss:
                continue
            defects[dst_id_int] = {
                "id": dst_id_int,
                "jira_key": iss.get("jira_key") or "",
                "summary": iss.get("summary") or "",
                "status": iss.get("status") or "",
                "priority": iss.get("priority") or "",
                "assignee": iss.get("assignee") or "",
                "linked_from": f"Relation({issue_id})",
            }

        # 2) 해당 이슈가 TEST_EXECUTION 인 경우, 그 하위 TCE 의 defects 문자열 기반
        row = None
        try:
            row = get_or_create_testexecution_for_issue(self.conn, issue_id)
        except Exception:
            row = None

        if row:
            tces = get_testcase_executions(self.conn, row["id"])
            from backend.db import get_issue_by_jira_key

            for tce in tces:
                defects_str = tce.get("defects") or ""
                if not defects_str:
                    continue
                keys = [x.strip() for x in defects_str.split(",") if x.strip()]
                for key in keys:
                    iss = get_issue_by_jira_key(self.conn, self.project.id, key)
                    if not iss:
                        continue
                    did = iss.get("id")
                    if not did:
                        continue
                    if did in defects:
                        lf = defects[did].setdefault("linked_from", "")
                        lf2 = f"TCE({tce.get('id')})"
                        if lf:
                            if lf2 not in lf.split(", "):
                                defects[did]["linked_from"] = lf + ", " + lf2
                        else:
                            defects[did]["linked_from"] = lf2
                        continue
                    defects[did] = {
                        "id": did,
                        "jira_key": iss.get("jira_key") or "",
                        "summary": iss.get("summary") or "",
                        "status": iss.get("status") or "",
                        "priority": iss.get("priority") or "",
                        "assignee": iss.get("assignee") or "",
                        "linked_from": f"TCE({tce.get('id')})",
                    }

        return list(defects.values())

    def _refresh_defects_tab_for_current_issue(self) -> None:
        """현재 선택된 이슈 기준으로 Defects 탭 내용을 갱신."""
        if getattr(self.left_panel, "issue_tabs", None) is None:
            return
        if self.current_issue_id is None:
            self.left_panel.issue_tabs.load_defects_for_issue([])
            return

        defects = self._collect_defects_for_issue(self.current_issue_id)
        self.left_panel.issue_tabs.load_defects_for_issue(defects)

    def on_refresh_defects_tab_clicked(self):
        """Defects 탭의 Refresh 버튼 핸들러."""
        self._refresh_defects_tab_for_current_issue()

    def on_open_selected_defect_clicked(self):
        """
        Defects 탭에서 선택된 결함을 좌측(Local) 트리/Details 에서 열어준다.
        - Local ID 를 기준으로 issues.id 를 찾는다.
        """
        tabs = self.left_panel.issue_tabs
        row = tabs.defects_table.currentRow()
        if row < 0:
            self.status_bar.showMessage("No defect selected.")
            return
        id_item = tabs.defects_table.item(row, 0)
        if not id_item:
            self.status_bar.showMessage("Selected row has no Local ID.")
            return
        try:
            defect_id = int(id_item.text().strip())
        except (TypeError, ValueError):
            self.status_bar.showMessage("Invalid Local ID for selected defect.")
            return

        # 좌측 트리에서 해당 이슈를 선택하도록 시도 (간단 구현: 현재 이슈만 Details 로 로드)
        defect = get_issue_by_id(self.conn, defect_id)
        if not defect:
            self.status_bar.showMessage(f"Defect id={defect_id} not found in DB.")
            return

        self.current_issue_id = defect_id
        self.left_panel.issue_tabs.set_issue(defect)
        # Relations/Requirements/Testcases/Executions 등은 필요 시 on_local_tree_selection_changed 경로를 재사용해도 된다.

    # --------------------------------------------------------------------- JIRA create/delete

    

    # --------------------------------------------------------------------- Local save (Details / Steps / Relations / Plans / Executions)

    def on_save_issue_clicked(self):
        """
        현재 좌측(Local) 패널에서 선택된 이슈에 대해
        - Details 탭 필드
        - Steps (TEST_CASE)
        - Relations (모든 이슈)
        - Test Plan - Test Case 매핑 (TEST_PLAN)
        - Test Execution 메타/케이스 (TEST_EXECUTION)
        를 로컬 SQLite DB에 저장한다.
        """
        if self.current_issue_id is None:
            self.status_bar.showMessage("No issue selected to save.")
            return

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage(f"Issue id={self.current_issue_id} not found in DB.")
            return

        issue_type = issue.get("issue_type")
        tabs = self.left_panel.issue_tabs

        # 1) Details 탭 → issues 테이블 필드 업데이트
        fields: Dict[str, Any] = {
            "summary": tabs.ed_summary.text().strip(),
            # Status / Priority 는 콤보박스 currentText 사용
            "status": tabs.ed_status.currentText().strip(),
            "priority": tabs.ed_priority.currentText().strip(),
            "assignee": tabs.ed_assignee.text().strip(),
            "reporter": tabs.ed_reporter.text().strip(),
            "labels": tabs.ed_labels.text().strip(),
            "components": tabs.ed_components.text().strip(),
            "security_level": tabs.ed_security_level.text().strip(),
            "fix_versions": tabs.ed_fix_versions.text().strip(),
            "affects_versions": tabs.ed_affects_versions.text().strip(),
            "epic_link": tabs.ed_epic_link.text().strip(),
            "sprint": tabs.ed_sprint.text().strip(),
            "rtm_environment": tabs.ed_rtm_env.currentText().strip(),
            "due_date": tabs.ed_due_date.text().strip(),
            "description": tabs.txt_description.toPlainText().strip(),
            "attachments": tabs.ed_attachments.text().strip(),
        }
        if issue_type == "TEST_CASE" and hasattr(tabs, "get_preconditions_text"):
            fields["preconditions"] = tabs.get_preconditions_text()
        # 빈 문자열만 있는 키는 그대로 둬도 무방하지만, 필요시 None 제거도 가능
        update_issue_fields(self.conn, self.current_issue_id, fields)

        # 2) Steps 저장 (TEST_CASE일 때만)
        if issue_type == "TEST_CASE":
            try:
                steps = tabs.collect_steps()
                replace_steps_for_issue(self.conn, self.current_issue_id, steps)
            except Exception as e_steps:
                print(f"[WARN] save steps failed: {e_steps}")

        # 3) Relations 저장 (모든 이슈 공통)
        try:
            rels = tabs.collect_relations()
            replace_relations_for_issue(self.conn, self.current_issue_id, rels)
        except Exception as e_rels:
            print(f"[WARN] save relations failed: {e_rels}")

        # 4) Test Plan - Test Case 매핑 저장
        if issue_type == "TEST_PLAN":
            try:
                tp_records = tabs.collect_testplan_testcases()
                replace_testplan_testcases(self.conn, self.current_issue_id, tp_records)
            except Exception as e_tp:
                print(f"[WARN] save testplan mappings failed: {e_tp}")

        # 5) Test Execution 메타/케이스 저장
        if issue_type == "TEST_EXECUTION":
            try:
                te_fields = tabs.collect_testexecution_meta()
                update_testexecution_for_issue(self.conn, self.current_issue_id, te_fields)

                tce_records = tabs.collect_testcase_executions()
                if tce_records:
                    te_row = get_or_create_testexecution_for_issue(self.conn, self.current_issue_id)
                    replace_testcase_executions(self.conn, te_row["id"], tce_records)
            except Exception as e_te:
                print(f"[WARN] save testexecution data failed: {e_te}")

        self.status_bar.showMessage("Local issue saved.")

        # 현재 이슈의 "편집됨" 상태 초기화
        self.current_issue_dirty = False
        if self.current_issue_id is not None:
            try:
                self.dirty_issue_ids.discard(int(self.current_issue_id))
            except Exception:
                pass

        # 트리 레이블(특히 Summary)이 갱신되도록 트리를 다시 로드
        self.reload_local_tree()

    def on_create_in_jira_clicked(self):
        """
        현재 선택된 로컬 이슈를 기준으로 JIRA RTM에 새 엔티티를 생성한다.

        전제:
          - 현재 이슈에 jira_key 가 비어 있어야 한다. (이미 연동된 경우는 update 사용)
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot create: Jira RTM not configured.")
            return
        if self.current_issue_id is None:
            self.status_bar.showMessage("No issue selected to create in JIRA.")
            return

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage(f"Issue id={self.current_issue_id} not found in DB.")
            return

        issue_type = issue.get("issue_type")
        jira_key = issue.get("jira_key")
        if jira_key:
            self.status_bar.showMessage(f"Issue already has JIRA key ({jira_key}); use Push instead.")
            return

        # 생성 전 로컬 저장 (Details/Steps 등)
        self.on_save_issue_clicked()
        issue = get_issue_by_id(self.conn, self.current_issue_id)

        try:
            self.status_bar.showMessage(f"Creating new {issue_type} in JIRA...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            payload = jira_mapping.build_jira_create_payload(issue_type, issue)
            resp = self.jira_client.create_entity(issue_type, payload)

            new_key = None
            if isinstance(resp, dict):
                # 일반 Jira create 응답: {"id": "...", "key": "KVHSICCU-123", ...}
                new_key = resp.get("key") or resp.get("jiraKey") or resp.get("issueKey")

            if not new_key:
                # RTM 환경에 따라 응답 구조가 다를 수 있으므로, 필요 시 로깅 후 사용자에게 알림
                self.status_bar.showMessage("Created entity in JIRA, but could not determine new issue key.")
                print("[WARN] create_entity response without 'key':", resp)
            else:
                from backend.db import update_issue_fields
                update_issue_fields(self.conn, self.current_issue_id, {"jira_key": new_key})
                self.status_bar.showMessage(f"Created in JIRA as {new_key}.")
                # 트리 갱신
                self.reload_local_tree()

        except Exception as e:
            self.status_bar.showMessage(f"Create in JIRA failed: {e}")
            print(f"[ERROR] Create in JIRA failed: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def on_delete_in_jira_clicked(self):
        """
        현재 선택된 로컬 이슈에 연결된 JIRA 엔티티를 삭제한다.

        - 실제로는 RTM / Jira 서버에서 이슈 삭제(또는 RTM 엔티티 삭제)를 시도한다.
        - 로컬 DB의 이슈는 삭제하지 않고, jira_key 만 비워서 "오프라인 전용"으로 남긴다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot delete: Jira RTM not configured.")
            return
        # 현재 단일 선택 기준 삭제는 멀티 삭제 helper를 재사용한다.
        self._delete_online_issues_in_jira()

    # --------------------------------------------------------------------- JIRA issue sync (Details only, skeleton)

    def _delete_online_issues_in_jira(self):
        """
        우측 온라인 트리에서 선택된 이슈들(jira_key 기반)을 JIRA 에서 삭제하고,
        로컬 DB 에서도 해당 jira_key 를 비운다.
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot delete: Jira RTM not configured.")
            return
        if not self.project:
            self.status_bar.showMessage("No project loaded; cannot delete in JIRA.")
            return

        view = self.right_panel.tree_view
        model = view.model()
        selection_model = view.selectionModel()
        if model is None or selection_model is None:
            self.status_bar.showMessage("No online tree to delete from.")
            return

        from backend.db import update_issue_fields

        indexes = selection_model.selectedIndexes()
        targets = []  # (issue_type, jira_key)
        for idx in indexes:
            item = model.itemFromIndex(idx)
            if not item or item.data(Qt.UserRole) != "ISSUE":
                continue
            jira_key = item.data(Qt.UserRole + 1) or ""
            issue_type = item.data(Qt.UserRole + 2) or ""
            if jira_key:
                targets.append((str(issue_type), str(jira_key)))

        if not targets:
            self.status_bar.showMessage("No online issue selected to delete.")
            return

        count = len(targets)
        ret = QMessageBox.question(
            self,
            "Delete in JIRA",
            f"Delete {count} issue(s) in JIRA?\n\n이 작업은 JIRA 서버에서 실제 이슈 삭제를 시도합니다.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        try:
            self.status_bar.showMessage(f"Deleting {count} issue(s) in JIRA...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            # JIRA 삭제 + 로컬 jira_key 비우기
            for issue_type, jira_key in targets:
                try:
                    self.jira_client.delete_entity(issue_type, jira_key)
                except Exception as e_del:
                    print(f"[ERROR] Failed to delete {jira_key} in JIRA: {e_del}")
                    continue

                # 로컬 DB 에서 jira_key 일치하는 이슈들의 jira_key 를 비운다.
                cur = self.conn.cursor()
                cur.execute(
                    "SELECT id FROM issues WHERE project_id = ? AND jira_key = ? AND is_deleted = 0",
                    (self.project.id, jira_key),
                )
                rows = cur.fetchall()
                for r in rows:
                    update_issue_fields(self.conn, int(r["id"]), {"jira_key": ""})

            self.status_bar.showMessage(f"Deleted {count} issue(s) in JIRA (local issues kept).")
            self.reload_local_tree()
        finally:
            QApplication.restoreOverrideCursor()

# --------------------------------------------------------------------- JIRA issue sync (Details only, skeleton)

    def on_pull_issue_clicked(self):
        """
        선택된 로컬 이슈에 대해 JIRA RTM에서 필드를 가져와 로컬 DB에 반영한다.
        - summary / description / status / priority / assignee / reporter
        - labels / components / security_level / fix_versions / affects_versions
        - rtm_environment / due_date / created / updated / attachments
        - (TEST_CASE일 경우) Steps 정보까지 testcase_steps 테이블에 동기화
        - (TEST_PLAN일 경우) Test Plan - Test Case 매핑 동기화
        - (TEST_EXECUTION일 경우) Test Execution 메타 + Test Case Execution 목록 동기화
        - Relations (Jira issue links) 를 relations 테이블로 동기화
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot pull: Jira RTM not configured.")
            return
        if self.current_issue_id is None:
            self.status_bar.showMessage("No issue selected to pull from JIRA.")
            return

        # 로컬 이슈 정보 조회
        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage(f"Issue id={self.current_issue_id} not found in DB.")
            return

        issue_type = issue.get("issue_type")
        jira_key = issue.get("jira_key")
        if not jira_key:
            self.status_bar.showMessage("Selected issue has no JIRA key; cannot pull.")
            return

        # 동기화 영향 설명 및 사용자 확인 (현재 모드에 따라 안내 문구 조정)
        sync_mode = getattr(self, "sync_mode", "server")
        if sync_mode == "local":
            mode_desc = "현재 Sync 모드는 '로컬 우선' 입니다.\nPull 을 실행하면 여전히 서버 값이 로컬을 덮어쓰지만,\nPush 를 통해 다시 로컬 내용을 서버에 반영하는 것을 전제로 합니다.\n\n"
        elif sync_mode == "merge":
            mode_desc = "현재 Sync 모드는 '병합(Merge)' 입니다.\n이 Pull 은 서버 기준으로 가져오되, 일부 필드는 추후 병합 규칙에 따라 처리됩니다.\n(현 단계에서는 서버 값이 우선 적용됩니다.)\n\n"
        else:
            mode_desc = "현재 Sync 모드는 '서버 우선(Server → Local)' 입니다.\n\n"

        ret = QMessageBox.question(
            self,
            "Pull from JIRA",
            (
                f"{mode_desc}"
                f"선택된 이슈 {jira_key} ({issue_type}) 에 대해 JIRA/RTM 의 값을 로컬로 가져옵니다.\n\n"
                "- Details / Steps / Relations / Test Plan / Test Execution 등의 정보가\n"
                "  현재 로컬 DB 값 위에 덮어써질 수 있습니다.\n"
                "- 아직 저장하지 않은 로컬 수정 내용이 있다면 사라질 수 있습니다.\n\n"
                "계속 진행하시겠습니까?"
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        try:
            self.status_bar.showMessage(f"Pulling from JIRA: {jira_key} ({issue_type})...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            # 1) JIRA RTM 엔티티 조회
            data = self.jira_client.get_entity(issue_type, jira_key)

            # 2) mapping layer 사용하여 로컬 필드 업데이트 dict 생성
            updates = jira_mapping.map_jira_to_local(issue_type, data)

            if updates:
                update_issue_fields(self.conn, self.current_issue_id, updates)

            # 2-1) 첨부파일: 설정에 따라 JIRA 첨부파일을 로컬 파일로 다운로드하고 메타를 저장
            attach_cfg = (self.local_settings or {}).get("attachments", {}) if hasattr(self, "local_settings") else {}
            if not attach_cfg.get("auto_download_on_pull", True):
                pass
            else:
                try:
                    from backend.attachments_fs import get_issue_attachments_dir
                    import json
                    from pathlib import Path
                    import requests
                    from requests.auth import HTTPBasicAuth

                    fields = data.get("fields") or {}
                    att_list = fields.get("attachment") or []
                    if isinstance(att_list, list) and att_list:
                        root = self._get_attachments_root()
                        dst_dir = get_issue_attachments_dir(
                            issue_type or "UNKNOWN", self.current_issue_id, root=root
                        )
                        items: list[dict] = []

                        for att in att_list:
                            if not isinstance(att, dict):
                                continue
                            url = (
                                att.get("content")
                                or att.get("contentUrl")
                                or att.get("self")
                            )
                            filename = att.get("filename") or att.get("fileName") or att.get("name")
                            att_id = att.get("id") or att.get("attachmentId")
                            size = att.get("size") or att.get("filesize")
                            if not url or not filename or not att_id:
                                continue

                            # 다운로드 대상 경로: attachments/<TYPE>/<ISSUE_ID>/<ATT_ID>/<filename>
                            sub_dir = dst_dir / str(att_id)
                            sub_dir.mkdir(parents=True, exist_ok=True)
                            dst = sub_dir / filename

                            # 동일 파일이 이미 있으면 다운로드를 생략할 수도 있으나,
                            # 여기서는 간단히 항상 덮어쓴다.
                            try:
                                auth = HTTPBasicAuth(
                                    self.jira_client.config.username,
                                    self.jira_client.config.api_token,
                                )
                                resp = requests.get(url, auth=auth, stream=True)
                                resp.raise_for_status()
                                with open(dst, "wb") as f:
                                    for chunk in resp.iter_content(chunk_size=8192):
                                        if chunk:
                                            f.write(chunk)
                            except Exception as e_dl:
                                print(f"[WARN] Failed to download attachment {att_id} from {url}: {e_dl}")
                                continue

                            try:
                                actual_size = dst.stat().st_size
                            except OSError:
                                actual_size = size

                            rel_path = str(dst.relative_to(root))
                            items.append(
                                {
                                    "filename": filename,
                                    "size": actual_size,
                                    "id": att_id,
                                    "content": url,
                                    "local_path": rel_path,
                                }
                            )

                        if items:
                            # 기존 로컬 attachments 메타와 병합 (로컬 전용 항목은 유지)
                            issue_before = get_issue_by_id(self.conn, self.current_issue_id)
                            merged: list[dict] = []
                            if issue_before:
                                raw = issue_before.get("attachments")
                                if raw:
                                    try:
                                        if isinstance(raw, str):
                                            prev_items = json.loads(raw)
                                        elif isinstance(raw, list):
                                            prev_items = raw
                                        else:
                                            prev_items = []
                                    except Exception:
                                        prev_items = []
                                    if isinstance(prev_items, list):
                                        for it in prev_items:
                                            if not isinstance(it, dict):
                                                continue
                                            # JIRA id 가 없는 순수 로컬 첨부만 유지
                                            if not it.get("id"):
                                                merged.append(it)
                            merged.extend(items)
                            json_text = json.dumps(merged, ensure_ascii=False)
                            update_issue_fields(self.conn, self.current_issue_id, {"attachments": json_text})
                            # UI 갱신
                            self.left_panel.issue_tabs._load_attachments_list(json_text)
                except Exception as e_att:
                    print(f"[WARN] Failed to sync attachments from JIRA: {e_att}")

            # 3) Test Case인 경우: Steps도 별도 endpoint를 통해 동기화
            if issue_type == "TEST_CASE":
                try:
                    steps_json = self.jira_client.get_testcase_steps(jira_key)
                    local_steps = jira_mapping.map_jira_testcase_steps_to_local(steps_json)
                    from backend.db import replace_steps_for_issue  # local import to avoid circular issues
                    replace_steps_for_issue(self.conn, self.current_issue_id, local_steps)
                    # UI 갱신
                    if hasattr(self.left_panel.issue_tabs, "load_steps"):
                        self.left_panel.issue_tabs.load_steps(local_steps)
                except Exception as e_steps:
                    # Steps 동기화 실패는 치명적 에러로 간주하지 않고 로그만 남김
                    print(f"[WARN] Failed to sync Test Case steps from JIRA: {e_steps}")

            # 4) Test Plan인 경우: Test Case 매핑 동기화
            if issue_type == "TEST_PLAN":
                try:
                    tp_json = self.jira_client.get_testplan_testcases(jira_key)
                    tp_items = jira_mapping.map_jira_testplan_testcases_to_local(tp_json)
                    from backend.db import replace_testplan_testcases, get_testplan_testcases
                    # testcase_key -> local testcase_id 로 변환
                    from backend.db import get_issue_by_jira_key
                    records = []
                    for item in tp_items:
                        tc_key = item.get("testcase_key")
                        if not tc_key:
                            continue
                        tc_issue = get_issue_by_jira_key(self.conn, self.project.id, tc_key)
                        if not tc_issue:
                            # 로컬에 없는 TC는 스킵
                            continue
                        records.append(
                            {
                                "order_no": item.get("order_no") or 0,
                                "testcase_id": tc_issue["id"],
                            }
                        )
                    if records:
                        replace_testplan_testcases(self.conn, self.current_issue_id, records)
                        tp_rels = get_testplan_testcases(self.conn, self.current_issue_id)
                        if hasattr(self.left_panel.issue_tabs, "load_testplan_testcases"):
                            self.left_panel.issue_tabs.load_testplan_testcases(tp_rels)
                except Exception as e_tp:
                    print(f"[WARN] Failed to sync Test Plan testcases from JIRA: {e_tp}")

            # 5) Test Execution인 경우: 메타 + Test Case Execution 동기화
            if issue_type == "TEST_EXECUTION":
                try:
                    te_json = self.jira_client.get_testexecution_details(jira_key)
                    from backend.db import get_or_create_testexecution_for_issue, update_testexecution_for_issue, get_testcase_executions, replace_testcase_executions
                    te_meta = jira_mapping.map_jira_testexecution_meta_to_local(te_json)
                    if te_meta:
                        # get_or_create 를 이용해 id 확보 후 메타 업데이트
                        te_row = get_or_create_testexecution_for_issue(self.conn, self.current_issue_id)
                        update_testexecution_for_issue(self.conn, self.current_issue_id, te_meta)
                    # Test Case Execution 목록
                    tce_json = self.jira_client.get_testexecution_testcases(jira_key)
                    tce_items = jira_mapping.map_jira_testexecution_testcases_to_local(tce_json)
                    if tce_items:
                        # testcase_key -> local testcase_id 매핑 필요
                        from backend.db import get_issue_by_jira_key
                        tce_records = []
                        for item in tce_items:
                            tc_key = item.get("testcase_key")
                            if not tc_key:
                                continue
                            tc_issue = get_issue_by_jira_key(self.conn, self.project.id, tc_key)
                            if not tc_issue:
                                continue
                            tce_records.append(
                                {
                                    "order_no": item.get("order_no") or 0,
                                    "testcase_id": tc_issue["id"],
                                    "assignee": item.get("assignee") or "",
                                    "result": item.get("result") or "",
                                    "rtm_environment": item.get("rtm_environment") or "",
                                    "defects": item.get("defects") or "",
                                    "actual_time": item.get("actual_time"),
                                    # RTM Test Case Execution key (Step API / Defect 링크에 필요)
                                    "tce_test_key": item.get("tce_test_key"),
                                }
                            )
                        if tce_records:
                            # 다시 get_or_create로 testexecution id 확보 후 replace
                            te_row = get_or_create_testexecution_for_issue(self.conn, self.current_issue_id)
                            replace_testcase_executions(self.conn, te_row["id"], tce_records)
                            # UI 갱신: Executions 탭
                            execs = get_testcase_executions(self.conn, te_row["id"])
                            if hasattr(self.left_panel.issue_tabs, "load_testcase_executions"):
                                self.left_panel.issue_tabs.load_testcase_executions(execs)
                except Exception as e_te:
                    print(f"[WARN] Failed to sync Test Execution from JIRA: {e_te}")

            # 6) Jira issue links + RTM Requirement coverage -> local relations 동기화
            try:
                rel_entries = jira_mapping.extract_relations_from_jira(data)

                # RTM Requirement 의 testCasesCovered 리스트를 Relations 로 병합
                if (issue_type or "").upper() == "REQUIREMENT":
                    extra_entries: list[Dict[str, Any]] = []
                    tc_covered = data.get("testCasesCovered") or []
                    if isinstance(tc_covered, list):
                        for item in tc_covered:
                            if isinstance(item, dict):
                                tc_key = (
                                    item.get("key")
                                    or item.get("testCaseKey")
                                    or item.get("jiraKey")
                                )
                            else:
                                tc_key = str(item) if item is not None else ""
                            if not tc_key:
                                continue
                            extra_entries.append(
                                {
                                    "relation_type": "Tests",  # Requirement 가 src, Test Case 가 dst
                                    "dst_jira_key": tc_key,
                                    "dst_summary": item.get("summary") if isinstance(item, dict) else "",
                                }
                            )

                    if extra_entries:
                        # (relation_type, dst_jira_key) 기준으로 중복 제거하여 기존 rel_entries 와 병합
                        merged: dict[tuple[str, str], Dict[str, Any]] = {}
                        for rel in rel_entries or []:
                            key = (str(rel.get("relation_type") or ""), str(rel.get("dst_jira_key") or ""))
                            merged[key] = rel
                        for rel in extra_entries:
                            key = (str(rel.get("relation_type") or ""), str(rel.get("dst_jira_key") or ""))
                            merged[key] = rel
                        rel_entries = list(merged.values())

                if rel_entries:
                    from backend.db import replace_relations_for_issue, get_relations_for_issue, get_issue_by_jira_key
                    # dst_jira_key 를 로컬 issue_id 로 변환
                    rel_records = []
                    for rel in rel_entries:
                        dst_key = rel.get("dst_jira_key")
                        rel_type = rel.get("relation_type") or ""
                        if not dst_key:
                            continue
                        dst_issue = get_issue_by_jira_key(self.conn, self.project.id, dst_key)
                        if not dst_issue:
                            # 아직 트리에 존재하지 않는 이슈는 스킵
                            continue
                        rel_records.append(
                            {
                                "dst_issue_id": dst_issue["id"],
                                "relation_type": rel_type,
                            }
                        )
                    if rel_records:
                        replace_relations_for_issue(self.conn, self.current_issue_id, rel_records)
                        # UI 갱신: Relations / Requirements / Test Cases 탭
                        rels = get_relations_for_issue(self.conn, self.current_issue_id)
                        if hasattr(self.left_panel.issue_tabs, "load_relations"):
                            link_types = self.jira_field_options.get("relation_types", [])
                            self.left_panel.issue_tabs.load_relations(rels, link_types)
                        if hasattr(self.left_panel.issue_tabs, "load_requirements"):
                            reqs = [r for r in rels if r.get("dst_issue_type") == "REQUIREMENT"]
                            self.left_panel.issue_tabs.load_requirements(reqs)
                        if hasattr(self.left_panel.issue_tabs, "load_linked_testcases"):
                            tcs = [
                                r
                                for r in rels
                                if (r.get("dst_issue_type") or "").upper() == "TEST_CASE"
                            ]
                            self.left_panel.issue_tabs.load_linked_testcases(tcs)
            except Exception as e_rel:
                print(f"[WARN] Failed to sync relations from JIRA/RTM: {e_rel}")

            if updates:
                self.status_bar.showMessage(f"Pulled from JIRA and updated local issue {jira_key}.")
                self.reload_local_tree()
            else:
                self.status_bar.showMessage("No mappable fields from JIRA response.")

        except Exception as e:
            self.status_bar.showMessage(f"Pull from JIRA failed: {e}")
            print(f"[ERROR] Pull from JIRA failed: {e}")
        finally:
            QApplication.restoreOverrideCursor()
    def on_push_issue_clicked(self):
        """
        선택된 로컬 이슈의 필드(Details 탭 기준 중 안전한 범위)를 JIRA RTM에 업데이트한다.
        - summary / description / labels / components / duedate / environment
        - (TEST_CASE일 경우) testcase_steps 내용을 RTM Test Case Steps로 함께 업데이트
        - (TEST_PLAN일 경우) Test Plan - Test Case 매핑을 RTM으로 업데이트
        - (TEST_EXECUTION일 경우) Test Execution 메타 + Test Case Execution 목록을 RTM으로 업데이트
        - 로컬 relations 를 기준으로 Jira issueLink 를 생성 (단순 skeleton, 중복 링크 체크는 미구현)
        (status / priority / assignee / reporter 등은 기본 매핑에서 제외)
        """
        if not self.jira_available or not self.jira_client:
            self.status_bar.showMessage("Cannot push: Jira RTM not configured.")
            return
        if self.current_issue_id is None:
            self.status_bar.showMessage("No issue selected to push to JIRA.")
            return

        # 먼저 로컬 DB에 저장(Details/Steps/Relations 등)
        self.on_save_issue_clicked()

        issue = get_issue_by_id(self.conn, self.current_issue_id)
        if not issue:
            self.status_bar.showMessage(f"Issue id={self.current_issue_id} not found in DB.")
            return

        issue_type = issue.get("issue_type")
        jira_key = issue.get("jira_key")
        if not jira_key:
            self.status_bar.showMessage("Selected issue has no JIRA key; cannot push.")
            return

        # 동기화 영향 설명 및 사용자 확인 (Sync 모드에 따라 안내 문구 조정)
        sync_mode = getattr(self, "sync_mode", "server")
        if sync_mode == "server":
            mode_desc = (
                "현재 Sync 모드는 '서버 우선(Server → Local)' 입니다.\n"
                "이 Push 는 예외적으로 로컬 변경을 서버에 반영하지만,\n"
                "일반적인 기준은 여전히 서버 값입니다.\n\n"
            )
        elif sync_mode == "local":
            mode_desc = (
                "현재 Sync 모드는 '로컬 우선(Local → Server)' 입니다.\n"
                "서버의 기존 값보다 현재 로컬 값을 우선시하여 반영합니다.\n\n"
            )
        else:
            mode_desc = (
                "현재 Sync 모드는 '병합(Merge)' 입니다.\n"
                "로컬/서버 데이터를 비교·병합하는 전략을 전제로 하지만,\n"
                "현 단계에서는 로컬 값을 우선 반영하는 동작에 가깝습니다.\n\n"
            )

        ret = QMessageBox.question(
            self,
            "Push to JIRA / RTM",
            (
                f"{mode_desc}"
                f"선택된 이슈 {jira_key} ({issue_type}) 의 로컬 값을 JIRA/RTM 으로 전송합니다.\n\n"
                "- summary / description / labels / components / 환경 등의 메타 필드가\n"
                "  JIRA 이슈에 반영됩니다.\n"
                "- Test Case 인 경우 Steps, Test Plan 인 경우 Test Case 매핑,\n"
                "  Test Execution 인 경우 TCE 메타/결과가 RTM 측 데이터에 영향을 줍니다.\n"
                "- JIRA/RTM 상의 기존 값은 되돌리기 어렵습니다.\n\n"
                "이 모드에서 Push 를 진행하시겠습니까?"
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        # 로컬 필드를 JIRA payload로 변환 (mapping layer 사용)
        payload = jira_mapping.build_jira_update_payload(issue_type, issue)

        try:
            self.status_bar.showMessage(f"Pushing to JIRA: {jira_key} ({issue_type})...")
            QApplication.setOverrideCursor(Qt.WaitCursor)

            # 1) 기본 필드 업데이트
            self.jira_client.update_entity(issue_type, jira_key, payload)

            # 2) Test Case인 경우 Steps도 별도 endpoint로 업데이트
            if issue_type == "TEST_CASE":
                from backend.db import get_steps_for_issue  # local import
                local_steps = get_steps_for_issue(self.conn, self.current_issue_id)
                steps_payload = jira_mapping.build_jira_testcase_steps_payload(local_steps)
                try:
                    self.jira_client.update_testcase_steps(jira_key, steps_payload)
                except Exception as e_steps:
                    print(f"[WARN] Failed to push Test Case steps to JIRA: {e_steps}")

            # 3) Test Plan인 경우: Test Case 매핑을 RTM에 업데이트
            if issue_type == "TEST_PLAN":
                try:
                    from backend.db import get_testplan_testcases
                    tp_rels = get_testplan_testcases(self.conn, self.current_issue_id)
                    # tp_rels 는 join 결과로 testcase_jira_key 또는 jira_key 포함하도록 설계되어 있음
                    tp_payload = jira_mapping.build_jira_testplan_testcases_payload(tp_rels)
                    self.jira_client.update_testplan_testcases(jira_key, tp_payload)
                except Exception as e_tp:
                    print(f"[WARN] Failed to push Test Plan testcases to JIRA: {e_tp}")

            # 4) Test Execution인 경우: 메타 + Test Case Execution 목록을 RTM에 업데이트
            if issue_type == "TEST_EXECUTION":
                try:
                    from backend.db import get_or_create_testexecution_for_issue, get_testcase_executions
                    te_row = get_or_create_testexecution_for_issue(self.conn, self.current_issue_id)
                    # te_row 에는 이미 최신 메타가 저장되어 있다고 가정
                    te_meta = {
                        "environment": te_row.get("environment"),
                        "start_date": te_row.get("start_date"),
                        "end_date": te_row.get("end_date"),
                        "result": te_row.get("result"),
                        "executed_by": te_row.get("executed_by"),
                    }
                    te_payload = jira_mapping.build_jira_testexecution_payload(te_meta)
                    self.jira_client.update_testexecution(jira_key, te_payload)

                    # Test Case Executions
                    tce_records = get_testcase_executions(self.conn, te_row["id"])
                    tce_payload = jira_mapping.build_jira_testexecution_testcases_payload(tce_records)
                    self.jira_client.update_testexecution_testcases(jira_key, tce_payload)
                except Exception as e_te:
                    print(f"[WARN] Failed to push Test Execution to JIRA: {e_te}")

            # 5) Relations -> Jira issueLink 생성 (skeleton)
            try:
                from backend.db import get_relations_for_issue
                rels = get_relations_for_issue(self.conn, self.current_issue_id)
                # Jira issueLink 생성 및 Requirement 의 testCasesCovered 동기화에 함께 사용
                tc_keys_for_req: set[str] = set()
                for rel in rels:
                    dst_issue_id = rel.get("dst_issue_id")
                    if not dst_issue_id:
                        continue
                    dst_issue = get_issue_by_id(self.conn, dst_issue_id)
                    if not dst_issue:
                        continue
                    dst_key = dst_issue.get("jira_key")
                    if not dst_key:
                        continue

                    rel_type = rel.get("relation_type") or ""
                    # relation_type 예: "Relates (out)", "Relates (in)"
                    base_type = rel_type
                    direction = "out"
                    if "(" in rel_type and rel_type.endswith(")"):
                        base_type, paren = rel_type.rsplit("(", 1)
                        base_type = base_type.strip()
                        direction = paren[:-1].strip()
                    if not base_type:
                        base_type = "Relates"

                    # Requirement 이슈에서 relation_type 이 "Tests" 이고
                    # 대상 이슈가 TEST_CASE 인 경우, RTM Requirement 의 testCasesCovered 로도 사용
                    if (issue_type or "").upper() == "REQUIREMENT":
                        if (dst_issue.get("issue_type") or "").upper() == "TEST_CASE":
                            if dst_key:
                                tc_keys_for_req.add(dst_key)

                    # 방향에 따라 inward/outward 설정
                    if direction == "in":
                        inward_key = dst_key
                        outward_key = jira_key
                    else:
                        inward_key = jira_key
                        outward_key = dst_key

                    try:
                        self.jira_client.create_issue_link(base_type, inward_key, outward_key)
                    except Exception as e_link:
                        print(f"[WARN] Failed to create issue link {base_type} between {jira_key} and {dst_key}: {e_link}")

                # Requirement 의 경우, Relations 정보로부터 RTM testCasesCovered 필드를 업데이트
                if tc_keys_for_req:
                    try:
                        rtm_payload = {
                            "testCasesCovered": [{"key": k} for k in sorted(tc_keys_for_req)]
                        }
                        self.jira_client.update_entity("REQUIREMENT", jira_key, rtm_payload)
                    except Exception as e_rtm:
                        print(f"[WARN] Failed to push Requirement testCasesCovered to RTM for {jira_key}: {e_rtm}")
            except Exception as e_rel:
                print(f"[WARN] Failed to push relations to JIRA: {e_rel}")

            # 6) 로컬 첨부파일을 JIRA 로 업로드 (아직 업로드되지 않은 파일만)
            try:
                import json
                from pathlib import Path

                raw = issue.get("attachments")
                items = []
                if raw:
                    try:
                        if isinstance(raw, str):
                            items = json.loads(raw)
                        elif isinstance(raw, list):
                            items = raw
                    except Exception:
                        items = []
                if isinstance(items, list) and items:
                    root = self._get_attachments_root()
                    changed = False
                    for att in items:
                        if not isinstance(att, dict):
                            continue
                        # 이미 JIRA id 가 있으면 건너뜀
                        if att.get("id"):
                            continue
                        local_path = att.get("local_path")
                        if not local_path:
                            continue
                        full_path = root / local_path
                        if not full_path.exists():
                            continue
                        try:
                            self.jira_client.add_issue_attachment_from_path(
                                jira_key, str(full_path)
                            )
                            # 성공 시 id/size 등을 다시 채우기 위해 JIRA 표준 이슈를 한번 더 조회
                            changed = True
                        except Exception as e_att:
                            print(f"[WARN] Failed to upload attachment {full_path} to JIRA: {e_att}")
                    if changed:
                        # 새로 업로드된 첨부를 반영하기 위해 JIRA 이슈를 다시 읽어
                        # attachments 메타를 업데이트하되, local_path 는 유지한다.
                        try:
                            from backend.attachments_fs import get_issue_attachments_dir
                            import requests
                            from requests.auth import HTTPBasicAuth

                            jira_issue_json = self.jira_client.get_jira_issue(jira_key)
                            fields = jira_issue_json.get("fields") or {}
                            att_list = fields.get("attachment") or []
                            if isinstance(att_list, list):
                                root = self._get_attachments_root()
                                dst_dir = get_issue_attachments_dir(issue_type or "UNKNOWN", self.current_issue_id, root=root)
                                merged: list[dict] = []
                                # 기존 메타에서 local_path 없는 항목(서버 전용) 제외
                                for old in items:
                                    if isinstance(old, dict) and not old.get("id"):
                                        merged.append(old)
                                auth = HTTPBasicAuth(
                                    self.jira_client.config.username,
                                    self.jira_client.config.api_token,
                                )
                                for att in att_list:
                                    if not isinstance(att, dict):
                                        continue
                                    url = (
                                        att.get("content")
                                        or att.get("contentUrl")
                                        or att.get("self")
                                    )
                                    filename = att.get("filename") or att.get("fileName") or att.get("name")
                                    att_id = att.get("id") or att.get("attachmentId")
                                    size = att.get("size") or att.get("filesize")
                                    if not url or not filename or not att_id:
                                        continue
                                    sub_dir = dst_dir / str(att_id)
                                    sub_dir.mkdir(parents=True, exist_ok=True)
                                    dst = sub_dir / filename
                                    try:
                                        resp = requests.get(url, auth=auth, stream=True)
                                        resp.raise_for_status()
                                        with open(dst, "wb") as f:
                                            for chunk in resp.iter_content(chunk_size=8192):
                                                if chunk:
                                                    f.write(chunk)
                                    except Exception:
                                        # 다운로드 실패 시에도 메타만 먼저 반영
                                        pass
                                    try:
                                        actual_size = dst.stat().st_size
                                    except OSError:
                                        actual_size = size
                                    rel_path = str(dst.relative_to(root))
                                    merged.append(
                                        {
                                            "filename": filename,
                                            "size": actual_size,
                                            "id": att_id,
                                            "content": url,
                                            "local_path": rel_path,
                                        }
                                    )
                                json_text = json.dumps(merged, ensure_ascii=False)
                                from backend.db import update_issue_fields

                                update_issue_fields(
                                    self.conn,
                                    self.current_issue_id,
                                    {"attachments": json_text},
                                )
                                # UI 갱신
                                self.left_panel.issue_tabs._load_attachments_list(json_text)
                        except Exception as e_sync_att:
                            print(f"[WARN] Failed to refresh attachments after upload: {e_sync_att}")

            except Exception as e_all_att:
                print(f"[WARN] Attachment push phase failed: {e_all_att}")

            self.status_bar.showMessage(f"Pushed local changes to JIRA for {jira_key}.")

        except Exception as e:
            self.status_bar.showMessage(f"Push to JIRA failed: {e}")
            print(f"[ERROR] Push to JIRA failed: {e}")
        finally:
            QApplication.restoreOverrideCursor()


class JiraLoginDialog(QDialog):
    """
    프로그램 최초 실행 시 JIRA 접속 정보를 입력받기 위한 간단한 로그인/설정 다이얼로그.
    - jira_config.json 이 존재하지 않을 때만 표시된다.
    - OK 를 누르면 JiraConfig 형식으로 파일을 저장하고 닫힌다.
    """

    def __init__(self, config_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.config_path = config_path

        self.setWindowTitle("JIRA 로그인 / 설정")

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.ed_base_url = QLineEdit()
        self.ed_base_url.setPlaceholderText("https://your-jira-server.example.com")

        self.ed_username = QLineEdit()
        self.ed_username.setPlaceholderText("jira 사용자 ID")

        self.ed_token = QLineEdit()
        self.ed_token.setEchoMode(QLineEdit.Password)
        self.ed_token.setPlaceholderText("비밀번호 또는 Personal Access Token")

        self.ed_project_key = QLineEdit()
        self.ed_project_key.setPlaceholderText("예: PROJ")

        self.ed_project_id = QLineEdit()
        self.ed_project_id.setPlaceholderText("예: 12345 (숫자)")

        form.addRow("Base URL", self.ed_base_url)
        form.addRow("Username", self.ed_username)
        form.addRow("API Token / Password", self.ed_token)
        form.addRow("Project Key", self.ed_project_key)
        form.addRow("Project ID", self.ed_project_id)

        layout.addLayout(form)

        hint = QLabel(
            "※ 이 설정은 jira_config.json 파일로 저장되며,\n"
            "   이후에는 Settings / JIRA Settings... 메뉴에서 다시 변경할 수 있습니다.\n"
            "   (입력을 건너뛰면 오프라인 모드로 시작합니다.)"
        )
        hint.setWordWrap(True)
        layout.addWidget(hint)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self._on_accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _on_accept(self) -> None:
        base_url = self.ed_base_url.text().strip()
        username = self.ed_username.text().strip()
        token = self.ed_token.text().strip()
        project_key = self.ed_project_key.text().strip()
        project_id_text = self.ed_project_id.text().strip()

        if not (base_url and username and token and project_key and project_id_text):
            QMessageBox.warning(
                self,
                "입력 오류",
                "모든 필드를 입력해 주세요.\n"
                "(나중에 Settings 에서도 변경할 수 있습니다.)",
            )
            return

        try:
            project_id = int(project_id_text)
        except ValueError:
            QMessageBox.warning(self, "입력 오류", "Project ID 는 숫자여야 합니다.")
            return

        cfg = JiraConfig(
            base_url=base_url,
            username=username,
            api_token=token,
            project_key=project_key,
            project_id=project_id,
        )
        try:
            save_config_to_file(self.config_path, cfg)
        except Exception as e:  # pragma: no cover - 파일 시스템 오류 방어
            QMessageBox.critical(
                self,
                "저장 오류",
                f"설정 파일을 저장하는 동안 오류가 발생했습니다.\n{e}",
            )
            return

        self.accept()


def run(
    db_path: str = "rtm_local.db",
    config_path: str = "jira_config.json",
    mode: str = "local",
) -> None:
    """
    애플리케이션 엔트리 포인트.

    - 최초 실행 시 jira_config.json 이 없으면 JIRA 로그인/설정 다이얼로그를 먼저 띄운다.
    - 이후에는 기존 동작과 동일하게 MainWindow 를 실행한다.
    """
    import os

    app = QApplication(sys.argv)

    # jira_config.json 이 없으면 JIRA 설정 다이얼로그를 표시 (선택 사항)
    try:
        if not os.path.exists(config_path):
            dlg = JiraLoginDialog(config_path=config_path)
            dlg.exec()
    except Exception:
        # 설정 다이얼로그 단계에서의 오류는 치명적이지 않으므로 무시하고 계속 진행
        pass

    win = MainWindow(db_path=db_path, config_path=config_path, mode=mode)
    win.show()
    sys.exit(app.exec())